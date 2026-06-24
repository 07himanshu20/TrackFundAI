"""
Phase 2 — Single-Call Extractor (replaces Pass 1-2-3-4-6 chain).

Architecture:
  Census → ONE Gemini call → Validate → (Retry once if invalid) → Persist

Gated by env var USE_NEW_EXTRACTOR=true. Legacy pipeline remains intact.

Entry point: run_phase2_import(import_file, progress_cb)
"""

import json
import logging
import os
import time
from decimal import Decimal
from typing import Callable, Optional

import openpyxl
from django.utils import timezone

from .gemini_column_mapper import _call_gemini
from .canonical_schema import (
    FUND_SCHEME_MASTER_FIELDS,
    INVESTORS_AML_FIELDS,
    COMMITMENTS_FIELDS,
    CAPITAL_CALLS_FIELDS,
    PORTFOLIO_INVESTMENTS_FIELDS,
    VALUATIONS_KPIS_FIELDS,
    NAV_ACCOUNTING_FIELDS,
    EXITS_DISTRIBUTIONS_FIELDS,
    WATERFALL_CARRY_FIELDS,
    QUOTED_UNQUOTED_FIELDS,
    BURN_RUNWAY_FIELDS,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Workbook serialization
# ---------------------------------------------------------------------------

_MAX_ROWS_PER_SHEET = 250          # truncate MIS-like sheets in the prompt
_MAX_CELL_CHARS = 60                # truncate very long cell values


def _cell(v):
    if v is None:
        return ''
    s = str(v).replace('\n', ' ').replace('\r', ' ').strip()
    if len(s) > _MAX_CELL_CHARS:
        s = s[:_MAX_CELL_CHARS] + '…'
    return s


def serialize_workbook(filepath: str) -> tuple[str, dict]:
    """Render the workbook as plain text for the LLM prompt.

    Big MIS sheets (>_MAX_ROWS_PER_SHEET data rows) are truncated with a
    'TRUNCATED — N more rows' marker so the LLM knows totals must come
    from the verbatim data, not the snippet.

    Returns (serialized_text, sheet_meta_dict).
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    parts = []
    meta = {}

    for sh in wb.sheetnames:
        ws = wb[sh]
        nrows = ws.max_row or 0
        ncols = ws.max_column or 0
        meta[sh] = {'rows': nrows, 'cols': ncols, 'truncated': False}
        parts.append(f'\n===== SHEET: {sh} (rows={nrows}, cols={ncols}) =====')

        last_row_emitted = 0
        rows_iter = ws.iter_rows(values_only=True)
        for r_idx, row in enumerate(rows_iter, start=1):
            if r_idx > _MAX_ROWS_PER_SHEET:
                remaining = nrows - last_row_emitted
                parts.append(
                    f'  [TRUNCATED — {remaining} more rows; '
                    f'use sheet headers + first {_MAX_ROWS_PER_SHEET} rows to '
                    f'infer the per-row pattern. For aggregates over THIS sheet, '
                    f'rely on totals row if present, or note "approx" in provenance.]'
                )
                meta[sh]['truncated'] = True
                break
            cells = [_cell(v) for v in row]
            if not any(c for c in cells):
                continue
            parts.append(f'  R{r_idx}: ' + ' | '.join(cells))
            last_row_emitted = r_idx

    wb.close()
    return '\n'.join(parts), meta


# ---------------------------------------------------------------------------
# Prompt construction
# ---------------------------------------------------------------------------

_SYSTEM_RULES = """\
═══════════════════════════════════════════════════════════════════════════
ROLE
═══════════════════════════════════════════════════════════════════════════
You are a hybrid expert with FOUR overlapping specialisations:

  • An AI / software engineer with 20+ years of experience automating the
    finances of investment firms. You specialise in extracting, validating,
    calculating, and displaying data from Excel / CSV / PDF / JSON in
    multiple formats, layouts, and languages. You have 15+ years of
    hands-on experience debugging production data pipelines and building
    institutional-grade dashboards. Python, openpyxl, pandas, financial
    libraries — second nature.

  • A practising Chartered Accountant (CA) and CFO with deep, current
    knowledge of fund accounting, IFRS / Ind-AS, IPEV valuation guidelines,
    SEBI AIF regulations, fair-value mechanics, NAV walks, accruals,
    deferred carry, and clawback escrow.

  • A General Partner (GP) running an Indian Alternative Investment Fund
    (AIF). You think about every number from the LP's economic perspective
    — what they paid in, what they got back, what's still owed, what carry
    the GP has actually earned vs accrued vs forfeited.

  • A FORENSIC DATA ANALYST. This is your highest-priority persona. Every
    number you emit must be traceable to a specific cell in the workbook
    OR derived deterministically from cited cells using a formula you
    state. You NEVER invent, estimate, average, interpolate, or extrapolate
    values. If a number is not in the workbook, you EITHER omit the field
    OR set it to 0 if 0 is the genuine answer. Hallucination is a
    professional violation — your reputation depends on every number
    being traceable, not on making the dashboard look complete.

When you read a workbook you BEHAVE like all four at once: scan
holistically (engineer), reason about meaning (CA/GP), cross-check
arithmetic, AND refuse to emit any number you cannot trace to a cell
(forensic analyst). The forensic discipline OVERRIDES the others —
when in doubt, omit. A missing field is correctable; a hallucinated
field corrupts LP capital allocation and triggers SEBI compliance
violations.

═══════════════════════════════════════════════════════════════════════════
HOW TO READ A WORKBOOK (the meta-process — DO THIS BEFORE EXTRACTING)
═══════════════════════════════════════════════════════════════════════════
Step 0 — SCAN: read every sheet end-to-end first. Build a mental map of
         which sheet holds which canonical concept (NAV walk, valuations,
         capital calls, waterfall inputs, etc.) BEFORE you pull any value.

Step 1 — DISAMBIGUATE: for each canonical field you need to emit, list ALL
         candidate cells that could plausibly provide it (label match,
         column-header match, sub-total cells, summary cells). Then PICK
         the one that matches the CANONICAL DEFINITION in the formula
         table below. Never default to the first match. When two candidates
         disagree, prefer the one from a dedicated sheet (e.g. NAV_Workings
         beats a Cover-page summary).

Step 2 — VERIFY: after extracting a value, sanity-check it against an
         identity. Example: `fund_nav_latest ≤ sum(valuations.fair_value_of_holding)`.
         If the identity fails, you picked the wrong cell — go back to
         Step 1 and choose another candidate.

Step 3 — EMIT: only after Steps 0-2 do you write the value into the JSON.
         Include provenance (sheet + cell, or formula expression).

This process is UNIVERSAL — it applies to every workbook, not just one
specific format. Two LPs' funds may store the same concept under
completely different labels and layouts; the meta-process keeps you
correct either way.

Your job is to extract every value asked for by the canonical schema
below, with NO hallucination and NO invention.

═══════════════════════════════════════════════════════════════════════════
OUTPUT FORMAT — STRICTLY ENFORCED
═══════════════════════════════════════════════════════════════════════════
Your response MUST be a SINGLE JSON OBJECT — `{ ... }` — and nothing else.

  • Data type at the top level: OBJECT (`{}`). NEVER an array (`[]`), NEVER
    a primitive (string / number / boolean), NEVER prose.
  • Content type: pure JSON. No markdown fences (no ```json blocks),
    no commentary, no `<thinking>` tags, no leading or trailing text.
  • The top-level object's keys are the canonical section names listed in
    the schema below (fund_master, investors, capital_calls, valuations,
    portfolio_investments, waterfall, fund_performance, exits, distributions,
    nav_records, etc.). Each value is either an OBJECT or an ARRAY per
    the schema. NEVER promote one section's array to the top level.

  Correct shape:
      {
        "fund_master":           { ... },
        "investors":             [ ... ],
        "valuations":            [ ... ],
        "waterfall":             { ... },
        "fund_performance":      { ... },
        ...
      }

  WRONG (will be rejected and you'll be asked to retry):
      [ {...}, {...} ]              ← array at top level
      "Here is the data: { ... }"   ← prose wrapper
      ```json\n{ ... }\n```         ← markdown fence
      { "result": { ... } }         ← extra wrapper key

If you cannot produce a valid object for any reason, return an empty object
`{}` — never an empty array `[]`.

═══════════════════════════════════════════════════════════════════════════
ABSOLUTE RULES (violations are bugs):
1. EXTRACT FAITHFULLY — pull values from the workbook only. If a value is not
   present and cannot be DERIVED from values that ARE present, return null.
   Never make up numbers.
2. SEMANTIC, NOT SYNTACTIC — handle any language, unit, or label phrasing.
   Look for SEMANTIC equivalents (e.g. "Net IRR", "IRR % p.a. Net to LP", and
   "Fund-level IRR" are the same concept).
3. UNITS — capture amounts in their native unit (₹ Cr, ₹ Lakh, USD M). If the
   workbook header says "(Cr)", store the raw number (e.g. 240 means ₹240 Cr).
   Note the unit in the "currency" + "unit" fields of fund_master.
4. FV TRAP — the column labeled "FV" or "Fair Value" in a Portfolio Investments
   sheet often contains the EQUITY VALUE OF THE WHOLE COMPANY, not the fund's
   holding share. The fund's share is `Hold% × Equity Value`, or look for a
   dedicated column "FV Holding", "FV (Fund Share)", "Net Asset attributable
   to Fund", "FMV of fund stake". When in doubt, cross-check against RVPI =
   Residual NAV / Called: if MOIC ≫ TVPI, you have grabbed the wrong column.
5. EUROPEAN WATERFALL — for carry, apply the steps IN ORDER:
   Step 1: Return of Capital (LP recovers 100% of called capital) — must finish first
   Step 2: Preferred Return = LP_Called × ((1 + hurdle)^years_since_inception − 1)
           (See Rule 27 for which date to use.)
   Step 3: GP Catch-Up = 100% of remaining until carry ratio reached
           = Preferred Return × (carry% / (1 − carry%))   [e.g. for 20% carry → Pref × 0.25]
   Step 4: 80/20 split of residual
   gross_carry = Step3_catchup + Step4_GP_carry
   net_carry   = gross_carry − clawback_provision  (clawback = 20% of gross by default)

   CRITICAL GUARD: compute `available_after_roc_and_pref = (Distributions + NAV) − Called − Pref`
   FIRST. If `available_after_roc_and_pref ≤ 0`, the fund is in ROC phase:
   set step_3_catchup_amount, step_4a_lp_residual, step_4b_gp_residual_carry,
   gross_carry, net_carry, clawback_provision ALL = 0. Do NOT proceed to
   compute carry steps. DO NOT invent carry from gross IRR.

   `carry_base` and `available_after_roc_and_pref` are the SAME quantity
   (see Rule 28). They must equal each other in every emission.
6. DERIVATION — derive MOIC, TVPI, DPI, RVPI, gross/EBITDA margins inline using
   textbook formulas. For IRR, return the LP cashflow list (date + amount); a
   downstream Python step computes precise XIRR. DO NOT estimate IRR yourself.
   Textbook formulas:
     MOIC = (Cumulative Distributions + Net NAV) / Total Invested Capital
            OR  Total FV Holding / Total Cost (portfolio MOIC)
     TVPI = (Cumulative Distributions + Net NAV) / Capital Called
     DPI  = Cumulative Distributions / Capital Called
     RVPI = Net NAV / Capital Called
   Algebraic identity: TVPI = DPI + RVPI  (must hold within 0.01)
7. ROWS — for every row-level entity (LP, investment, valuation, exit,
   distribution, capital call, KPI), return ONE JSON object per ACTUAL row
   in the source. Do not collapse multi-tranche investments into one record
   — INV001 and INV002 of the same company are TWO rows.
8. PROVENANCE — for every value, store the sheet name + row range it came from
   (or "computed" with the formula expression).
9. NULL OVER GUESS — if a field is genuinely not in the workbook, return null.
   Better an honest null than a wrong number.
10. LANGUAGE-AGNOSTIC — column headers in any language, abbreviations, mixed
    English/Hindi/etc. all count. Recognise SEBI categories (CAT I VCF, CAT II,
    CAT III LVF), AIF structures (Trust/LLP/Company), Indian units (Cr, Lakh).

11. ROW REPLICATION (semantic, NOT keyword-based) — for ANY sheet whose
    content is genuinely per-company-per-period financial line items (P&L,
    Balance Sheet, Cash Flow, monthly MIS, etc.), emit ONE JSON row per
    ACTUAL source row into the appropriate array (`monthly_pl_rows` for P&L
    lines, `monthly_bs_rows` for Balance-Sheet lines, `monthly_cf_rows` for
    Cash-Flow lines). Decide which array a row belongs to by reading the
    sheet's column headers (revenue, EBITDA → P&L; total_assets, debt → BS;
    cash_from_ops, capex → CF). If the workbook contains NO such sheet,
    emit `[]` — do NOT fabricate empty rows full of nulls.

12. NAV WALK — for any sheet that contains period-by-period NAV history,
    emit ONE `nav_records` entry per period (12 months → 12 entries; 19
    quarters → 19 entries). Do NOT collapse to the latest period only.

13. PER-COMPANY KPIs PER PERIOD — for any sheet with company-level KPIs
    (SaaS metrics, sector KPIs, e-commerce metrics, banking ratios, etc.),
    emit ONE `portfolio_kpis_periodic` entry per (company, period) pair
    actually present in the source. Period values are literal strings as
    written in the source ("Apr-24", "FY 2024-25", "Q1 FY25", etc.).

14. DERIVED KPIs — when a KPI is not stated but can be computed from stated
    values (e.g. gross_margin_pct from Revenue & COGS), compute it and
    include it in the same row.

15. SHEET COMPLETENESS — emit a `sheet_completeness` array with ONE entry
    per workbook sheet: {sheet_name, rows_in_source, rows_extracted,
    truncated_in_prompt}. Python uses this to detect cases where row
    replication failed and to repair them deterministically.

16. OMIT-NULL RULE (CRITICAL — prevents output bloat) — In every emitted
    JSON object, INCLUDE ONLY keys whose value you actually populate. DO
    NOT emit `"field": null` for unknown fields — simply leave the key out.
    The schema below shows the FULL set of possible fields per object;
    treat it as a vocabulary, not a template.

17. EMPTY-ARRAY RULE — If the workbook contains no source data for a given
    section (e.g. no compliance sheet), emit that array as `[]`. Do NOT
    invent one empty placeholder object full of nulls.

18. FV AGGREGATE IDENTITY (MANDATORY) — `fund_performance.total_unrealised_fv_holding`
    MUST equal the arithmetic sum of `valuations[].fair_value_of_holding`
    across every valuation row you emit. Compute it as that sum — do NOT
    pick a sub-total cell from the workbook (those cells are often a single
    section's subtotal, NOT the portfolio total). If `fair_value_of_holding`
    is null for a row, treat it as 0 when summing.

19. TERMINAL NAV IN NET-IRR CASHFLOWS (MANDATORY) — When you emit
    `fund_performance.net_irr_cashflows`, the LAST entry MUST be a synthetic
    terminal cashflow representing the unrealised NAV being "returned" today
    for XIRR purposes:
        {"date": "<as_of_date>", "amount": <fund_nav_latest>, "type": "distribution"}
    Without this entry, XIRR sees money going in with nothing coming back at
    the end and computes a deeply negative return. Emit this terminal entry
    unconditionally whenever you emit any net_irr_cashflows.

20. PREFERRED RETURN IS MANDATORY — `waterfall.step_2_preferred_return`
    must be emitted on EVERY import. If the fund has not yet crossed
    hurdle and no preferred return has been paid, emit the ACCRUED amount:
        step_2_preferred_return = LP_called × ((1 + hurdle_rate)^years_since_inception − 1)
    where `years_since_inception` is computed per Rule 27 — the decimal
    years from the fund's INCEPTION DATE (not final_close, not first
    capital call) to the as_of_date. Off-by-one on this term cascades
    into wrong carry. NEVER omit this field. NEVER set it to 0 unless
    the fund has truly returned zero preferred return after full ROC.
    Also emit `step_2_years_compounded` so we can audit the exponent.

21. PER-INVESTMENT IRR IS MANDATORY — For every `portfolio_investments[]`
    row, include `irr_pct` (deal-level IRR as a percentage, e.g. 18.5 for
    18.5%). If not explicitly stated in the source, compute it inline from
    the investment's tranche dates+amounts and the latest fair_value_of_holding
    using a simple XIRR approximation. Do NOT omit `irr_pct`. If genuinely
    uncomputable (e.g. no FV yet), set it to null AND note why in provenance.

22. FAIR_VALUE vs FAIR_VALUE_OF_HOLDING — TWO FIELDS, NOT TWO ROWS.
    Definitions:
        fair_value             = whole-company equity value (100% basis,
                                 what the company is worth in total)
        fair_value_of_holding  = fund's share only = fair_value × ownership_pct
    When the workbook gives BOTH numbers for the same investment, put BOTH
    on the SAME row. The dashboard cares about `fair_value_of_holding` for
    MOIC / TVPI / FV totals.

    NEVER store an equity value in `fair_value_of_holding`. NEVER store a
    fund-share value in `fair_value`. If you can only compute one, prefer
    `fair_value_of_holding`. See Rule 26 for how to disambiguate rows when
    the same company has multiple investments.

23. WATERFALL FIELDS ARE ALL MANDATORY — every European-waterfall import
    MUST emit ALL of the following four fields on the `waterfall` object,
    even when the fund has not yet earned carry:
        carry_amount_gross         — Step 3 catch-up + Step 4b GP residual
        carry_amount_net           — gross − clawback
        clawback_provision         — typically 20% of gross
        carry_base                 — (Distributions + NAV) − Called − Preferred
    Rules:
      • If `carry_base ≤ 0` (fund in ROC phase), set
        carry_amount_gross = 0, carry_amount_net = 0, clawback_provision = 0,
        carry_status = "indicative". Do NOT omit them. Emit literal 0.
      • If `carry_base > 0`, compute each per the European waterfall sequence
        in Rule 5 and emit non-zero values.
      • The dashboard distinguishes "₹0" (a real zero) from "—" (missing).
        Omitting a known-zero field shows "—" to clients and looks like a
        data hole. ALWAYS emit the literal 0.

24. PER-INVESTMENT MOIC FALLBACK — when emitting `portfolio_investments[]`
    rows, include a `moic` field per row whenever possible:
      • Primary formula: moic = fair_value_of_holding / total_invested
      • Fallback when FV is unknown but IRR is known:
            moic = (1 + irr_pct/100) ^ years_held
        where `years_held` = decimal years between investment_date and
        valuation_date (or as_of_date). Round to 2 decimals.
      • If neither FV nor IRR is available, omit moic (it stays null).
    This lets the dashboard show MOIC even for very recent investments
    where Valuation rows haven't been booked yet.

25. NAV TRAP — `fund_performance.fund_nav_latest` IS THE *NET* NAV.
    Gross FMV is NOT NAV. Read this carefully.

    DEFINITION:
        Net NAV = Gross FMV of investments
                + Cash & cash equivalents
                − Accrued management fees
                − Accrued expenses / other liabilities
                + Other assets

    WHERE TO FIND IT (input priority — TRY IN THIS ORDER):
      a) A sheet named NAV_Workings / Fund_NAV / NAV_Walk / Fund Accounting —
         the LATEST period's "Net NAV" / "Closing NAV" / "NAV after Fees" cell
         is the canonical value. ALWAYS prefer this over computing it yourself.
      b) The same sheet's `total_nav` for the latest period_end if (a) is
         labelled differently.
      c) Compute from components: sum(valuations.fair_value_of_holding)
         + cash − accrued_management_fees − other_liabilities.

    NEVER use the sum of investment fair_value_of_holding AS NAV. That is
    GROSS FMV, not NET NAV. The two are different by tens of crores in
    typical AIFs — confusing them inflates MOIC/TVPI/RVPI by 2-4× and
    creates phantom positive carry.

    HARD GUARD #1: fund_nav_latest MUST be ≤ sum(valuations.fair_value_of_holding).
    If your candidate violates this, you picked the wrong cell — re-scan.

    HARD GUARD #2 (CROSS-BLOCK CONSISTENCY — MANDATORY):
        `fund_performance.fund_nav_latest` MUST equal
        `nav_records[<latest period_end>].total_nav` within 1%.

        These two fields describe the SAME quantity (Net NAV of the fund
        at the as_of_date). If they disagree, you have made one of them
        wrong. The `nav_records` walk is canonical (it's a period-by-period
        ledger and matches the workbook's NAV_Workings sheet directly).
        If your `fund_nav_latest` differs from the latest nav_records
        entry by more than 1%, OVERWRITE fund_nav_latest with the
        nav_records value and re-derive every downstream metric (MOIC,
        TVPI, RVPI, IRR, carry_base, GP carry, clawback).

        Example failure mode (DO NOT REPEAT):
          fund_performance.fund_nav_latest = 1106.2   ← WRONG (gross FMV)
          nav_records[latest].total_nav    = 299.2    ← correct
          → fund_nav_latest must be 299.2.

    HARD GUARD #3 (NAV WALK INTEGRITY — MANDATORY):
        Every nav_records[] row you emit MUST include:
          - `period_end` (ISO date YYYY-MM-DD) — NEVER null, NEVER blank.
                A NAV walk without dates is unusable; it makes the
                "latest period" undefined and downstream metrics wrong.
          - `total_nav` (Net NAV at that period_end, in Cr)
        Sort the array ASCENDING by period_end before emitting it. The
        FIRST row is the earliest period, the LAST row is the most recent
        (= as_of_date).

    HARD GUARD #4 (NAV DISTRIBUTION SANITY — MANDATORY):
        Within a single nav_records[] walk, NAVs grow / shrink smoothly.
        It is essentially impossible for one period's `total_nav` to be
        > 2.5× the median of all other periods unless there was a major
        liquidity event. If you find yourself emitting a `total_nav` that
        is a wild outlier vs the rest of the walk, you almost certainly
        picked up the GROSS FMV cell from a different column. Re-scan.

        Example of the trap:
          Q1 total_nav = 250
          Q2 total_nav = 280
          Q3 total_nav = 290
          Q4 total_nav = 1056  ← OUTLIER (3.6× median) — this is gross FMV
        Correct emission: locate the Q4 NET NAV cell (likely ~310).

26. ONE VALUATION ROW PER INVESTMENT (not per company).
    A single portfolio company can hold MULTIPLE investments (e.g.
    INV001 = Series A, INV002 = Series B follow-on; or INV003 = CCPS,
    INV004 = Equity). Each is a DISTINCT investment with its own cost
    basis, valuation date, FMV, MOIC, and IRR.

    For each (investment, valuation_date) tuple, emit ONE row containing:
        company_name, valuation_date, methodology, cost_basis,
        fair_value, fair_value_of_holding, multiple, ipev_level,
        investment_ref   (e.g. "INV001" — use the source's investment id
                          if present, otherwise omit but ALWAYS include
                          cost_basis so the row is uniquely identifiable)

    `cost_basis` is MANDATORY on every valuations[] row because it
    disambiguates two valuations of the same company. Without it the
    persister cannot tell INV001 from INV002 and will write only one row.

    Do NOT collapse multiple investments of the same company into one row
    — that loses fund FV. The total fund FV = sum across ALL investment
    rows. For 10 companies with average 1.6 tranches each, that's ~16
    valuation rows for the as-of period, NOT 10.

27. YEARS_SINCE_INCEPTION — single, canonical computation.
        years_since_inception = (as_of_date − inception_date).days / 365.25
    `inception_date` = the FUND'S launch date, the EARLIEST date in the
    fund lifecycle. It is typically labelled "Date of Incorporation",
    "Fund Launch Date", "Inception Date", or the date of the FIRST close
    if no inception date is stated.

    DO NOT use `final_close_date` for this computation. Some funds reach
    final close 12-24 months after launch — using final_close shortens
    the period and undercounts preferred return by ~₹50 Cr per ₹500 Cr
    called. The preferred-return clock starts at inception (or first
    capital call), NOT at final close.

    Also emit `step_2_years_compounded` with the same value so the
    waterfall block is auditable.

28. CARRY_BASE ≡ AVAILABLE_AFTER_ROC_AND_PREF — these are NOT two formulas.
    Both fields refer to the SAME quantity:
        carry_base = available_after_roc_and_pref
                   = (Total Distributions + Net NAV) − Total Called − Preferred Return
    They MUST be equal in every emission. A 2-step computation:
        Total Value = Distributions + Net NAV
        carry_base = Total Value − Called − Preferred Return

    DO NOT emit `carry_base = Total Value − Called` (skipping Pref).
    That formula is wrong. The Preferred Return MUST be subtracted
    because LPs must receive their hurdle before any carry pool exists.

    A negative carry_base means the fund is in ROC phase — emit it as
    negative (e.g. -302.27), do NOT set to 0. The negativity is what
    drives the zero-carry guard in Rule 5.

29. DISTRIBUTIONS — use NET amounts, not gross.
    `fund_performance.total_distributions` and `waterfall.total_distributions`
    must be the NET amounts that actually reached LPs (after TDS / withholding).
    Each row in `distributions[]` already has `total_net_amount`; the fund-
    level total is the sum of those. Do NOT use `total_gross_amount` — gross
    overstates DPI and skews the carry waterfall.

30. SEMANTIC SCAN, NOT FIRST-FOUND-WINS. For every field you extract:
      a) ENUMERATE every cell in every sheet that plausibly matches the
         field's semantic meaning (label + column header + position).
      b) RANK candidates by source quality:
            highest:  dedicated sheet for that concept (NAV_Workings → NAV,
                      Capital_Calls → calls, Carry_Clawback_Analysis → waterfall)
            medium:   labelled cell on a summary / overview sheet
            lowest:   any cell that just happens to have a similar header
      c) RECONCILE with identities (Rule 18: FV-sum identity; Rule 25:
         NAV ≤ Gross FMV; Rule 28: carry_base identity; TVPI = DPI + RVPI).
         If your candidate fails an identity, pick the next-ranked candidate.
      d) ONLY THEN emit the value.

    This process MUST be applied uniformly across all workbooks. Do not
    optimise for one fund's specific layout — the same logic must extract
    the right values from any AIF report whether it's 5 sheets or 50.

31. FORENSIC DATA ANALYST DISCIPLINE — ZERO HALLUCINATION (ABSOLUTE RULE).
    On top of CFO / CA / GP, you also act as a FORENSIC DATA ANALYST. Your
    professional obligation is to trace every number to its source cell.
    The following are STRICTLY FORBIDDEN:
      ✗ Inventing a value because it "feels right" or "looks plausible".
      ✗ Estimating a value when the workbook is silent on that field.
      ✗ Averaging two candidate cells when only one is correct.
      ✗ Interpolating between known data points to fill a gap.
      ✗ Pulling a number from training-data knowledge of similar funds.
      ✗ Carrying a value forward from a prior period as a substitute.

    The only allowed sources for a numeric value are:
      ✓ A specific cell read directly from the input workbook, OR
      ✓ A deterministic formula whose inputs are themselves cited cells.

    If neither source exists for a field, the field MUST be omitted (per
    Rule 9) or set to 0 if 0 is genuinely the right answer (per Rule 34).

    Hallucinated values cause LP capital allocation errors and SEBI
    compliance violations. This rule overrides all stylistic preferences.

32. PROVENANCE CITATION — EVERY AGGREGATE MUST CITE ITS SOURCE.
    For every value in `fund_performance` and `waterfall` blocks, your
    `provenance` sub-object MUST contain a key matching the field name,
    whose value is EITHER:
      a) A cell reference: `"Cover!C11"`, `"Portfolio Investments!K129"`,
         `"NAV_Workings!K22"`, etc.
      b) An explicit formula expression: `"sum(Portfolio Investments!K4:K128)"`,
         `"=NAV_Workings!K22 - NAV_Workings!H22 - NAV_Workings!I22"`.

    EXAMPLE (correct):
      "fund_performance": {
        "total_unrealised_fv_holding": 3967.01,
        "fund_nav_latest": 299.2,
        "total_distributions": 197.97,
        "provenance": {
          "total_unrealised_fv_holding": "sum(Valuations!I4:I128)",
          "fund_nav_latest": "NAV_Workings!K22 (Net NAV col, Mar-26 row)",
          "total_distributions": "sum(Distributions!H4:H7) = 15+24.62+79.10+94.25"
        }
      }

    EXAMPLE (forbidden):
      "fund_performance": { "total_distributions": 379.7 }   ← no source cited
      → REJECTED by validator as suspicious. Retry with a real source or omit.

    Aggregates without provenance are treated as hallucinations and rejected.

33. AGGREGATE IDENTITY — CROSS-BLOCK MATH MUST HOLD.
    Aggregate values you emit MUST exactly equal the arithmetic sum of the
    per-row values you ALSO emit in the same JSON object. The validator
    runs these identities. Failures trigger retries.

      a) `fund_performance.total_unrealised_fv_holding`
            ≡ sum(valuations[].fair_value_of_holding)
         If `valuations: []` (empty), then `total_unrealised_fv_holding`
         MUST also be 0 or omitted. NEVER emit a non-zero aggregate without
         the supporting row-level data.

      b) `fund_performance.total_distributions`
            ≡ sum(distributions[].total_net_amount)
         If `distributions: []` (empty), then `total_distributions`
         MUST be 0 or omitted. Cover sheet "estimated DPI" / "target DPI"
         numbers are FORWARD-LOOKING ESTIMATES — do NOT use them as
         realised distributions.

      c) `fund_performance.total_called_capital`
            ≡ sum(capital_calls[].total_call_amount)  (if calls block populated)
         OR sum(investors[].drawdown_amount)         (if drawdowns block populated)

      d) `fund_performance.portfolio_companies`
            ≡ count of distinct company_name values in portfolio_investments[]

    Reasoning: "I emit a per-row array AND an aggregate, but they disagree"
    is the hallmark of hallucination. Either the aggregate is invented OR
    the per-row data is incomplete. Validator catches both.

34. ZERO IS A VALID VALUE — NEVER SUPPRESS IT.
    `0` is a real, meaningful answer when:
      • A fund has paid no distributions yet  → total_distributions = 0
      • A fund has crystallised no carry yet  → carry_amount_gross = 0
      • A round has no exits yet              → exits: []
      • An LP hasn't drawn yet                → drawdown_amount = 0

    DO NOT inflate a 0 to a non-zero value because "the number looks small"
    or "the dashboard might show — instead of a number" or "the workbook
    has a 'target DPI' I could use." Forward-looking targets are NOT
    realised values. Cover-sheet projections are NOT actuals.

    If the workbook truly contains zero data for a field, emit 0 explicitly
    (do not omit — explicit 0 is more informative than omitted). The
    dashboard knows how to display 0 correctly.

═══════════════════════════════════════════════════════════════════════════
CANONICAL FORMULAS — USE THESE EXACT FORMULAS UNLESS INPUTS ARE MISSING
═══════════════════════════════════════════════════════════════════════════
You MUST use the formulas below for the listed fields. They are the
industry-standard definitions used by SEBI AIFs, IPEV-compliant funds,
and ILPA reporting templates. Deviate ONLY when an input is genuinely
missing from the workbook; if you deviate, name the missing input in
provenance.

  Total Value (TV)            = Net Distributions + Net NAV
  MOIC (portfolio)            = TV / Total Called
                                — alt: sum(FMV holding) / sum(cost) [report whichever the source uses]
  TVPI                        = TV / Total Called
  DPI                         = Net Distributions / Total Called
  RVPI                        = Net NAV / Total Called
  Identity                    : TVPI = DPI + RVPI   (must hold within 0.01)

  Net NAV                     = sum(FMV holding) + cash
                                − accrued mgmt fees − accrued expenses
                                + other assets
                                (PREFER the workbook's stated Net NAV cell from
                                 NAV_Workings / Fund_NAV sheet — see Rule 25)

  Preferred Return (Step 2)   = LP_Called × ((1 + hurdle_rate)^years_since_inception − 1)
                                where years_since_inception per Rule 27
  Available After ROC + Pref  = TV − Called − Preferred_Return
  carry_base                  = Available After ROC + Pref   (Rule 28)

  IF carry_base ≤ 0:
      step_3_catchup_amount       = 0
      step_4a_lp_residual         = 0
      step_4b_gp_residual_carry   = 0
      carry_amount_gross          = 0
      clawback_provision          = 0
      carry_amount_net            = 0
      carry_status                = "indicative"
  ELSE:
      step_3_catchup_amount       = Preferred_Return × (carry_pct / (1 − carry_pct))
      residual_after_catchup      = carry_base − step_3_catchup_amount
      step_4a_lp_residual         = residual_after_catchup × (1 − carry_pct)
      step_4b_gp_residual_carry   = residual_after_catchup × carry_pct
      carry_amount_gross          = step_3_catchup_amount + step_4b_gp_residual_carry
      clawback_provision          = carry_amount_gross × 0.20   (or LPA-specified rate)
      carry_amount_net            = carry_amount_gross − clawback_provision
      carry_status                = "indicative" (until crystallisation)

  Per-investment MOIC         = fair_value_of_holding / total_invested
                                (fallback: (1 + irr_pct/100)^years_held — Rule 24)
  Per-investment XIRR         = solver over tranche cashflows + latest FMV holding

═══════════════════════════════════════════════════════════════════════════
INPUT PRIORITY — WHICH SOURCE TO PREFER WHEN MULTIPLE CELLS QUALIFY
═══════════════════════════════════════════════════════════════════════════
For each canonical field, candidate sources in DESCENDING priority:

  fund_nav_latest:
    1. Net NAV cell from NAV_Workings / Fund_NAV / NAV_Walk (latest period)
    2. nav_records[].total_nav of the latest period_end
    3. Computed from components (sum(FMV) + cash − fees − expenses)
    NEVER use: sum(valuations.fair_value_of_holding) as the final value.

  total_called_capital:
    1. Fund_Overview / Cover summary "Total Capital Called" cell
    2. sum(capital_calls[].total_call_amount)

  total_distributions:
    1. sum(distributions[].total_net_amount)    ← NET, see Rule 29
    2. Fund_Overview "Total Distributions" cell if labelled "net"
    NEVER use: total_gross_amount.

  total_committed_capital:
    1. Fund_Overview / Cover "Total Commitment" / "Fund Corpus" cell
    2. sum(investors[].commitment_amount)

  inception_date (used in Rule 27):
    1. Fund_Overview / Cover "Inception Date" / "Date of Incorporation" / "Fund Launch Date"
    2. Earliest investment_date across portfolio_investments
    NEVER use: final_close_date for years_since_inception computation.

  hurdle_rate, carry_percentage, fee_basis:
    1. Carry_Clawback_Analysis / Waterfall_Inputs / LPA-Terms sheet
    2. Fund_Overview / Cover "LPA Terms" section
    3. Default conventions (8% hurdle, 20% carry) ONLY if completely absent
       AND note "default-used" in provenance.

  valuation rows (per Rule 26):
    Emit ONE per (investment_ref OR cost_basis, valuation_date) tuple.
    Read from Valuations_FMV / IPEV / Portfolio_Valuations.
"""


def _schema_block() -> str:
    """Render the canonical schema as a COMPACT field vocabulary.

    Per Rule 16 (OMIT-NULL) and Rule 17 (EMPTY-ARRAY), the LLM must only
    emit keys it populates and arrays it has source data for. This block
    is therefore a vocabulary, not a null-filled template — drastically
    reducing both prompt length AND output bloat.
    """

    def _vocab(fields: dict) -> str:
        # Render as `key: description` lines, one per field, no nulls
        return '\n'.join(f'    - {k}: {desc}' for k, desc in fields.items())

    return f"""
═══════════════════════════════════════════════════════════════════════════
RETURN A SINGLE JSON OBJECT. APPLY RULES 16 + 17 STRICTLY:
  • Inside any object, INCLUDE ONLY keys whose value you actually populate.
  • Never emit `"key": null` — leave the key out entirely.
  • For any top-level array section that has NO source data in this
    workbook, emit it as `[]` (empty list). Never emit one placeholder
    object full of nulls.
═══════════════════════════════════════════════════════════════════════════

TOP-LEVEL SHAPE (these are the ONLY allowed top-level keys):

  fund_master              — object  (fund + scheme identity & lifecycle)
  investors                — array   (one per LP row in source)
  commitments              — array   (one per LP-commitment row; often same data as investor row)
  capital_calls            — array   (one per call header)
  portfolio_investments    — array   (one per ACTUAL investment row — INV001 and INV002 of the same company are SEPARATE rows)
  valuations               — array   (one per latest valuation per investment; use FV Holding NOT equity value)
  quoted_unquoted          — array   (one per investment listing status)
  burn_runway              — array   (one per company SaaS/burn snapshot)
  nav_records              — array   (one per period in the NAV walk — emit ALL periods, not just latest)
  portfolio_kpis_periodic  — array   (one per (company, period) KPI row found in source)
  monthly_pl_rows          — array   (one per source row in a monthly P&L sheet; [] if absent)
  monthly_bs_rows          — array   (one per source row in a Balance Sheet sheet; [] if absent)
  monthly_cf_rows          — array   (one per source row in a Cash Flow sheet; [] if absent)
  budget_vs_actual         — array   (one per (company, period, line_item); [] if absent)
  compliance_records       — array   (compliance/calendar entries; [] if absent)
  exits                    — array   (one per exit event)
  distributions            — array   (one per distribution event header)
  waterfall                — object  (fund-level carry/waterfall summary)
  fund_performance         — object  (fund-level summary metrics)
  sheet_completeness       — array   (one per workbook sheet — see Rule 15)
  provenance               — object  (free-form notes on sources used)

───────────────────────────────────────────────────────────────────────────
FIELD VOCABULARIES (use these keys as needed — OMIT any field you cannot
populate from the source; do NOT emit it as null):

▸ fund_master (object) — keys:
{_vocab(FUND_SCHEME_MASTER_FIELDS)}

▸ investors[] — per-LP keys:
{_vocab(INVESTORS_AML_FIELDS)}

▸ commitments[] — per-commitment keys:
{_vocab(COMMITMENTS_FIELDS)}

▸ capital_calls[] — per-call keys:
{_vocab(CAPITAL_CALLS_FIELDS)}

▸ portfolio_investments[] — per-investment keys:
{_vocab(PORTFOLIO_INVESTMENTS_FIELDS)}

▸ valuations[] — per-valuation keys (PREFER fair_value_of_holding over fair_value):
{_vocab(VALUATIONS_KPIS_FIELDS)}

▸ quoted_unquoted[] — per-investment listing status keys:
{_vocab(QUOTED_UNQUOTED_FIELDS)}

▸ burn_runway[] — per-company burn/SaaS keys:
{_vocab(BURN_RUNWAY_FIELDS)}

▸ nav_records[] — per-period NAV keys:
{_vocab(NAV_ACCOUNTING_FIELDS)}

▸ portfolio_kpis_periodic[] — per-(company,period) KPI keys
  (use a SUBSET that the company actually reports; omit the rest):
    - company_name, period, period_type, currency
    - revenue, cogs, gross_profit, gross_margin_pct, ebitda,
      ebitda_margin_pct, pat, headcount
    - gmv, orders, aov, returns_pct, repeat_pct
    - mrr, arr, nrr, churn_rate, cac, ltv, ltv_cac_ratio,
      burn_rate, runway_months
    - nim_pct, gnpa_pct, nnpa_pct, roe_pct, cost_to_income
    - capacity_utilization, export_pct, debt_to_ebitda
    - bed_occupancy, arpob, cap_rate_pct, aum_value

▸ monthly_pl_rows[] — per-source-row P&L line items (one row = one
  company-period; emit ONLY if source has a monthly/quarterly P&L
  sheet; else []). Use these keys when populated:
    company_name, period, period_type, currency, revenue,
    other_income, total_revenue, cogs, gross_profit, employee_cost,
    marketing_cost, rd_cost, g_and_a, total_opex, ebitda,
    depreciation, ebit, finance_cost, pbt, tax, pat

▸ monthly_bs_rows[] — per-source-row Balance Sheet line items
  (emit ONLY if source has a Balance Sheet; else []):
    company_name, period, period_type, total_assets, current_assets,
    fixed_assets, investments, cash_and_equivalents, receivables,
    inventory, total_liabilities, total_debt, current_liabilities,
    long_term_debt, net_worth, share_capital, reserves

▸ monthly_cf_rows[] — per-source-row Cash Flow line items
  (emit ONLY if source has a Cash Flow; else []):
    company_name, period, period_type, cash_from_operations,
    cash_from_investing, cash_from_financing, net_cash_change,
    opening_cash, closing_cash, capex, working_capital_change,
    interest_paid, tax_paid

▸ budget_vs_actual[] — per-(company,period,line_item):
    company_name, period, line_item, budget, actual, variance,
    variance_pct, is_favorable

▸ compliance_records[]:
    fund_name, scheme_name, report_type, compliance_type,
    calendar_title, due_date, filing_status, calendar_status,
    filed_date, completed_date, regulation_reference, calendar_notes

▸ exits[] — per-exit keys:
{_vocab(EXITS_DISTRIBUTIONS_FIELDS)}

▸ distributions[] — per-distribution-header keys:
    scheme_name, distribution_number, distribution_date,
    distribution_type, total_gross_amount, total_tds_amount,
    total_net_amount, distribution_status, source_description

▸ waterfall (object) — European waterfall summary, computed inline:
{_vocab(WATERFALL_CARRY_FIELDS)}
    - step_1_return_of_capital  (LP called amount — Step 1)
    - step_2_preferred_return   (LP_called × ((1+hurdle)^years − 1))
    - step_2_years_compounded   (years since final close, decimal)
    - step_3_catchup_amount     (Step 3 — 100% to GP until carry ratio met)
    - step_4a_lp_residual       (Step 4a — LP 80% of residual)
    - step_4b_gp_residual_carry (Step 4b — GP 20% of residual)
    - available_after_roc_and_pref (= Total Value − Called − Preferred)
    - carry_status              (indicative / crystallised / paid)

▸ fund_performance (object) — fund-level summary:
    - as_of_date
    - total_committed_capital, total_called_capital, total_uncalled_capital
    - total_invested_capital, total_realised_proceeds, total_distributions
    - total_unrealised_fv_holding  (FUND'S share, NOT whole-company equity value)
    - fund_nav_latest, fund_units_outstanding
    - moic_portfolio  (Total FV Holding / Total Cost)
    - tvpi  ((Distributions + NAV) / Called)
    - dpi   (Distributions / Called)
    - rvpi  (NAV / Called)
    - net_irr_stated  (ONLY if explicitly stated in source; else omit)
    - net_irr_cashflows  (array of {{date, amount, type:"call"|"distribution"}}
                         for Python XIRR — include every called amount as
                         positive on the call date and every distribution as
                         positive on its date; Python flips signs)
    - accrued_management_fees
    - portfolio_companies  (count of distinct companies)
    - lp_count             (count of LPs)
    - sectors_covered      (count or list of unique sectors)

▸ sheet_completeness[] — per workbook sheet:
    - sheet_name
    - rows_in_source        (integer)
    - rows_extracted        (integer; rows you emitted from this sheet across all arrays)
    - truncated_in_prompt   (boolean — true if the sheet had a "[TRUNCATED]" marker)
    - target_array          (which JSON array(s) this sheet feeds — comma-separated)

▸ provenance (object) — free-form keys you choose, e.g.:
    {{"fund_master": "Cover sheet", "investors": "LP Register",
      "valuations": "Valuations (IPEV) — used FV Holding column",
      "waterfall": "computed via European waterfall from inputs",
      "derived_kpis": "gross_margin_pct=1-COGS/Revenue per company",
      "notes": "any unusual aspects of this file"}}

═══════════════════════════════════════════════════════════════════════════
FINAL REMINDERS:
  • Return ONLY the JSON object. No prose, no markdown fences.
  • OMIT every null key. The vocabulary above is what's ALLOWED, not REQUIRED.
  • Empty arrays [] are encouraged for sections with no source data.
═══════════════════════════════════════════════════════════════════════════
"""


def _build_prompt(workbook_text: str, retry_hint: str = '') -> str:
    schema = _schema_block()
    hint_block = ''
    if retry_hint:
        hint_block = f"""
RETRY HINT (your previous output failed validation):
{retry_hint}

Re-examine the workbook carefully and fix the violation. Specifically check
the FV column trap (rule #4) and the European waterfall sequence (rule #5)
before re-emitting.
"""

    return f"""{_SYSTEM_RULES}

{hint_block}

WORKBOOK CONTENT (sheet by sheet, row by row):
{workbook_text}

{schema}

Return ONLY the JSON object. No prose, no markdown fences.
"""


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def run_phase2_import(import_file, progress_cb: Optional[Callable] = None):
    """Phase 2 import pipeline. Called from views.py when USE_NEW_EXTRACTOR=true.

    Args:
        import_file: ImportFile model instance
        progress_cb: function(pct, msg) for SSE progress events

    Returns: result dict matching the legacy FundImportService.import_file contract.
    """
    from .phase2_validator import validate_extraction
    from .phase2_persister import persist_phase2

    def _p(pct, msg):
        if progress_cb:
            try:
                progress_cb(pct, msg)
            except Exception:
                pass
        logger.info(f'[Phase2 {pct}%] {msg}')

    started_at = time.time()
    import_file.status = 'importing'
    import_file.save(update_fields=['status'])

    _p(5, 'Phase 2: Reading workbook…')
    filepath = import_file.file.path
    workbook_text, sheet_meta = serialize_workbook(filepath)

    _p(15, f'Phase 2: Serialized {len(sheet_meta)} sheets ({len(workbook_text):,} chars). Calling Gemini…')

    # ---------- Single Gemini call ----------
    # Three attempts: 1 initial + 2 retries. The extra retry gives one more
    # chance to fix the rare "Gemini returned a list at top level" crash mode
    # without losing the whole import.
    retry_hint = ''
    extracted = None
    MAX_ATTEMPTS = 3
    for attempt in range(1, MAX_ATTEMPTS + 1):
        prompt = _build_prompt(workbook_text, retry_hint)
        _p(20 + attempt * 5, f'Phase 2: Gemini extraction attempt {attempt}…')
        try:
            extracted = _call_gemini(prompt, context_label=f'phase2_extract_attempt_{attempt}')
        except Exception as e:
            logger.exception(f'Phase 2 Gemini call failed on attempt {attempt}')
            if attempt == MAX_ATTEMPTS:
                raise
            time.sleep(2)
            continue

        # ---------- Validation ----------
        _p(60, f'Phase 2: Validating extraction (attempt {attempt})…')
        ok, violations, hint = validate_extraction(extracted)
        if ok:
            _p(70, f'Phase 2: Validation PASSED on attempt {attempt}')
            break

        logger.warning(f'Phase 2 validation FAILED on attempt {attempt}: {violations}')
        if attempt == MAX_ATTEMPTS:
            _p(70, f'Phase 2: Validation still failing after {MAX_ATTEMPTS} attempts; persisting best-effort with {len(violations)} violations')
            break

        retry_hint = (
            f'Cross-check violations: {"; ".join(violations[:5])}. '
            f'Most likely cause: {hint}'
        )

    # ── Hard fail-fast guard ──────────────────────────────────────────
    # If after all retries the top-level shape is still wrong (Gemini stubbornly
    # returned a list instead of an object), refuse to call the persister —
    # it would crash on the first `data.get(...)`. Mark the import failed with
    # a clear, actionable message instead of a 500 traceback.
    if not isinstance(extracted, dict):
        actual = type(extracted).__name__
        msg = (
            f'Phase 2: Gemini returned a {actual} at the top level after '
            f'{MAX_ATTEMPTS} attempts (expected a JSON object). Aborting persistence '
            f'to avoid corrupt DB writes. Please retry the import; if this recurs, '
            f'try a different Gemini model or simplify the workbook.'
        )
        logger.error(msg)
        import_file.status = 'failed'
        import_file.error_detail = msg
        import_file.save(update_fields=['status', 'error_detail'])
        return {'status': 'failed', 'error': 'gemini_returned_non_object', 'detail': msg}

    # ---- ALWAYS save Gemini output to disk for inspection ----
    # Saved BEFORE persistence so we have a record even if persist crashes.
    try:
        from django.conf import settings as _dj_settings
        out_dir = os.path.join(_dj_settings.MEDIA_ROOT, 'dataimport',
                               '_gemini_outputs')
        os.makedirs(out_dir, exist_ok=True)
        out_path = os.path.join(
            out_dir,
            f'{import_file.id}_{import_file.original_filename}.json',
        )
        with open(out_path, 'w', encoding='utf-8') as _f:
            json.dump(extracted, _f, indent=2, default=str, ensure_ascii=False)
        logger.info(f'Phase 2: Gemini output saved → {out_path}')
    except Exception as _e:
        logger.warning(f'Phase 2: could not save Gemini output to disk: {_e}')

    if extracted is None:
        import_file.status = 'failed'
        import_file.error_detail = 'Phase 2 Gemini extraction returned no data'
        import_file.save(update_fields=['status', 'error_detail'])
        return {'status': 'failed', 'error': 'no_extraction'}

    # ---------- Compute precise XIRR in Python ----------
    _p(75, 'Phase 2: Computing precise XIRR from LP cashflows…')
    try:
        fp_block = extracted.setdefault('fund_performance', {})
        cashflows = list(fp_block.get('net_irr_cashflows') or [])
        # Safety net for Rule 19 — if Gemini omitted the terminal NAV
        # cashflow, append it here so XIRR isn't biased deeply negative.
        if cashflows:
            nav_latest = fp_block.get('fund_nav_latest') or fp_block.get('total_unrealised_fv_holding')
            as_of_date = fp_block.get('as_of_date')
            last_entry = cashflows[-1] if cashflows else None
            has_terminal = (
                last_entry
                and (last_entry.get('type') or '').lower() == 'distribution'
                and as_of_date
                and str(last_entry.get('date') or '') == str(as_of_date)
            )
            if nav_latest and as_of_date and not has_terminal:
                cashflows.append({
                    'date': str(as_of_date),
                    'amount': float(nav_latest),
                    'type': 'distribution',
                    'synthetic_terminal': True,
                })
                fp_block['net_irr_cashflows'] = cashflows
                logger.info('Phase 2: appended synthetic terminal NAV cashflow for XIRR')
        net_irr = _compute_xirr_from_cashflows(cashflows)
        if net_irr is not None:
            fp_block['net_irr_computed'] = net_irr
    except Exception as e:
        logger.warning(f'Phase 2 XIRR compute failed (non-fatal): {e}')

    # ---------- Python row-replication for MIS sheets truncated in prompt ----------
    _p(77, 'Phase 2: Filling MIS rows from sheets truncated in the prompt…')
    try:
        _fill_truncated_mis_rows(extracted, filepath, sheet_meta)
    except Exception as e:
        logger.warning(f'Phase 2 MIS row-replication skipped (non-fatal): {e}')

    # ---------- Persist ----------
    _p(80, 'Phase 2: Persisting to database…')
    organization = import_file.job.organization
    user = import_file.job.uploaded_by
    persist_result = persist_phase2(extracted, import_file, organization, user, progress_cb=_p)

    import_file.status = 'completed'
    import_file.completed_at = timezone.now()
    import_file.column_mapping = {'phase2': True, 'sheet_meta': sheet_meta}
    import_file.save(update_fields=['status', 'completed_at', 'column_mapping'])

    elapsed = time.time() - started_at
    _p(100, f'Phase 2 complete in {elapsed:.1f}s — {persist_result.get("summary", "")}')

    return {
        'status': 'completed',
        'phase': 'phase2',
        'elapsed_seconds': elapsed,
        **persist_result,
    }


# ---------------------------------------------------------------------------
# MIS row-replication fallback (Python, no LLM)
# ---------------------------------------------------------------------------

_MIS_TARGETS = {
    # marker keyword in sheet name (lowercased) → JSON array key
    'monthly p&l': 'monthly_pl_rows',
    'monthly p & l': 'monthly_pl_rows',
    'p&l': 'monthly_pl_rows',
    'profit and loss': 'monthly_pl_rows',
    'balance sheet': 'monthly_bs_rows',
    'cash flow': 'monthly_cf_rows',
    'cashflow': 'monthly_cf_rows',
    'budget vs actual': 'budget_vs_actual',
    'budget v actual': 'budget_vs_actual',
    'bva': 'budget_vs_actual',
}


def _sheet_to_target(sheet_name: str) -> Optional[str]:
    name = (sheet_name or '').lower()
    for marker, target in _MIS_TARGETS.items():
        if marker in name:
            return target
    return None


def _fill_truncated_mis_rows(extracted: dict, filepath: str, sheet_meta: dict):
    """For sheets truncated in the LLM prompt, replicate ALL rows here via
    deterministic openpyxl reading + Gemini's column header mapping.

    Strategy: re-read the sheet, use the first non-empty row as header row,
    emit one dict per data row with keys = lowercased+slugified headers.
    Then append into the matching JSON array (monthly_pl_rows etc.).
    """
    truncated_sheets = [name for name, m in sheet_meta.items() if m.get('truncated')]
    if not truncated_sheets:
        return

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    for sh in truncated_sheets:
        target = _sheet_to_target(sh)
        if not target:
            continue
        if sh not in wb.sheetnames:
            continue
        ws = wb[sh]

        # Find header row: first row with ≥3 non-empty cells
        header = None
        header_row_idx = None
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            non_empty = [c for c in row if c not in (None, '')]
            if len(non_empty) >= 3:
                header = [_slug(c) for c in row]
                header_row_idx = r_idx
                break
        if not header:
            continue

        # Emit data rows
        rows_added = 0
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if r_idx <= header_row_idx:
                continue
            row_dict = {h: row[i] if i < len(row) else None
                        for i, h in enumerate(header) if h}
            # Skip blank rows
            if not any(v not in (None, '') for v in row_dict.values()):
                continue
            row_dict['_source_sheet'] = sh
            row_dict['_source_row'] = r_idx
            extracted.setdefault(target, []).append(row_dict)
            rows_added += 1

        if rows_added:
            logger.info(f'Phase 2 row-replication: {sh} → {target} (+{rows_added} rows)')
    wb.close()


def _slug(v) -> str:
    """Slugify a header cell value to a JSON-safe key."""
    if v is None:
        return ''
    s = str(v).strip().lower()
    # Replace common separators with underscore
    import re as _re
    s = _re.sub(r'[\s\-/\\&]+', '_', s)
    # Strip non-alphanum-underscore
    s = _re.sub(r'[^a-z0-9_]', '', s)
    # Collapse repeated underscores
    s = _re.sub(r'_+', '_', s).strip('_')
    return s or 'col'


# ---------------------------------------------------------------------------
# XIRR (precise IRR computed in Python, not LLM)
# ---------------------------------------------------------------------------

def _compute_xirr_from_cashflows(cashflows: list) -> Optional[float]:
    """Bisection XIRR. Cashflows: [{date, amount, type}, ...] where calls are
    negative-from-LP perspective (we flip sign so calls are -ve, distributions +ve).
    Returns annualised IRR as a decimal (0.158 = 15.8%) or None.
    """
    from datetime import date as _date, datetime as _dt

    def _to_date(s):
        if isinstance(s, _date):
            return s
        if isinstance(s, _dt):
            return s.date()
        if not s:
            return None
        for fmt in ('%Y-%m-%d', '%d-%b-%Y', '%d/%m/%Y', '%Y/%m/%d', '%d-%m-%Y'):
            try:
                return _dt.strptime(str(s)[:10], fmt).date()
            except ValueError:
                continue
        return None

    flows = []
    for cf in cashflows:
        if not isinstance(cf, dict):
            continue
        d = _to_date(cf.get('date'))
        amt = cf.get('amount')
        if d is None or amt is None:
            continue
        try:
            amt = float(amt)
        except (ValueError, TypeError):
            continue
        # Sign convention: calls are LP outflows (negative), distributions are positive
        kind = (cf.get('type') or '').lower()
        if 'call' in kind or 'contribution' in kind or 'drawdown' in kind:
            amt = -abs(amt)
        elif 'dist' in kind or 'realiz' in kind or 'realis' in kind:
            amt = abs(amt)
        flows.append((d, amt))

    if len(flows) < 2:
        return None

    flows.sort(key=lambda x: x[0])
    t0 = flows[0][0]

    def npv(r):
        return sum(amt / ((1 + r) ** ((d - t0).days / 365.25)) for d, amt in flows)

    # Bisection over [-0.99, 10]
    lo, hi = -0.99, 10.0
    f_lo, f_hi = npv(lo), npv(hi)
    if f_lo * f_hi > 0:
        return None  # no sign change → IRR doesn't exist in this range
    for _ in range(200):
        mid = (lo + hi) / 2
        f_mid = npv(mid)
        if abs(f_mid) < 1e-6:
            return round(mid, 6)
        if f_lo * f_mid < 0:
            hi, f_hi = mid, f_mid
        else:
            lo, f_lo = mid, f_mid
    return round((lo + hi) / 2, 6)
