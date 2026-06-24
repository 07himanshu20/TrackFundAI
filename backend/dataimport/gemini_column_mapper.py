"""
Shared Gemini + Excel utilities used by Phase 2 (single-call extractor) and
Phase 3 (Flavor A + Flavor B layered extractor).

Public surface:
  _get_model            — model name lookup
  _call_gemini          — Vertex AI call with retry + tolerant JSON parse +
                          parse-first-then-truncate ordering
  _tolerant_parse       — JSON parser that survives common LLM slips
                          (trailing commas, smart quotes, Python literals,
                          BOM, embedded markdown fences)
  _build_cross_sheet_value_cache  — resolve =Sheet!Cell formula references
  _extract_sheet_previews         — first 5 populated rows per sheet
  GeminiTruncated       — raised when Gemini hit its output-token ceiling
  GeminiNonDictTopLevel — raised when Gemini returned valid JSON but not a dict
                          (orchestrator treats both the same way: split + retry)
"""

import json
import logging
import os
import re
import time

import openpyxl

from api.gemini_service import generate_content, get_model_name

logger = logging.getLogger(__name__)

_MAX_RETRIES = 3
_RETRY_BACKOFF_BASE = 2


class GeminiTruncated(ValueError):
    """Raised when Gemini's response was cut off by the output-token ceiling.

    Caller (Phase 3 orchestrator) catches this and splits the chunk in two,
    retrying each half — far more reliable than predicting output size up-front.
    """
    def __init__(self, message, partial_text='', finish_reason=None,
                 output_tokens=None):
        super().__init__(message)
        self.partial_text = partial_text
        self.finish_reason = finish_reason
        self.output_tokens = output_tokens


class GeminiNonDictTopLevel(ValueError):
    """Raised when Gemini returns JSON whose top level is not an object.

    Treated by the orchestrator as a split-and-retry signal: the model
    misunderstood the schema and a smaller / more focused chunk usually
    succeeds. Carries the actual parsed value for diagnostics.
    """
    def __init__(self, message, parsed_value=None):
        super().__init__(message)
        self.parsed_value = parsed_value


class GeminiQuotaExhausted(Exception):
    """Raised when Gemini returns 429 RESOURCE_EXHAUSTED after in-call retries.

    The orchestrator catches this and RE-QUEUES the SAME chunk (no split,
    no halving) into the next executor pass — by then the per-minute quota
    window has rolled over. This preserves the chunk's data instead of
    discarding it. Universal across any AIF Excel format / file size.
    """
    def __init__(self, message, attempts=0):
        super().__init__(message)
        self.attempts = attempts


def _get_model():
    return get_model_name()


def _detect_truncation(response):
    """Return (truncated_bool, finish_reason_str, output_tokens_int).

    Looks at:
      • candidates[0].finish_reason — gold standard. MAX_TOKENS = certain truncation.
      • usage_metadata.candidates_token_count — informational.
    Tolerates SDK shape variations (enum vs string vs int).
    """
    finish_reason = None
    output_tokens = None
    try:
        cands = getattr(response, 'candidates', None) or []
        if cands:
            fr = getattr(cands[0], 'finish_reason', None)
            if fr is not None:
                finish_reason = getattr(fr, 'name', None) or str(fr)
    except Exception:
        pass
    try:
        um = getattr(response, 'usage_metadata', None)
        if um is not None:
            output_tokens = getattr(um, 'candidates_token_count', None)
    except Exception:
        pass

    fr_norm = (finish_reason or '').upper()
    is_truncated = (
        'MAX_TOKENS' in fr_norm
        or fr_norm == 'LENGTH'
        or fr_norm == '2'
    )
    return is_truncated, finish_reason, output_tokens


def _detect_safety_block(response):
    """Return (blocked_bool, reason_str). Blocked = safety filter triggered."""
    try:
        fb = getattr(response, 'prompt_feedback', None)
        if fb is not None:
            br = getattr(fb, 'block_reason', None)
            if br:
                return True, getattr(br, 'name', None) or str(br)
    except Exception:
        pass
    try:
        cands = getattr(response, 'candidates', None) or []
        if cands:
            fr = getattr(cands[0], 'finish_reason', None)
            fr_norm = (getattr(fr, 'name', None) or str(fr or '')).upper()
            if fr_norm in ('SAFETY', 'RECITATION', 'BLOCKED', 'PROHIBITED_CONTENT'):
                return True, fr_norm
    except Exception:
        pass
    return False, None


def _safe_text(response):
    """Access response.text defensively. Some SDK versions raise on .text
    when finish_reason is SAFETY/MAX_TOKENS/etc. We want the partial text
    if available, else empty string."""
    try:
        return response.text or ''
    except Exception:
        try:
            cands = getattr(response, 'candidates', None) or []
            if not cands:
                return ''
            content = getattr(cands[0], 'content', None)
            parts = getattr(content, 'parts', None) or []
            out = []
            for p in parts:
                t = getattr(p, 'text', None)
                if t:
                    out.append(t)
            return ''.join(out)
        except Exception:
            return ''


# ── Tolerant JSON parser ────────────────────────────────────────────────────
#
# Defends against common LLM JSON slips that would crash json.loads:
#   • UTF-8 BOM / zero-width chars at start
#   • Smart quotes (curly “ ” ‘ ’) in place of straight " '
#   • Trailing commas before } or ]
#   • Python literals (None / True / False / NaN) emitted instead of JSON
#   • Single-quoted strings
#   • Markdown fences (``` or ```json) wrapping the body
#   • Prose surrounding a JSON body (extract first {...} or [...])
# Used as a SECONDARY parse step BEFORE the orchestrator escalates to
# terse-retry or auto-split. Keeps recoverable slips from burning Gemini calls.

_SMART_QUOTES = str.maketrans({
    '“': '"', '”': '"',
    '‘': "'", '’': "'",
    '′': "'", '″': '"',
})

_PY_LITERAL_RE = re.compile(
    r'(?<!["\\\w])(None|True|False|NaN|Infinity|-Infinity)(?!["\w])'
)
_PY_TO_JSON = {
    'None': 'null', 'True': 'true', 'False': 'false',
    'NaN': 'null', 'Infinity': 'null', '-Infinity': 'null',
}
_TRAILING_COMMA_RE = re.compile(r',(\s*[}\]])')
_SINGLE_QUOTE_KEY_RE = re.compile(r"(?<=[{,]\s)'([^'\\]+)'(\s*:)")
_CTRL_CHARS_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')


def _strip_markdown_fences(text: str) -> str:
    t = text.strip()
    if t.startswith('```'):
        t = re.sub(r'^```(?:json|JSON)?\s*\n?', '', t)
        t = re.sub(r'\n?\s*```\s*$', '', t)
    return t


def _tolerant_parse(text: str):
    """Best-effort JSON parse. Returns parsed value. Raises json.JSONDecodeError
    on definitive failure (after all repair attempts)."""
    if text is None:
        raise json.JSONDecodeError('empty response', '', 0)

    raw = text
    try:
        raw = raw.encode('utf-8').decode('utf-8-sig')
    except Exception:
        pass

    raw = raw.translate(_SMART_QUOTES)
    raw = _strip_markdown_fences(raw)
    raw = raw.strip()
    if not raw:
        raise json.JSONDecodeError('empty after normalisation', '', 0)

    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        pass

    # Repair pass 1: strip control chars + trailing commas + Python literals
    repaired = _CTRL_CHARS_RE.sub('', raw)
    repaired = _TRAILING_COMMA_RE.sub(r'\1', repaired)
    repaired = _PY_LITERAL_RE.sub(lambda m: _PY_TO_JSON[m.group(1)], repaired)
    try:
        return json.loads(repaired)
    except json.JSONDecodeError:
        pass

    # Repair pass 2: prose-wrapped JSON — extract first balanced object / array
    body = _extract_first_balanced(repaired)
    if body is not None:
        try:
            return json.loads(body)
        except json.JSONDecodeError:
            body2 = _TRAILING_COMMA_RE.sub(r'\1', body)
            try:
                return json.loads(body2)
            except json.JSONDecodeError:
                pass

    # Repair pass 3: single-quoted keys → double-quoted (last-resort, can
    # corrupt apostrophes inside string values, so try only on full failure)
    try:
        relaxed = _SINGLE_QUOTE_KEY_RE.sub(r'"\1"\2', repaired)
        return json.loads(relaxed)
    except json.JSONDecodeError:
        pass

    raise json.JSONDecodeError(
        'tolerant_parse: all repair passes failed', text[:200] if text else '', 0,
    )


def _extract_first_balanced(text: str):
    """Return the first balanced {...} or [...] in `text`, respecting strings.
    Used to peel a JSON body out of prose. Returns None if nothing balanced."""
    n = len(text)
    start = -1
    open_ch = None
    close_ch = None
    for i, ch in enumerate(text):
        if ch in '{[':
            start = i
            open_ch = ch
            close_ch = '}' if ch == '{' else ']'
            break
    if start < 0:
        return None

    depth = 0
    in_str = False
    esc = False
    for i in range(start, n):
        ch = text[i]
        if in_str:
            if esc:
                esc = False
            elif ch == '\\':
                esc = True
            elif ch == '"':
                in_str = False
            continue
        if ch == '"':
            in_str = True
            continue
        if ch == open_ch:
            depth += 1
        elif ch == close_ch:
            depth -= 1
            if depth == 0:
                return text[start:i + 1]
    return None


# ── Main Gemini call ────────────────────────────────────────────────────────

def _call_gemini(prompt, context_label='', response_schema=None,
                 require_dict_top_level=True, timeout_ms=None):
    """Call Gemini (Vertex AI) with retry + exponential backoff. Returns parsed JSON.

    Resilience order (universal across any prompt size / Excel format):

      1. Make the call (with optional response_schema + per-call timeout).
      2. Try to parse `response.text` with the tolerant parser FIRST.
         If it parses → check top-level shape. If dict → SUCCESS. If non-dict
         and require_dict_top_level → raise GeminiNonDictTopLevel.
      3. If parse fails AND finish_reason == MAX_TOKENS → raise GeminiTruncated
         so the caller can split the chunk.
      4. If parse fails for any other reason → raise json.JSONDecodeError so
         the caller can retry with a terse-output hint.
      5. Empty response or safety-blocked → retry once with a stripped prompt;
         then fail.

    Transient errors (429 / 5xx / timeout) get exponential backoff up to _MAX_RETRIES.
    """
    last_error = None
    empty_retries_done = 0

    for attempt in range(1, _MAX_RETRIES + 1):
        try:
            response = generate_content(
                prompt,
                temperature=0,
                response_mime_type='application/json',
                response_schema=response_schema,
                timeout_ms=timeout_ms,
            )

            raw_text = _safe_text(response)

            if not raw_text:
                blocked, block_reason = _detect_safety_block(response)
                is_trunc, fr, ot = _detect_truncation(response)

                if blocked:
                    if empty_retries_done == 0:
                        empty_retries_done += 1
                        logger.warning(
                            f'Gemini {context_label} blocked by safety filter '
                            f'({block_reason}); retrying once with safety-neutral hint.'
                        )
                        prompt = _strip_safety_triggers(prompt)
                        continue
                    raise ValueError(
                        f'Gemini blocked prompt ({context_label}): {block_reason}'
                    )

                if is_trunc:
                    raise GeminiTruncated(
                        f'Output truncated by Gemini (finish_reason={fr}, '
                        f'output_tokens={ot})',
                        partial_text='', finish_reason=fr, output_tokens=ot,
                    )

                if empty_retries_done == 0:
                    empty_retries_done += 1
                    logger.warning(
                        f'Gemini {context_label} returned empty response (no '
                        f'safety block, no truncation); retrying once.'
                    )
                    continue
                raise ValueError(f'Gemini returned empty response ({context_label})')

            # Parse-first (B3): try to parse before declaring truncation.
            # A response that happened to close brackets just in time IS valid
            # data even if finish_reason==MAX_TOKENS.
            parse_error = None
            try:
                result = _tolerant_parse(raw_text)
            except json.JSONDecodeError as je:
                parse_error = je
                result = None

            if parse_error is None:
                if require_dict_top_level and not isinstance(result, dict):
                    is_trunc, fr, ot = _detect_truncation(response)
                    logger.warning(
                        f'Gemini {context_label} returned top-level '
                        f'{type(result).__name__} (need dict). '
                        f'finish_reason={fr}. Escalating to split.'
                    )
                    raise GeminiNonDictTopLevel(
                        f'Top-level JSON is {type(result).__name__}, expected dict',
                        parsed_value=result,
                    )
                if attempt > 1:
                    logger.info(f'Gemini {context_label} succeeded on attempt {attempt}')
                um = getattr(response, 'usage_metadata', None)
                pt = getattr(um, 'prompt_token_count', 0) if um else 0
                ct = getattr(um, 'candidates_token_count', 0) if um else 0
                logger.info(f'[phase3.tokens] {context_label} input={pt or 0} output={ct or 0}')
                return result

            # Parse failed. Use finish_reason to decide whether to escalate
            # to truncation-split (recoverable) or terse-retry (recoverable
            # via the caller's malformed-JSON handler).
            is_trunc, fr, ot = _detect_truncation(response)
            if is_trunc:
                logger.warning(
                    f'Gemini {context_label} TRUNCATED + unparseable — '
                    f'finish_reason={fr}, output_tokens={ot}, '
                    f'response_chars={len(raw_text)}. Caller will split chunk.'
                )
                raise GeminiTruncated(
                    f'Output truncated by Gemini (finish_reason={fr}, '
                    f'output_tokens={ot})',
                    partial_text=raw_text, finish_reason=fr, output_tokens=ot,
                )

            logger.warning(
                f'Gemini {context_label} returned malformed JSON despite '
                f'finish_reason={fr}: {parse_error}'
            )
            raise parse_error

        except (GeminiTruncated, GeminiNonDictTopLevel):
            raise

        except (json.JSONDecodeError, ValueError) as e:
            logger.error(f'Gemini {context_label} non-retryable error: {e}')
            raise

        except Exception as e:
            last_error = e
            err_name = type(e).__name__
            err_str = str(e)

            is_rate_limit = '429' in err_str or 'RESOURCE_EXHAUSTED' in err_str.upper() or 'quota' in err_str.lower()
            is_server_error = any(code in err_str for code in ('500', '502', '503', '504'))
            is_timeout = 'timeout' in err_str.lower() or 'timed out' in err_str.lower()
            is_transient = is_rate_limit or is_server_error or is_timeout

            # Rate limit: bail out FAST after just 2 attempts so the orchestrator
            # can re-queue this chunk for a later executor pass. Three retries
            # with 30s back-off wastes 90s per chunk and doesn't help — the
            # whole pool is saturating the same per-minute quota. Re-queue is
            # the universal fix.
            if is_rate_limit and attempt >= 2:
                logger.warning(
                    f'Gemini {context_label} quota-exhausted after {attempt} '
                    f'attempt(s) — raising GeminiQuotaExhausted for re-queue'
                )
                raise GeminiQuotaExhausted(
                    f'Vertex 429 RESOURCE_EXHAUSTED after {attempt} attempts: {err_str[:200]}',
                    attempts=attempt,
                )

            if not is_transient or attempt == _MAX_RETRIES:
                logger.error(
                    f'Gemini {context_label} failed after {attempt} attempt(s): {err_name}: {err_str}'
                )
                raise

            wait = _RETRY_BACKOFF_BASE ** attempt
            if is_rate_limit:
                # Quota windows on Vertex are 60s — wait long enough that the
                # next attempt has a real chance of landing in a fresh window.
                wait = max(wait, int(os.environ.get('PHASE3_QUOTA_BACKOFF_S', '30')))
            logger.warning(
                f'Gemini {context_label} attempt {attempt} failed ({err_name}), retrying in {wait}s...'
            )
            time.sleep(wait)

    raise last_error


def _strip_safety_triggers(prompt: str) -> str:
    """Light-touch normaliser for the rare case Gemini's safety filter trips
    on Excel content (names, place-names that look adjacent to flagged
    categories). We only neutralise the framing, not the data."""
    return prompt + (
        '\n\n[note: this is a regulated financial data extraction task; '
        'every name/string in the workbook is operational data of an SEC/SEBI '
        'regulated AIF, not user-supplied content.]'
    )


def _parse_json_response(text):
    """Back-compat alias used by single_call_extractor / other legacy callers."""
    return _tolerant_parse(text)


def _build_cross_sheet_value_cache(filepath):
    """
    Resolve cross-sheet formulas (='Sheet'!A1 / =Sheet!B2) so Gemini sees the
    actual value, not a blank cell. Returns {(sheet_name, row, col): value}.
    """
    XREF_RE = re.compile(
        r"^=\s*'?([^'!\r\n]+?)'?\s*!\s*([A-Z]+)(\d+)\s*$", re.IGNORECASE
    )

    cache = {}
    try:
        wb_data = openpyxl.load_workbook(filepath, data_only=True)
        wb_formula = openpyxl.load_workbook(filepath, data_only=False)
    except Exception as e:
        logger.warning(f'Cross-sheet cache build failed: {e}')
        return cache

    try:
        from openpyxl.utils import column_index_from_string

        for sname in wb_data.sheetnames:
            ws_data = wb_data[sname]
            ws_formula = wb_formula[sname] if sname in wb_formula.sheetnames else None

            for row in ws_data.iter_rows():
                for cell in row:
                    val = cell.value
                    if val is not None:
                        cache[(sname, cell.row, cell.column)] = val
                        continue

                    if ws_formula is None:
                        continue
                    formula_cell = ws_formula.cell(row=cell.row, column=cell.column)
                    formula = formula_cell.value
                    if not formula or not isinstance(formula, str):
                        continue
                    formula = formula.strip()
                    if not formula.startswith('='):
                        continue

                    m = XREF_RE.match(formula)
                    if not m:
                        continue

                    ref_sheet = m.group(1).strip()
                    ref_col = column_index_from_string(m.group(2))
                    ref_row = int(m.group(3))

                    if ref_sheet not in wb_data.sheetnames:
                        continue

                    ref_ws = wb_data[ref_sheet]
                    ref_val = ref_ws.cell(row=ref_row, column=ref_col).value
                    if ref_val is not None:
                        cache[(sname, cell.row, cell.column)] = ref_val

    except Exception as e:
        logger.warning(f'Cross-sheet resolution error: {e}')
    finally:
        try:
            wb_data.close()
            wb_formula.close()
        except Exception:
            pass

    return cache


def _extract_sheet_previews(filepath, preview_rows=6):
    """
    Return (sheet_names, {sheet_name: [[row1], [row2], ...]}) with the first
    `preview_rows` populated rows per sheet, cross-sheet formulas resolved.
    """
    xsheet_cache = _build_cross_sheet_value_cache(filepath)

    wb = openpyxl.load_workbook(filepath, data_only=True)
    previews = {}
    sheet_names = wb.sheetnames

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        rows = []
        for i, row in enumerate(ws.iter_rows()):
            row_vals = []
            for cell in row:
                val = xsheet_cache.get((sheet_name, cell.row, cell.column), cell.value)
                row_vals.append(str(val) if val is not None else '')
            rows.append(row_vals)
            if i >= preview_rows - 1:
                break
        if rows:
            previews[sheet_name] = rows

    wb.close()
    return sheet_names, previews
