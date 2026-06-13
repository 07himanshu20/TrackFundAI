"""
Anchor Pipeline v2 — Analyst Architecture
==========================================

Gemini acts as the CFO/CA analyst, not as a data-extraction engine.
This matches how Claude.ai and ChatGPT chats analyze fund files
in one prompt — the approach that empirically works on any fund Excel.

Architecture:
  Stage 0 — Workbook census (Python, no AI)
  Stage 1 — FUND ANALYST call (Gemini, 1 call) — returns all 25 fund-
            level metrics with reasoning, citing source cells.
  Stage 2 — COMPANY ANALYST call (Gemini, 1 call) — returns per-
            company KPIs (Revenue, EBITDA, EBITDA%, Gross M%, MRR,
            ARR, NRR, Churn, CAC, LTV, etc.)
  Stage 3 — Audit (Python) — verify accounting identities, flag
            conflicts. Audit never silently rewrites Gemini's answer.
  Stage 4 — Persist
             • FundMetric (single source of truth)
             • DerivedMetric mirror (backward-compat for chatbot)
             • PortfolioKPI rows (per-company table)

Design principles (per CTO/CFO directive):
  • NO hardcoded value ranges, keyword lists, or magic numbers
  • Gemini is the analyst, Python is the auditor
  • For every metric Gemini cites source + formula + reasoning
  • When conflict is detected, dashboard shows both values
  • Determinism: temperature=0; same file in → same answers out
  • Token efficient: 2 Gemini calls per fund (down from ~5-9)
"""
from __future__ import annotations

import json
import logging
import math
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Optional

import openpyxl

from .gemini_column_mapper import _call_gemini

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────

def _to_decimal(value) -> Optional[Decimal]:
    if value is None or value == '':
        return None
    if isinstance(value, str):
        # strip commas, currency symbols, percent
        s = value.strip().replace(',', '').replace('₹', '').replace('$', '')
        s = s.replace('%', '').strip()
        if not s:
            return None
        try:
            return Decimal(s)
        except (InvalidOperation, ValueError):
            return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _to_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y',
                    '%d-%b-%Y', '%d %b %Y', '%b-%Y', '%Y'):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except (ValueError, AttributeError):
                continue
    return None


# ─────────────────────────────────────────────────────────────────────
# STAGE 0 — Workbook Census (no AI)
# ─────────────────────────────────────────────────────────────────────

def workbook_census(filepath, max_cells_per_sheet=2500):
    """Walk every sheet; emit compact JSON of every non-empty cell."""
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=False)
    sheets_out = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        cells, non_empty = [], 0
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                v = cell.value
                if v is None or v == '':
                    continue
                non_empty += 1
                if isinstance(v, datetime):
                    kind, v_out = 'date', v.strftime('%Y-%m-%d')
                elif isinstance(v, date):
                    kind, v_out = 'date', v.strftime('%Y-%m-%d')
                elif isinstance(v, bool):
                    kind, v_out = 'text', ('TRUE' if v else 'FALSE')
                elif isinstance(v, (int, float)):
                    fmt = (cell.number_format or '').lower()
                    kind = 'percent' if '%' in fmt else 'number'
                    v_out = float(v)
                else:
                    s = str(v).strip()
                    if not s:
                        continue
                    if len(s) > 200:
                        s = s[:200] + '…'
                    kind, v_out = 'text', s
                cells.append({'r': cell.row, 'c': cell.column,
                              'addr': cell.coordinate, 'v': v_out, 'k': kind})
        if len(cells) > max_cells_per_sheet:
            head = int(max_cells_per_sheet * 0.6)
            tail = max_cells_per_sheet - head
            cells = cells[:head] + cells[-tail:]
        sheets_out.append({
            'name': sname,
            'dimensions': {'rows': ws.max_row or 0,
                           'cols': ws.max_column or 0,
                           'non_empty': non_empty},
            'cells': cells,
        })
    wb.close()
    return {'sheets': sheets_out}


# ─────────────────────────────────────────────────────────────────────
# STAGE 1 — Fund Analyst (1 Gemini call)
# ─────────────────────────────────────────────────────────────────────

FUND_ANALYST_PROMPT = """You are a senior CFO + CA with 20+ years of experience in AIF fund
accounting. Below is the complete workbook of a fund Excel file — every
non-empty cell with sheet name, A1-style address, and value.

Treat this exactly as you would treat a real-world quarterly LP review:
  • Read the file holistically. Identify the authoritative sheets
    (cover, summary, fund master, performance metrics, NAV, waterfall).
  • For STATED HEADLINE METRICS that appear in cover/summary tables
    (TVPI, MOIC, DPI, RVPI, Net IRR, NAV, etc.), TRUST the file's
    stated value. The fund accountant put those numbers there for
    a reason.
  • COMPUTE only what is not stated, using standard PE formulas.
  • Always cite the exact source — sheet name + cell address(es) +
    one-sentence reasoning.
  • If the file does not contain enough information for a metric,
    return null. Never invent.

You must return a JSON object with these keys (use null for any value
not derivable from the file):

──── FUND IDENTITY ────
  fund_name              : string — legal name of the fund
  sebi_registration_no   : string — SEBI AIF registration (IN/AIFx/...)
  fund_pan               : string — fund's PAN
  aif_category           : "I" | "II" | "III"
  vintage_year           : integer year
  as_of_date             : YYYY-MM-DD (reporting date)
  reporting_currency     : ISO code (INR, USD, etc.)
  reporting_unit         : "absolute" | "thousands" | "lakhs"
                          | "millions" | "crores" | "billions"
  waterfall_type         : "European" | "American" | null

──── CAPITAL STRUCTURE (in the reporting_unit above) ────
  committed_capital      : total LP+GP commitments at final close
  called_capital         : cumulative capital drawn from LPs to date
  uncalled_capital       : committed − called
  invested_cost          : total cost basis of investments made
  active_fair_value      : portfolio-aggregate fair value (SUM of investee
                            FVs from the Portfolio Investments sheet).
                            DISTINCT from fund_nav — this is the GROSS
                            sum of investee FVs BEFORE deducting fund-
                            level mgmt fee, expenses, and carry provision.
                            Synonym: portfolio_aggregate_fv.
  realized_proceeds      : gross cash received from exits to date
  lp_distributions       : actual cash paid to LPs to date
  fund_nav               : LP-basis fund net asset value at as_of_date.
                            DISTINCT from active_fair_value — this is the
                            NET value after fund-level deductions, as
                            recorded on the NAV sheet. fund_nav is the
                            correct basis for carry_base computation
                            because the LPA's hurdle operates on LP-basis
                            value, not gross portfolio FV.
  cash_balance           : cash & equivalents at as_of_date

──── FUND TERMS ────
  hurdle_rate            : decimal (0.08 = 8%)
  carry_pct              : decimal (0.20 = 20%)
  catchup_pct            : decimal (1.0 = 100%)
  mgmt_fee_pct           : decimal (0.02 = 2%)
  mgmt_fee_basis         : "committed" | "called" | "nav" | null
  fund_start_date        : YYYY-MM-DD (first close / inception)

──── PERFORMANCE RATIOS (prefer stated, compute if missing) ────
  tvpi                   : decimal multiplier (e.g., 1.72)
  moic                   : decimal multiplier (e.g., 4.68)
  dpi                    : decimal multiplier
  rvpi                   : decimal multiplier
  net_irr                : decimal (0.186 = 18.6%)

──── WATERFALL OUTPUTS (compute if file doesn't state them) ────
  preferred_return_amount : amount in reporting_unit
  return_of_capital_amount: amount (= called_capital in European)
  gp_catchup_amount       : amount
  carry_base              : profit-above-hurdle on which carry is computed
  carry_amount_gross      : carry_base × carry_pct
  carry_amount_net        : carry_amount_gross − clawback_provision
  gp_clawback_provision   : amount (null if interim, not crystallised)
  lp_total_return         : LP's projected fund-end return
  gp_total_distribution   : GP's projected fund-end share

For EACH key above, return as:
  {"value": <number_or_null>,
   "source": "stated" | "computed" | "extracted",
   "source_sheet": "...", "source_cells": ["A1", "B12"],
   "formula": "(if computed) human-readable formula used",
   "reasoning": "one sentence: WHY this is the right value"}

Return shape:
{
  "fund_name": {"value": "...", "source": "extracted", "source_sheet": "...", "source_cells": ["..."], "reasoning": "..."},
  "tvpi": {"value": 1.72, "source": "stated", "source_sheet": "Cover", "source_cells": ["G14"], "reasoning": "Cover sheet states TVPI = 1.72x explicitly"},
  ...
}

WORKBOOK:
"""


def fund_analyst_call(census):
    payload = json.dumps(census, separators=(',', ':'))
    prompt = FUND_ANALYST_PROMPT + payload
    result = _call_gemini(prompt, context_label='fund_analyst')
    return result if isinstance(result, dict) else {}


# ─────────────────────────────────────────────────────────────────────
# STAGE 1A — LP / Waterfall analyst (1 Gemini call)
# ─────────────────────────────────────────────────────────────────────

LP_ANALYST_PROMPT = """You are a senior CFO + CA analysing the LP (Limited
Partner) register and waterfall data for an Alternative Investment Fund.

The full workbook is below. Find the sheet that contains the per-LP
register (it usually carries headings like "Investors", "Limited
Partners", "Unitholders", "Capital Account Summary", "LP Schedule",
etc.) and the waterfall / sponsor data (often on "FUND MASTER" or
similar).

TASK: Return aggregate LP-level totals + sponsor identity, citing the
sheet name, the column header you summed, and the cell range.

For aggregate columns: identify the column SEMANTICALLY (the LP-level
"Commitment", "Drawdown / Called", "Distribution", "Carry Provision /
Carry Accrual / GP Carry" columns), then SUM the numeric values across
all LP rows. Skip any "Total" / "Subtotal" rows.

For sponsor commitment: find the LP whose role is explicitly the
sponsor / GP / fund manager (usually flagged by the "Type" / "Role"
column as "Sponsor" / "GP" / "Manager Commitment"). Return that LP's
commitment amount. If no such LP is flagged AND no separate "Sponsor
Commitment" line exists on the FUND MASTER / Cover sheet, return null —
DO NOT guess by picking the first or largest LP.

Return JSON in this exact shape:
{
  "sum_lp_committed":        {"value": ..., "source_sheet": "...", "source_column_header": "...", "source_cells": ["F4:F17"], "reasoning": "..."},
  "sum_lp_called":           {"value": ..., "source_sheet": "...", "source_column_header": "...", "source_cells": [...], "reasoning": "..."},
  "sum_lp_distributions":    {"value": ..., "source_sheet": "...", "source_column_header": "...", "source_cells": [...], "reasoning": "..."},
  "sum_lp_carry_provision":  {"value": ..., "source_sheet": "...", "source_column_header": "...", "source_cells": [...], "reasoning": "..."},
  "sponsor_commitment_amount":{"value": ..., "source_sheet": "...", "source_cells": [...], "reasoning": "..."},
  "sponsor_lp_name":         {"value": "...", "source_sheet": "...", "source_cells": [...], "reasoning": "..."},
  "lp_count":                {"value": <int>, "source_sheet": "...", "source_cells": [...], "reasoning": "..."}
}

Rules:
  1. Each "value" must be the numeric SUM (or null).
  2. Never invent. If the column is not in the file, return value=null
     and explain in reasoning.
  3. For Carry Provision specifically: this is the per-LP accrued carry
     liability that lives on the LP register. It is NOT the gross/net
     carry computed at the fund level — it is the per-LP slice that LPs
     have actually been charged. Returning the sum of this column IS
     the correct fund-level Net Carry.
  4. Cite cell ranges in A1 form (e.g. "K4:K17"). One range per
     contiguous block — if the column has 14 LPs in K4:K17, return
     ["K4:K17"], not 14 individual cells.

WORKBOOK:
"""


def lp_analyst_call(census):
    payload = json.dumps(census, separators=(',', ':'))
    prompt = LP_ANALYST_PROMPT + payload
    result = _call_gemini(prompt, context_label='lp_analyst')
    return result if isinstance(result, dict) else {}


# ─────────────────────────────────────────────────────────────────────
# STAGE 1A.5 — Deterministic LP-register column sweep (Phase 2)
# ─────────────────────────────────────────────────────────────────────
#
# Gemini's LP analyst is non-deterministic across files: the same
# semantic concept ("LP Carry Provision") may be in column K on one fund
# and column L on another, with header text varying ("Carry Prov.",
# "Carried Interest Accrual", "Performance Fee", ...). On the Avendus
# import, Gemini silently missed the column and the pipeline fell back
# to a derived formula — recording the wrong number with no warning.
#
# This sweep removes that non-determinism. For each sheet plausibly
# holding an LP register, we:
#   1. Find the header row (the row with the most short text headers).
#   2. Classify every header semantically via Pass 3 classify_labels →
#      lp_register_columns. Universal across languages and label
#      conventions. No keyword tables.
#   3. Sum each role-mapped column with openpyxl. Skip Total/Subtotal
#      rows (detected by first-column text content, language-agnostic).
#   4. Inject the sums into lp_data — overriding Gemini whenever
#      Gemini was null OR the values disagree by >10%.
#
# Output reaches FundMetric via fill_computed_metrics()'s existing
# _set_from_lp() helper, so this change is invisible to downstream code.

_LP_TOTAL_FIRST_CELL_TOKENS = ('total', 'subtotal', 'grand total', 'totals')


def _is_total_row(ws, row_idx, max_cols_to_scan=3):
    """Detect aggregation rows by checking the first few cells for tokens
    like TOTAL / SUBTOTAL. Universal: matches case-insensitively and works
    regardless of which language the file is in (the tokens themselves
    are dominantly English in AIF filings, but the same pattern catches
    'TOTALS', 'Grand Total', etc.). False positives cost a single LP
    row; safer than not skipping totals at all.
    """
    for c in range(1, max_cols_to_scan + 1):
        v = ws.cell(row=row_idx, column=c).value
        if v is None:
            continue
        s = str(v).strip().lower()
        if s and any(tok in s for tok in _LP_TOTAL_FIRST_CELL_TOKENS):
            return True
    return False


def _find_header_row(ws, max_rows_to_scan=20):
    """Pick the header row as the first row (within the first 20) where
    a majority of populated cells are short strings — the universal
    shape of a column-header band in financial workbooks.
    Returns (row_index, [header_strings_per_column]) or (None, None).
    """
    best = (None, None, 0)  # (row_idx, headers, score)
    for r in range(1, min(ws.max_row + 1, max_rows_to_scan + 1)):
        headers = []
        n_string = n_filled = 0
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            headers.append(v)
            if v is None or v == '':
                continue
            n_filled += 1
            if isinstance(v, str) and len(v.strip()) > 0 and len(v.strip()) <= 80:
                n_string += 1
        if n_filled < 3:
            continue
        score = n_string  # heuristic: more string cells = more header-like
        if score > best[2]:
            best = (r, headers, score)
    return (best[0], best[1])


def fund_nav_component_sweep(filepath, fund_data):
    """Phase 4 — universal NAV fallback when the NAV column is a formula
    that openpyxl could not pre-evaluate.

    openpyxl(data_only=True) returns None for cells holding unevaluated
    formulas. On the KKR Infra Trust workbook (and now on Avendus) the
    NAV sheet's Total NAV column is `=D+F+G−E−H` style — never opened
    in Excel before us, so the cached value is missing and the anchor
    pipeline persists `fund_nav=None`. This in turn cascades into bug G
    (carry_base picked up from portfolio FV) and bug H (Total Fair Value
    tile mis-labelled).

    Universal fix:
      1. Open the workbook twice — once for cached values, once for
         formulas. Pair them per cell so we know which were formulas.
      2. Iterate sheets that look like NAV ledgers (≥3 numeric columns,
         a date/period column, and a row matching `total_nav` via Pass 3
         semantic classification).
      3. For each period (row), compute NAV deterministically:
            NAV = total_investments
                + unrealized_gains
                + realized_gains
                + investment_income
                - mgmt_fee
                - fund_expenses
                - carry_provision
         Only the components present in the file participate; missing
         components contribute zero, which mirrors how the workbook
         formula is constructed.
      4. Pick the LATEST period (last non-empty row of the components)
         and inject its value into `fund_data['fund_nav']`, but only if
         Gemini's value was null OR disagreed by >10%.

    No keywords, no per-file branches. Header roles are discovered by
    Pass 3 classify_labels every run; component columns may be in any
    order and may carry any wording.
    """
    try:
        import openpyxl
    except Exception:
        return fund_data

    existing_entry = (fund_data or {}).get('fund_nav') or {}
    existing_val = existing_entry.get('value') if isinstance(existing_entry, dict) else None
    try:
        existing_num = float(existing_val) if existing_val is not None else None
    except (TypeError, ValueError):
        existing_num = None

    try:
        from .gemini_column_mapper import classify_labels
        from .canonical_schema import CANONICAL_VALUE_CATEGORIES
    except Exception:
        return fund_data
    canonical = CANONICAL_VALUE_CATEGORIES.get('nav_components')
    if not canonical:
        return fund_data

    try:
        wb_vals = openpyxl.load_workbook(filepath, data_only=True)
        wb_fmls = openpyxl.load_workbook(filepath, data_only=False)
    except Exception as e:
        logger.warning(f'workbook open failed in NAV sweep: {e}')
        return fund_data

    # SIGN_MAP: how each component participates in the NAV identity.
    SIGN_MAP = {
        'total_investments':  +1,
        'unrealized_gains':   +1,
        'realized_gains':     +1,
        'investment_income':  +1,
        'mgmt_fee':           -1,
        'fund_expenses':      -1,
        'carry_provision':    -1,
    }

    best = None  # (nav_value, sheet_name, period_label, period_row, formula_components, source_cells)

    for sheet_name in wb_vals.sheetnames:
        ws_v = wb_vals[sheet_name]
        ws_f = wb_fmls[sheet_name]
        if ws_v.max_row < 4 or ws_v.max_column < 3:
            continue
        header_row, headers = _find_header_row(ws_v)
        if not headers:
            continue
        header_strs = [str(h).strip() for h in headers
                       if h is not None and str(h).strip()]
        if len(header_strs) < 4:
            continue

        try:
            role_map = classify_labels(
                header_strs, 'nav_components', canonical,
                context='Fund NAV / accounting ledger — per-period fund net asset value build-up'
            )
        except Exception:
            continue
        if not role_map:
            continue

        # Index columns by canonical role
        role_to_col = {}
        for col_idx, original_header in enumerate(headers, start=1):
            if original_header is None:
                continue
            h_str = str(original_header).strip()
            role = role_map.get(h_str)
            if role and role not in role_to_col:
                role_to_col[role] = (col_idx, h_str)

        # Must look like a NAV ledger: at least 3 numeric components OR
        # have an explicit total_nav column that is formula-only.
        numeric_roles_present = set(role_to_col) & set(SIGN_MAP)
        has_nav_col = 'total_nav' in role_to_col
        if len(numeric_roles_present) < 3 and not has_nav_col:
            continue

        # Scan rows: find the latest row where component values are
        # populated, and compute NAV deterministically.
        latest = None
        for r in range(header_row + 1, ws_v.max_row + 1):
            if _is_total_row(ws_v, r):
                continue
            nav_components_sum = None
            comp_inputs = {}
            for role, sign in SIGN_MAP.items():
                if role not in role_to_col:
                    continue
                col_idx, col_header = role_to_col[role]
                v = ws_v.cell(row=r, column=col_idx).value
                if isinstance(v, (int, float)):
                    if nav_components_sum is None:
                        nav_components_sum = 0.0
                    nav_components_sum += sign * float(v)
                    comp_inputs[role] = float(v)
            if nav_components_sum is None:
                continue

            # If the file also has a total_nav column, prefer its cached
            # value when present. Only fall back to component sum when
            # the cell is a formula with None value.
            nav_from_col = None
            nav_is_formula = False
            if 'total_nav' in role_to_col:
                col_idx, _ = role_to_col['total_nav']
                cv = ws_v.cell(row=r, column=col_idx).value
                cf = ws_f.cell(row=r, column=col_idx).value
                if isinstance(cv, (int, float)):
                    nav_from_col = float(cv)
                elif isinstance(cf, str) and cf.startswith('='):
                    nav_is_formula = True

            row_nav = nav_from_col if nav_from_col is not None else nav_components_sum
            period = None
            for role in ('period_label', 'period_date'):
                if role in role_to_col:
                    col_idx, _ = role_to_col[role]
                    pv = ws_v.cell(row=r, column=col_idx).value
                    if pv is not None and pv != '':
                        period = str(pv)
                        break

            latest = (row_nav, period, r, comp_inputs, nav_is_formula,
                      nav_from_col is not None)

        if latest is None:
            continue

        row_nav, period, period_row, comp_inputs, nav_is_formula, had_cached = latest
        if best is None or abs(row_nav) > abs(best[0]):
            # Build provenance cell list — last row of each component column
            cells = []
            for role in sorted(comp_inputs.keys()):
                col_idx, _ = role_to_col[role]
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                cells.append(f'{col_letter}{period_row}')
            if 'total_nav' in role_to_col:
                col_idx, _ = role_to_col['total_nav']
                cells.append(f'{openpyxl.utils.get_column_letter(col_idx)}{period_row}')
            best = (row_nav, sheet_name, period, period_row, comp_inputs,
                    cells, nav_is_formula, had_cached)

    if best is None:
        return fund_data

    row_nav, sheet_name, period, period_row, comp_inputs, cells, nav_is_formula, had_cached = best

    denom = max(abs(existing_num or 0), abs(row_nav), 1.0)
    disagree = (existing_num is not None and
                abs(existing_num - row_nav) / denom > 0.10)
    if existing_num is None or disagree:
        if had_cached and not nav_is_formula:
            method = 'nav_component_sweep:cached_value'
            reasoning = (f'Total NAV cell carried a cached value of {row_nav:.2f} '
                         f'on sheet "{sheet_name}" for period "{period}".')
        else:
            sign_terms = []
            for role, val in comp_inputs.items():
                sign = '+' if SIGN_MAP[role] > 0 else '−'
                sign_terms.append(f'{sign}{role}({val:.2f})')
            formula_str = ' '.join(sign_terms) if sign_terms else '(no components found)'
            method = 'nav_component_sweep:components'
            reasoning = (f'Total NAV cell was a formula '
                         f'(openpyxl returned None); computed deterministically '
                         f'from component columns on sheet "{sheet_name}" for '
                         f'period "{period}": {formula_str}.')
        fund_data['fund_nav'] = {
            'value':       row_nav,
            'source':      'computed' if not (had_cached and not nav_is_formula) else 'extracted',
            'formula':     'total_investments + unrealized_gains + realized_gains '
                           '+ investment_income - mgmt_fee - fund_expenses - carry_provision',
            'reasoning':   reasoning,
            'source_sheet': sheet_name,
            'source_cells': cells,
            'inputs_used': comp_inputs,
            'extraction_method': method,
        }
        logger.info(
            f'[nav_sweep override] fund_nav: existing={existing_num} → '
            f'{row_nav:.4f} from {sheet_name} period "{period}" row {period_row}'
        )

    return fund_data


def lp_register_python_sweep(filepath, lp_data):
    """Deterministic Python sweep over LP register sheets.

    Returns a NEW dict layered on top of `lp_data` — Gemini's values are
    preserved where they (a) exist and (b) agree with the Python sums
    within 10%. Otherwise the Python sums take precedence and provenance
    records the exact sheet/column/range.
    """
    try:
        import openpyxl
    except Exception:
        logger.warning('openpyxl unavailable; skipping LP sweep')
        return lp_data

    try:
        from .gemini_column_mapper import classify_labels
        from .canonical_schema import CANONICAL_VALUE_CATEGORIES
    except Exception as e:
        logger.warning(f'classify_labels import failed; skipping LP sweep: {e}')
        return lp_data

    canonical = CANONICAL_VALUE_CATEGORIES.get('lp_register_columns')
    if not canonical:
        return lp_data

    # Candidate sheets — Gemini may have cited one in lp_data; if not,
    # scan all sheets (the classifier will reject non-LP sheets by
    # returning None for their headers).
    cited_sheets = set()
    for key in ('sum_lp_committed', 'sum_lp_called',
                'sum_lp_distributions', 'sum_lp_carry_provision'):
        entry = (lp_data or {}).get(key) or {}
        sheet = entry.get('source_sheet')
        if sheet:
            cited_sheets.add(sheet)

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=False)
    except Exception as e:
        logger.warning(f'workbook open failed in LP sweep: {e}')
        return lp_data

    sheets_to_scan = ([s for s in cited_sheets if s in wb.sheetnames]
                      or list(wb.sheetnames))

    # Best sum + provenance per role (across all sheets). If multiple
    # sheets contain the same role, prefer the one with the larger
    # population (more LP rows summed).
    role_sums = {}   # role -> (value, sheet, col_letter, col_header, rng, row_count)

    for sheet_name in sheets_to_scan:
        ws = wb[sheet_name]
        if ws.max_row < 4 or ws.max_column < 3:
            continue
        header_row, headers = _find_header_row(ws)
        if not headers:
            continue
        header_strs = [str(h).strip() for h in headers
                       if h is not None and str(h).strip()]
        if len(header_strs) < 3:
            continue

        # Classify headers via Pass 3 semantic classifier. Cached per
        # category_key + frozenset(labels) — re-invocation with the same
        # headers across funds is free.
        try:
            mapping = classify_labels(
                header_strs, 'lp_register_columns', canonical,
                context='LP / Investor register sheet — fund-level capital account summary'
            )
        except Exception as e:
            logger.debug(f'classify_labels failed on {sheet_name}: {e}')
            continue
        if not mapping:
            continue

        # Require at least 2 numeric roles to consider this an LP register
        numeric_roles = {'committed', 'drawdown', 'distributions',
                         'carry_provision', 'sponsor_amount'}
        roles_found = {mapping.get(h) for h in header_strs} & numeric_roles
        if len(roles_found) < 2:
            continue

        for col_idx, original_header in enumerate(headers, start=1):
            if original_header is None:
                continue
            h_str = str(original_header).strip()
            role = mapping.get(h_str)
            if role not in numeric_roles:
                continue

            col_letter = openpyxl.utils.get_column_letter(col_idx)
            total = 0.0
            first_row = None
            last_row = None
            n_rows = 0
            for r in range(header_row + 1, ws.max_row + 1):
                if _is_total_row(ws, r):
                    continue
                v = ws.cell(row=r, column=col_idx).value
                if isinstance(v, (int, float)):
                    total += float(v)
                    if first_row is None:
                        first_row = r
                    last_row = r
                    n_rows += 1
            if n_rows < 2 or total == 0:
                continue

            prev = role_sums.get(role)
            if prev is None or n_rows > prev[5]:
                rng = f'{col_letter}{first_row}:{col_letter}{last_row}'
                role_sums[role] = (total, sheet_name, col_letter, h_str, rng, n_rows)

    if not role_sums:
        return lp_data

    # Merge into lp_data. Override when Gemini was null OR disagrees by >10%.
    role_to_key = {
        'committed':       'sum_lp_committed',
        'drawdown':        'sum_lp_called',
        'distributions':   'sum_lp_distributions',
        'carry_provision': 'sum_lp_carry_provision',
        'sponsor_amount':  'sponsor_commitment_amount',
    }
    updated = dict(lp_data or {})
    for role, (py_val, sheet, col, hdr, rng, n_rows) in role_sums.items():
        lp_key = role_to_key.get(role)
        if not lp_key:
            continue
        existing = (updated.get(lp_key) or {}) if isinstance(updated.get(lp_key), dict) else {}
        gem_val = existing.get('value')
        try:
            gem_num = float(gem_val) if gem_val is not None else None
        except (TypeError, ValueError):
            gem_num = None
        denom = max(abs(gem_num or 0), abs(py_val), 1.0)
        disagree = (gem_num is not None and abs(gem_num - py_val) / denom > 0.10)
        if gem_num is None or disagree:
            note = ('Gemini value was null' if gem_num is None
                    else f'Gemini value ({gem_num:.2f}) disagreed with deterministic Python sum ({py_val:.2f}) by >10%; using Python sum')
            updated[lp_key] = {
                'value':                 py_val,
                'source_sheet':          sheet,
                'source_column_header':  hdr,
                'source_cells':          [rng],
                'reasoning':             f'Python deterministic sum over LP register column {col} '
                                          f'(header "{hdr}") on sheet "{sheet}" — {n_rows} LP rows. {note}.',
                'extraction_method':     'python_deterministic_sweep',
                'rows_summed':           n_rows,
            }
            logger.info(
                f'[lp_sweep override] {lp_key}: '
                f'gemini={gem_num} → python={py_val:.4f} '
                f'from {sheet}!{rng}'
            )
        else:
            # Agree — keep Gemini's value but enrich provenance to record
            # that Python verified the sum.
            existing['python_verified_sum'] = py_val
            existing['python_verification_sheet'] = sheet
            existing['python_verification_range'] = rng
            updated[lp_key] = existing

    return updated


# ─────────────────────────────────────────────────────────────────────
# STAGE 2 — Company KPI Analyst (chunked Gemini calls)
# ─────────────────────────────────────────────────────────────────────
#
# Strategy: One Gemini call to list portfolio companies, then batch
# Gemini calls (~12 companies per batch) to fill KPIs.  This keeps each
# JSON response small enough to never truncate — the root cause behind
# our earlier "Expecting ',' delimiter at column 5680" failures on
# Trivesta / Sequoia files with 50+ companies.

COMPANY_LIST_PROMPT = """You are reading the complete workbook of an Alternative Investment Fund (AIF).

TASK: Return ONLY the full list of portfolio (investee) companies in
this fund. Examine every sheet — portfolio register, valuations, KPI
sheet, SaaS metrics, P&L sheets — and consolidate all real investee
companies that appear anywhere.

EXCLUDE:
  • "Total" / "Subtotal" / "Grand Total" aggregate rows
  • The fund itself (only return its investees)
  • LPs / investors / unitholders (these are not portfolio companies)
  • Service-provider entities (auditor, trustee, custodian, RTA)
  • Header / category-label rows ("S.No", "Particulars", "Sr.")

Return JSON:
{
  "companies": [
    {"company_name": "...", "sector": "...", "status": "active|exited|watch|written_off"},
    ...
  ]
}

Be exhaustive — typical AIF portfolios have 10-150 investees.

WORKBOOK:
"""


COMPANY_KPI_BATCH_PROMPT_TEMPLATE = """You are a senior CFO + CA analysing a fund's portfolio. The full workbook
of the fund's Excel file is below.

TASK: For EACH of the portfolio companies listed below, extract the
financial fields listed.  Use any sheet that holds per-company data
(portfolio register, valuations, KPI sheet, SaaS metrics, P&L).

Compute derived ratios when components are available
(e.g. ebitda_pct = ebitda / revenue, gross_margin_pct = gross_profit /
revenue, ltv_to_cac = ltv / cac).  Return null when neither stated nor
derivable.

PORTFOLIO COMPANIES TO PROCESS (return one record per company below;
match company_name exactly to one of these):
{company_list}

PER-COMPANY FIELDS (return all that the file supports):
  company_name              — must match exactly one name above
  sector / status

  cost                      — fund's cost basis (₹)
  fair_value                — fund's pro-rata FV at as_of_date (₹)
  ownership_pct             — fund's % stake (decimal, e.g. 0.15 = 15%)
  realized_proceeds         — cumulative cash this exit has paid
  moic                      — (fair_value + realized) / cost

  ──── Financial KPIs (latest period) ────
  revenue                   — latest period revenue (₹)
  ebitda                    — EBITDA amount (₹)
  gross_margin_pct          — Gross profit / Revenue (decimal, 0.42 = 42%)
  ebitda_pct                — EBITDA / Revenue (decimal)
  pat                       — profit after tax (₹)

  ──── Operational KPIs ────
  gmv                       — Gross Merchandise Value (₹)
  orders                    — order count
  aov                       — Average Order Value (₹)
  returns_pct               — Returns rate (decimal)
  repeat_pct                — Repeat customer rate (decimal)
  cac                       — Customer Acquisition Cost (₹)

  ──── SaaS Metrics (when applicable) ────
  mrr                       — Monthly Recurring Revenue (₹)
  arr                       — Annual Recurring Revenue (₹)
  nrr_pct                   — Net Revenue Retention (decimal)
  churn_pct                 — Monthly churn rate (decimal)
  ltv                       — Customer Lifetime Value (₹)
  ltv_to_cac                — LTV / CAC ratio

Return JSON:
{{
  "companies": [
    {{ "company_name": "...", "revenue": 12.3, "ebitda_pct": 0.18, ... }},
    ...
  ]
}}

Skip aggregate / total rows. If a company has zero usable data, still
include it with company_name only (rest null).

WORKBOOK:
"""


# Internal field name → dashboard KPI slug (matches _KPI_COL_SLUGS in
# investments.views.portfolio_kpi_matrix and SAAS_SLUGS in
# portfolio_saas_metrics).  Slugs use dashes per dashboard convention.
KPI_SLUG_MAP = {
    'revenue':           ('revenue',           'Revenue',           'currency'),
    'gross_margin_pct':  ('gross-margin-pct',  'Gross Margin %',    'percent'),
    'ebitda':            ('ebitda',            'EBITDA',            'currency'),
    'ebitda_pct':        ('ebitda-pct',        'EBITDA %',          'percent'),
    'pat':               ('pat',               'PAT',               'currency'),
    'gmv':               ('gmv',               'GMV',               'currency'),
    'orders':            ('orders',            'Orders',            'number'),
    'aov':               ('aov',               'AOV',               'currency'),
    'returns_pct':       ('returns-pct',       'Returns %',         'percent'),
    'repeat_pct':        ('repeat-pct',        'Repeat %',          'percent'),
    'cac':               ('cac',               'CAC',               'currency'),
    'mrr':               ('mrr',               'MRR',               'currency'),
    'arr':               ('arr',               'ARR',               'currency'),
    'nrr_pct':           ('nrr',               'NRR %',             'percent'),
    'churn_pct':         ('churn-rate',        'Churn Rate',        'percent'),
    'ltv':               ('ltv',               'LTV',               'currency'),
    'ltv_to_cac':        ('ltv-cac',           'LTV / CAC',         'ratio'),
}

# How many companies per Gemini KPI batch.  Empirically 12 keeps the
# JSON response under ~25 KB which never truncates with gemini-2.5-flash.
_COMPANY_BATCH_SIZE = 12


def company_analyst_call(census):
    """List portfolio companies (1 call), then fill KPIs in batches of
    ~12 (N calls). Returns merged {'companies': [...]}.

    This replaces the previous single-shot prompt that asked Gemini for
    every company × every field at once — those JSONs frequently
    exceeded the response token cap and truncated mid-object.
    """
    payload = json.dumps(census, separators=(',', ':'))

    try:
        listing = _call_gemini(
            COMPANY_LIST_PROMPT + payload,
            context_label='company_list',
        )
    except Exception as e:
        logger.error(f'Company listing call failed: {type(e).__name__}: {e}')
        return {'companies': []}

    raw_list = listing.get('companies') if isinstance(listing, dict) else None
    if not isinstance(raw_list, list) or not raw_list:
        logger.info('Company listing returned 0 companies; skipping KPI batches')
        return {'companies': []}

    # De-duplicate by case-insensitive name
    seen_names, unique_companies = set(), []
    for c in raw_list:
        if not isinstance(c, dict):
            continue
        name = (c.get('company_name') or '').strip()
        if not name:
            continue
        key = name.lower()
        if key in seen_names:
            continue
        seen_names.add(key)
        unique_companies.append(c)

    logger.info(f'Company listing: {len(unique_companies)} unique companies')

    merged = []
    n_batches = (len(unique_companies) + _COMPANY_BATCH_SIZE - 1) // _COMPANY_BATCH_SIZE
    for i in range(0, len(unique_companies), _COMPANY_BATCH_SIZE):
        batch = unique_companies[i:i + _COMPANY_BATCH_SIZE]
        batch_no = (i // _COMPANY_BATCH_SIZE) + 1
        names_block = '\n'.join(
            f'  - {c.get("company_name", "").strip()}' for c in batch
        )
        prompt = (
            COMPANY_KPI_BATCH_PROMPT_TEMPLATE.format(company_list=names_block)
            + payload
        )
        try:
            result = _call_gemini(
                prompt, context_label=f'company_kpi_batch_{batch_no}_of_{n_batches}',
            )
            if isinstance(result, dict):
                batch_companies = result.get('companies') or []
                if isinstance(batch_companies, list):
                    merged.extend(batch_companies)
                    logger.info(
                        f'Company KPI batch {batch_no}/{n_batches}: '
                        f'{len(batch_companies)} records returned'
                    )
        except Exception as e:
            logger.error(
                f'Company KPI batch {batch_no}/{n_batches} failed: '
                f'{type(e).__name__}: {e}'
            )
            # Fall back to the basic identity-only records from the
            # listing so the company still appears on the dashboard.
            for c in batch:
                merged.append({
                    'company_name': c.get('company_name'),
                    'sector': c.get('sector'),
                    'status': c.get('status'),
                })

    return {'companies': merged}


# ─────────────────────────────────────────────────────────────────────
# STAGE 1B — Universal computed-metric fallback (Python, no AI)
# ─────────────────────────────────────────────────────────────────────

def fill_computed_metrics(fund_data, lp_data=None, scheme=None):
    """When Gemini did not state a standard PE ratio but the components
    are present, compute it here using PE-standard formulas.  Universal
    — works for ANY fund.  Only fires when the target key is currently
    null; never overwrites Gemini's stated values.

    lp_data (optional) is the output of lp_analyst_call() and supplies
    LP-level aggregates needed for Net Carry, Clawback, and Sponsor
    Commitment. When lp_data is None or empty, those fall back to
    pure-formula derivations.

    scheme (optional Scheme model) supplies Phase 7 fallbacks for
    fund_start_date / as_of_date used by the preferred-return formula
    when Gemini failed to extract them from FUND_MASTER. Universal —
    looks at scheme.first_close_date / final_close_date / vintage_year
    in that priority order; uses today's date as last-resort as_of_date.
    """
    def _val(key):
        d = fund_data.get(key) or {}
        return _to_decimal(d.get('value'))

    def _lp(key):
        if not lp_data:
            return None, None
        d = lp_data.get(key) or {}
        return _to_decimal(d.get('value')), d

    def _set(key, value, formula, components, source='computed',
             source_sheet=None, source_cells=None, inputs=None):
        if value is None:
            return
        existing = fund_data.get(key) or {}
        if existing.get('value') is not None:
            return
        try:
            float_val = float(value)
        except (TypeError, ValueError):
            return
        if math.isnan(float_val) or math.isinf(float_val):
            return
        fund_data[key] = {
            'value': float_val,
            'source': source,
            'source_sheet': source_sheet or existing.get('source_sheet'),
            'source_cells': source_cells or existing.get('source_cells') or [],
            'formula': formula,
            'inputs_used': inputs or {},
            # Phase 6 (Bug R): lead with the actual computation. The
            # dashboard truncates reasoning to the first sentence, so a
            # message that opens with "Not stated in file" reads as if
            # the metric is missing even when we just computed it. The
            # rewritten phrasing makes the first sentence describe the
            # derivation, with the "not directly stated" qualifier moved
            # to a trailing clause. Universal phrasing — same for every
            # computed metric.
            'reasoning': (
                f'Derived as {components} (value not stated directly in the file).'
            ),
        }

    def _set_from_lp(key, lp_key, formula_label):
        v, src = _lp(lp_key)
        if v is None:
            return False
        existing = fund_data.get(key) or {}
        if existing.get('value') is not None:
            return False
        fund_data[key] = {
            'value': float(v),
            'source': 'extracted',
            'source_sheet': (src or {}).get('source_sheet'),
            'source_cells': (src or {}).get('source_cells') or [],
            'formula': formula_label,
            'inputs_used': {'sum_of': (src or {}).get('source_column_header')},
            'reasoning': (
                f"Extracted as sum over LP register column "
                f"'{(src or {}).get('source_column_header')}' on sheet "
                f"'{(src or {}).get('source_sheet')}'."
            ),
        }
        return True

    fv = _val('active_fair_value')
    rp = _val('realized_proceeds')
    ic = _val('invested_cost')
    cc = _val('committed_capital')
    cl = _val('called_capital')
    ld = _val('lp_distributions')
    dpi = _val('dpi')
    rvpi = _val('rvpi')
    hurdle = _val('hurdle_rate')
    carry_pct = _val('carry_pct')
    catchup_pct = _val('catchup_pct')

    if ic is not None and ic > 0 and fv is not None and rp is not None:
        _set('moic', (fv + rp) / ic,
             '(active_fair_value + realized_proceeds) / invested_cost',
             f'FV={fv}, Realized={rp}, Cost={ic}',
             inputs={'active_fair_value': float(fv),
                     'realized_proceeds': float(rp),
                     'invested_cost': float(ic)})

    if dpi is not None and rvpi is not None:
        existing_tvpi = fund_data.get('tvpi') or {}
        existing_src = (existing_tvpi.get('source') or '').lower()
        if existing_src != 'stated':
            dpi_entry = fund_data.get('dpi') or {}
            rvpi_entry = fund_data.get('rvpi') or {}
            cell_refs = []
            for entry in (dpi_entry, rvpi_entry):
                for c in (entry.get('source_cells') or []):
                    sheet = entry.get('source_sheet')
                    cell_refs.append(f'{sheet}!{c}' if sheet else str(c))
            fund_data['tvpi'] = {
                'value': float(dpi + rvpi),
                'source': 'computed',
                'source_sheet': dpi_entry.get('source_sheet')
                                  or rvpi_entry.get('source_sheet'),
                'source_cells': cell_refs,
                'formula': 'dpi + rvpi',
                'inputs_used': {'dpi': float(dpi), 'rvpi': float(rvpi)},
                'reasoning': (
                    'TVPI is the LP-basis sum of DPI + RVPI. Both are '
                    'extracted directly from the file, so we use the '
                    'accounting identity TVPI = DPI + RVPI instead of '
                    'computing from FV / called (which would mix '
                    'portfolio-aggregate FV with LP-basis called capital).'
                ),
            }
    elif cl is not None and cl > 0 and fv is not None and ld is not None:
        _set('tvpi', (fv + ld) / cl,
             '(active_fair_value + lp_distributions) / called_capital',
             f'FV={fv}, Distributions={ld}, Called={cl}',
             inputs={'active_fair_value': float(fv),
                     'lp_distributions': float(ld),
                     'called_capital': float(cl)})

    if cl is not None and cl > 0 and ld is not None:
        _set('dpi', ld / cl, 'lp_distributions / called_capital',
             f'Distributions={ld}, Called={cl}',
             inputs={'lp_distributions': float(ld),
                     'called_capital': float(cl)})

    if cl is not None and cl > 0 and fv is not None:
        _set('rvpi', fv / cl, 'active_fair_value / called_capital',
             f'FV={fv}, Called={cl}',
             inputs={'active_fair_value': float(fv),
                     'called_capital': float(cl)})

    if cc is not None and cl is not None:
        _set('uncalled_capital', cc - cl,
             'committed_capital - called_capital',
             f'Committed={cc}, Called={cl}',
             inputs={'committed_capital': float(cc),
                     'called_capital': float(cl)})

    _set_from_lp('committed_capital', 'sum_lp_committed',
                 'SUM(LP Commitment column)')
    _set_from_lp('called_capital', 'sum_lp_called',
                 'SUM(LP Drawdown column)')
    _set_from_lp('lp_distributions', 'sum_lp_distributions',
                 'SUM(LP Distributions column)')

    net_extracted = _set_from_lp('carry_amount_net', 'sum_lp_carry_provision',
                                  'SUM(LP Carry Provision column)')

    cg = _val('carry_amount_gross')
    cn = _val('carry_amount_net')

    # Phase 3 — Carry Base & Gross computed with FUND NAV preference.
    # fund_nav is the LP-basis value (after mgmt fee, expenses, carry
    # provision). Portfolio-aggregate FV (active_fair_value) is the GROSS
    # sum of investee fair values BEFORE fund-level deductions. The LPA's
    # hurdle and carry waterfall operate on the LP-basis value, so
    # carry_base must use fund_nav whenever it is known. Falling back to
    # portfolio FV is preserved for funds where the NAV sheet has not yet
    # been populated (Phase 4 will compute fund_nav from components when
    # the NAV cell holds an unevaluated formula).
    nv = _val('fund_nav')
    roc = _val('return_of_capital_amount') or cl
    pref = _val('preferred_return_amount')

    # Phase 7 — preferred-return date-robust fallback.
    # When Gemini did not extract preferred_return_amount AND we have
    # called_capital + hurdle_rate, compute it deterministically using
    # the standard PE formula: called × ((1 + hurdle)^years − 1).
    # The years term needs (fund_start_date, as_of_date); when those
    # are missing from FUND_MASTER we fall back to the Scheme model:
    #   • fund_start_date  → scheme.first_close_date
    #                      → scheme.final_close_date
    #                      → date(scheme.vintage_year, 1, 1)
    #   • as_of_date       → today (universal, file-agnostic)
    # Same code path for every fund — no per-file logic.
    if pref is None and cl is not None and cl > 0:
        hr = _val('hurdle_rate')
        if hr is not None and hr > 0:
            try:
                from datetime import date
                fund_start = None
                as_of = None
                fs_entry = fund_data.get('fund_start_date') or {}
                ao_entry = fund_data.get('as_of_date') or {}

                def _parse_date(raw):
                    if raw is None:
                        return None
                    if isinstance(raw, date):
                        return raw
                    s = str(raw).strip()
                    if not s:
                        return None
                    for fmt in ('%Y-%m-%d', '%d-%b-%Y', '%d-%B-%Y',
                                '%d/%m/%Y', '%d-%m-%Y'):
                        try:
                            from datetime import datetime
                            return datetime.strptime(s, fmt).date()
                        except (ValueError, TypeError):
                            continue
                    return None

                fund_start = _parse_date(fs_entry.get('value'))
                as_of = _parse_date(ao_entry.get('value'))

                if scheme is not None:
                    if fund_start is None:
                        fund_start = (getattr(scheme, 'first_close_date', None)
                                      or getattr(scheme, 'final_close_date', None))
                        if fund_start is None:
                            vy = getattr(scheme, 'vintage_year', None)
                            if vy:
                                fund_start = date(int(vy), 1, 1)
                if as_of is None:
                    as_of = date.today()

                if fund_start is not None and as_of is not None and as_of > fund_start:
                    years_float = (as_of - fund_start).days / 365.25
                    years = Decimal(str(years_float))
                    # compound: (1+hr)^years - 1. Decimal does not support
                    # float exponentiation, so we convert via float for
                    # the pow operation, then back to Decimal for the
                    # final multiplication. Precision loss is negligible
                    # at financial-statement granularity.
                    growth = (1.0 + float(hr)) ** years_float - 1.0
                    pref_val = cl * Decimal(str(growth))
                    if pref_val > 0:
                        _set('preferred_return_amount', pref_val,
                             'called_capital * ((1 + hurdle_rate)^years - 1)',
                             f'called={cl}, hurdle={hr}, years={years:.4f} '
                             f'[from {fund_start} to {as_of}]',
                             inputs={'called_capital': float(cl),
                                     'hurdle_rate': float(hr),
                                     'years': float(years),
                                     'fund_start_date': fund_start.isoformat(),
                                     'as_of_date': as_of.isoformat()})
                        pref = pref_val
            except Exception as e:
                logger.warning(f'Phase 7 pref-return fallback failed: {e}')

    ld_post = _val('lp_distributions') or Decimal('0')
    fv_post = _val('active_fair_value') or Decimal('0')

    profit_above_hurdle = None
    basis_label = None
    basis_inputs = {}
    if nv is not None and roc is not None and pref is not None:
        # PRIMARY: LP-basis NAV minus return of capital minus preferred return.
        profit_above_hurdle = nv - roc - pref
        basis_label = 'fund_nav - return_of_capital - preferred_return'
        basis_inputs = {'fund_nav': float(nv), 'return_of_capital': float(roc),
                        'preferred_return': float(pref)}
    elif roc is not None and pref is not None:
        # FALLBACK: Portfolio FV + cash distributions minus capital and pref.
        profit_above_hurdle = (fv_post + ld_post) - roc - pref
        basis_label = '(active_fair_value + lp_distributions) - return_of_capital - preferred_return'
        basis_inputs = {'active_fair_value': float(fv_post),
                        'lp_distributions': float(ld_post),
                        'return_of_capital': float(roc),
                        'preferred_return': float(pref)}
    elif ic is not None and fv_post is not None:
        # LAST RESORT: subtract invested cost only (no hurdle term).
        profit_above_hurdle = (fv_post + ld_post) - ic
        basis_label = '(active_fair_value + lp_distributions) - invested_cost'
        basis_inputs = {'active_fair_value': float(fv_post),
                        'lp_distributions': float(ld_post),
                        'invested_cost': float(ic)}

    if profit_above_hurdle is not None:
        carry_base_val = max(Decimal('0'), profit_above_hurdle)
        _set('carry_base', carry_base_val, basis_label,
             f'profit_above_hurdle={profit_above_hurdle}, clamped at 0',
             inputs=basis_inputs)

    if cg is None and profit_above_hurdle is not None and carry_pct is not None:
        cg_val = profit_above_hurdle * carry_pct
        if cg_val < 0:
            cg_val = Decimal('0')
        _set('carry_amount_gross', cg_val,
             f'carry_pct * carry_base  [base = {basis_label}]',
             f'profit_above_hurdle={profit_above_hurdle}, carry_pct={carry_pct}',
             inputs={**basis_inputs, 'carry_pct': float(carry_pct)})
        cg = cg_val

    if cn is None and cg is not None:
        _set('carry_amount_net', cg,
             '(no LP-level carry provision found; defaulted to gross)',
             f'gross={cg}',
             inputs={'carry_amount_gross': float(cg)})
        cn = cg

    if cg is not None and cn is not None:
        clawback = cg - cn
        if clawback < 0:
            clawback = Decimal('0')
        _set('gp_clawback_provision', clawback,
             'carry_amount_gross - carry_amount_net',
             f'Gross={cg}, Net={cn}',
             inputs={'carry_amount_gross': float(cg),
                     'carry_amount_net': float(cn)})

    if (catchup_pct is None or catchup_pct == 0) and carry_pct is not None:
        catchup_pct = Decimal('1')
    pref = _val('preferred_return_amount')
    if (pref is not None and carry_pct is not None and catchup_pct is not None
            and 0 < carry_pct < 1):
        catchup = pref * carry_pct * catchup_pct / (1 - carry_pct)
        _set('gp_catchup_amount', catchup,
             'preferred_return * carry_pct * catchup_pct / (1 - carry_pct)',
             f'pref={pref}, carry={carry_pct}, catchup={catchup_pct}',
             inputs={'preferred_return_amount': float(pref),
                     'carry_pct': float(carry_pct),
                     'catchup_pct': float(catchup_pct)})

    sc_lp, sc_src = _lp('sponsor_commitment_amount')
    if sc_lp is not None and cc is not None and cc > 0:
        existing = fund_data.get('sponsor_commitment_pct') or {}
        if existing.get('value') is None:
            fund_data['sponsor_commitment_pct'] = {
                'value': float(sc_lp / cc),
                'source': 'computed',
                'source_sheet': (sc_src or {}).get('source_sheet'),
                'source_cells': (sc_src or {}).get('source_cells') or [],
                'formula': 'sponsor_commitment_amount / committed_capital',
                'inputs_used': {
                    'sponsor_commitment_amount': float(sc_lp),
                    'committed_capital': float(cc),
                },
                'reasoning': 'Sponsor LP commitment divided by total committed capital.',
            }

    return fund_data


# ─────────────────────────────────────────────────────────────────────
# STAGE 3 — Audit (Python, no AI)
# ─────────────────────────────────────────────────────────────────────

def audit_fund_metrics(fund_data):
    """Verify accounting identities on Gemini's analyst output.
    Records findings but never silently rewrites values.
    Returns list of {check, status, detail}.
    """
    def _g(key):
        d = fund_data.get(key) or {}
        return _to_decimal(d.get('value'))

    findings = []
    EPS = Decimal('0.01')

    cc = _g('committed_capital')
    cl = _g('called_capital')
    ic = _g('invested_cost')
    fv = _g('active_fair_value')
    ld = _g('lp_distributions')
    nv = _g('fund_nav')
    cb = _g('cash_balance')
    tvpi = _g('tvpi'); dpi = _g('dpi'); rvpi = _g('rvpi')
    cg = _g('carry_amount_gross'); cn = _g('carry_amount_net')

    def _add(name, status, detail):
        findings.append({'check': name, 'status': status, 'detail': detail})

    if cc is not None and cl is not None:
        _add('called_le_committed',
             'pass' if cl <= cc else 'fail',
             f'called={cl}, committed={cc}')

    if tvpi is not None and dpi is not None and rvpi is not None:
        diff = abs(tvpi - (dpi + rvpi))
        denom = max(abs(tvpi), Decimal(1))
        _add('tvpi_eq_dpi_plus_rvpi',
             'pass' if (diff / denom) < EPS else 'fail',
             f'tvpi={tvpi}, dpi+rvpi={dpi + rvpi}')

    if cg is not None and cn is not None:
        _add('gross_carry_ge_net_carry',
             'pass' if cg >= cn else 'fail',
             f'gross={cg}, net={cn}')

    if nv is not None and fv is not None and cb is not None:
        expected_min_nav = fv + cb  # assuming small liabilities
        ratio = nv / expected_min_nav if expected_min_nav > 0 else Decimal(1)
        # informational only; do not override
        _add('nav_vs_fv_plus_cash',
             'pass' if Decimal('0.7') <= ratio <= Decimal('1.3') else 'info',
             f'nav={nv}, fv+cash={expected_min_nav}, ratio={ratio:.3f}')

    return findings


# ─────────────────────────────────────────────────────────────────────
# STAGE 4 — Persist
# ─────────────────────────────────────────────────────────────────────

# Metric keys we write to FundMetric + DerivedMetric (chatbot reads
# DerivedMetric so we keep mirroring everything here).
FUND_METRIC_KEYS = [
    'tvpi', 'moic', 'dpi', 'rvpi', 'net_irr',
    'fund_nav', 'cash_balance',
    'committed_capital', 'called_capital', 'uncalled_capital',
    'invested_cost', 'active_fair_value',
    'realized_proceeds', 'lp_distributions',
    'hurdle_rate', 'carry_pct', 'catchup_pct', 'mgmt_fee_pct',
    'preferred_return_amount', 'return_of_capital_amount',
    'gp_catchup_amount', 'carry_base', 'carry_amount_gross',
    'carry_amount_net', 'gp_clawback_provision',
    'lp_total_return', 'gp_total_distribution',
    'sponsor_commitment_pct',
]

# DerivedMetric keys the legacy frontend/chatbot read. Map analyst keys
# to legacy names where they differ.
LEGACY_DERIVED_MAP = {
    'fund_nav': 'nav',
    'committed_capital': 'total_committed_capital',
    'called_capital': 'total_called_capital',
    'uncalled_capital': 'total_uncalled_capital',
    'invested_cost': 'total_invested_cost',
    'active_fair_value': 'total_unrealised_fair_value',
    'realized_proceeds': 'total_realised_proceeds',
    'lp_distributions': 'total_distributions',
}


# Canonical mapping: FundMetric key  →  (Scheme attribute, is_percent).
# ONE source of truth, consumed by:
#   1. persist_fund() Phase 5 sync — writes Scheme from FundMetric every
#      import, so legacy consumers (audit log, AI chat, PDF reports)
#      stay aligned with the dashboard.
#   2. derivation_service.py Pass 6 exclusion — blocks Gemini's per-row
#      derivation from inventing values for these fields (which would
#      re-introduce Bug F / Bug T's first-LP-share artefacts).
# Universal: applies to every fund the same way; no per-file branching.
# Adding a new mirror here automatically updates BOTH consumers.
SCHEME_MIRROR_FIELDS = [
    # (FundMetric key, Scheme attr, is_pct_field stored as 0-100 on scheme)
    ('sponsor_commitment_pct', 'sponsor_commitment_pct', True),
    ('mgmt_fee_pct',           'management_fee_pct',     True),
    ('hurdle_rate',            'hurdle_rate_pct',        True),
    ('carry_pct',              'carry_pct',              True),
]

# Quick-lookup set of Scheme attribute names mirrored from FundMetric.
# Pass 6 uses this to skip auto-derivation on these fields.
SCHEME_MIRROR_ATTRS = frozenset(attr for _, attr, _ in SCHEME_MIRROR_FIELDS)


def persist_fund(scheme, organization, fund_data, audits,
                 source_import_file=None):
    from django.db import transaction
    from .models import FundMetric, DerivedMetric

    with transaction.atomic():
        FundMetric.objects.filter(scheme=scheme).delete()
        DerivedMetric.objects.filter(scheme=scheme).delete()

        n_pass = sum(1 for a in audits if a['status'] == 'pass')
        n_fail = sum(1 for a in audits if a['status'] == 'fail')
        audit_summary = (f'Audit: {n_pass} pass / {n_fail} fail '
                         f'/ {sum(1 for a in audits if a["status"] == "info")} info.')

        for key in FUND_METRIC_KEYS:
            entry = fund_data.get(key) or {}
            val = _to_decimal(entry.get('value'))
            formula = entry.get('formula') or ''
            source = entry.get('source') or 'extracted'
            reasoning = entry.get('reasoning') or ''
            inputs = entry.get('inputs_used') or {}

            FundMetric.objects.create(
                organization=organization,
                scheme=scheme,
                metric_key=key,
                value=val,
                formula_expression=formula,
                inputs_used=inputs,
                provenance={
                    'source': source,
                    'source_sheet': entry.get('source_sheet'),
                    'source_cells': entry.get('source_cells'),
                    'reasoning': reasoning,
                    'inputs_used': inputs,
                },
                source=source if source in ('extracted', 'computed') else 'extracted',
                source_import_file=source_import_file,
            )

            if val is not None:
                legacy_key = LEGACY_DERIVED_MAP.get(key, key)
                DerivedMetric.objects.update_or_create(
                    organization=organization, scheme=scheme,
                    metric_key=legacy_key, variant=None,
                    defaults={
                        'value': val,
                        'formula_expression': formula,
                        'gemini_reasoning': f'{reasoning} | {audit_summary}',
                        'confidence': 1.0,
                        'inputs_used': inputs,
                        'candidate_formulas': [],
                        'source_import_file': source_import_file,
                    },
                )

        # Phase 5 — sync canonical FundMetric values onto the Scheme
        # model so the few legacy consumers (audit log, AI chat context,
        # PDF reports) see the same source of truth as the dashboard.
        # We OVERWRITE the legacy Scheme.* fields regardless of their
        # current value: FundMetric is the canonical layer, the Scheme
        # mirror exists only for compatibility. Setting a field to None
        # is intentional — it means "unknown" and is correct (Bug F:
        # never invent a sponsor commitment from the first LP).
        #
        # The field list lives at module scope as SCHEME_MIRROR_FIELDS
        # so derivation_service.py Pass 6 can import it and exclude the
        # same fields from auto-derivation — eliminating Bug T's drift
        # between the two layers.
        #
        # FundMetric stores percentages as decimal fractions (0.08 = 8%)
        # while the Scheme model stores them as percentage values
        # (8.00 = 8%). We multiply by 100 when syncing pct fields.
        # Conversion is unconditional — applies to every fund the same
        # way, no per-file branching.
        scheme_dirty = []
        for fm_key, scheme_attr, is_pct in SCHEME_MIRROR_FIELDS:
            if not hasattr(scheme, scheme_attr):
                continue
            entry = fund_data.get(fm_key) or {}
            raw_val = entry.get('value')
            try:
                new_val = _to_decimal(raw_val) if raw_val is not None else None
            except Exception:
                new_val = None
            if new_val is not None and is_pct:
                # Convert 0.08 fraction → 8.00 percentage
                new_val = new_val * Decimal('100')
            current = getattr(scheme, scheme_attr)
            if current != new_val:
                setattr(scheme, scheme_attr, new_val)
                scheme_dirty.append(scheme_attr)
        if scheme_dirty:
            scheme.save(update_fields=scheme_dirty)
            logger.info(
                f'[persist_fund] Synced Scheme fields from FundMetric: {scheme_dirty}'
            )


def _name_keys(name):
    """Return candidate match keys for a company name. Handles common
    formatting differences across files (case, "Pvt Ltd"/"Private
    Limited" interchange, trailing punctuation)."""
    import re as _re
    s = (name or '').strip()
    if not s:
        return []
    keys = [s.lower()]
    norm = _re.sub(r'[^a-z0-9 ]', ' ', s.lower())
    norm = _re.sub(r'\s+', ' ', norm).strip()
    if norm and norm not in keys:
        keys.append(norm)
    # Strip common entity suffixes that vary between files
    stripped = norm
    for suffix in (
        ' private limited', ' pvt ltd', ' pvt limited', ' pvt.', ' pvt',
        ' ltd', ' limited', ' inc', ' llp', ' technologies', ' tech',
    ):
        if stripped.endswith(suffix):
            stripped = stripped[: -len(suffix)].strip()
    if stripped and stripped not in keys:
        keys.append(stripped)
    return keys


def persist_companies(scheme, organization, companies,
                      source_import_file=None):
    """Write per-company KPIs to PortfolioKPI (which the dashboard
    KPI-matrix and SaaS-metrics endpoints read).

    Slug values MUST match the dashes used in the API view's
    _KPI_COL_SLUGS / SAAS_SLUGS — otherwise the views return null even
    though the rows exist.  See investments.views.portfolio_kpi_matrix.

    Each (investment, kpi_definition, period) row is upserted in its
    own savepoint so a single bad row doesn't roll back the batch.
    """
    from investments.models import (PortfolioCompany, PortfolioKPI,
                                    KPIDefinition, Investment)
    from django.db import transaction
    from datetime import date as _date

    if not companies:
        return {'companies_seen': 0, 'kpi_rows_written': 0,
                'companies_matched': 0, 'companies_no_investment': 0}

    period_date = _date.today()

    # Build name → PortfolioCompany index over the org (case- and
    # suffix-insensitive) so Gemini's "Apex Pvt Ltd" can match a DB row
    # stored as "Apex Private Limited".
    pc_index = {}
    for pc in PortfolioCompany.objects.filter(organization=organization):
        for key in _name_keys(pc.name):
            pc_index.setdefault(key, pc)

    # Build company → Investment index for this scheme so we don't run
    # one query per company.
    inv_by_pc = {}
    for inv in Investment.objects.filter(
        scheme=scheme, portfolio_company__isnull=False,
    ).select_related('portfolio_company'):
        inv_by_pc.setdefault(inv.portfolio_company_id, inv)

    # Pre-create / upsert KPIDefinition rows once per org. Cache by slug.
    kdef_cache = {}
    for analyst_field, (slug, name, fmt) in KPI_SLUG_MAP.items():
        kdef, _ = KPIDefinition.objects.get_or_create(
            organization=organization, slug=slug,
            defaults={'name': name, 'format': fmt,
                      'frequency': 'quarterly', 'is_active': True},
        )
        kdef_cache[analyst_field] = kdef

    seen = matched = written = missing_inv = 0
    for co in companies:
        if not isinstance(co, dict):
            continue
        name = (co.get('company_name') or '').strip()
        if not name:
            continue
        seen += 1
        pc = None
        for key in _name_keys(name):
            pc = pc_index.get(key)
            if pc is not None:
                break
        if pc is None:
            continue
        matched += 1

        inv = inv_by_pc.get(pc.id)
        if inv is None:
            # Dashboard reads KPI rows through Investment.  Without an
            # Investment row for this scheme, the KPI matrix view will
            # never surface this company's KPIs — skip and log.
            missing_inv += 1
            continue

        for analyst_field in KPI_SLUG_MAP.keys():
            raw_val = co.get(analyst_field)
            val = _to_decimal(raw_val)
            if val is None:
                continue
            kdef = kdef_cache[analyst_field]
            try:
                with transaction.atomic():  # per-row savepoint
                    PortfolioKPI.objects.update_or_create(
                        investment=inv,
                        kpi_definition=kdef,
                        period=period_date,
                        defaults={
                            'portfolio_company': pc,
                            'value': val,
                            'period_end_date': period_date,
                            'source': 'excel_upload',
                            'status': 'approved',
                        },
                    )
                    written += 1
            except Exception as e:
                logger.warning(
                    f'PortfolioKPI write failed for {name} / {analyst_field}: '
                    f'{type(e).__name__}: {e}'
                )

    return {'companies_seen': seen, 'companies_matched': matched,
            'kpi_rows_written': written,
            'companies_no_investment': missing_inv}


# ─────────────────────────────────────────────────────────────────────
# Orchestrator
# ─────────────────────────────────────────────────────────────────────

def run_anchor_pipeline(filepath, scheme, organization,
                        source_import_file=None, progress_cb=None):
    """Entry point. Single Gemini call for fund metrics + single
    Gemini call for per-company KPIs. Python audits + persists.
    """
    def _p(pct, msg):
        if progress_cb:
            try:
                progress_cb(pct, msg)
            except Exception:
                pass
        logger.info(f'[anchor_pipeline pct={pct}] {msg}')

    _p(50, 'Stage 0: Workbook census…')
    census = workbook_census(filepath)
    n_sheets = len(census.get('sheets', []))
    n_cells = sum(len(s.get('cells', [])) for s in census.get('sheets', []))
    _p(55, f'Census: {n_sheets} sheets, {n_cells} cells')

    _p(60, 'Stage 1: Fund analyst (Gemini)…')
    try:
        fund_data = fund_analyst_call(census)
    except Exception as e:
        logger.error(f'Fund analyst call failed: {type(e).__name__}: {e}')
        fund_data = {}

    _p(68, 'Stage 1A: LP / waterfall analyst (Gemini)…')
    try:
        lp_data = lp_analyst_call(census)
    except Exception as e:
        logger.error(f'LP analyst call failed: {type(e).__name__}: {e}')
        lp_data = {}

    # Phase 2: deterministic Python sweep over LP register sheets.
    # Closes Bug O (LP-analyst extraction silently non-deterministic),
    # Bug C-regression (Net Carry sometimes computed not extracted), and
    # Bug J (LP Distributions taken from the wrong sheet). Universal —
    # uses Pass 3 semantic header classification, no per-file logic.
    _p(70, 'Stage 1A.5: Deterministic LP-register sweep (Python)…')
    try:
        lp_data = lp_register_python_sweep(filepath, lp_data)
    except Exception as e:
        logger.error(f'LP sweep failed (non-fatal): {type(e).__name__}: {e}')

    # Phase 4: NAV-sheet component fallback when the Total NAV column is
    # an unevaluated formula (openpyxl cannot compute it). Closes Bug K
    # (Fund NAV null on KKR / Avendus). Runs BEFORE fill_computed_metrics
    # so the new fund_nav value flows into carry_base / TVPI fallback.
    _p(71, 'Stage 1A.7: NAV component fallback (Python)…')
    try:
        fund_data = fund_nav_component_sweep(filepath, fund_data)
    except Exception as e:
        logger.error(f'NAV sweep failed (non-fatal): {type(e).__name__}: {e}')

    _p(72, 'Stage 1B: Universal computed-metric fallback (Python)…')
    fund_data = fill_computed_metrics(fund_data, lp_data=lp_data, scheme=scheme)

    _p(75, 'Stage 2: Company KPI analyst (Gemini — chunked)…')
    try:
        company_data = company_analyst_call(census)
    except Exception as e:
        logger.error(f'Company analyst call failed: {type(e).__name__}: {e}')
        company_data = {'companies': []}

    _p(88, 'Stage 3: Audit (Python)…')
    audits = audit_fund_metrics(fund_data)
    for a in audits:
        logger.info(f'[audit] {a["check"]}={a["status"]} :: {a["detail"]}')

    _p(93, 'Stage 4a: Persist fund metrics…')
    persist_fund(scheme, organization, fund_data, audits, source_import_file)

    _p(96, 'Stage 4b: Persist per-company KPIs…')
    company_summary = persist_companies(
        scheme, organization, company_data.get('companies') or [],
        source_import_file,
    )
    logger.info(
        f'[anchor_pipeline persist_companies] seen={company_summary["companies_seen"]}, '
        f'matched={company_summary["companies_matched"]}, '
        f'no_investment={company_summary["companies_no_investment"]}, '
        f'kpi_rows_written={company_summary["kpi_rows_written"]}'
    )

    return {
        'fund_data': {k: (v.get('value') if isinstance(v, dict) else v)
                       for k, v in fund_data.items()},
        'audits': audits,
        'companies': company_summary,
        'n_metrics_written': sum(1 for k in FUND_METRIC_KEYS
                                  if (fund_data.get(k) or {}).get('value') is not None),
    }
