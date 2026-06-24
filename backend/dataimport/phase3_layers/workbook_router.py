"""
Workbook router — classifies each sheet of an Excel workbook into one of
three Phase 3 layers WITHOUT calling Gemini. Runs in milliseconds.

Universal across AIF formats: matches sheet name + first-row text against
domain-specific anchor terms drawn from canonical_schema.SHEET_DOMAINS.
Sheets that fail to classify are assigned to Layer 1 (smallest marginal
output cost; Layer 1's prompt will simply emit `[]` for unrelated content).

Sheet → Layer assignment (universal):

  Layer 1 (Identity):
    fund_scheme_master, organization_users, investors_aml, commitments,
    capital_calls, nav_accounting, nav_calculation, waterfall_carry,
    compliance, fees_register, lp_capital_accounts, fund_pl_bs,
    distributions-side of exits_distributions sheets

  Layer 2 (Universe):
    portfolio_investments, valuations_kpis (per-investment), quoted_unquoted,
    portfolio_hierarchy, exits-side of exits_distributions sheets

  Layer 3 (Time-series):
    financials_pl_bva, burn_runway, per-company KPI matrix sheets
"""

import re
import logging

import openpyxl

logger = logging.getLogger(__name__)


# ── Layer assignment for each canonical domain (universal mapping) ──────────
DOMAIN_TO_LAYER = {
    # Layer 1 — Identity / fund-level
    'fund_scheme_master':    'L1',
    'organization_users':    'L1',
    'investors_aml':         'L1',
    'commitments':           'L1',
    'capital_calls':         'L1',
    'distributions':         'L1',
    'nav_accounting':        'L1',
    'nav_calculation':       'L1',
    'waterfall_carry':       'L1',
    'compliance':            'L1',
    'fees_register':         'L1',
    'lp_capital_accounts':   'L1',
    'fund_pl_bs':            'L1',
    'entities':              'L1',

    # Layer 2 — Investment universe
    'portfolio_investments': 'L2',
    'valuations_kpis':       'L2',
    'quoted_unquoted':       'L2',
    'portfolio_hierarchy':   'L2',
    'exits':                 'L2',

    # Layer 3 — Time-series per company
    'financials_pl_bva':     'L3',
    'burn_runway':           'L3',
    'kpi_matrix':            'L3',
}


# ── Anchor terms per domain (universal — match against sheet name + header) ─
# Order matters for tie-break: first domain whose anchors match wins.
# Tuned to be SPECIFIC, not greedy — e.g., "carry" hits waterfall before
# "portfolio" so a "Carry Schedule" sheet routes correctly.
_DOMAIN_ANCHORS = [
    ('waterfall_carry',     ['waterfall', 'carried interest', 'carry schedule', 'gp economics', 'distribution waterfall', 'hurdle', 'clawback', 'catch-up', 'catchup']),
    ('nav_calculation',     ['nav computation', 'nav calc', 'nav build', 'nav buildup', 'closing nav per unit']),
    ('nav_accounting',      ['nav workings', 'nav record', 'nav walk', 'nav history', 'nav by period', 'monthly nav', 'quarterly nav', 'fund accounting', 'nav & accounting', 'nav and accounting', 'nav accounting', 'nav statement']),
    ('exits',               ['exit register', 'exit event', 'realisation', 'realization', 'exit summary', 'exit ipo', 'realised proceeds', 'realized proceeds']),
    ('distributions',       ['distribution schedule', 'distribution register', 'lp distribution', 'distribution to lp', 'distributions']),
    ('quoted_unquoted',     ['quoted', 'unquoted', 'listed share', 'unlisted share', 'ipev level']),
    ('valuations_kpis',     ['valuation', 'fmv', 'fair value of holding', 'ipev', 'price book']),
    ('portfolio_investments',['portfolio investment', 'portfolio companies', 'investee', 'investments register', 'investment register', 'portfolio register', 'portfolio summary', 'cost & fmv', 'company master']),
    ('portfolio_hierarchy', ['portfolio hierarchy', 'portfolio tree', 'sector segment', 'fund hierarchy']),
    ('capital_calls',       ['capital call', 'drawdown notice', 'drawdown schedule', 'call notice']),
    ('commitments',         ['commitment register', 'lp commitment', 'commitment schedule', 'lp register', 'investor register', 'lp commitments']),
    ('investors_aml',       ['investor master', 'lp master', 'aml', 'kyc', 'investor list', 'limited partner']),
    ('lp_capital_accounts', ['capital account', 'lp account', 'partner account', 'lp statement', 'lp ledger']),
    ('fees_register',       ['fee schedule', 'fee register', 'management fee', 'mgmt fee', 'gst on fee']),
    ('compliance',          ['compliance', 'sebi report', 'sebi filing', 'qar', 'aar', 'compliance calendar', 'ppm amendment', 'compliance test']),
    ('burn_runway',         ['burn', 'runway', 'cash burn', 'mrr', 'arr', 'saas metric', 'churn', 'nrr', 'arpu', 'ltv/cac']),
    ('financials_pl_bva',   ['monthly p&l', 'monthly pl', 'monthly p & l', 'company p&l', 'company pl',
                             'budget vs actual', 'budget v actual', 'monthly mis', 'quarterly mis',
                             'balance sheet', 'cash flow', 'profit & loss', 'profit and loss',
                             'portfolio financials', 'company financials']),
    ('fund_pl_bs',          ['fund p&l', 'fund pl', 'fund balance sheet', 'fund financials',
                             'fund-level p&l', 'fund-level pl', 'consolidated p&l']),
    ('fund_scheme_master',  ['fund master', 'scheme master', 'fund overview', 'cover', 'fund identity',
                             'fund details', 'scheme details', 'lpa terms', 'fund summary']),
    ('organization_users',  ['organization', 'org master', 'user master', 'gp user']),
    ('entities',            ['service entities', 'sponsor', 'trustee', 'custodian', 'auditor entity']),
]


_DATE_HEADER_RE = re.compile(
    r"\b("
    r"jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec"   # month names
    r"|q[1-4]"                                                 # Q1..Q4
    r"|fy\s?\d{2,4}"                                           # FY24, FY 2024
    r"|\d{4}-\d{2}-\d{2}"                                      # ISO dates
    r"|\d{2}/\d{2}/\d{4}"                                      # DD/MM/YYYY
    r")\b", re.IGNORECASE,
)


def _normalise(text) -> str:
    """Lowercase and replace underscores/hyphens with spaces so anchor phrases
    like 'portfolio companies' match sheet names like 'Portfolio_Companies'
    or 'portfolio-companies' equally."""
    if text is None:
        return ''
    s = str(text).lower().strip()
    return s.replace('_', ' ').replace('-', ' ').replace('.', ' ')


def _header_text_of_sheet(ws, scan_rows: int = 12) -> str:
    """Concatenate the first `scan_rows` rows' cell text. Lowercased.

    Scans 12 rows instead of 6 to survive banner / merged-cell / title-band
    layouts where the real header sits at row 7-10 (common in MIS files).
    """
    parts = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= scan_rows:
            break
        for v in row:
            if v is None:
                continue
            parts.append(_normalise(v))
    return ' | '.join(parts)


def _count_date_headers(ws, scan_rows: int = 12) -> int:
    """How many date-like tokens appear in the header rows? > 3 → time-series."""
    text = _header_text_of_sheet(ws, scan_rows=scan_rows)
    return len(_DATE_HEADER_RE.findall(text))


def _classify_domain(sheet_name: str, header_text: str) -> str:
    """Return canonical domain key, or 'unknown'.

    Match priority: sheet-name match (strongest signal) > header-text match.
    First matching domain in _DOMAIN_ANCHORS wins.
    """
    name = _normalise(sheet_name)
    for domain, anchors in _DOMAIN_ANCHORS:
        for a in anchors:
            if a in name:
                return domain
    for domain, anchors in _DOMAIN_ANCHORS:
        for a in anchors:
            if a in header_text:
                return domain
    return 'unknown'


def classify_workbook(filepath: str) -> dict:
    """Walk the workbook once and return a routing plan.

    Returns:
        {
          'L1': [sheet_name, ...],
          'L2': [...],
          'L3': [...],
          'classification_detail': {sheet_name: {'domain': ..., 'layer': ...,
                                                 'rows': N, 'cols': M,
                                                 'time_series_hint': bool}}
        }
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    routing = {'L1': [], 'L2': [], 'L3': []}
    detail = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.max_row or 0
        cols = ws.max_column or 0

        header_text = _header_text_of_sheet(ws)
        date_hits = _count_date_headers(ws)
        time_series_hint = date_hits >= 3 and rows >= 10

        domain = _classify_domain(sheet_name, header_text)
        layer = DOMAIN_TO_LAYER.get(domain, 'L1')   # safe default

        # Time-series boost: any sheet with strong date-header pattern + many
        # rows is per-period data → Layer 3 regardless of the domain inferred
        # from anchor terms. Exceptions: identity-domains (fund_master,
        # entities, organization_users) and valuations_kpis (per-investment
        # snapshots, not per-period) stay in their natural layer.
        _IDENTITY_DOMAINS_STAY = {
            'fund_scheme_master', 'organization_users', 'entities',
            'valuations_kpis',
        }
        if time_series_hint and domain not in _IDENTITY_DOMAINS_STAY:
            layer = 'L3'

        routing[layer].append(sheet_name)
        detail[sheet_name] = {
            'domain': domain,
            'layer': layer,
            'rows': rows,
            'cols': cols,
            'time_series_hint': time_series_hint,
        }

    wb.close()

    n = sum(len(v) for v in routing.values())
    logger.info(
        f'[phase3.router] {n} sheets → L1={len(routing["L1"])} '
        f'L2={len(routing["L2"])} L3={len(routing["L3"])}'
    )

    return {**routing, 'classification_detail': detail}
