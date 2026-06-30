"""
Token estimator + chunk planner — ONE SHEET PER GEMINI CALL.

Universal architecture (2026-06-30 rewrite).

The previous bin-packing strategy piled multiple sheets into one Gemini call
to "minimise API calls". In practice that forced Gemini to juggle N output
target arrays in a single JSON response, and it routinely failed by
SILENTLY DROPPING entire sheets to stay inside its output budget. That
forced the validator to detect drops + retry with row-range splits — but
row-splitting doesn't reduce the SHEET COUNT in the prompt, so the same
multi-sheet drop pattern repeated at every depth, cascading into long
multi-pass runs (5 passes × 32+ chunks on Bharatcrest).

The new contract:

  • Every populated sheet in a layer gets its OWN dedicated Gemini call.
  • Within a single sheet, if est_output > _CHUNK_TARGET, that one sheet is
    row-range split into N sub-chunks (same Flavor B mechanism as before).
  • Empty sheets are skipped (no point sending nothing to Gemini).
  • All chunks across all layers run in PARALLEL via the shared
    ThreadPoolExecutor — see parallel_executor._MAX_WORKERS.

Why this works:
  • Each prompt is small, focused on one target schema → fits in budget,
    no silent multi-sheet drops.
  • sheet_completeness self-reports map 1:1 to a chunk → the validator
    has zero ambiguity about which sheet's numbers are which.
  • Wall time = MAX(single call), not SUM, because all calls run in
    parallel. Bharatcrest 14 sheets → 14 parallel calls finishing in
    roughly the time of the slowest one.

Tuning constants are env-overrideable so ops can adjust without redeploy.
"""

import math
import os
import logging

import openpyxl

logger = logging.getLogger(__name__)


_TOKENS_PER_CELL = float(os.environ.get('PHASE3_TOKENS_PER_CELL', '8.0'))
_FIXED_OVERHEAD = int(os.environ.get('PHASE3_FIXED_OVERHEAD_TOKENS', '3000'))
_CHUNK_TARGET = int(os.environ.get('PHASE3_CHUNK_TARGET_TOKENS', '30000'))


def _per_sheet_stats(filepath: str, sheets: list[str]) -> dict:
    """Walk only the given sheets. Returns {sheet: {'cells': N, 'rows': R}}.

    Reads the workbook through the shared cache — no disk access after first
    load per import. Universal across all funds/formats. The file on disk can
    vanish after the cache is populated and this still works.
    """
    if not sheets:
        return {}
    from .workbook_cache import load_workbook
    cached = load_workbook(filepath)
    available = set(cached['sheets'])
    stats: dict = {}
    for sname in sheets:
        if sname not in available:
            stats[sname] = {'cells': 0, 'rows': 0}
            continue
        cells = 0
        populated_rows = 0
        for row in cached['data'][sname]['rows']:
            row_has_value = False
            for v in row:
                if v is not None and v != '':
                    cells += 1
                    row_has_value = True
            if row_has_value:
                populated_rows += 1
        stats[sname] = {'cells': cells, 'rows': populated_rows}
    return stats


def _estimate_output_tokens(cells: int) -> int:
    return _FIXED_OVERHEAD + int(cells * _TOKENS_PER_CELL)


def estimate_layer_output(filepath: str, sheets: list[str]) -> int:
    """Coarse estimate of output tokens for a layer covering the given sheets."""
    stats = _per_sheet_stats(filepath, sheets)
    total_cells = sum(s['cells'] for s in stats.values())
    return _estimate_output_tokens(total_cells)


def _split_sheet_into_ranges(sheet_name: str, total_rows: int,
                             cells_per_row: float, layer: str) -> list[dict]:
    """Split one large sheet into row-range chunks so each chunk's est
    output stays ≤ _CHUNK_TARGET. Each chunk owns rows [start, end)
    (0-indexed populated-row positions; see orchestrator._serialize_sheets)."""
    est = _estimate_output_tokens(int(cells_per_row * total_rows))
    n = max(2, math.ceil(est / _CHUNK_TARGET))
    rows_per_chunk = max(1, math.ceil(total_rows / n))

    chunks = []
    for i in range(n):
        start = i * rows_per_chunk
        end = min(total_rows, start + rows_per_chunk)
        if start >= end:
            break
        chunks.append({
            'chunk_id': f'{layer}.{sheet_name[:20]}.rows_{start}_{end}',
            'layer': layer,
            'sheets': [sheet_name],
            'row_ranges': {sheet_name: (start, end)},
            'est_output_tokens': _estimate_output_tokens(int(cells_per_row * (end - start))),
            'flavor_b': True,
        })
    return chunks


def plan_chunks(filepath: str, layer: str, sheets: list[str]) -> list[dict]:
    """ONE SHEET PER GEMINI CALL — universal across any AIF format.

    Returns a list of chunk dicts of shape:
      {
        'chunk_id':            str,
        'layer':               str,
        'sheets':              list[str],          # always length 1
        'row_ranges':          dict[sheet_name, (start, end)]  # {} unless row-split
        'est_output_tokens':   int,
        'flavor_b':            bool,               # True only when row-split
      }

    Behaviour:
      • Every populated sheet gets exactly ONE chunk (its own Gemini call).
      • If a single sheet's est_output exceeds _CHUNK_TARGET, that one sheet
        is row-range split into N sub-chunks via _split_sheet_into_ranges.
        Each sub-chunk still owns ONLY that sheet — no cross-sheet mixing.
      • Sheets with zero populated cells are skipped — no useless Gemini call.

    Bin-packing has been removed. Calling Gemini with N sheets in one prompt
    triggered silent per-sheet drops that the row-range split path could not
    fix (halving rows doesn't reduce sheet count). One-sheet-per-call is the
    only structurally safe shape.
    """
    if not sheets:
        return []

    stats = _per_sheet_stats(filepath, sheets)
    chunks: list[dict] = []
    skipped_empty: list[str] = []
    row_split_sheets: list[str] = []

    for sname in sheets:
        s = stats.get(sname, {})
        cells = s.get('cells', 0)
        rows = s.get('rows', 0)

        if cells == 0:
            skipped_empty.append(sname)
            continue

        est = _estimate_output_tokens(cells)

        if est <= _CHUNK_TARGET or rows < 2:
            # Comfortable single-sheet call. No row splitting needed.
            chunks.append({
                'chunk_id': f'{layer}.{sname[:40]}',
                'layer': layer,
                'sheets': [sname],
                'row_ranges': {},
                'est_output_tokens': est,
                'flavor_b': False,
            })
        else:
            # Single sheet too big for one call → row-range split.
            cells_per_row = (cells / rows) if rows else 1.0
            sub_chunks = _split_sheet_into_ranges(sname, rows, cells_per_row, layer)
            chunks.extend(sub_chunks)
            row_split_sheets.append(f'{sname}({len(sub_chunks)})')

    logger.info(
        f'[phase3.token_estimator] {layer}: {len(sheets)} sheet(s) → '
        f'{len(chunks)} chunk(s) '
        f'(1-per-sheet; row-split: {row_split_sheets or "none"}; '
        f'skipped empty: {skipped_empty or "none"})'
    )
    return chunks


def estimate_workbook_input(filepath: str) -> int:
    """Rough total-input token estimate for the entire workbook (all sheets).
    Used for capacity-planning logging only — not for routing decisions.
    """
    from .workbook_cache import load_workbook
    cached = load_workbook(filepath)
    total = 0
    for sname in cached['sheets']:
        for row in cached['data'][sname]['rows']:
            for v in row:
                if v is not None and v != '':
                    total += 1
    return int(total * _TOKENS_PER_CELL)
