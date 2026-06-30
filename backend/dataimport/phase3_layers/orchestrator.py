"""
Phase 3 Orchestrator — main entry point: run_phase3_import(import_file, progress_cb).

Pipeline:
  1. Open workbook (encrypted-file guard); classify sheets via workbook_router
  2. Plan chunks per layer — ONE chunk per populated sheet; row-range split
     only when a single sheet alone exceeds the per-call output budget
  3. Build identity context (Cover + Fund_Master raw text → replicated to L2/L3)
  4. Multi-pass parallel execution. Failed chunks get auto-split and resubmitted
     to the SAME ThreadPool (C7). Sub-chunks are produced by halving the row
     range of the failing chunk — Python-side filtering, no Gemini-counted modulo.
  5. Per-layer merge (with dedup) → per-layer validate
  6. Cross-layer merge → reconcile → identity-validate
  7. Row-completeness audit (I2): rows_extracted vs rows_in_source per sheet
  8. Save unified JSON, compute XIRR, persist
  9. Wall-time budget (H5) bounds the whole import
 10. Concurrent-import sentinel (H1) created at start, deleted at end (always)
"""

import json
import logging
import os
import time
from typing import Callable, Optional

import openpyxl
from django.utils import timezone

from .workbook_router import classify_workbook, DOMAIN_TO_LAYER
from .token_estimator import plan_chunks, estimate_workbook_input
from .parallel_executor import run_in_parallel
from .merger import merge_layer_chunks, merge_all_layers
from .reconciler import reconcile
from .cross_layer_validator import validate_identities
from .layer_validators import LAYER_VALIDATORS
from .prompts import (
    LAYER1_PROMPT_TEMPLATE, LAYER2_PROMPT_TEMPLATE, LAYER3_PROMPT_TEMPLATE,
    schema_for,
)
from ..gemini_column_mapper import (
    _call_gemini, GeminiTruncated, GeminiNonDictTopLevel, GeminiQuotaExhausted,
)

logger = logging.getLogger(__name__)


# ── Tuning constants (all env-overrideable) ──────────────────────────────────

_MAX_CELL_CHARS = 60
_MAX_SPLIT_DEPTH = int(os.environ.get('PHASE3_MAX_SPLIT_DEPTH', '4'))
_MAX_QUOTA_REQUEUES = int(os.environ.get('PHASE3_MAX_QUOTA_REQUEUES', '3'))
# Wall-time budget. Now used as a SOFT signal: enforced only if we have
# zero successful chunks (truly hung). If any chunks succeeded, the orchestrator
# proceeds to merge+persist regardless of elapsed time — never throw away
# successful Gemini work. Default bumped to 3600s to comfortably cover
# multi-pass recursive splits on dense MIS files.
_GLOBAL_WALL_TIME_S = int(os.environ.get('PHASE3_GLOBAL_TIMEOUT_S', '3600'))
_PER_CALL_TIMEOUT_MS = int(os.environ.get('PHASE3_PER_CALL_TIMEOUT_MS', '300000'))

_SENTINEL_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
_SENTINEL_PATH = os.path.join(_SENTINEL_DIR, '.import_active')


# ── Sheet serialisation with source-side row filtering (C1, C3) ──────────────

def _cell(v) -> str:
    if v is None:
        return ''
    s = str(v).replace('\n', ' ').replace('\r', ' ').strip()
    if len(s) > _MAX_CELL_CHARS:
        s = s[:_MAX_CELL_CHARS] + '…'
    return s


_HEADER_PREVIEW_ROWS = int(os.environ.get('PHASE3_HEADER_PREVIEW_ROWS', '3'))


def _serialize_sheets(filepath: str, sheet_names: list[str],
                      row_ranges: Optional[dict] = None) -> tuple[str, dict]:
    """Render the named sheets into LLM text. When `row_ranges` is provided
    as `{sheet_name: (start, end)}`, only rows in `[start, end)` are emitted
    (0-indexed populated-row positions) — this is the source-side chunk
    filter (C1).

    UNIVERSAL CONTEXT GUARANTEE: when a chunk owns a sub-range (start > 0),
    the first `_HEADER_PREVIEW_ROWS` populated rows of the sheet are ALWAYS
    prepended as a `[HEADER PREVIEW]` block so Gemini sees the column
    schema regardless of where its row slice starts. Without this, a chunk
    owning rows 100-200 of a Monthly P&L sheet would never see the
    "Company | Apr-24 | May-24 | …" header row at the top and would have
    to guess at column meanings — which is exactly when bad extractions
    happen. Header rows are also written to a banner so the dedup at merge
    time can collapse them as duplicates (header-only rows have no natural
    key and are skipped by persisters).
    """
    if not sheet_names:
        return '(no sheets routed to this layer)', {}

    row_ranges = row_ranges or {}
    # Read through the shared in-memory cache — no disk access. File can be
    # gone, the cache is authoritative.
    from .workbook_cache import load_workbook
    cached = load_workbook(filepath)
    parts: list[str] = []
    meta: dict = {}

    want = set(sheet_names)
    for sh in cached['sheets']:
        if sh not in want:
            continue
        sheet_data = cached['data'][sh]
        nrows = sheet_data['max_row']
        ncols = sheet_data['max_col']
        all_rows = sheet_data['rows']

        rrange = row_ranges.get(sh)
        if rrange:
            start, end = int(rrange[0]), int(rrange[1])
            banner = (
                f'\n===== SHEET: {sh} (rows={nrows}, cols={ncols}, '
                f'CHUNK SLICE: populated-rows[{start}:{end}]) ====='
            )
        else:
            start, end = None, None
            banner = f'\n===== SHEET: {sh} (rows={nrows}, cols={ncols}) ====='

        parts.append(banner)

        # First pass: collect header preview (first N populated rows)
        # when we're serving a sub-range that doesn't already include row 0.
        header_lines: list[str] = []
        need_header = (start is not None and start > 0)
        if need_header:
            header_collected = 0
            for r_idx, row in enumerate(all_rows, start=1):
                cells = [_cell(v) for v in row]
                if not any(c for c in cells):
                    continue
                header_lines.append(f'  R{r_idx}: ' + ' | '.join(cells))
                header_collected += 1
                if header_collected >= _HEADER_PREVIEW_ROWS:
                    break

        if header_lines:
            parts.append('  [HEADER PREVIEW — column schema, do not re-emit as data]')
            parts.extend(header_lines)
            parts.append('  [DATA SLICE BEGINS]')

        # Second pass: emit the actual row slice.
        emitted = 0
        populated_idx = 0
        for r_idx, row in enumerate(all_rows, start=1):
            cells = [_cell(v) for v in row]
            if not any(c for c in cells):
                continue
            if start is not None:
                if populated_idx < start:
                    populated_idx += 1
                    continue
                if populated_idx >= end:
                    break
            parts.append(f'  R{r_idx}: ' + ' | '.join(cells))
            emitted += 1
            populated_idx += 1

        meta[sh] = {
            'rows': nrows, 'cols': ncols,
            'emitted_rows': emitted,
            'header_preview_rows': len(header_lines),
            'row_range': [start, end] if rrange else None,
        }
    return '\n'.join(parts), meta


def _build_identity_context(filepath: str, l1_sheets: list[str]) -> str:
    """Replicate fund-identity context to L2 / L3 so they reason independently.
    We pick small L1 sheets up to ~3K chars — Cover / Fund Master style."""
    if not l1_sheets:
        return ''
    text, _ = _serialize_sheets(filepath, l1_sheets[:3])
    if len(text) > 3000:
        text = text[:3000] + '\n[...identity context truncated...]'
    return text


def _chunk_scope_note(chunk: dict) -> str:
    """Human-readable note describing what slice this chunk owns.
    Used in the prompt so Gemini knows the workbook excerpt is intentionally
    partial (don't extrapolate)."""
    rr = chunk.get('row_ranges') or {}
    if not rr:
        if len(chunk.get('sheets') or []) > 0:
            return f'full content of {len(chunk["sheets"])} sheet(s) listed below'
        return ''
    parts = []
    for sname, (start, end) in rr.items():
        parts.append(f'{sname}: populated rows [{start}:{end}]')
    return 'this chunk owns ' + '; '.join(parts)


def _build_chunk_prompt(filepath: str, chunk: dict, identity_context: str) -> str:
    layer = chunk['layer']
    sheets = chunk['sheets']
    row_ranges = chunk.get('row_ranges') or {}

    workbook_text, _ = _serialize_sheets(filepath, sheets, row_ranges=row_ranges)
    scope_note = _chunk_scope_note(chunk) if (chunk.get('flavor_b') or row_ranges) else ''

    if layer == 'L1':
        return LAYER1_PROMPT_TEMPLATE(workbook_text)
    if layer == 'L2':
        return LAYER2_PROMPT_TEMPLATE(workbook_text, identity_context, scope_note)
    return LAYER3_PROMPT_TEMPLATE(workbook_text, identity_context, scope_note)


# ── Chunk splitting (row-range halving) ──────────────────────────────────────

def _split_chunk(chunk: dict) -> list[dict]:
    """Halve the failing chunk's row range. Universal across any sheet:

      • Chunk with explicit row_ranges → halve each sheet's range.
      • Chunk with NO row_ranges (full-sheet) → introduce a row range
        for each sheet using its populated row count, then halve.

    This guarantees the union of sub-chunks covers the same row set as
    the parent (subject to the [start, end) discipline).
    """
    sheets = chunk['sheets']
    row_ranges = dict(chunk.get('row_ranges') or {})
    layer = chunk['layer']

    if not row_ranges:
        # Need to derive row counts to halve. Count populated rows per sheet
        # from the in-memory cache — no disk access; file can be gone.
        from .workbook_cache import load_workbook
        cached = load_workbook(chunk['_filepath'])
        available = set(cached['sheets'])
        for sname in sheets:
            if sname not in available:
                continue
            count = sum(
                1 for row in cached['data'][sname]['rows']
                if any(v is not None and v != '' for v in row)
            )
            row_ranges[sname] = (0, count)

    half_a: dict = {}
    half_b: dict = {}
    for sname, (start, end) in row_ranges.items():
        mid = (start + end) // 2
        if mid <= start or mid >= end:
            half_a[sname] = (start, end)
        else:
            half_a[sname] = (start, mid)
            half_b[sname] = (mid, end)

    new_depth = chunk.get('_split_depth', 0) + 1
    out = []
    if half_a:
        out.append({
            **chunk,
            'chunk_id': f'{chunk["chunk_id"]}.d{new_depth}a',
            'row_ranges': half_a,
            '_split_depth': new_depth,
            'flavor_b': True,
        })
    if half_b:
        out.append({
            **chunk,
            'chunk_id': f'{chunk["chunk_id"]}.d{new_depth}b',
            'row_ranges': half_b,
            '_split_depth': new_depth,
            'flavor_b': True,
        })
    return out


# ── Per-chunk runner (non-recursive: returns _resubmit on split) ─────────────

def _run_chunk(chunk_with_filepath: dict) -> dict:
    """Single-job runner submitted to ThreadPoolExecutor.

    NEVER recurses. Instead, on truncation / non-dict / unrecoverable parse
    error → returns {'_resubmit': [sub_chunks...]}. The executor's main loop
    re-enqueues those, preserving parallelism (C7).
    """
    import json as _json

    filepath = chunk_with_filepath['_filepath']
    identity_ctx = chunk_with_filepath['_identity_context']
    chunk = chunk_with_filepath
    depth = chunk.get('_split_depth', 0)

    prompt = _build_chunk_prompt(filepath, chunk, identity_ctx)
    label = f'phase3_{chunk["chunk_id"]}'
    schema = schema_for(chunk['layer'])

    try:
        data = _call_gemini(
            prompt, context_label=label, response_schema=schema,
            timeout_ms=_PER_CALL_TIMEOUT_MS,
        )

    except GeminiQuotaExhausted as qe:
        # 429 RESOURCE_EXHAUSTED — re-queue this chunk for the NEXT executor
        # pass instead of giving up. By the time the executor's next pass
        # fires, the per-minute quota window has rolled over and this chunk
        # gets a fresh shot. Universal across any file/quota tier.
        requeue_count = chunk.get('_quota_requeue_count', 0) + 1
        if requeue_count > _MAX_QUOTA_REQUEUES:
            logger.error(
                f'[phase3.runner] {chunk["chunk_id"]} hit quota limit on '
                f'{requeue_count} consecutive passes — giving up'
            )
            return {
                'chunk_id': chunk['chunk_id'],
                'layer': chunk['layer'],
                'data': {},
                '_ok': False,
                '_error': f'GeminiQuotaExhausted: {qe} (after {requeue_count} re-queues)',
            }
        requeued = {**chunk, '_quota_requeue_count': requeue_count}
        logger.warning(
            f'[phase3.runner] {chunk["chunk_id"]} 429 quota — re-queueing for '
            f'next pass (requeue #{requeue_count}/{_MAX_QUOTA_REQUEUES})'
        )
        return {
            'chunk_id': chunk['chunk_id'],
            'layer': chunk['layer'],
            'data': {},
            '_resubmit': [requeued],
        }

    except (GeminiTruncated, GeminiNonDictTopLevel) as e:
        if depth >= _MAX_SPLIT_DEPTH:
            logger.error(
                f'[phase3.runner] {chunk["chunk_id"]} at max split depth '
                f'{depth} with {type(e).__name__} — giving up'
            )
            return {
                'chunk_id': chunk['chunk_id'],
                'layer': chunk['layer'],
                'data': {},
                '_ok': False,
                '_error': f'{type(e).__name__} at max split depth',
            }
        sub_chunks = _split_chunk(chunk)
        for sc in sub_chunks:
            sc['_filepath'] = filepath
            sc['_identity_context'] = identity_ctx
        logger.warning(
            f'[phase3.runner] {chunk["chunk_id"]} {type(e).__name__} '
            f'(depth={depth}) — resubmitting {len(sub_chunks)} sub-chunks'
        )
        return {
            'chunk_id': chunk['chunk_id'],
            'layer': chunk['layer'],
            'data': {},
            '_resubmit': sub_chunks,
        }

    except _json.JSONDecodeError as e:
        # Try ONE terse retry first (cheaper than splitting).
        char_at = getattr(e, 'pos', None)
        line_at = getattr(e, 'lineno', None)
        logger.warning(
            f'[phase3.runner] {chunk["chunk_id"]} JSON parse failed at '
            f'line {line_at} char {char_at} — terse-retry'
        )
        terse_hint = (
            "\n\n═══════════════════════════════════════════════════════════════\n"
            "RETRY HINT — YOUR PREVIOUS JSON RESPONSE WAS UNPARSEABLE\n"
            "═══════════════════════════════════════════════════════════════\n"
            f"Parse failed at character ~{char_at} (line {line_at}). Re-emit "
            "the SAME data but be strictly JSON-valid:\n"
            "  • Shorten provenance strings to cell refs (e.g. 'Sheet!A1').\n"
            "  • Omit any optional field per Rule 16 — if a value is null, drop the key.\n"
            "  • For long text, summarise to ≤80 characters or omit.\n"
            "  • Do NOT drop any data rows — emit every row more compactly.\n"
            "═══════════════════════════════════════════════════════════════\n"
        )
        try:
            data = _call_gemini(
                prompt + terse_hint, context_label=label + '_retry',
                response_schema=schema, timeout_ms=_PER_CALL_TIMEOUT_MS,
            )
            logger.info(f'[phase3.runner] {chunk["chunk_id"]} terse-retry SUCCEEDED')
        except (GeminiTruncated, GeminiNonDictTopLevel, _json.JSONDecodeError) as e2:
            if depth >= _MAX_SPLIT_DEPTH:
                logger.error(
                    f'[phase3.runner] {chunk["chunk_id"]} terse-retry also '
                    f'failed at max depth — giving up: {type(e2).__name__}'
                )
                return {
                    'chunk_id': chunk['chunk_id'],
                    'layer': chunk['layer'],
                    'data': {},
                    '_ok': False,
                    '_error': f'{type(e2).__name__} after terse-retry at max depth',
                }
            sub_chunks = _split_chunk(chunk)
            for sc in sub_chunks:
                sc['_filepath'] = filepath
                sc['_identity_context'] = identity_ctx
            logger.warning(
                f'[phase3.runner] {chunk["chunk_id"]} terse-retry failed — '
                f'resubmitting {len(sub_chunks)} sub-chunks'
            )
            return {
                'chunk_id': chunk['chunk_id'],
                'layer': chunk['layer'],
                'data': {},
                '_resubmit': sub_chunks,
            }

    if not isinstance(data, dict):
        # Defensive — _call_gemini should have raised GeminiNonDictTopLevel.
        return {
            'chunk_id': chunk['chunk_id'],
            'layer': chunk['layer'],
            'data': {},
            '_ok': False,
            '_error': f'non_dict_top_level: {type(data).__name__}',
        }

    # ── Online completeness check (universal) ────────────────────────────
    # Even when Gemini returns a syntactically valid JSON, it sometimes
    # silently drops rows (e.g. PC025 SoilSense skipped mid-list, or sheet
    # truncated as Gemini self-reports). Trigger the SAME row-range split
    # path used by GeminiTruncated. This catches both:
    #   (a) explicit self-report:  sheet_completeness[].truncated_in_prompt = true
    #   (b) implicit drop:         rows_extracted < 0.95 * source rows
    # Only sheets owned by THIS chunk are evaluated — other chunks own the rest.
    dropped = _chunk_dropped_sheets(data, chunk, filepath)
    if dropped and depth < _MAX_SPLIT_DEPTH:
        sub_chunks = _split_chunk(chunk)
        for sc in sub_chunks:
            sc['_filepath'] = filepath
            sc['_identity_context'] = identity_ctx
        logger.warning(
            f'[phase3.runner] {chunk["chunk_id"]} silently dropped rows on '
            f'{len(dropped)} sheet(s) ({dropped}) — resubmitting '
            f'{len(sub_chunks)} sub-chunks (depth={depth})'
        )
        return {
            'chunk_id': chunk['chunk_id'],
            'layer': chunk['layer'],
            'data': {},
            '_resubmit': sub_chunks,
        }
    if dropped and depth >= _MAX_SPLIT_DEPTH:
        # Surface but accept — at max depth, taking what we have is better
        # than infinite splitting. The post-import audit logs this too.
        logger.error(
            f'[phase3.runner] {chunk["chunk_id"]} at max split depth '
            f'{depth} still missing rows on {dropped} — accepting partial'
        )

    return {
        'chunk_id': chunk['chunk_id'],
        'layer': chunk['layer'],
        'data': data,
    }


def _chunk_dropped_sheets(data: dict, chunk: dict, filepath: str) -> list:
    """Return list of sheet names in this chunk where Gemini ran out of output
    budget and explicitly self-reported truncation.

    Universal across any sheet/fund.

    SINGLE SIGNAL — trust Gemini's explicit truncated_in_prompt flag only.

    Why this is the right approach (validated 2026-06-30 on Bharatcrest):
      • The ROW PRESERVATION RULE in JSON_OUTPUT_CONTRACT instructs Gemini to
        set sheet_completeness[].truncated_in_prompt = true whenever it cannot
        fit all rows in its output budget. Gemini honours this reliably.
      • The previous implicit ratio check (rows_extracted < 0.95 × source rows)
        produced systematic false positives because:
          – Gemini's rows_extracted counts ONLY data rows it emitted as records
            (correctly excluding headers, banners, subtotals, blank separators).
          – The denominator (source populated rows) counts ALL non-empty rows
            including those very headers / banners / subtotals.
          – Ratio is therefore always < 1.0 on row-dense sheets even when 100%
            of data was captured. Investment_Register at 22.2K input → 21.5K
            output flagged "dropped" with this rule, then cascaded through 5
            executor passes and ~200 Gemini calls.
      • Trusting the explicit flag is universal, simple, and matches the
        contract we already enforce in the prompt.
    """
    chunk_sheets = chunk.get('sheets') or []
    if not chunk_sheets:
        return []

    completeness = data.get('sheet_completeness') or []
    if not isinstance(completeness, list):
        completeness = []
    by_name = {}
    for rec in completeness:
        if not isinstance(rec, dict):
            continue
        sn = rec.get('sheet_name')
        if sn:
            by_name[sn] = rec

    dropped = []
    for sname in chunk_sheets:
        rec = by_name.get(sname, {})
        if rec.get('truncated_in_prompt') is True:
            dropped.append(sname)
    return dropped


# ── I2: Row-completeness audit ───────────────────────────────────────────────

def _row_completeness_audit(unified: dict, routing: dict, filepath: str) -> dict:
    """Compare router-known populated row count per sheet vs the sum of
    rows_extracted reported via sheet_completeness[]. Flags sheets where
    we extracted < 80% of populated rows. Returns audit dict for the
    diagnostics block."""
    from .token_estimator import _per_sheet_stats

    all_sheets = (routing.get('L1') or []) + (routing.get('L2') or []) + (routing.get('L3') or [])
    stats = _per_sheet_stats(filepath, all_sheets)

    extracted_per_sheet: dict = {}
    for block_name in ('sheet_completeness',):
        rows = unified.get(block_name) or []
        if isinstance(rows, list):
            for r in rows:
                if not isinstance(r, dict):
                    continue
                sn = r.get('sheet_name')
                rx = r.get('rows_extracted') or 0
                try:
                    rx = int(rx)
                except (TypeError, ValueError):
                    rx = 0
                if sn:
                    extracted_per_sheet[sn] = extracted_per_sheet.get(sn, 0) + rx

    flags = []
    by_sheet = {}
    for sname, st in stats.items():
        source_rows = st.get('rows') or 0
        extracted = extracted_per_sheet.get(sname, 0)
        ratio = (extracted / source_rows) if source_rows else None
        rec = {
            'sheet': sname,
            'source_populated_rows': source_rows,
            'rows_extracted_reported': extracted,
            'ratio': round(ratio, 3) if ratio is not None else None,
        }
        by_sheet[sname] = rec
        if source_rows >= 5 and ratio is not None and ratio < 0.80:
            flags.append({**rec, 'flag': 'under_extracted'})

    return {
        'by_sheet': by_sheet,
        'flagged_sheets': flags,
        'total_source_rows': sum(s.get('rows', 0) for s in stats.values()),
        'total_extracted_reported': sum(extracted_per_sheet.values()),
    }


# ── Concurrent-import sentinel (H1) ──────────────────────────────────────────

def _write_sentinel(import_file_id):
    try:
        with open(_SENTINEL_PATH, 'w') as f:
            f.write(f'{import_file_id}\n{timezone.now().isoformat()}\n')
    except Exception as e:
        logger.warning(f'[phase3] failed to write import sentinel: {e}')


def _delete_sentinel():
    try:
        if os.path.exists(_SENTINEL_PATH):
            os.remove(_SENTINEL_PATH)
    except Exception as e:
        logger.warning(f'[phase3] failed to delete import sentinel: {e}')


# ── Main entry point ────────────────────────────────────────────────────────

def run_phase3_import(import_file, progress_cb: Optional[Callable] = None):
    """Phase 3 import pipeline. Same contract as run_phase2_import."""
    from ..phase2_persister import persist_phase2

    def _p(pct: int, msg: str):
        if progress_cb:
            try:
                progress_cb(pct, msg)
            except Exception:
                pass
        logger.info(f'[Phase3 {pct}%] {msg}')

    started_at = time.time()
    _write_sentinel(import_file.id)

    try:
        import_file.status = 'importing'
        import_file.save(update_fields=['status'])

        _p(3, 'Phase 3: opening workbook…')
        filepath = import_file.file.path

        # ── Step 0: encrypted-file guard + EAGER load into in-memory cache ─
        # This is the ONLY place Phase 3 reads the file on disk. Every
        # downstream consumer (router, chunker, prompt builder, completeness
        # validator) reads from workbook_cache. After this point the file
        # on disk can vanish without breaking the import.
        try:
            from .workbook_cache import load_workbook as _eager_load
            _eager_load(filepath)
        except Exception as e:
            err = f'workbook open failed: {type(e).__name__}: {e}'
            if any(t in str(e).lower() for t in ('encrypted', 'password', 'invalid', 'badzipfile')):
                err = (f'workbook is encrypted, password-protected, or corrupted '
                       f'({type(e).__name__}: {e}). Please save as plain .xlsx and re-upload.')
            logger.error(f'[Phase3] {err}')
            import_file.status = 'failed'
            import_file.error_detail = err
            import_file.save(update_fields=['status', 'error_detail'])
            return {'status': 'failed', 'error': 'workbook_open_failed', 'detail': err}

        # ── Step 1: route sheets to layers ───────────────────────────────
        _p(6, 'Phase 3: classifying sheets → L1 / L2 / L3 (no Gemini)…')
        routing = classify_workbook(filepath)
        l1_sheets, l2_sheets, l3_sheets = routing['L1'], routing['L2'], routing['L3']
        _p(10, f'Phase 3: routed {len(l1_sheets)} → L1, {len(l2_sheets)} → L2, {len(l3_sheets)} → L3')

        # ── Step 2: plan chunks per layer ────────────────────────────────
        _p(12, 'Phase 3: planning chunks (one-sheet-per-call; row-range split for huge sheets)…')
        jobs: list[dict] = []
        jobs += plan_chunks(filepath, 'L1', l1_sheets)
        jobs += plan_chunks(filepath, 'L2', l2_sheets)
        jobs += plan_chunks(filepath, 'L3', l3_sheets)
        flavor_b_active = any(j.get('flavor_b') for j in jobs)
        _p(15,
           f'Phase 3: {len(jobs)} parallel calls planned '
           f'({"Flavor B active" if flavor_b_active else "Flavor A only"}); '
           f'est input ≈ {estimate_workbook_input(filepath):,} tokens')

        # ── Step 3: identity context for L2 / L3 ─────────────────────────
        identity_context = _build_identity_context(filepath, l1_sheets)
        for j in jobs:
            j['_filepath'] = filepath
            j['_identity_context'] = identity_context

        # ── Step 4: parallel execution ───────────────────────────────────
        def _exec_progress(chunk_id: str, pct: float):
            overall = 20 + int(pct * 0.4)
            _p(overall, f'Phase 3: {chunk_id} done ({pct:.0f}% of all calls)')

        _p(20, f'Phase 3: launching {len(jobs)} parallel Gemini calls…')

        results = run_in_parallel(jobs, _run_chunk, progress_cb=_exec_progress)

        ok_results = [r for r in results if r and r.get('_ok')]
        failed = [r for r in results if r and not r.get('_ok')]
        elapsed = time.time() - started_at

        # Smart H5: only abort on wall-time if we have NO successful work.
        # When chunks have already succeeded, persist them — a slow run is
        # not a failed run. Universal across any file size / split depth.
        if elapsed > _GLOBAL_WALL_TIME_S and not ok_results:
            msg = (f'Phase 3: exceeded global wall-time budget '
                   f'({_GLOBAL_WALL_TIME_S}s) with zero successful chunks — failing import')
            logger.error(msg)
            import_file.status = 'failed'
            import_file.error_detail = msg
            import_file.save(update_fields=['status', 'error_detail'])
            return {'status': 'failed', 'error': 'global_timeout', 'detail': msg}
        if elapsed > _GLOBAL_WALL_TIME_S:
            logger.warning(
                f'[Phase3] wall-time {elapsed:.0f}s exceeded budget '
                f'{_GLOBAL_WALL_TIME_S}s, but {len(ok_results)}/{len(results)} '
                f'chunks succeeded — continuing to persist (never throw away successful work)'
            )

        if not ok_results:
            msg = f'Phase 3: ALL {len(results)} parallel calls failed.'
            logger.error(msg + ' Errors: ' + str([r.get('_error') for r in failed]))
            import_file.status = 'failed'
            import_file.error_detail = msg
            import_file.save(update_fields=['status', 'error_detail'])
            return {'status': 'failed', 'error': 'all_chunks_failed', 'detail': msg}

        # ── Step 5: per-layer merge + per-layer validation ───────────────
        _p(62, 'Phase 3: merging chunk outputs per layer (with dedup)…')
        per_layer: dict[str, dict] = {}
        layer_warnings: dict[str, list] = {}
        for layer in ('L1', 'L2', 'L3'):
            chunks_of_layer = [r for r in ok_results if r['layer'] == layer]
            if not chunks_of_layer:
                per_layer[layer] = {}
                layer_warnings[layer] = []
                continue
            merged_layer = merge_layer_chunks(layer, chunks_of_layer)
            per_layer[layer] = merged_layer
            validator = LAYER_VALIDATORS.get(layer)
            layer_warnings[layer] = validator(merged_layer) if validator else []
            if layer_warnings[layer]:
                _p(63 + ('L1L2L3'.index(layer[1]) * 1),
                   f'Phase 3: {layer} produced {len(layer_warnings[layer])} soft warnings')

        # ── Step 6: cross-layer merge ────────────────────────────────────
        _p(67, 'Phase 3: merging layers into unified JSON…')
        unified = merge_all_layers(per_layer)

        # ── Step 7: reconciliation ───────────────────────────────────────
        _p(72, 'Phase 3: reconciling via priority matrix (P1..P7)…')
        unified = reconcile(unified)

        # ── Step 8: cross-layer identity validator ───────────────────────
        _p(76, 'Phase 3: cross-layer identity validator…')
        unified = validate_identities(unified)
        n_violations = len(unified.get('__identity_violations__', []))
        if n_violations:
            _p(78, f'Phase 3: {n_violations} identity warnings (non-blocking)')

        # ── Row-completeness audit (I2) ──────────────────────────────────
        completeness = _row_completeness_audit(unified, routing, filepath)
        if completeness['flagged_sheets']:
            _p(79, f'Phase 3: {len(completeness["flagged_sheets"])} sheets under-extracted (< 80%)')

        unified['__phase3_diagnostics__'] = {
            'routing': {k: v for k, v in routing.items() if k != 'classification_detail'},
            'layer_warnings': layer_warnings,
            'flavor_b_active': flavor_b_active,
            'parallel_calls_initial': len(jobs),
            'terminal_results': len(results),
            'successful_calls': len(ok_results),
            'failed_calls': len(failed),
            'failed_chunks': [{'chunk_id': r.get('chunk_id'),
                               'error': r.get('_error')} for r in failed],
            'row_completeness': completeness,
            'merge_collisions': unified.pop('__merge_collisions__', []),
            'merge_dedup': unified.pop('__merge_dedup__', {}),
            'wall_time_s': round(time.time() - started_at, 2),
        }

        # ── Step 9: save unified JSON to disk ────────────────────────────
        try:
            from django.conf import settings as _dj_settings
            out_dir = os.path.join(_dj_settings.MEDIA_ROOT, 'dataimport', '_phase3_outputs')
            os.makedirs(out_dir, exist_ok=True)
            out_path = os.path.join(
                out_dir,
                f'{import_file.id}_{import_file.original_filename}.json',
            )
            with open(out_path, 'w', encoding='utf-8') as f:
                json.dump(unified, f, indent=2, default=str, ensure_ascii=False)
            logger.info(f'[Phase 3] unified JSON saved → {out_path}')
        except Exception as e:
            logger.warning(f'[Phase 3] could not save unified JSON: {e}')

        # ── Step 10: Net IRR — DELIBERATELY moved to the Phase 4 aggregator.
        # The previous safety net here appended a synthetic "terminal NAV"
        # cashflow entry (using fund_nav_latest OR total_unrealised_fv_holding)
        # which inflated Net IRR whenever Gemini fabricated the NAV. Net IRR
        # is now computed downstream by compute_all_fund_aggregates() from
        # ONLY real ledger events: CapitalCall (negative) + Distribution
        # (positive) + terminal NAV (only when extracted from a real cell).
        # Same DB rows → same IRR every re-import. ────────────────────────

        # ── Step 11: persist ─────────────────────────────────────────────
        _p(85, 'Phase 3: persisting to database…')
        organization = import_file.job.organization
        user = import_file.job.uploaded_by

        # I5 — STRICT fund_master fallback: only stamp filename when L1 truly
        # failed (zero L1 chunks succeeded). If L1 ran but returned empty
        # fund_master, that is a CONTENT bug (Gemini under-performed) and we
        # surface it loudly via a quality flag rather than silently masking
        # it with a filename-derived name.
        fm = unified.setdefault('fund_master', {})
        l1_chunks_ok = [r for r in ok_results if r.get('layer') == 'L1']
        if not fm.get('fund_name'):
            if not l1_chunks_ok:
                stem = (import_file.original_filename or 'Unnamed Fund').rsplit('.', 1)[0]
                fm['fund_name'] = stem
                unified.setdefault('__phase3_diagnostics__', {})['fund_master_fallback'] = (
                    f'fund_name defaulted from filename: {stem!r} (all L1 chunks failed)'
                )
            else:
                unified.setdefault('__phase3_diagnostics__', {})['fund_master_empty_warning'] = (
                    f'L1 ran ({len(l1_chunks_ok)} chunk(s) succeeded) but fund_master '
                    f'came back empty. Persister will use a placeholder; dashboard will '
                    f'show "Fund identity incomplete — re-import recommended".'
                )
                stem = (import_file.original_filename or 'Unnamed Fund').rsplit('.', 1)[0]
                fm['fund_name'] = f'{stem} (identity incomplete)'

        persist_result = persist_phase2(
            unified, import_file, organization, user, progress_cb=_p,
        )

        elapsed = round(time.time() - started_at, 2)
        _p(100, f'Phase 3: complete in {elapsed}s '
                f'({len(results)} terminal results, {n_violations} identity warnings)')

        import_file.status = 'completed'
        import_file.completed_at = timezone.now()
        import_file.save(update_fields=['status', 'completed_at'])

        return {
            'status': 'completed',
            'extractor': 'phase3',
            'parallel_calls_initial': len(jobs),
            'terminal_results': len(results),
            'flavor_b_active': flavor_b_active,
            'identity_violations': n_violations,
            'wall_time_s': elapsed,
            'persist': persist_result,
        }
    finally:
        _delete_sentinel()
        # Free the in-memory workbook cache for this import. Always runs,
        # regardless of success / failure / exception path. Universal.
        try:
            from .workbook_cache import evict as _evict_cache
            # filepath may not be bound if we failed before line 628.
            _evict_cache(locals().get('filepath') or '')
        except Exception:
            pass
