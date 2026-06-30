"""
Phase 3 — workbook in-memory cache.

Universal architectural fix: load the ENTIRE Excel workbook into memory ONCE
per import, then never touch the file on disk again.

Every Phase 3 code path that previously called `openpyxl.load_workbook(...)`
now goes through this module. Once `load_workbook(filepath)` returns, the
source file can be deleted, moved, renamed, cancelled, or affected by a
Django auto-reload — none of that breaks any in-flight chunk, because all
subsequent operations read from the in-memory dict.

Design properties (deliberate):
  • READS THE ENTIRE WORKBOOK. No row slicing, no column slicing, no
    "first N rows", no "skip blanks", no hardcoded start/end cells. Every
    cell of every sheet is preserved exactly as openpyxl yields it via
    `iter_rows(values_only=True)`. Downstream consumers do their own
    slicing if they need to (e.g. Flavor B row-range chunks slice the
    cached list, not the file).
  • Thread-safe via a single lock on the cache dict. Loading happens
    outside the lock so concurrent imports don't block each other.
  • Keyed by filepath. Two different imports (different uploaded files)
    get separate cache entries. Same import re-using the same filepath
    hits cache.
  • Manual eviction via `evict(filepath)` at end of import (success OR
    failure). Memory footprint is bounded by the number of concurrent
    imports.

Memory footprint: a typical AIF Excel (50KB–5MB on disk) sits at roughly
the same size in this in-memory dict — openpyxl returns Python primitives
(str/int/float/datetime), so it's compact. Bharatcrest's 67KB / 14-sheet
workbook → ~250KB resident.
"""
import logging
import threading

import openpyxl

logger = logging.getLogger(__name__)


# {filepath: {'sheets': [name, ...], 'data': {name: {'rows': [...], 'max_row': N, 'max_col': M}}}}
_CACHE: dict = {}
_LOCK = threading.Lock()


def load_workbook(filepath: str) -> dict:
    """Read the entire workbook from disk ONCE per filepath. Subsequent calls
    return the cached structure instantly.

    Returns:
        {
          'sheets':  list[str],                           # source order preserved
          'data':    dict[sheet_name, {
              'rows':    list[tuple],                     # every populated + blank row
              'max_row': int,
              'max_col': int,
          }]
        }
    """
    # Fast path — already cached
    with _LOCK:
        if filepath in _CACHE:
            return _CACHE[filepath]

    # Heavy read OUTSIDE the lock so concurrent imports don't serialise.
    # openpyxl is read-only here; safe to run from multiple threads on
    # different files at once.
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    cached = {'sheets': list(wb.sheetnames), 'data': {}}
    total_rows = 0
    total_cells = 0
    for sname in wb.sheetnames:
        ws = wb[sname]
        # iter_rows(values_only=True) returns the FULL grid, every row,
        # every column in source order. No filtering, no truncation.
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        max_row = ws.max_row or len(rows)
        # max_col from openpyxl can be None on empty sheets — fall back to
        # the widest row we actually saw.
        max_col = ws.max_column or (max((len(r) for r in rows), default=0))
        cached['data'][sname] = {
            'rows': rows,
            'max_row': max_row,
            'max_col': max_col,
        }
        total_rows += len(rows)
        total_cells += sum(
            1 for r in rows for v in r if v is not None and v != ''
        )
    wb.close()

    with _LOCK:
        # Another thread may have populated the same key while we were
        # loading; honour theirs to avoid a double-load race.
        if filepath in _CACHE:
            return _CACHE[filepath]
        _CACHE[filepath] = cached
        logger.info(
            f'[workbook_cache] loaded {filepath} — '
            f'{len(cached["sheets"])} sheets, {total_rows} rows total, '
            f'{total_cells} populated cells'
        )
        return cached


def get_sheet_names(filepath: str) -> list:
    """Sheet names in source order."""
    return list(load_workbook(filepath)['sheets'])


def get_sheet_rows(filepath: str, sheet_name: str) -> list:
    """All rows of a sheet as tuples. Returns [] if sheet doesn't exist."""
    cached = load_workbook(filepath)
    if sheet_name not in cached['data']:
        return []
    return cached['data'][sheet_name]['rows']


def get_sheet_dims(filepath: str, sheet_name: str) -> tuple:
    """Returns (max_row, max_col) for the sheet. (0, 0) when missing."""
    cached = load_workbook(filepath)
    if sheet_name not in cached['data']:
        return (0, 0)
    d = cached['data'][sheet_name]
    return (d['max_row'], d['max_col'])


def evict(filepath: str) -> None:
    """Drop the cached workbook for this filepath, freeing memory.
    Call at the end of an import (success OR failure path)."""
    with _LOCK:
        if filepath in _CACHE:
            del _CACHE[filepath]
            logger.info(f'[workbook_cache] evicted {filepath}')


def cache_size() -> int:
    """Number of workbooks currently in memory (diagnostic)."""
    with _LOCK:
        return len(_CACHE)


import re as _re


def _parse_cell_ref(cell_ref: str):
    """Convert an A1-style cell reference into (col_idx_1based, row_idx_1based).

    Accepts: 'B60', 'AA12', 'b60' (case-insensitive).
    Returns: (col_idx, row_idx) — both 1-based, openpyxl convention.
              None on malformed input.

    Universal across any A1-style reference; supports multi-letter columns
    up to ZZ (which is more than any realistic workbook).
    """
    if not cell_ref:
        return None
    s = str(cell_ref).strip().upper()
    # Strip any leading '$' (Excel absolute references like '$B$60').
    s = s.replace('$', '')
    m = _re.match(r'^([A-Z]+)(\d+)$', s)
    if not m:
        return None
    col_letters, row_str = m.group(1), m.group(2)
    col_idx = 0
    for ch in col_letters:
        col_idx = col_idx * 26 + (ord(ch) - ord('A') + 1)
    row_idx = int(row_str)
    if row_idx < 1:
        return None
    return col_idx, row_idx


def get_cell_value(filepath: str, sheet_name: str, cell_ref: str):
    """Read a single cell's value from the cached workbook.

    Universal — works with any sheet name (case-sensitive, must match what
    openpyxl returned) and any A1-style cell ref ('B60', 'AA12', '$B$60').

    Used by Option C (cell-verified aggregate overrides): when Gemini claims
    a labeled aggregate lives at, e.g., `Fund_Overview!B60`, Python calls
    this to read the actual cell value and compare.

    Returns the raw cell value (str/int/float/datetime/None) so caller can
    apply its own type coercion + tolerance. Returns None when:
      • the sheet doesn't exist
      • the cell reference is malformed
      • the row/col is out of bounds
      • the cell is genuinely empty
    """
    if not filepath or not sheet_name or not cell_ref:
        return None
    try:
        cached = load_workbook(filepath)
    except Exception as e:
        logger.warning(f'[workbook_cache] get_cell_value({sheet_name}!{cell_ref}) '
                       f'workbook load failed: {e}')
        return None
    sheet_data = cached['data'].get(sheet_name)
    if not sheet_data:
        return None
    parsed = _parse_cell_ref(cell_ref)
    if not parsed:
        return None
    col_idx, row_idx = parsed
    rows = sheet_data['rows']
    if row_idx > len(rows):
        return None
    row = rows[row_idx - 1]
    if col_idx > len(row):
        return None
    return row[col_idx - 1]
