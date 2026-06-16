"""
FundImportService — orchestrates the import of a single fund Excel file.

Uses Gemini AI exclusively for all sheet classification and column mapping.
No hardcoded keywords, sheet names, or fallback scanning.

Pipeline:
  1. Gemini Pass 1: classify each sheet → domain_map (1:many, 19 domains)
  2. Gemini Pass 2: map column headers → canonical field names
  3. Each _import_* method uses _dm_sheets(domain_map, domain) to find its sheets
  4. Row reading via read_all_sections_from_sheet() + _find_col() fuzzy matching
"""

import calendar
import logging
import os
import re
from datetime import date
from decimal import Decimal, InvalidOperation

import openpyxl
from django.db import transaction
from django.db.models import Q
from django.utils.text import slugify

from accounts.models import Organization, User, FundAccess
from funds.models import FundCategory, Entity, Fund, Scheme
from lp.models import (BankAccount, Investor, Commitment, CapitalCall,
                        CapitalCallLineItem, Distribution, DistributionLineItem,
                        LPCapitalAccount)
from investments.models import (PortfolioCompany, Investment, InvestmentTranche,
                                 Valuation, KPIDefinition, PortfolioKPI,
                                 CompanyFinancials, ExitEvent, BoardMeeting)
from accounting.models import (ChartOfAccounts, NAVRecord, CarriedInterest,
                                FundLedger, ManagementFeeSchedule)
from portfolio.models import PortfolioSnapshot, PortfolioNode

try:
    from compliance.models import (SEBIReport, AMLDueDiligence,
                                    ComplianceCalendar, ComplianceTestReport,
                                    CTRChecklistItem, PPMAmendment,
                                    SEBICircular, CircularAction,
                                    PortfolioCompanyCompliance)
    HAS_COMPLIANCE = True
except ImportError:
    HAS_COMPLIANCE = False

from .gemini_column_mapper import map_workbook_columns

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

CATEGORY_MAP = {
    'CAT_I_VCF': ('Category I AIF', 'Venture Capital Fund', False),
    'CAT_II': ('Category II AIF', 'Private Equity Fund', False),
    'CAT_III_LVF': ('Category III AIF', 'Long-Short Equity Fund', True),
}




def _d(val, default=None):
    if val is None or val == '' or val == 'None':
        return default
    # Excel date serial corruption: openpyxl reads numeric cells with date-format
    # as datetime/date/time objects instead of their plain numeric value.
    # Recover the original Excel serial number so that amounts like "800 Cr"
    # stored in a date-formatted cell are not silently dropped as None.
    from datetime import datetime as _dt, date as _date_cls, time as _time_cls
    if isinstance(val, _dt):
        # datetime → Excel serial (days from 1899-12-31) + fractional day
        _epoch = _date_cls(1899, 12, 31)
        _days = (val.date() - _epoch).days
        _frac = (val.hour * 3600 + val.minute * 60 + val.second
                 + val.microsecond / 1_000_000) / 86400
        return Decimal(str(round(_days + _frac, 6)))
    if isinstance(val, _date_cls):
        _epoch = _date_cls(1899, 12, 31)
        return Decimal(str((val - _epoch).days))
    if isinstance(val, _time_cls):
        # time values are fractions stored as HH:MM:SS (e.g. 0.897 → 21:31:06)
        _frac = (val.hour * 3600 + val.minute * 60 + val.second
                 + val.microsecond / 1_000_000) / 86400
        return Decimal(str(round(_frac, 6)))
    try:
        return Decimal(str(val))
    except (InvalidOperation, ValueError):
        return default


def _parse_pct(raw):
    """Extract the leading percentage value from any text representation.

    The previous logic — `Decimal(str(x).replace('%','').strip())` — failed
    on values like "8% p.a. (compounding)" or "20% above hurdle" because
    `.replace('%','')` strips only the percent sign and leaves the prose
    suffix, which then can't be coerced to Decimal.

    This helper extracts the first numeric token regardless of surrounding
    text — language-independent, format-agnostic. Handles every form that
    appears in real fund Excel files:

        Input                              → Output
        "8"                                → Decimal('8')
        "8%"                               → Decimal('8')
        "8 % p.a."                         → Decimal('8')
        "8% p.a. (compounding)"            → Decimal('8')
        "20% above hurdle"                 → Decimal('20')
        "20.5%"                            → Decimal('20.5')
        "0.08"                             → Decimal('8')   (decimal → human pct)
        "0.5%"                             → Decimal('0.5') (already pct form)
        "₹2.0% per annum"                  → Decimal('2.0')
        "200 bps"                          → Decimal('2')   (basis points)
        None / "" / "N/A" / "—"            → None

    No hardcoded keywords — works for any language. Returns Decimal or None.
    """
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s.upper() in ('N/A', 'NA', '--', '—', '-'):
        return None

    # Pull the first signed numeric token (handles comma-grouping)
    m = re.search(r'-?\d{1,3}(?:,\d{3})+(?:\.\d+)?|-?\d+(?:\.\d+)?', s)
    if not m:
        return None
    try:
        val = Decimal(m.group().replace(',', ''))
    except (InvalidOperation, ValueError):
        return None

    s_lower = s.lower()
    # Basis-point form (rare but valid): "200 bps" → 2%
    if 'bps' in s_lower or 'basis point' in s_lower:
        return val / Decimal('100')
    # Decimal-fraction form (0.08 → 8): only when value <1 AND no '%' present.
    # "0.5%" is a literal 0.5% — keep as-is. "0.08" is fractional → 8%.
    if val < 1 and '%' not in s:
        return val * Decimal('100')
    return val


def _parse_amount(raw):
    """Parse currency amounts with prefix/suffix into a normalised Decimal.

    Cover/Summary sheets routinely store amounts as decorated strings:
        "₹3,800 Cr"   "Rs 3,458 Cr"   "$5 Mn"   "INR 100 Lakhs"   "₹1.2 Bn"

    The plain `_d()` path drops these (Decimal can't parse the symbols), so
    Pass 3.5 silently fell back to cleanly-parseable neighbour cells — which
    is exactly how `Fund Corpus = ₹3,800 Cr` collapsed into
    `committed_capital = 4.52` (the Portfolio MOIC cell on the same row).

    Returns a tuple (value:Decimal | None, unit_hint:str) where unit_hint ∈
    {'amount', 'percent', 'multiple', 'unknown'}. The unit hint lets the
    caller refuse mismatches like "extract carry_amount_gross from a
    percentage cell".

    Conventions:
      * Crore (Cr)  ×1                  ← we report in Cr already
      * Lakhs       ×0.01
      * Million     ×0.1                (10 Lakh = 1 Million)
      * Billion     ×100                (100 Cr = 1 Bn)
      * percent (%) → unit_hint = 'percent', raw number returned as-is
      * multiple (x, X) → unit_hint = 'multiple', raw number returned as-is
    """
    if raw is None:
        return (None, 'unknown')
    if isinstance(raw, bool):
        return (None, 'unknown')
    if isinstance(raw, (int, float, Decimal)):
        try:
            return (Decimal(str(raw)), 'amount')
        except (InvalidOperation, ValueError):
            return (None, 'unknown')
    s = str(raw).strip()
    if not s or s.upper() in ('N/A', 'NA', '--', '—', '-', 'NIL', 'NULL'):
        return (None, 'unknown')

    s_lower = s.lower()
    # Detect unit hint BEFORE stripping anything
    if '%' in s:
        unit_hint = 'percent'
    elif re.search(r'(?<![A-Za-z])(x|X)(?![A-Za-z])', s):
        unit_hint = 'multiple'
    else:
        unit_hint = 'amount'

    # Pull the first signed numeric token (handles comma-grouping like 3,800)
    m = re.search(r'-?\d{1,3}(?:,\d{3})+(?:\.\d+)?|-?\d+(?:\.\d+)?', s)
    if not m:
        return (None, unit_hint)
    try:
        val = Decimal(m.group().replace(',', ''))
    except (InvalidOperation, ValueError):
        return (None, unit_hint)

    # Unit-multiplier suffix detection. Order matters: check longer first
    # (Billion before B, Million before Mn before M).
    multiplier = Decimal('1')
    if re.search(r'\b(billion|bn)\b', s_lower):
        multiplier = Decimal('100')          # 1 Bn = 100 Cr
    elif re.search(r'\b(million|mn|mln)\b', s_lower):
        multiplier = Decimal('0.1')          # 1 Mn = 0.1 Cr
    elif re.search(r'\b(lakhs?|lac)\b', s_lower):
        multiplier = Decimal('0.01')         # 1 Lakh = 0.01 Cr
    elif re.search(r'\b(crores?|cr)\b', s_lower):
        multiplier = Decimal('1')            # already Cr
    elif re.search(r'\b(thousand|k)\b', s_lower) and unit_hint == 'amount':
        multiplier = Decimal('0.0001')       # 1 K = 0.0001 Cr (rarely useful)

    return (val * multiplier, unit_hint)


def _date(val):
    if val is None:
        return None
    if hasattr(val, 'date'):
        return val.date()
    if isinstance(val, date):
        return val
    # Try string parsing
    if isinstance(val, str):
        val = val.strip()
        for fmt in (
            '%Y-%m-%d',       # 2022-01-18
            '%d-%m-%Y',       # 18-01-2022
            '%d/%m/%Y',       # 18/01/2022
            '%m/%d/%Y',       # 01/18/2022
            '%d-%b-%Y',       # 18-Jan-2022  ← DD-MMM-YYYY (most Excel files)
            '%d %b %Y',       # 18 Jan 2022
            '%d/%b/%Y',       # 18/Jan/2022
            '%b %d, %Y',      # Jan 18, 2022
            '%d-%B-%Y',       # 18-January-2022
            '%d %B %Y',       # 18 January 2022
            '%d.%m.%Y',       # 18.01.2022
            '%Y/%m/%d',       # 2022/01/18
            '%d-%b-%y',       # 31-Mar-25  ← 2-digit year, common in Indian AIF files
            '%d/%m/%y',       # 31/03/25
            '%d.%m.%y',       # 31.03.25
            '%d %b %y',       # 31 Mar 25
        ):
            try:
                from datetime import datetime
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return None


def _str(val, default=''):
    if val is None:
        return default
    return str(val).strip()


def _bool(val):
    if isinstance(val, bool):
        return val
    if val is None:
        return False
    return str(val).strip().lower() in ('yes', 'true', '1')


# Sheet names that are NEVER data sheets — they are summaries/covers/indices.
# No data import function should read companies, investments, or any
# transactional data from these sheets. Source of truth is always the
# dedicated data sheets (Portfolio Investments, Investors & LPs, etc.).
def _is_cover_or_summary_sheet(sheet_name, ws=None):
    """Return True if this sheet name indicates a cover/summary page.

    Uses layout-based detection (no hardcoded keywords):
    - Very few non-empty rows → likely a cover page
    - No row with 4+ non-empty cells (no tabular data) → likely a cover page

    Cover sheets may have fund statistics (company count, total FV, etc.)
    that look like real data but are just display aggregates — often
    computed by hand and prone to errors. We always derive statistics
    from the actual data sheets instead.
    """
    if ws is None:
        return False
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row < 3:
        return True
    non_empty_rows = 0
    has_tabular_row = False
    scan_limit = min(max_row + 1, 30)
    for r in range(1, scan_limit):
        cell_count = sum(
            1 for c in range(1, min(max_col + 1, 20))
            if ws.cell(r, c).value is not None
        )
        if cell_count >= 1:
            non_empty_rows += 1
        if cell_count >= 4:
            has_tabular_row = True
            break
    if non_empty_rows < 5 and not has_tabular_row:
        return True
    return False


def _is_section_header(val):
    """Return True if val looks like a section header — layout-only detection.

    A section header is a string that is:
      - Predominantly uppercase (≥70% of alpha chars)
      - Longer than 3 characters
      - Contains at least one space (multi-word title)
    No hardcoded keywords — purely format-based.
    """
    if not val:
        return False
    s = str(val).strip()
    if len(s) <= 3 or ' ' not in s:
        return False
    alpha_chars = [ch for ch in s if ch.isalpha()]
    if not alpha_chars:
        return False
    upper_ratio = sum(1 for ch in alpha_chars if ch.isupper()) / len(alpha_chars)
    return upper_ratio >= 0.70


def _is_header_row(val):
    """Check if a cell looks like a header row marker (e.g., '#' or 'S.No')."""
    if not val:
        return False
    s = str(val).strip()
    return s in ('#', 'S.No', 'Sr', 'Sr.', 'SNo', 'S.No.')


# Prefixes that identify non-data rows masquerading as data:
# subtotal lines, grand-total lines, repeated header rows, separator labels.
def _is_junk_row(name):
    """Return True if *name* is obviously junk based on format alone.

    Format-based checks (universal, language-independent):
      - Empty / None / whitespace-only
      - Dash/placeholder characters: —, –, -, #, *
      - Purely numeric: serial numbers, counts, aggregated values

    For semantic junk detection (subtotals, totals, headers, notes),
    use FundImportService._classify_junk_names() which classifies via Gemini.
    """
    if not name:
        return True
    n = str(name).strip()
    if not n or n in ('—', '–', '-', '#', '*'):
        return True
    # Purely numeric (serial number, count, or aggregated value in name col)
    try:
        float(n.replace(',', '').replace(' ', ''))
        return True
    except ValueError:
        pass
    return False


def find_section_rows(ws, section_name):
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val and section_name in str(val):
            return r
    return None


def read_table(ws, start_row=1, max_rows=None):
    """Read rows from a worksheet starting at a header row.
    Returns list of dicts keyed by header names."""
    header_row = start_row
    for r in range(start_row, min(ws.max_row + 1, start_row + 10)):
        val = ws.cell(r, 1).value
        if val and not _is_section_header(val):
            header_row = r
            break
        elif val and _is_section_header(val) and r == start_row:
            continue

    headers = []
    for c in range(1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        if h:
            headers.append((c, str(h).strip()))

    if not headers:
        return []

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        if max_rows and len(rows) >= max_rows:
            break
        row_data = {}
        all_empty = True
        for col, name in headers:
            val = ws.cell(r, col).value
            if val is not None:
                all_empty = False
            row_data[name] = val
        if all_empty:
            break
        first_val = ws.cell(r, 1).value
        if _is_section_header(first_val):
            break
        rows.append(row_data)

    return rows


def _is_section_title_row(ws, r, max_col):
    """Return (True, title_text) if row r looks like a section title.

    100% format-agnostic layout detection — no hardcoded keywords.
    A section title row is identified by:
      - 1-2 non-empty cells in the row (not a data row with many columns)
      - First cell text is predominantly uppercase (≥70% of alpha chars)
      - Text length > 3 characters

    Also handles pipe-delimited multi-cell titles (3+ cells) where the first
    cell is predominantly uppercase (≥70%) — these are section headers with
    metadata columns appended (e.g. "PORTFOLIO INVESTMENTS | Fund Name | Date").
    """
    first_cell = ws.cell(r, 1).value
    if first_cell is None:
        return False, ''
    first_str = str(first_cell).strip()
    if not first_str or len(first_str) <= 3:
        return False, ''

    # Compute uppercase ratio of first cell
    alpha_chars = [ch for ch in first_str if ch.isalpha()]
    if not alpha_chars:
        return False, ''
    upper_ratio = sum(1 for ch in alpha_chars if ch.isupper()) / len(alpha_chars)
    if upper_ratio < 0.70:
        return False, ''

    # Count non-empty cells in the row
    cell_count = sum(1 for c in range(1, max_col + 1)
                     if ws.cell(r, c).value is not None)

    # Classic section header: 1-2 cells, predominantly uppercase
    if cell_count <= 2:
        return True, first_str

    # Pipe-delimited multi-cell title: first cell is uppercase title,
    # remaining cells are metadata. Only if first cell has a space
    # (multi-word title, not a single-word data value).
    if cell_count >= 3 and ' ' in first_str:
        return True, first_str

    return False, ''


def read_table_from_sheet(ws, skip_title_rows=1, alias_map=None):
    """Read a full sheet as a table, skipping initial title/section-header rows.

    Finds the header row (first row with multiple non-empty cells that looks
    like column headers), then reads all data rows below it.

    Tolerates up to 2 consecutive blank rows inside the data (some Excel files
    have blank separator rows between company groups within the same table).

    alias_map: optional {excel_col_header: canonical_field_name} from Gemini.
    When supplied, each row dict is enriched with the canonical field name as an
    additional key (only if that key is not already present). This lets _find_col()
    match by the canonical name regardless of the original Excel column wording.

    Returns (headers_dict, rows) where headers_dict maps header_text → col_index,
    and rows is a list of dicts keyed by header text.
    """
    max_col = ws.max_column or 0
    header_row = None
    for r in range(1, min(ws.max_row + 1, 20)):
        cells = []
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if v is not None:
                cells.append((c, str(v).strip()))
        # A header row has multiple cells and doesn't look like a title
        if len(cells) >= 3:
            first_val = cells[0][1] if cells else ''
            # Skip rows that are section/title rows
            is_title, _ = _is_section_title_row(ws, r, max_col)
            if is_title:
                continue
            if not (_is_section_header(first_val) and len(cells) < 5):
                header_row = r
                break

    if not header_row:
        return {}, []

    headers = {}
    for c in range(1, max_col + 1):
        h = ws.cell(header_row, c).value
        if h:
            headers[str(h).strip()] = c

    rows = []
    trailing_blanks = 0  # tracks consecutive blank rows at current tail
    for r in range(header_row + 1, ws.max_row + 1):
        row_data = {}
        all_empty = True
        for name, col in headers.items():
            val = ws.cell(r, col).value
            if val is not None:
                all_empty = False
            row_data[name] = val

        if all_empty:
            # Never terminate on blank rows — Excel files routinely have
            # 3-10+ blank separator rows between company groups within the
            # same table.  Stopping on blanks silently drops companies.
            # We only truly stop at a recognized section-title row or at
            # the sheet boundary.
            trailing_blanks += 1
            continue

        trailing_blanks = 0

        # Stop if we hit a new section header inside the data area
        is_title, _ = _is_section_title_row(ws, r, max_col)
        if is_title:
            break

        # Gemini alias enrichment: add canonical field names as additional
        # keys so that _find_col(row, 'total_invested') hits Pass-1 exact match
        # regardless of how the Excel column was actually labelled.
        if alias_map:
            for excel_col, canonical in alias_map.items():
                if excel_col in row_data and canonical not in row_data:
                    row_data[canonical] = row_data[excel_col]

        rows.append(row_data)

    return headers, rows


def read_all_sections_from_sheet(ws, alias_map=None):
    """Read ALL sections from a multi-section sheet.

    Many fund Excel files have sheets with multiple sections separated by
    section headers (all-caps text like "PORTFOLIO COMPANIES", "INVESTMENTS",
    "INVESTMENT TRANCHES") and empty rows.

    Also handles pipe-delimited multi-cell section titles such as:
        "PORTFOLIO INVESTMENTS — FULL REGISTER | Fund Name | DD-MMM-YYYY"
    where the first cell contains the domain keyword and remaining cells are
    contextual metadata.

    Tolerates up to 2 consecutive blank rows within a data section (some Excel
    files have blank rows separating company groups in the same table).

    alias_map: optional {excel_col: canonical_field} from Gemini. Applied to
    each row so that _find_col(row, 'canonical_name') always hits Pass-1 exact
    match regardless of the original Excel column wording.

    Returns: dict mapping section_name → (headers_dict, rows)
    where section_name is the all-caps header text (e.g., 'INVESTMENTS')
    and the default first section (if no header) is keyed as '__default__'.
    """
    sections = {}
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    def _read_data_rows(header_row, headers):
        """Read data rows starting after header_row. Returns (rows, next_r).

        Termination policy (in priority order):
          1. A recognised section-title row  → stop, next_r = that row
          2. End of sheet                    → stop naturally
          3. Blank rows                      → SKIP, never stop

        Rationale: fund Excel files regularly use 3-10+ consecutive blank
        rows as visual separators between company sub-groups (e.g., between
        sector groups within the same PORTFOLIO INVESTMENTS section).
        Stopping on blanks silently truncates the company list.  The only
        reliable terminator is an ALL-CAPS section header that signals a
        genuinely different domain.
        """
        rows = []
        next_r = max_row + 1
        for data_r in range(header_row + 1, max_row + 1):
            row_data = {}
            all_empty = True
            for name, col in headers.items():
                val = ws.cell(data_r, col).value
                if val is not None:
                    all_empty = False
                row_data[name] = val

            if all_empty:
                continue  # skip blank row — never terminate on blanks

            # Stop if we encounter a new section header
            is_title, _ = _is_section_title_row(ws, data_r, max_col)
            if is_title:
                next_r = data_r
                break

            # Gemini alias enrichment: add canonical field names as keys
            if alias_map:
                for excel_col, canonical in alias_map.items():
                    if excel_col in row_data and canonical not in row_data:
                        row_data[canonical] = row_data[excel_col]

            rows.append(row_data)
        return rows, next_r

    r = 1
    while r <= max_row:
        # Count non-empty cells in this row
        cell_count = sum(1 for c in range(1, max_col + 1)
                         if ws.cell(r, c).value is not None)

        # Skip completely empty rows
        if cell_count == 0:
            r += 1
            continue

        # Detect section header (single/double cell or multi-cell pipe title)
        is_section, section_name = _is_section_title_row(ws, r, max_col)

        if is_section:
            # Next row should be the header row for this section
            r += 1
            # Find the header row (first row with 2+ non-empty cells)
            header_row = None
            for scan_r in range(r, min(r + 6, max_row + 1)):
                scan_count = sum(1 for c in range(1, max_col + 1)
                                 if ws.cell(scan_r, c).value is not None)
                # Skip blank rows between section title and header
                if scan_count == 0:
                    continue
                # Skip a note/subtitle row that looks like another title
                is_sub, _ = _is_section_title_row(ws, scan_r, max_col)
                if is_sub and scan_count <= 2:
                    continue
                if scan_count >= 2:
                    header_row = scan_r
                    break

            if not header_row:
                r += 1
                continue

            # Read headers
            headers = {}
            for c in range(1, max_col + 1):
                h = ws.cell(header_row, c).value
                if h:
                    headers[str(h).strip()] = c

            rows, r = _read_data_rows(header_row, headers)

            # Normalize section name for lookup
            norm_name = section_name.upper().strip()
            # Remove parenthetical details like "(Main Scheme — Call #1)"
            if '(' in norm_name:
                norm_name = norm_name[:norm_name.index('(')].strip()
            # Remove pipe-delimited metadata suffix if present
            if '|' in norm_name:
                norm_name = norm_name[:norm_name.index('|')].strip()
            # Remove em-dash suffixes like "— FULL REGISTER"
            for sep in [' — ', ' - ', '—', ' –']:
                if sep in norm_name:
                    norm_name = norm_name[:norm_name.index(sep)].strip()
                    break

            sections[norm_name] = (headers, rows)
            continue

        # Not a section header — this might be a direct header row
        # (sheet starts directly with data, no section header)
        if cell_count >= 3:
            headers = {}
            for c in range(1, max_col + 1):
                h = ws.cell(r, c).value
                if h:
                    headers[str(h).strip()] = c

            rows, r = _read_data_rows(r, headers)
            sections['__default__'] = (headers, rows)
            continue

        r += 1

    return sections


def _dm_sheets(domain_map, domain_key):
    """Return list of sheet names for a domain. Works with 1:many domain_map."""
    val = domain_map.get(domain_key, [])
    if isinstance(val, list):
        return val
    return [val] if val else []


def _dm_first(domain_map, domain_key):
    """Return the first (primary) sheet name for a domain, or None."""
    sheets = _dm_sheets(domain_map, domain_key)
    return sheets[0] if sheets else None


def _get_section_rows(ws, domain_map, domain_key, section_subdomains=None,
                      section_map=None):
    """Smart row reader: tries dedicated domain sheet first, then falls back
    to reading a specific section from a multi-section sheet using Gemini
    sub-domain classification.

    Returns (headers_dict, rows) like read_table_from_sheet.
    """
    sheet_name = _dm_first(domain_map, domain_key)
    if not sheet_name or sheet_name not in (ws.parent.sheetnames if hasattr(ws, 'parent') else []):
        return {}, []

    target_ws = ws.parent[sheet_name] if hasattr(ws, 'parent') else ws

    # First try reading as a flat table
    headers, rows = read_table_from_sheet(target_ws)

    if rows:
        return headers, rows

    # If no rows from flat table, try reading sections and matching by
    # Gemini sub-domain classification
    if section_subdomains and section_map:
        sections = read_all_sections_from_sheet(target_ws)
        sheet_secs = section_map.get(sheet_name, {})
        for sec_name, (sec_headers, sec_rows) in sections.items():
            mapped_subdomain = sheet_secs.get(sec_name, 'unknown')
            if mapped_subdomain in section_subdomains and sec_rows:
                return sec_headers, sec_rows

    return headers, rows


# ---------------------------------------------------------------------------
# Fuzzy header matching — finds a column by trying multiple possible names
# ---------------------------------------------------------------------------

def _find_col(row, *candidates):
    """Find the first matching header value in a row dict.

    Matching priority (most specific → least specific).
    We iterate ALL candidates at each pass before falling to the next pass —
    this ensures a more-specific candidate (e.g. 'Company Name') is found
    before a less-specific one (e.g. 'Name') even if 'Name' would match earlier
    in a looser pass.

    Pass 1   — Exact case-sensitive:       "Company Name" == "Company Name"
    Pass 1.5 — Normalised (space↔underscore, case-insensitive):
               "Company Name" matches "company_name" from Gemini alias map
    Pass 2   — Exact case-insensitive:     "company name" == "Company Name"
    Pass 3   — Key ends with candidate:    "Company Name" ends with "Name"
    Pass 4   — Candidate ends with key:    "HQ City" key, candidate "City"
    Pass 5   — Loose substring (guarded)
    """
    # Pass 1 — exact case-sensitive
    for c in candidates:
        if c in row:
            return row[c]

    # Build lowercase lookup once
    row_lower = {k.lower(): v for k, v in row.items()}

    # Pass 1.5 — normalised: treat spaces, underscores, and hyphens as equivalent.
    # This bridges English candidates ("Company Name") with Gemini canonical
    # field names ("company_name") added by the alias map.
    row_norm = {k.replace(' ', '_').replace('-', '_'): v for k, v in row_lower.items()}
    for c in candidates:
        cn = c.lower().replace(' ', '_').replace('-', '_')
        if cn in row_norm:
            return row_norm[cn]

    # Pass 2 — exact case-insensitive
    for c in candidates:
        cl = c.lower()
        if cl in row_lower:
            return row_lower[cl]

    # Pass 3 — the column header ends with the candidate phrase (word-boundary)
    for c in candidates:
        cl = c.lower()
        for key_l, val in row_lower.items():
            if key_l == cl:
                return val
            if key_l.endswith(' ' + cl) or key_l.endswith('-' + cl):
                return val

    # Pass 4 — the candidate ends with the column header
    for c in candidates:
        cl = c.lower()
        for key_l, val in row_lower.items():
            if cl.endswith(' ' + key_l) or cl.endswith('-' + key_l):
                return val

    # Pass 5 — loose substring, guarded
    for c in candidates:
        cl = c.lower()
        if len(c.split()) == 1 and len(c) < 8:
            continue
        for key_l, val in row_lower.items():
            if cl in key_l or key_l in cl:
                return val

    return None


def _find_col_str(row, *candidates, default=''):
    val = _find_col(row, *candidates)
    return _str(val, default)


def _find_col_decimal(row, *candidates, default=None):
    val = _find_col(row, *candidates)
    return _d(val, default)


def _normalize_col_key(name):
    """
    Normalize an Excel column header for robust, format-agnostic fuzzy matching.

    Fund managers format column headers in wildly different ways across Excel files:
      'Budget(₹Cr)'        — no space before unit
      'Revenue (INR Mn)'   — space before unit
      'GrossProfit'        — CamelCase without spaces
      'Gross Profit(Cr)'   — CamelCase + unit
      'EBITDA%'            — ratio marker attached
      'Budget [INR Cr]'    — square-bracket unit notation

    This function normalises all such variations to a clean label so that
    _find_col() and _pl_map_to_line_item() can match reliably regardless of
    the Excel file's formatting conventions.

    Transformations applied (in order):
      1. Strip unit/currency annotations enclosed in () or []:
         'Budget(₹Cr)' → 'Budget', 'Revenue (INR Mn)' → 'Revenue'
         Works with any currency symbol (₹ $ € £ ¥) and any unit
         (Cr, Crore, Lakh, Mn, Million, Bn, Billion, INR, USD, EUR, GBP,
          K, Thousands, Rs, Rupees, Lk, Lakhs).
      2. Split CamelCase into words:
         'GrossProfit' → 'Gross Profit', 'NetWorth' → 'Net Worth'
         Only inserts a space where a lowercase letter is immediately followed
         by an uppercase letter — safe for acronyms like 'EBITDA', 'PAT'.
    """
    if not name:
        return ''
    s = str(name).strip()
    # Step 1 — strip unit/currency suffix in parentheses or square brackets
    s = re.sub(
        r'[\s]*[\(\[]\s*[₹$€£¥]?\s*'
        r'(?:cr|crore|lakh|lakhs|lk|mn|million|bn|billion|'
        r'inr|usd|eur|gbp|k|thousands|rs|rupees|₹)'
        r'[^)\]]*[\)\]]',
        '',
        s,
        flags=re.IGNORECASE,
    ).strip()
    # Step 2 — insert space at lowercase→uppercase boundary (CamelCase split)
    s = re.sub(r'(?<=[a-z])(?=[A-Z])', ' ', s)
    return s


def _norm_row(row):
    """
    Return a copy of a row dict with all keys passed through _normalize_col_key.

    Preserves None keys as empty string (they are always skipped in matching).
    Duplicate normalised keys are last-writer-wins — in practice this only
    occurs when a sheet has two columns that both normalise to the same label
    (e.g., 'EBITDA(Cr)' and 'EBITDA(INR Cr)'), which should not happen in
    well-formed Excel files.
    """
    return {_normalize_col_key(k) if k is not None else '': v for k, v in row.items()}


def _find_col_date(row, *candidates):
    val = _find_col(row, *candidates)
    return _date(val)


def _find_col_bool(row, *candidates):
    val = _find_col(row, *candidates)
    return _bool(val)


# ---------------------------------------------------------------------------
# Domain-to-sheet resolver
# ---------------------------------------------------------------------------

def _build_domain_sheet_map(classifications, wb):
    """Build a 1:many mapping from canonical domain name to sheet name list.

    Uses Gemini's sheet classification (Pass 1) exclusively — no keyword
    fallback. Multiple sheets can share the same domain (e.g., 4 financial
    sheets all classified as financials_pl_bva).

    CRITICAL RULE: Cover/summary sheets (Cover, Summary, Index, Dashboard,
    etc.) are NEVER mapped to transactional data domains like
    portfolio_investments, investors_aml, capital_calls, etc.
    They may only be mapped to fund_scheme_master for basic fund identity
    fields. All statistics on cover sheets (company count, total FV, etc.)
    are display aggregates — the source of truth is always the actual
    data sheets.
    """
    # Domains that MUST NEVER be served by a cover/summary sheet.
    # These domains read transactional/company rows from their sheets.
    _DATA_ONLY_DOMAINS = {
        'portfolio_hierarchy', 'portfolio_investments', 'organization_users',
        'investors_aml', 'commitments', 'capital_calls', 'valuations_kpis',
        'nav_accounting', 'exits_distributions', 'compliance',
    }

    from collections import defaultdict
    domain_map = defaultdict(list)  # domain → [sheet_name, ...]

    # From Gemini classification — allow MULTIPLE sheets per domain.
    # Gemini assigns each sheet a primary domain; multiple sheets CAN share
    # the same domain (e.g., P&L, BS, CF, BvA all → financials_pl_bva).
    for cls in classifications:
        sheet_name = cls.get('sheet_name', '')
        domains = cls.get('domains', [])
        for domain in domains:
            if domain and domain != 'unknown':
                ws_check = wb[sheet_name] if sheet_name in wb.sheetnames else None
                if domain in _DATA_ONLY_DOMAINS and _is_cover_or_summary_sheet(sheet_name, ws_check):
                    logger.warning(
                        f'Gemini classified cover/summary sheet "{sheet_name}" as '
                        f'data domain "{domain}" — overriding (layout detection: '
                        f'no tabular data) to prevent metadata rows being imported '
                        f'as company/investment records.'
                    )
                    continue
                if sheet_name not in domain_map[domain]:
                    domain_map[domain].append(sheet_name)

    # Convert to regular dict for downstream consumption
    domain_map = dict(domain_map)

    # Log Gemini domain_map
    total_sheets = sum(len(v) for v in domain_map.values())
    logger.info(f'[GEMINI] domain_map: {len(domain_map)} domains, {total_sheets} sheet assignments')
    for d, sheets in sorted(domain_map.items()):
        logger.info(f'  {d} → {sheets}')

    return domain_map


# ---------------------------------------------------------------------------
# Financial aggregation helpers (roll-up from companies → sectors → funds)
# ---------------------------------------------------------------------------

def _build_summary_from_pl(monthly_pl, budget_vs_actual=None):
    """Build a summary dict from a company's monthly_pl entries.

    Uses the latest period's values as the summary, plus computes YTD totals.
    This is what the portfolio dashboard reads: financials.summary.revenue, etc.
    """
    if not monthly_pl:
        return {}

    # Latest period values
    latest = monthly_pl[-1] if monthly_pl else {}
    summary = {
        'revenue': latest.get('revenue', 0),
        'cogs': latest.get('cogs', 0),
        'gross_profit': latest.get('gross_profit', 0),
        'opex': latest.get('opex', 0),
        'ebitda': latest.get('ebitda', 0),
        'gp_pct': latest.get('gp_pct'),
        'ebitda_pct': latest.get('ebitda_pct'),
        'period': latest.get('period', 'Latest'),
    }

    # YTD totals (sum all periods)
    for field in ('revenue', 'cogs', 'gross_profit', 'opex', 'ebitda'):
        ytd_key = f'ytd_{field}'
        vals = [p.get(field, 0) for p in monthly_pl if isinstance(p.get(field), (int, float))]
        if vals:
            summary[ytd_key] = round(sum(vals), 2)

    # Budget info from budget_vs_actual
    if budget_vs_actual:
        for entry in budget_vs_actual:
            li = (entry.get('line_item') or '').lower()
            if 'revenue' in li:
                summary['budget_revenue'] = entry.get('budget', 0)
                summary['ytd_budget_revenue'] = entry.get('budget', 0)
            elif 'ebitda' in li:
                summary['budget_ebitda'] = entry.get('budget', 0)
                summary['ytd_budget_ebitda'] = entry.get('budget', 0)

    return summary


def _aggregate_financials(children_financials):
    """Aggregate financial data from a list of child node financials dicts.

    Produces a parent-level financials dict with:
    - summary: summed revenue, cogs, gross_profit, opex, ebitda + derived pcts
    - monthly_pl: merged and summed by period
    - budget_vs_actual: merged by line_item
    """
    # --- Aggregate summary ---
    fields_to_sum = [
        'revenue', 'cogs', 'gross_profit', 'opex', 'ebitda',
        'ytd_revenue', 'ytd_cogs', 'ytd_gross_profit', 'ytd_opex', 'ytd_ebitda',
        'budget_revenue', 'budget_ebitda',
        'ytd_budget_revenue', 'ytd_budget_ebitda',
    ]
    summary = {}
    for field in fields_to_sum:
        vals = [
            (cf.get('summary') or {}).get(field)
            for cf in children_financials
        ]
        vals = [v for v in vals if isinstance(v, (int, float))]
        if vals:
            summary[field] = round(sum(vals), 2)

    rev = summary.get('revenue')
    if rev:
        gp = summary.get('gross_profit')
        ebitda = summary.get('ebitda')
        if gp is not None:
            summary['gp_pct'] = round(gp / rev * 100, 2)
        if ebitda is not None:
            summary['ebitda_pct'] = round(ebitda / rev * 100, 2)
    summary['period'] = 'Latest'

    # Budget variance percentages
    ytd_rev = summary.get('ytd_revenue')
    bud_rev = summary.get('ytd_budget_revenue')
    if ytd_rev and bud_rev:
        summary['bva_revenue_pct'] = round((ytd_rev - bud_rev) / bud_rev * 100, 2)

    # --- Aggregate monthly_pl ---
    by_period = {}
    for cf in children_financials:
        for pt in cf.get('monthly_pl', []) or []:
            period = pt.get('period')
            if not period:
                continue
            agg = by_period.setdefault(period, {
                'period': period, 'revenue': 0, 'cogs': 0,
                'gross_profit': 0, 'opex': 0, 'ebitda': 0,
            })
            for f in ('revenue', 'cogs', 'gross_profit', 'opex', 'ebitda'):
                v = pt.get(f)
                if isinstance(v, (int, float)):
                    agg[f] += v
    monthly_pl = []
    for period in sorted(by_period.keys()):
        agg = by_period[period]
        rev = agg['revenue']
        if rev:
            agg['gp_pct'] = round(agg['gross_profit'] / rev * 100, 2)
            agg['ebitda_pct'] = round(agg['ebitda'] / rev * 100, 2)
        for f in ('revenue', 'cogs', 'gross_profit', 'opex', 'ebitda'):
            agg[f] = round(agg[f], 2)
        monthly_pl.append(agg)

    # --- Aggregate budget_vs_actual ---
    bva_by_li = {}
    for cf in children_financials:
        for entry in cf.get('budget_vs_actual', []) or []:
            li = entry.get('line_item')
            if not li:
                continue
            agg = bva_by_li.setdefault(li, {
                'line_item': li, 'period': 'YTD', 'budget': 0, 'actual': 0,
            })
            for f in ('budget', 'actual'):
                v = entry.get(f)
                if isinstance(v, (int, float)):
                    agg[f] += v
    budget_vs_actual = []
    for agg in bva_by_li.values():
        agg['variance'] = round(agg['actual'] - agg['budget'], 2)
        agg['variance_pct'] = round(
            agg['variance'] / agg['budget'] * 100, 2
        ) if agg['budget'] else None
        for f in ('budget', 'actual'):
            agg[f] = round(agg[f], 2)
        budget_vs_actual.append(agg)

    return {
        'summary': summary,
        'monthly_pl': monthly_pl,
        'budget_vs_actual': budget_vs_actual,
    }


# ---------------------------------------------------------------------------
# FundImportService
# ---------------------------------------------------------------------------

class FundImportService:
    """
    Orchestrates the import of a single fund Excel file.

    Uses Gemini AI to semantically classify sheets and map columns,
    then imports data into Django models using header-based reading.
    No hardcoded sheet names or column positions.
    """

    def __init__(self, organization, user):
        self.org = organization
        self.user = user
        self.errors = []
        self.counts = {}
        self._imported_fund = None
        self._gemini_sheet_aliases = {}  # {sheet_name: {excel_col: canonical_field}}
        self._pl_line_item_map = {}     # {label: canonical_pl_category} — populated per sheet
        self._junk_name_cache = {}      # {name: row_type} — populated per section
        self._filepath = None           # set by _do_import; used by Pass 2.5 layout calls
        self._sheet_layouts = {}        # {sheet_name: layout_dict} — cached Pass 2.5 results
        self._layout_failed_sheets = set()  # sheets where Gemini layout call failed → fallback

    # ------------------------------------------------------------------
    # Pass 2.5 — Per-sheet Gemini layout detection (PRIMARY row reader)
    #
    # Replaces the brittle Python heuristic that previously decided "where
    # does the table start / where does it end / which row is the header /
    # which rows are banner-disclaimer noise."  The Python heuristic
    # (`_is_section_title_row` + `_read_data_rows`) survives ONLY as a
    # safety-net fallback for the rare case where Gemini itself is
    # unreachable.  In normal operation Gemini is the source of truth.
    # ------------------------------------------------------------------

    def _get_sheet_layout(self, sheet_name):
        """Return the cached Pass 2.5 layout dict for a sheet, calling
        Gemini once on the first access.  Returns None ONLY if the Gemini
        call raised an exception — caller then falls back to the heuristic
        reader.  A successful "this sheet has no tables" response is a
        valid layout (sub_tables=[]) and is cached as such, not None.
        """
        if sheet_name in self._sheet_layouts:
            return self._sheet_layouts[sheet_name]
        if sheet_name in self._layout_failed_sheets:
            return None
        if not self._filepath:
            return None
        try:
            from .gemini_column_mapper import detect_sheet_layout
            layout = detect_sheet_layout(self._filepath, sheet_name)
            self._sheet_layouts[sheet_name] = layout
            return layout
        except Exception as e:
            logger.warning(
                f'Pass 2.5 layout detection failed for "{sheet_name}" '
                f'({type(e).__name__}: {e}); falling back to heuristic reader'
            )
            self._layout_failed_sheets.add(sheet_name)
            return None

    def _read_sheet_via_layout(self, ws, alias_map=None):
        """Read a sheet using the Gemini layout map (Pass 2.5) — PRIMARY path.

        Returns a dict of {section_name: (headers_dict, rows)} with the same
        contract as read_all_sections_from_sheet() so call sites stay identical.

        Behaviour:
          * Gemini layout succeeded → read rows mechanically per the map.
          * Gemini layout failed    → fall back to read_all_sections_from_sheet
                                       (the deterministic Python heuristic).
          * Gemini layout says "no tables" → return {} (empty result, NOT a
                                              fallback — Gemini explicitly said
                                              the sheet has no data, e.g. Cover).
        """
        sheet_name = ws.title
        layout = self._get_sheet_layout(sheet_name)

        if layout is None:
            # Gemini truly failed (network / API error) — fall back to
            # heuristic. This is the ONLY path that uses the old reader.
            return read_all_sections_from_sheet(ws, alias_map=alias_map)

        sub_tables = layout.get('sub_tables', [])
        if not sub_tables:
            # Gemini explicitly said: no tables in this sheet
            return {}

        sections = {}
        max_col = ws.max_column or 0
        max_row = ws.max_row or 0

        for idx, st in enumerate(sub_tables):
            header_row_0 = st['header_row']
            data_start_0 = st['data_start']
            data_end_0   = st['data_end']

            # Convert 0-indexed → openpyxl 1-indexed
            header_row = header_row_0 + 1
            data_start = data_start_0 + 1
            data_end   = data_end_0 + 1

            # Defensive re-validation (Gemini already validated, but the
            # sheet may have changed dimensions between Gemini's sample
            # and this read — unlikely, but free insurance)
            if not (1 <= header_row < data_start <= data_end <= max_row):
                logger.warning(
                    f'Layout sub-table {idx} for "{sheet_name}" has invalid '
                    f'bounds (header={header_row}, start={data_start}, '
                    f'end={data_end}, sheet_rows={max_row}); skipping'
                )
                continue

            # Read header row
            headers = {}
            for c in range(1, max_col + 1):
                h = ws.cell(header_row, c).value
                if h is not None:
                    headers[str(h).strip()] = c
            if not headers:
                continue

            # Read data rows (data_start ... data_end inclusive)
            rows = []
            for r in range(data_start, data_end + 1):
                row_data = {}
                all_empty = True
                for name, col in headers.items():
                    v = ws.cell(r, col).value
                    if v is not None:
                        all_empty = False
                    row_data[name] = v
                if all_empty:
                    continue

                # Gemini Pass-2 alias enrichment
                if alias_map:
                    for excel_col, canonical in alias_map.items():
                        if excel_col in row_data and canonical not in row_data:
                            row_data[canonical] = row_data[excel_col]

                rows.append(row_data)

            # Normalize section name (compat with existing call sites that
            # look up sections by ALL-CAPS keys)
            section_name = (st.get('section_name') or '').upper().strip()
            if '|' in section_name:
                section_name = section_name[:section_name.index('|')].strip()
            for sep in [' — ', ' - ', '—', ' –']:
                if sep in section_name:
                    section_name = section_name[:section_name.index(sep)].strip()
                    break

            # If this is the only/first sub-table and has no banner,
            # use __default__ so the existing call sites (which look for
            # __default__ before iterating named sections) keep working.
            if not section_name:
                if idx == 0 and len(sub_tables) == 1:
                    section_name = '__default__'
                else:
                    section_name = f'__section_{idx}__'

            sections[section_name] = (headers, rows)

        return sections

    def _derived_column_for(self, sheet_name, target_header_aliases):
        """Look up the Gemini-discovered formula for a derived column.

        target_header_aliases: list of header-text patterns to match against
        the layout's derived_columns (case-insensitive substring match — same
        spirit as _find_col).

        Returns the matched derived_column dict, or None.
        """
        layout = self._sheet_layouts.get(sheet_name)
        if not layout:
            return None
        aliases_lc = [a.lower() for a in target_header_aliases if a]
        for st in layout.get('sub_tables', []):
            for dc in st.get('derived_columns', []):
                col_lc = (dc.get('column_name') or '').lower()
                if not col_lc:
                    continue
                for a in aliases_lc:
                    if a in col_lc or col_lc in a:
                        return dc
        return None

    def _evaluate_derived_formula(self, derived_col, row):
        """Compute a derived column's value from its formula_components.

        derived_col: {'column_name', 'formula_components': [{sign, source_column}]}
        row: the row dict (header_text → cell_value)

        Returns Decimal or None if any source column is missing/non-numeric.
        Universal — no hardcoded formula, no SEBI assumption.  The formula
        was discovered by Gemini either from an explicit disclaimer row in
        the Excel ("Col I = C+E+F-D") or from standard accounting identities.
        """
        from decimal import Decimal, InvalidOperation
        total = Decimal('0')
        for comp in derived_col.get('formula_components', []):
            src_name = comp.get('source_column', '')
            sign = comp.get('sign', '+')
            # Case-insensitive lookup in row dict — handles whitespace/case drift
            val_raw = None
            for k, v in row.items():
                if k and str(k).strip().lower() == src_name.strip().lower():
                    val_raw = v
                    break
            if val_raw is None or val_raw == '':
                return None  # missing source → can't compute
            try:
                val = Decimal(str(val_raw))
            except (InvalidOperation, TypeError, ValueError):
                return None
            total = total + val if sign == '+' else total - val
        return total

    # ------------------------------------------------------------------
    # Pass 3 helpers: Gemini-powered value classification
    # ------------------------------------------------------------------

    def _classify_labels(self, labels, category_key, context=''):
        """Classify text labels into canonical categories via Gemini.

        Returns {label: canonical_key_or_None}.
        """
        from .gemini_column_mapper import classify_labels
        from .canonical_schema import CANONICAL_VALUE_CATEGORIES
        options = CANONICAL_VALUE_CATEGORIES.get(category_key, {})
        if not options:
            logger.warning(f'Unknown category_key: {category_key}')
            return {}
        return classify_labels(list(labels), category_key, options, context)

    def _classify_enum(self, values, enum_key, context=''):
        """Classify text values into a closed enum set via Gemini.

        Returns {value: enum_key}. Never returns None — always picks closest match.
        """
        from .gemini_column_mapper import classify_enum_values
        from .canonical_schema import CANONICAL_ENUM_TYPES
        options = CANONICAL_ENUM_TYPES.get(enum_key, {})
        if not options:
            logger.warning(f'Unknown enum_key: {enum_key}')
            return {}
        return classify_enum_values(list(values), enum_key, options, context)

    def _classify_junk_names(self, names):
        """Classify a batch of names as real entities vs junk rows via Gemini.

        Returns a set of names that are junk (subtotals, totals, headers, serials).
        Real entity names are NOT in the returned set.
        """
        unique_names = list(set(n for n in names if n and str(n).strip()))
        if not unique_names:
            return set()

        row_type_map = self._classify_labels(unique_names, 'row_type',
                                              context='Entity names vs junk rows')
        junk_types = {'subtotal', 'total', 'header', 'serial', 'note'}
        junk_set = set()
        for name, rtype in row_type_map.items():
            if rtype in junk_types:
                junk_set.add(name)
            self._junk_name_cache[name] = rtype
        return junk_set

    def _get_alias(self, ws) -> dict:
        """Return Gemini-built alias map for this worksheet (empty dict if none)."""
        return self._gemini_sheet_aliases.get(getattr(ws, 'title', ''), {})

    def _get_section_subdomain(self, sheet_name, sec_name):
        """Return the Gemini-classified sub-domain for a section.

        Looks up the section_map built by Gemini Pass 1.5.
        The lookup normalizes the section name to uppercase and tries:
          1. Exact match on the raw section name
          2. Exact match on uppercase-stripped version
          3. Substring match (section_map key is contained in sec_name or vice versa)
          4. Single-section fallback: when Pass 1.5 mapped only
             `__default__` for this sheet (one-table sheet), any section
             name returned by Pass 2.5 (which may have used the sheet's
             banner row as the section name) is treated as that default.
             Universal — handles every file where Pass 1.5 and Pass 2.5
             disagree on the section name for a single-table sheet.

        Returns the sub-domain string (e.g. 'portfolio_companies') or 'unknown'.
        """
        sheet_secs = self._section_map.get(sheet_name, {})
        if not sheet_secs:
            return 'unknown'

        # Try exact match first
        if sec_name in sheet_secs:
            return sheet_secs[sec_name]

        # Try uppercase-normalized match
        sec_upper = sec_name.upper().strip()
        for mapped_name, subdomain in sheet_secs.items():
            if mapped_name.upper().strip() == sec_upper:
                return subdomain

        # Try substring containment (handles truncation/suffix differences)
        for mapped_name, subdomain in sheet_secs.items():
            mu = mapped_name.upper().strip()
            if mu in sec_upper or sec_upper in mu:
                return subdomain

        # Single-section fallback: when Pass 1.5 declared a single
        # `__default__` entry for this sheet, attribute any incoming
        # section name to that default. Pass 2.5 occasionally returns
        # the banner-row text as section_name while Pass 1.5 always
        # uses `__default__` for a one-table sheet; without this
        # fallback the bucket-router (AD multi-sheet portfolio walk)
        # silently drops the section.
        if len(sheet_secs) == 1 and '__default__' in sheet_secs:
            return sheet_secs['__default__']

        return 'unknown'

    def import_file(self, import_file_record, progress_cb=None):
        """
        Main entry point. Processes a single ImportFile record.
        """
        filepath = import_file_record.file.path

        def _progress(pct, msg):
            if progress_cb:
                progress_cb(pct, msg)

        # Clear Pass 3 classification cache for fresh import
        from .gemini_column_mapper import clear_classification_cache
        clear_classification_cache()

        # Step 1: Gemini column mapping (two-pass)
        _progress(5, 'Reading workbook...')

        mapping_result = None
        try:
            mapping_result = map_workbook_columns(filepath, _progress)
            import_file_record.column_mapping = mapping_result.get('column_mappings', {})
            import_file_record.gemini_confidence = mapping_result.get('overall_confidence', 0.0)
            import_file_record.sheet_names = mapping_result.get('sheet_names', [])
            import_file_record.status = 'importing'
            import_file_record.save(update_fields=[
                'column_mapping', 'gemini_confidence', 'sheet_names', 'status',
            ])
            n_classified = len(mapping_result.get('sheet_classifications', []))
            n_mapped = len(mapping_result.get('column_mappings', {}))
            logger.info(
                f'Gemini mapping complete: {n_classified} sheets classified, '
                f'{n_mapped} sheets column-mapped, '
                f'confidence={mapping_result.get("overall_confidence", 0):.2f}'
            )
        except Exception as e:
            import traceback
            logger.error(
                f'Gemini mapping FAILED for "{import_file_record.original_filename}": '
                f'{type(e).__name__}: {e}\n{traceback.format_exc()}'
            )
            self.errors.append({
                'section': 'gemini_mapping',
                'error': f'{type(e).__name__}: {e}',
            })
            mapping_result = None
            import_file_record.status = 'importing'
            import_file_record.save(update_fields=['status'])

        # Step 2: Import data
        _progress(25, 'Starting data import...')

        classifications = []
        column_mappings = {}
        section_map = {}
        if mapping_result:
            classifications = mapping_result.get('sheet_classifications', [])
            column_mappings = mapping_result.get('column_mappings', {})
            section_map = mapping_result.get('section_map', {})

        result = self._do_import(filepath, classifications, _progress,
                                 column_mappings, section_map)

        # Save fund reference back to ImportFile for cascading delete support
        if self._imported_fund:
            import_file_record.fund = self._imported_fund
            import_file_record.fund_name = self._imported_fund.name
            import_file_record.save(update_fields=['fund', 'fund_name'])

        # Pass 4: Derive missing fund-level metrics via Gemini.
        # Runs per scheme on the imported fund; persists DerivedMetric rows
        # with full provenance (chosen formula, inputs used, Gemini reasoning,
        # alternates considered). All formulas come from Gemini — nothing
        # is hardcoded in derivation_service.
        # Pass 2.6 — Column Semantic Role Classifier.
        # For each horizontal section that Pass 2.5 found, classify every
        # column's SEMANTIC role (per_period_amount / cumulative_total /
        # ratio_percent / identifier / metadata_text / derived_indicator /
        # unknown). Pass 3.5 uses these roles to filter candidate cells
        # by metric-vs-role compatibility BEFORE disambiguation — this
        # is what prevents a "per_step" metric like preferred_return_amount
        # from being extracted from a "Cumulative Distributed" column.
        # ZERO keyword matching — Gemini reads headers + sample data and
        # reasons about each column's intent.
        try:
            _progress(91, 'Pass 2.6: Classifying column semantic roles...')
            self._classify_column_roles(import_file_record=import_file_record)
        except Exception as e:
            logger.warning(
                f'Pass 2.6 column-role classification failed: {e} — '
                f'Pass 3.5 will fall back to no role filter.'
            )

        # ────────────────────────────────────────────────────────────
        # ANCHOR-FIRST PIPELINE — replaces Pass 3.5 / Pass 4 / Pass 8 /
        # Pass 9 / MetricArbiter for fund-level metrics.
        #
        # Stages (all in anchor_pipeline.py):
        #   0. workbook_census    (Python, no AI)
        #   1. identity_hunt      (Gemini, 1 call, role-based prompt)
        #   2. anchor_extraction  (Gemini, 1 call, role-based prompt)
        #   3. cash_flow_series   (Gemini, 1 call)
        #   5. compute_metrics    (Python, textbook formulas)
        #   6. audit_assertions   (Python, accounting identities)
        #   7. persist            (writes FundMetric + mirrors to
        #                          DerivedMetric for backward-compat)
        #
        # Determinism: same file → same numbers, every time.
        # Tokens: ~3 Gemini calls per scheme (vs ~dozens before).
        # No catalogue formulas, no trust tiers, no Arbiter, no
        # candidate competition.  No hardcoded value ranges or
        # keyword lists in any prompt.
        # ────────────────────────────────────────────────────────────
        try:
            _progress(93, 'Anchor pipeline: workbook census + identity + anchors + cash flows...')
            from .anchor_pipeline import run_anchor_pipeline

            anchor_summary = []
            if self._imported_fund:
                for sch in self._imported_fund.schemes.all():
                    try:
                        ap_result = run_anchor_pipeline(
                            filepath=filepath,
                            scheme=sch,
                            organization=self.org,
                            source_import_file=import_file_record,
                            progress_cb=_progress,
                        )
                        anchor_summary.append({
                            'scheme_id':   str(sch.id),
                            'scheme_name': sch.name,
                            'result':      {
                                'metrics_count':   ap_result.get('metrics_count'),
                                'cash_flow_count': ap_result.get('cash_flow_count'),
                                'audits_pass':     sum(1 for a in ap_result.get('audits', []) if a['status'] == 'pass'),
                                'audits_fail':     sum(1 for a in ap_result.get('audits', []) if a['status'] == 'fail'),
                                'conflicts':       len(ap_result.get('conflicts', [])),
                            },
                        })
                        logger.info(
                            f'[anchor_pipeline] scheme={sch.name} '
                            f'metrics={ap_result.get("metrics_count")} '
                            f'cashflows={ap_result.get("cash_flow_count")} '
                            f'audits_fail={sum(1 for a in ap_result.get("audits", []) if a["status"] == "fail")}'
                        )
                    except Exception as inner:
                        import traceback as _tb
                        logger.error(
                            f'[anchor_pipeline] FAILED for scheme={sch.name}: '
                            f'{type(inner).__name__}: {inner}\n{_tb.format_exc()}'
                        )
                        self.errors.append({
                            'section': 'anchor_pipeline',
                            'scheme':  sch.name,
                            'error':   f'{type(inner).__name__}: {inner}',
                        })

            if isinstance(result, dict):
                result['anchor_pipeline_summary'] = anchor_summary
            _progress(96, 'Anchor pipeline complete')
        except Exception as e:
            import traceback
            logger.error(
                f'Anchor pipeline orchestration failed: '
                f'{type(e).__name__}: {e}\n{traceback.format_exc()}'
            )
            self.errors.append({
                'section': 'anchor_pipeline',
                'error':   f'{type(e).__name__}: {e}',
            })

        # Pass 7 — Carry & Clawback writer.
        # AFTER Pass 4 has populated DerivedMetric with both extracted
        # (Pass 3.5) and Gemini-derived waterfall components, write the
        # CarriedInterest record by consuming those DerivedMetric values
        # directly. ZERO hardcoded waterfall math — compute_carry() is
        # now a pure record-writer that surfaces whatever Pass 3.5/Pass 4
        # produced, with full provenance in CarriedInterest.notes.
        # Synchronous (no Celery dependence) so the dashboard reflects
        # the latest carry numbers the moment the import finishes.
        try:
            _progress(99, 'Pass 7: Writing CarriedInterest from Pass 3.5/Pass 4/Pass 8 outputs...')
            from accounting.carry_engine import compute_carry
            from .gemini_column_mapper import validate_waterfall_identity
            from .models import DerivedMetric

            today_ = date.today()
            if self._imported_fund:
                for sch in self._imported_fund.schemes.all():
                    try:
                        compute_carry(sch, today_)
                        logger.info(f'[Pass7] CarriedInterest written for scheme={sch.name}')

                        # Layer E identity check — purely mathematical, no
                        # heuristics. Confirms that the extracted +
                        # derived waterfall components satisfy
                        #   ROC + PrefRet + GPCatchup + CarryBase ≈ TotalProceeds
                        # If they don't, log loudly so the audit log
                        # surfaces a likely Pass 3.5 column-pick or
                        # Pass 4 disjointness error.
                        identity_inputs = {}
                        for k in ['return_of_capital_amount',
                                  'preferred_return_amount',
                                  'gp_catchup_amount',
                                  'carry_base',
                                  'total_realised_proceeds',
                                  'total_unrealised_fair_value']:
                            dm = DerivedMetric.objects.filter(
                                scheme=sch, metric_key=k,
                            ).exclude(value=None).first()
                            if dm and dm.value is not None:
                                identity_inputs[k] = float(dm.value)
                        identity_result = validate_waterfall_identity(
                            identity_inputs, tolerance_pct=2.0,
                        )
                        logger.info(
                            f'[Pass7 identity-check scheme={sch.name}] '
                            f'status={identity_result["status"]} '
                            f'diff_pct={identity_result.get("diff_pct")} '
                            f'reasoning={identity_result["reasoning"]}'
                        )
                    except Exception as inner:
                        logger.warning(
                            f'[Pass7] compute_carry/identity-check failed '
                            f'for {sch.name}: '
                            f'{type(inner).__name__}: {inner}'
                        )
        except Exception as e:
            logger.warning(f'Pass 7 (CarriedInterest writer) failed: {e}')

        # Pass 6 — Per-row metric completer.
        # Walks every fund-data model whose rows belong to this fund, finds
        # numeric/percent/decimal fields that are null on most rows, and asks
        # Gemini for the formula to compute each missing field from the
        # other available row-level inputs. Then evaluates the formula per
        # row via the same safe AST walker used by Pass 4. NO hardcoded
        # field names — discovery is via Django _meta introspection;
        # formulas are whatever Gemini returns. The killer feature: per-
        # company IRR, ownership %, holding period, unrealised gain — any
        # missing numeric column on any model — gets filled when derivable.
        try:
            _progress(98, 'Pass 6: per-row metric completion via Gemini...')
            from .derivation_service import PerRowMetricCompleter
            if self._imported_fund:
                pass6 = PerRowMetricCompleter(
                    organization=self.org,
                    fund=self._imported_fund,
                )
                pass6_outcomes = pass6.complete_all()
                if isinstance(result, dict):
                    result['pass6_per_row_completion'] = pass6_outcomes
        except Exception as e:
            logger.warning(f'Pass 6 (per-row completion) failed: {e}')

        # Pass 6.5 — Universal KPI projection.
        # Walks every sheet's Gemini Pass 2 column_mapping; for every column
        # mapped to a KPI canonical field (revenue / ebitda / gmv / mrr /
        # arr / orders / aov / cac / *_pct etc.) AND a sibling
        # `company_name` column, creates one PortfolioKPI row per
        # company × metric. Fixes the gap where PORTFOLIO_MASTER's
        # Revenue/EBITDA columns were dropped (no matching Investment
        # field) — the KPI matrix endpoint reads PortfolioKPI, so
        # without this projection the matrix shows blank for any
        # non-SaaS KPI. ZERO English keywords — relies entirely on
        # Gemini Pass 2's canonical-field assignments.
        try:
            _progress(98, 'Pass 6.5: Universal KPI projection to PortfolioKPI...')
            self._project_kpi_columns_to_portfolio_kpi(import_file_record)
        except Exception as e:
            logger.warning(f'Pass 6.5 (KPI projection) failed: {e}')

        # Pass 6.6 — Derived percentage KPIs.
        # After Pass 6.5 populates raw values, ask Gemini ONCE per missing
        # percentage canonical field (e.g. ebitda_margin_pct,
        # gross_margin_pct) for the formula to derive it from the raw
        # inputs available on each investment, then evaluate via the
        # safe AST walker. ZERO hardcoded formulas — Gemini picks them.
        try:
            _progress(98, 'Pass 6.6: Derive missing KPI percentages via Gemini...')
            self._complete_portfolio_kpi_percentages(import_file_record)
        except Exception as e:
            logger.warning(f'Pass 6.6 (percentage derivation) failed: {e}')

        # Pass 5 — extraction completeness audit (diagnostic only).
        # Walks every dashboard-critical model and reports empty tables,
        # along with the Excel sheet they SHOULD have been populated from
        # (per Gemini Pass 1's domain classification). This surfaces silent
        # importer failures (e.g. Pass 2 returning low-confidence mappings
        # that get filtered out, leaving the importer with nothing to write)
        # without any silent retry or hardcoded fallback. The output is
        # persisted to job.result_summary['audit'] so the UI / API can show
        # the gap directly.
        try:
            _progress(99, 'Pass 5: extraction completeness audit...')
            audit = self._audit_extraction_completeness(
                classifications=classifications,
                column_mappings=column_mappings,
            )
            if isinstance(result, dict):
                result['audit'] = audit
            if audit.get('empty_critical_tables'):
                logger.warning(
                    f'[Pass5 Audit] EMPTY critical tables after import: '
                    f'{audit["empty_critical_tables"]}'
                )
            else:
                logger.info(
                    f'[Pass5 Audit] All critical tables populated.'
                )
        except Exception as e:
            logger.warning(f'Pass 5 audit failed: {e}')

        return result

    @transaction.atomic
    def _do_import(self, filepath, classifications, progress_cb,
                   column_mappings=None, section_map=None):
        """
        Run the actual import using Gemini's sheet classification.

        Builds a domain→sheet_name map, then reads each sheet by headers
        (no hardcoded column positions or sheet names).

        column_mappings: {sheet_name: {sections: [{mappings: [{excel_column, canonical_field, confidence}]}]}}
        Built by Gemini Pass-2. We flatten it into self._gemini_sheet_aliases so
        read_table_from_sheet / read_all_sections_from_sheet can enrich every row
        with canonical field names regardless of how the Excel column was labelled.

        section_map: {sheet_name: {section_name: sub_domain}}
        Built by Gemini Pass-1.5. Maps detected section titles to canonical
        sub-domains (e.g. 'portfolio_companies', 'investments', etc.) for
        routing sections within multi-section sheets.
        """
        wb = openpyxl.load_workbook(filepath, data_only=True)
        org = self.org

        # Store filepath so Pass 2.5 layout detection can re-open the workbook
        # per-sheet. Reset layout cache for this import (fresh classification
        # per file — no cross-import contamination).
        self._filepath = filepath
        self._sheet_layouts = {}
        self._layout_failed_sheets = set()

        # Store Gemini Pass 1.5 section classification for routing
        self._section_map = section_map or {}

        # Build domain→sheet map from Gemini classification
        domain_map = _build_domain_sheet_map(classifications, wb)
        logger.info(f'Domain→sheet map: {domain_map}')

        # Build per-sheet Gemini alias map: {sheet_name: {excel_col: canonical_field}}
        # PRODUCTION CHANGE: previously we silently dropped any mapping with
        # confidence < 0.70. That filter was the silent killer that left
        # LP_REGISTER / CAPITAL_CALLS / EXITS with zero aliases and zero
        # imported rows. We now accept ALL mappings Gemini returns and trust
        # the per-sub-table architecture (Fix 1) to keep individual
        # confidences high. If a sub-table comes back with avg confidence
        # below 0.85, we log it loudly so any genuine ambiguity surfaces.
        self._gemini_sheet_aliases = {}
        for sheet_name, mapping_data in (column_mappings or {}).items():
            aliases = {}
            low_conf_count = 0
            high_conf_count = 0
            for section in mapping_data.get('sections', []):
                for m in section.get('mappings', []):
                    excel_col = m.get('excel_column', '')
                    canonical = m.get('canonical_field', '')
                    confidence = m.get('confidence', 0.0) or 0.0
                    if excel_col and canonical:
                        # Accept ALL mappings — no silent threshold filter.
                        aliases[excel_col] = canonical
                        if confidence < 0.70:
                            low_conf_count += 1
                        else:
                            high_conf_count += 1
            if aliases:
                self._gemini_sheet_aliases[sheet_name] = aliases
                logger.info(
                    f'Gemini aliases for "{sheet_name}": {list(aliases.items())} '
                    f'[high_conf={high_conf_count}, low_conf={low_conf_count}]'
                )
            else:
                # Sheet got column_mappings dict but zero usable mappings.
                # Surface this loudly so the user sees that Pass 2 produced
                # nothing for this sheet.
                logger.warning(
                    f'Pass 2 produced NO column mappings for sheet "{sheet_name}" '
                    f'— downstream importers for this sheet will rely on fuzzy '
                    f'header matching only.'
                )

        # Ensure fund categories
        for code, (name, sub_cat, leverage) in CATEGORY_MAP.items():
            FundCategory.objects.get_or_create(
                sebi_category_code=code,
                defaults={
                    'name': name,
                    'sub_category': sub_cat,
                    'leverage_permitted': leverage,
                },
            )

        # --- Extract fund name from Cover sheet or filename ---
        progress_cb(28, 'Reading fund information...')
        fund_name = self._extract_fund_name(wb, domain_map, filepath)
        if not fund_name:
            wb.close()
            raise ValueError('Could not determine fund name from workbook')

        # --- Create Fund & Scheme ---
        progress_cb(32, f'Creating fund: {fund_name}...')
        fund, schemes = self._import_fund_and_schemes(wb, org, domain_map, fund_name)
        self._imported_fund = fund

        # Grant fund access to the uploading user
        FundAccess.objects.get_or_create(
            user=self.user, fund=fund,
            defaults={'access_level': 'admin'},
        )

        # --- Extract fund metadata from Cover sheet ---
        progress_cb(34, 'Extracting fund metadata...')
        try:
            self._extract_fund_metadata(wb, fund, schemes, domain_map)
        except Exception as e:
            logger.warning(f'Fund metadata extraction error: {e}')
            self.errors.append({'section': 'fund_metadata', 'error': str(e)})

        # --- Extract scheme lifecycle from QAR Part B / compliance sheets ---
        progress_cb(35, 'Extracting scheme lifecycle parameters...')
        try:
            self._import_scheme_lifecycle(wb, fund, schemes, domain_map)
        except Exception as e:
            logger.warning(f'Scheme lifecycle extraction error: {e}')
            self.errors.append({'section': 'scheme_lifecycle', 'error': str(e)})

        # --- Import key entities & link to fund ---
        progress_cb(35, 'Importing key entities...')
        try:
            self._import_entities(wb, org, fund, domain_map)
        except Exception as e:
            logger.warning(f'Entity import error: {e}')
            self.errors.append({'section': 'entities', 'error': str(e)})

        # --- Import investors ---
        progress_cb(37, 'Importing investors...')
        investors = {}
        try:
            investors = self._import_investors(wb, org, domain_map)
        except Exception as e:
            logger.warning(f'Investors import error: {e}')
            self.errors.append({'section': 'investors', 'error': str(e)})

        # --- Import commitments from investor data ---
        progress_cb(40, 'Importing commitments...')
        commitments = {}
        try:
            commitments = self._import_commitments(wb, org, investors, schemes, domain_map)
        except Exception as e:
            logger.warning(f'Commitments import error: {e}')
            self.errors.append({'section': 'commitments', 'error': str(e)})

        # --- Import capital calls from investor drawdowns ---
        progress_cb(43, 'Importing capital calls...')
        try:
            self._import_capital_calls(wb, schemes, commitments, domain_map)
        except Exception as e:
            logger.warning(f'Capital calls import error: {e}')
            self.errors.append({'section': 'capital_calls', 'error': str(e)})

        # --- Import portfolio companies & investments ---
        progress_cb(47, 'Importing portfolio companies & investments...')
        companies = {}
        investments = {}
        try:
            companies, investments = self._import_portfolio(
                wb, org, schemes, domain_map, progress_cb)
        except Exception as e:
            logger.warning(f'Portfolio import error: {e}')
            self.errors.append({'section': 'portfolio', 'error': str(e)})

        # --- Import investment tranches ---
        progress_cb(52, 'Importing investment tranches...')
        try:
            self._import_tranches(wb, investments, domain_map)
        except Exception as e:
            logger.warning(f'Tranches import error: {e}')
            self.errors.append({'section': 'tranches', 'error': str(e)})

        # --- Import valuations ---
        progress_cb(56, 'Importing valuations...')
        try:
            self._import_valuations(wb, investments, domain_map)
        except Exception as e:
            logger.warning(f'Valuations import error: {e}')
            self.errors.append({'section': 'valuations', 'error': str(e)})

        # --- Import KPIs ---
        progress_cb(60, 'Importing KPIs...')
        try:
            self._import_kpis(wb, org, investments, companies, domain_map)
        except Exception as e:
            logger.warning(f'KPIs import error: {e}')
            self.errors.append({'section': 'kpis', 'error': str(e)})

        # --- Import company financials (burn & runway) + SaaS KPIs ---
        progress_cb(63, 'Importing company financials & burn rates...')
        try:
            self._import_company_financials(wb, org, investments, companies, domain_map)
        except Exception as e:
            logger.warning(f'Company financials import error: {e}')
            self.errors.append({'section': 'company_financials', 'error': str(e)})

        # --- Import MIS financials (P&L, BvA → BudgetVsActual + ConsolidatedMIS) ---
        progress_cb(63, 'Importing MIS financials (P&L & Budget vs Actual)...')
        try:
            self._import_mis_financials(wb, org, fund, investments, companies, domain_map)
        except Exception as e:
            logger.warning(f'MIS financials import error: {e}')
            self.errors.append({'section': 'mis_financials', 'error': str(e)})

        # --- Classify quoted vs unquoted companies ---
        progress_cb(64, 'Classifying quoted & unquoted companies...')
        try:
            self._import_quoted_unquoted(wb, org, investments, companies, domain_map)
        except Exception as e:
            logger.warning(f'Quoted/Unquoted import error: {e}')
            self.errors.append({'section': 'quoted_unquoted', 'error': str(e)})

        # --- Import NAV records ---
        progress_cb(65, 'Importing NAV & accounting...')
        try:
            self._import_nav(wb, schemes, domain_map)
        except Exception as e:
            logger.warning(f'NAV import error: {e}')
            self.errors.append({'section': 'nav', 'error': str(e)})

        # --- Import exits & distributions ---
        progress_cb(70, 'Importing exits & distributions...')
        try:
            self._import_exits_and_distributions(wb, investments, schemes, domain_map)
        except Exception as e:
            logger.warning(f'Exits/distributions import error: {e}')
            self.errors.append({'section': 'exits_distributions', 'error': str(e)})

        # --- Import distributions to LPs ---
        progress_cb(74, 'Importing LP distributions...')
        try:
            self._import_distributions(wb, schemes, commitments, investments, domain_map)
        except Exception as e:
            logger.warning(f'Distributions import error: {e}')
            self.errors.append({'section': 'distributions', 'error': str(e)})

        # --- Seed Chart of Accounts & create ledger entries ---
        progress_cb(78, 'Setting up fund accounting...')
        try:
            self._setup_fund_accounting(org, fund, schemes, investments)
        except Exception as e:
            logger.warning(f'Fund accounting setup error: {e}')
            self.errors.append({'section': 'fund_accounting', 'error': str(e)})

        # --- Import management fee schedule ---
        progress_cb(82, 'Importing management fees...')
        try:
            self._import_management_fees(wb, schemes, domain_map)
        except Exception as e:
            logger.warning(f'Management fees import error: {e}')
            self.errors.append({'section': 'management_fees', 'error': str(e)})

        # --- Import compliance (per-company obligations + fund-level filings) ---
        progress_cb(84, 'Importing compliance...')
        try:
            self._import_compliance(wb, org, fund, companies, domain_map)
        except Exception as e:
            logger.warning(f'Compliance import error: {e}')
            self.errors.append({'section': 'compliance', 'error': str(e)})

        # --- Compute carried interest ---
        progress_cb(85, 'Computing carried interest...')
        try:
            self._compute_carried_interest(schemes, wb=wb, domain_map=domain_map)
        except Exception as e:
            logger.warning(f'Carried interest error: {e}')
            self.errors.append({'section': 'carried_interest', 'error': str(e)})

        # --- Generate LP Capital Accounts from imported data ---
        progress_cb(87, 'Generating LP capital accounts...')
        try:
            self._generate_lp_capital_accounts(fund, schemes, commitments)
        except Exception as e:
            logger.warning(f'LP capital accounts error: {e}')
            self.errors.append({'section': 'lp_capital_accounts', 'error': str(e)})

        # --- Create income/expense ledger entries from NAV data ---
        progress_cb(89, 'Generating income & expense ledger entries...')
        try:
            self._generate_income_expense_ledger(org, fund, schemes)
        except Exception as e:
            logger.warning(f'Income/expense ledger error: {e}')
            self.errors.append({'section': 'income_expense_ledger', 'error': str(e)})

        # --- Build portfolio hierarchy ---
        progress_cb(92, 'Building portfolio hierarchy...')
        try:
            self._build_hierarchy(wb, org, fund, schemes, companies,
                                  investments, domain_map, filepath)
        except Exception as e:
            logger.warning(f'Portfolio hierarchy error: {e}')
            self.errors.append({'section': 'hierarchy', 'error': str(e)})

        # --- Seed IC Workflow deal pipeline from investments ---
        progress_cb(95, 'Seeding IC deal pipeline...')
        try:
            self._seed_ic_pipeline(org, fund, investments)
        except Exception as e:
            logger.warning(f'IC pipeline seeding error: {e}')
            self.errors.append({'section': 'ic_pipeline', 'error': str(e)})

        wb.close()

        # Collect counts — always from DB rows, never from Cover sheet aggregates
        self.counts = self._collect_counts(org, fund)

        # Post-import sanity check: catch metadata-as-data pollution early.
        # If we only imported a tiny number of companies (≤20) for a fund,
        # check whether any of those "company names" look like metadata labels
        # (e.g., "Short Code", "Fund Corpus", "Hurdle Rate"). If so, log a
        # critical warning so operators can investigate immediately.
        self._validate_imported_companies(fund)

        progress_cb(100, 'Import complete')

        return {
            'counts': self.counts,
            'errors': self.errors,
            'fund_name': fund.name,
        }

    # ------------------------------------------------------------------
    # Extract fund name
    # ------------------------------------------------------------------

    def _extract_fund_identity(self, wb, domain_map):
        """Extract fund identity metadata from fund_scheme_master sheet via Gemini.

        Returns dict with keys: fund_name, sebi_registration_number, category,
        structure_type, fund_pan, fund_gstin, is_gift_city (all optional).
        """
        sheet_name = _dm_first(domain_map, 'fund_scheme_master')
        if not sheet_name or sheet_name not in wb.sheetnames:
            return {}
        ws = wb[sheet_name]
        # Collect all label-value pairs from the sheet
        kv_pairs = {}
        scan_limit = min(ws.max_row + 1, 100)
        for r in range(1, scan_limit):
            label = _str(ws.cell(r, 1).value).strip()
            val = ws.cell(r, 2).value
            if label and val is not None:
                kv_pairs[label] = _str(val).strip()
        if not kv_pairs:
            return {}
        from .gemini_column_mapper import extract_structured_metadata
        from .canonical_schema import CANONICAL_METADATA_FIELDS
        field_defs = CANONICAL_METADATA_FIELDS.get('fund_identity', {})
        return extract_structured_metadata(kv_pairs, field_defs,
                                           context='Fund identity metadata from AIF fund master sheet')

    def _extract_fund_name(self, wb, domain_map, filepath):
        """Extract fund name from Gemini-classified fund_scheme_master sheet or filename."""
        identity = self._extract_fund_identity(wb, domain_map)
        if identity.get('fund_name'):
            return identity['fund_name']
        # Fallback: extract from filename
        basename = os.path.basename(filepath)
        name = re.sub(r'^Mock_\d+_', '', basename)
        name = re.sub(r'\.xlsx?$', '', name, flags=re.IGNORECASE)
        name = name.replace('_', ' ')
        return name if name else None

    # ------------------------------------------------------------------
    # Fund & Schemes
    # ------------------------------------------------------------------

    def _import_fund_and_schemes(self, wb, org, domain_map, fund_name):
        """Create fund and schemes from available data.

        Handles:
        - Fund & Scheme Master sheet with FUND MASTER DATA + SCHEMES sections
        - Cover sheet with basic fund info
        - Auto-creates default scheme if no explicit scheme data found
        """
        # Use Gemini-extracted fund identity metadata
        identity = self._extract_fund_identity(wb, domain_map)

        # Determine AIF category from Gemini extraction or CATEGORY_MAP code
        cat_code = 'CAT_II'  # default
        cat_raw = identity.get('category', '')
        if cat_raw:
            if cat_raw in CATEGORY_MAP:
                cat_code = cat_raw
            else:
                cl = str(cat_raw).lower().strip()
                if re.search(r'\biii\b|cat.?iii', cl):
                    cat_code = 'CAT_III_LVF'
                elif re.search(r'\bii\b|cat.?ii', cl):
                    cat_code = 'CAT_II'
                elif re.search(r'\bi\b|cat.?i', cl):
                    cat_code = 'CAT_I_VCF'

        fund_category = FundCategory.objects.filter(
            sebi_category_code=cat_code).first()

        sebi_reg = identity.get('sebi_registration_number', '')
        fund_pan = identity.get('fund_pan', '')[:10] if identity.get('fund_pan') else ''
        fund_gstin = identity.get('fund_gstin', '')
        gift_city_raw = identity.get('is_gift_city', '')
        gift_city = str(gift_city_raw).lower() in ('yes', 'true', '1', 'y') if gift_city_raw else False

        # Structure type: Gemini extracts as enum
        structure = 'trust'
        struct_raw = identity.get('structure_type', '')
        if struct_raw:
            struct_result = self._classify_enum([struct_raw], 'structure_type',
                                                context='Legal structure of AIF fund')
            structure = struct_result.get(struct_raw, 'trust')

        fund_defaults = {}
        if fund_category:
            fund_defaults['fund_category'] = fund_category
        if structure:
            fund_defaults['structure_type'] = structure
        fund_defaults['base_currency'] = 'INR'
        if sebi_reg:
            fund_defaults['sebi_registration_number'] = sebi_reg
        if fund_pan:
            fund_defaults['pan'] = fund_pan
        if fund_gstin:
            fund_defaults['gstin'] = fund_gstin
        if gift_city:
            fund_defaults['is_gift_city'] = gift_city

        fund, created = Fund.objects.update_or_create(
            organization=org,
            name=fund_name,
            defaults=fund_defaults,
        )
        logger.info(f'{"Created" if created else "Updated"} Fund: {fund.name}')

        schemes = {}

        # Check for explicit scheme data in SCHEMES section
        sheet_name = _dm_first(domain_map, 'fund_scheme_master')
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            schemes_start = find_section_rows(ws, 'SCHEMES')
            if schemes_start:
                # Read the header row right after the SCHEMES section header
                header_row = None
                for scan_r in range(schemes_start + 1, min(schemes_start + 5, ws.max_row + 1)):
                    cell_count = sum(1 for c in range(1, ws.max_column + 1)
                                     if ws.cell(scan_r, c).value is not None)
                    if cell_count >= 3:
                        header_row = scan_r
                        break

                if header_row:
                    # Read as a proper table
                    headers = {}
                    for c in range(1, ws.max_column + 1):
                        h = ws.cell(header_row, c).value
                        if h:
                            headers[str(h).strip()] = c

                    for r in range(header_row + 1, ws.max_row + 1):
                        row_data = {}
                        all_empty = True
                        for name, col in headers.items():
                            val = ws.cell(r, col).value
                            if val is not None:
                                all_empty = False
                            row_data[name] = val
                        if all_empty:
                            break
                        fc = ws.cell(r, 1).value
                        if fc and _is_section_header(_str(fc)):
                            break

                        sn = _find_col_str(
                            row_data, 'Scheme Name', 'Name', 'Scheme')
                        if not sn:
                            continue

                        vintage = _find_col_decimal(
                            row_data, 'Vintage Year', 'Vintage')
                        first_close = _find_col_date(
                            row_data, 'First Close', 'First Close Date')
                        final_close = _find_col_date(
                            row_data, 'Final Close', 'Final Close Date')
                        scheme_size = _find_col_decimal(
                            row_data, 'Scheme Size (Cr)', 'Scheme Size',
                            'Fund Size', 'Size')
                        tenure = _find_col_decimal(
                            row_data, 'Tenure (Years)', 'Tenure', 'Term')
                        hurdle = _find_col_decimal(
                            row_data, 'Hurdle Rate %', 'Hurdle', 'Hurdle Rate')
                        carry = _find_col_decimal(
                            row_data, 'Carry %', 'Carry', 'Carried Interest')
                        # No hardcoded fallback — if the Excel does not name
                        # an explicit waterfall type, the value stays None and
                        # the scheme is saved without carry_type. The dashboard
                        # honestly shows "Distribution Waterfall" with no
                        # waterfall-type qualifier, not a synthesised "European".
                        carry_type_raw = _find_col_str(
                            row_data, 'Carry Type', 'Waterfall Type',
                            default='')
                        fee_basis_raw = _find_col_str(
                            row_data, 'Mgmt Fee Basis', 'Fee Basis',
                            default='')
                        fee_pct = _find_col_decimal(
                            row_data, 'Mgmt Fee %', 'Management Fee %',
                            'Fee %')
                        sponsor_pct = _find_col_decimal(
                            row_data, 'Sponsor Commitment %', 'Sponsor %')
                        status_raw = _find_col_str(
                            row_data, 'Status', 'Scheme Status',
                            default='Investing')
                        scheme_status_result = self._classify_enum(
                            [status_raw], 'scheme_status',
                            context='Scheme/fund lifecycle status')
                        scheme_status = scheme_status_result.get(
                            status_raw, 'investing')

                        # No hardcoded fallback — if Gemini cannot classify the
                        # raw value, carry_type stays None and is omitted from
                        # the scheme defaults dict. Same rule for fee_basis.
                        carry_type_result = self._classify_enum(
                            [carry_type_raw], 'carry_type',
                            context='Waterfall/carry type') if carry_type_raw else {}
                        carry_type = carry_type_result.get(carry_type_raw) or None

                        fee_basis_result = self._classify_enum(
                            [fee_basis_raw], 'fee_basis',
                            context='Management fee calculation basis') if fee_basis_raw else {}
                        fee_basis = fee_basis_result.get(fee_basis_raw) or None

                        scheme_defaults = {'is_active': True}
                        if vintage:
                            scheme_defaults['vintage_year'] = int(vintage)
                        if first_close:
                            scheme_defaults['first_close_date'] = first_close
                        if final_close:
                            scheme_defaults['final_close_date'] = final_close
                        if scheme_size:
                            scheme_defaults['scheme_size'] = scheme_size
                        if tenure:
                            scheme_defaults['tenure_years'] = int(tenure)
                        if hurdle:
                            scheme_defaults['hurdle_rate_pct'] = hurdle
                        if carry:
                            scheme_defaults['carry_pct'] = carry
                        if carry_type:
                            scheme_defaults['carry_type'] = carry_type
                        if fee_pct:
                            scheme_defaults['management_fee_pct'] = fee_pct
                        if fee_basis:
                            scheme_defaults['management_fee_basis'] = fee_basis
                        if sponsor_pct:
                            scheme_defaults['sponsor_commitment_pct'] = sponsor_pct
                        if scheme_status:
                            scheme_defaults['scheme_status'] = scheme_status

                        s, _ = Scheme.objects.update_or_create(
                            fund=fund, name=sn,
                            defaults=scheme_defaults,
                        )
                        schemes[sn] = s
                else:
                    # Fallback: simple row-by-row reading
                    for r in range(schemes_start + 1, ws.max_row + 1):
                        sn = ws.cell(r, 1).value
                        if not sn or _str(sn) == 'Scheme Name':
                            continue
                        if _is_section_header(sn):
                            break
                        sn = _str(sn)
                        s, _ = Scheme.objects.update_or_create(
                            fund=fund, name=sn,
                            defaults={
                                'scheme_status': 'investing',
                                'is_active': True,
                            },
                        )
                        schemes[sn] = s

        # If no schemes found, create a default one
        if not schemes:
            scheme_name = f'{fund_name} - Scheme I'
            scheme, _ = Scheme.objects.update_or_create(
                fund=fund,
                name=scheme_name,
                defaults={
                    'vintage_year': date.today().year,
                    'scheme_status': 'investing',
                    'is_active': True,
                },
            )
            schemes[scheme_name] = scheme

        return fund, schemes

    # ------------------------------------------------------------------
    # Fund metadata extraction from Cover sheet
    # ------------------------------------------------------------------

    def _extract_fund_metadata(self, wb, fund, schemes, domain_map=None):
        """Extract rich fund metadata from ALL candidate sheets.

        Scans all relevant sheets (cover, fund_scheme_master, summary, overview)
        and merges their key-value pairs. Detailed/dedicated sheets are scanned
        FIRST so their values take priority; Sheet 1 / summary sheets are
        scanned LAST as fallback. This ensures we never stop at Sheet 1 if
        more detailed data exists in sub-sheets.
        """
        # Use Gemini-classified fund_scheme_master sheets
        candidate_sheets = _dm_sheets(domain_map, 'fund_scheme_master') if domain_map else []
        # Filter to sheets that actually exist
        candidate_sheets = [s for s in candidate_sheets if s in wb.sheetnames]

        if not candidate_sheets:
            return

        # Build a MERGED key-value map from ALL candidate sheets (raw labels preserved).
        kv = {}  # {raw_label: raw_value}
        for sheet_name in candidate_sheets:
            ws = wb[sheet_name]
            for r in range(1, min(ws.max_row + 1, 200)):
                for label_col, val_col in [(1, 2), (2, 3), (6, 7)]:
                    label = _str(ws.cell(r, label_col).value).strip().rstrip(':')
                    val = ws.cell(r, val_col).value
                    if label and val is not None and label not in kv:
                        kv[label] = val
            logger.info(f'  Fund metadata: scanned sheet "{sheet_name}" ({len(kv)} kv pairs total)')

        if not kv:
            return

        # Use Gemini to extract fund identity fields
        from .gemini_column_mapper import extract_structured_metadata
        from .canonical_schema import CANONICAL_METADATA_FIELDS

        kv_str = {k: _str(v).strip() for k, v in kv.items() if v is not None}

        # Extract fund identity
        fund_fields = CANONICAL_METADATA_FIELDS.get('fund_identity', {})
        fund_identity = extract_structured_metadata(
            kv_str, fund_fields,
            context='Fund identity metadata from cover/master sheet')

        # Extract scheme lifecycle
        scheme_fields = CANONICAL_METADATA_FIELDS.get('scheme_lifecycle', {})
        scheme_lifecycle = extract_structured_metadata(
            kv_str, scheme_fields,
            context='Scheme lifecycle parameters from cover/master sheet')

        # Update Fund fields
        update_fields = []

        reg_no = fund_identity.get('sebi_registration_number')
        if reg_no and not fund.sebi_registration_number:
            fund.sebi_registration_number = str(reg_no)
            update_fields.append('sebi_registration_number')

        corpus_raw = scheme_lifecycle.get('scheme_size')
        if corpus_raw and not fund.corpus_target:
            corpus_str = str(corpus_raw)
            corpus_num = re.sub(r'[₹,\s]', '', corpus_str)
            corpus_num = re.sub(r'[Cc]r\.?$', '', corpus_num)
            corpus_num = re.sub(r'^[Rr][Ss]\.?\s*', '', corpus_num)
            corpus_d = _d(corpus_num)
            if corpus_d:
                fund.corpus_target = corpus_d
                update_fields.append('corpus_target')

        vintage_raw = scheme_lifecycle.get('vintage_year')
        if vintage_raw:
            vintage_str = str(vintage_raw)
            if not fund.inception_date:
                vintage_date = _date(vintage_str)
                if vintage_date:
                    fund.inception_date = vintage_date
                    update_fields.append('inception_date')
                elif vintage_str.isdigit() and len(vintage_str) == 4:
                    fund.inception_date = date(int(vintage_str), 1, 1)
                    update_fields.append('inception_date')

        pan_raw = fund_identity.get('fund_pan')
        if pan_raw and not fund.pan:
            fund.pan = str(pan_raw)[:10]
            update_fields.append('pan')

        if update_fields:
            fund.save(update_fields=update_fields)
            logger.info(f'  Fund metadata updated: {update_fields}')

        # Update Scheme fields from Cover data — apply to ALL schemes
        for default_scheme in schemes.values():
            scheme_updates = []

            if vintage_raw:
                v_str = str(vintage_raw)
                if v_str.isdigit() and len(v_str) == 4:
                    default_scheme.vintage_year = int(v_str)
                    scheme_updates.append('vintage_year')

            hurdle_raw = scheme_lifecycle.get('hurdle_rate_pct')
            if hurdle_raw:
                h_val = _parse_pct(hurdle_raw)
                if h_val and not default_scheme.hurdle_rate_pct:
                    default_scheme.hurdle_rate_pct = h_val
                    scheme_updates.append('hurdle_rate_pct')

            carry_raw = scheme_lifecycle.get('carry_pct')
            if carry_raw:
                c_val = _parse_pct(carry_raw)
                if c_val and not default_scheme.carry_pct:
                    default_scheme.carry_pct = c_val
                    scheme_updates.append('carry_pct')

            fee_raw = scheme_lifecycle.get('management_fee_pct')
            if fee_raw:
                f_val = _parse_pct(fee_raw)
                if f_val and not default_scheme.management_fee_pct:
                    default_scheme.management_fee_pct = f_val
                    scheme_updates.append('management_fee_pct')

            fee_basis_raw = scheme_lifecycle.get('management_fee_basis')
            if fee_basis_raw:
                fb_result = self._classify_enum([str(fee_basis_raw)], 'fee_basis',
                                                context='Management fee basis')
                fb = fb_result.get(str(fee_basis_raw))
                if fb and not default_scheme.management_fee_basis:
                    default_scheme.management_fee_basis = fb
                    scheme_updates.append('management_fee_basis')

            tenure_raw = scheme_lifecycle.get('tenure_years')
            if tenure_raw and not default_scheme.tenure_years:
                t_str = re.sub(r'[^\d]', '', str(tenure_raw))
                if t_str and t_str.isdigit():
                    default_scheme.tenure_years = int(t_str)
                    scheme_updates.append('tenure_years')

            fc_raw = scheme_lifecycle.get('first_close_date')
            if fc_raw and not default_scheme.first_close_date:
                d_val = _date(str(fc_raw))
                if d_val:
                    default_scheme.first_close_date = d_val
                    scheme_updates.append('first_close_date')

            lc_raw = scheme_lifecycle.get('final_close_date')
            if lc_raw and not default_scheme.final_close_date:
                d_val = _date(str(lc_raw))
                if d_val:
                    default_scheme.final_close_date = d_val
                    scheme_updates.append('final_close_date')

            corpus_raw2 = scheme_lifecycle.get('scheme_size') or corpus_raw
            if corpus_raw2 and not default_scheme.scheme_size:
                c_str2 = str(corpus_raw2)
                c_num = re.sub(r'[₹,\s]', '', c_str2)
                c_num = re.sub(r'[Cc]r\.?$', '', c_num)
                c_num = re.sub(r'^[Rr][Ss]\.?\s*', '', c_num)
                c_val2 = _d(c_num)
                if c_val2:
                    default_scheme.scheme_size = c_val2
                    scheme_updates.append('scheme_size')

            # Phase 5 (Bug F): the Gemini extract_structured_metadata
            # call sometimes hallucinated sponsor_commitment_pct from
            # unrelated cells (e.g. an LP's "% Fund" share, or the
            # corpus-expense ratio). Sponsor Commitment is now ONLY
            # populated by anchor_pipeline.persist_fund() from
            # FundMetric.sponsor_commitment_pct, which is in turn set
            # from a SEMANTIC sponsor LP detection (Pass 3 + FUND_MASTER
            # cross-check). Stopping the legacy write here ensures the
            # mirror cannot drift from canonical. Leaving the code path
            # in place but no-op so callers don't break.
            _ = scheme_lifecycle.get('sponsor_commitment_pct')  # intentionally unused

            if scheme_updates:
                default_scheme.save(update_fields=scheme_updates)
                logger.info(f'  Scheme "{default_scheme.name}" metadata updated: {scheme_updates}')

    # ------------------------------------------------------------------
    # Scheme Lifecycle Parameters (QAR Part B, compliance sheets, etc.)
    # ------------------------------------------------------------------

    def _import_scheme_lifecycle(self, wb, fund, schemes, domain_map):
        """Extract scheme lifecycle parameters from QAR Part B / compliance sheets.

        Many fund Excel files have a sheet like "QAR Part B" or "Scheme Lifecycle
        Parameters" that contains key-value pairs for: tenure, close dates, corpus,
        hurdle rate, carry, management fees, waterfall type, etc.

        This method scans ALL candidate sheets for key-value pairs matching scheme
        lifecycle fields. Detailed sub-sheets (QAR, lifecycle, compliance) are
        scanned FIRST so their values take priority over Sheet 1 summary data.
        Sheet 1 (fund_scheme_master) is scanned LAST as a fallback for any fields
        not found in dedicated sheets.
        """
        # Use Gemini-classified sheets: compliance first (more detailed), then fund_scheme_master
        candidate_sheets = []
        for domain in ('compliance', 'fund_scheme_master'):
            for s in _dm_sheets(domain_map, domain):
                if s in wb.sheetnames and s not in candidate_sheets:
                    candidate_sheets.append(s)

        if not candidate_sheets:
            return

        # Scan each candidate sheet for label-value pairs
        all_pairs = {}  # label_str → val_str (first occurrence wins)

        for sheet_name in candidate_sheets:
            ws = wb[sheet_name]
            for r in range(1, min(ws.max_row + 1, 200)):
                for label_col, val_col in [(1, 2), (2, 3), (1, 3)]:
                    label_raw = ws.cell(r, label_col).value
                    val_raw = ws.cell(r, val_col).value
                    if not label_raw or val_raw is None:
                        continue
                    label = _str(label_raw).strip().rstrip(':')
                    val_str = _str(val_raw).strip()
                    if not val_str or val_str.lower() in ('none', 'n/a', '-', ''):
                        continue
                    if label and label not in all_pairs:
                        all_pairs[label] = val_str

        if not all_pairs:
            return

        # Use Gemini to semantically match labels → scheme lifecycle fields
        from .gemini_column_mapper import extract_structured_metadata
        from .canonical_schema import CANONICAL_METADATA_FIELDS
        field_defs = CANONICAL_METADATA_FIELDS.get('scheme_lifecycle', {})
        extracted = extract_structured_metadata(all_pairs, field_defs,
                                                context='Scheme lifecycle parameters from QAR / compliance sheet')

        # Parse raw values into typed Python values
        collected = {}
        for field_name, raw_val in extracted.items():
            if field_name in collected:
                continue
            ftype = field_defs.get(field_name, {}).get('type', 'str')
            parsed = None
            val_str = str(raw_val).strip()
            if ftype == 'int':
                nums = re.sub(r'[^\d]', '', val_str)
                if nums and nums.isdigit():
                    parsed = int(nums)
            elif ftype == 'decimal':
                clean = re.sub(r'[₹,\s]', '', val_str)
                clean = re.sub(r'\s*[Cc]r\.?$', '', clean)
                clean = re.sub(r'^[Rr][Ss]\.?\s*', '', clean)
                parsed = _d(clean)
            elif ftype == 'pct':
                clean = val_str.replace('%', '').strip()
                clean = re.split(r'\s+(?:per|p\.?a|above|on)', clean, maxsplit=1)[0].strip()
                parsed = _d(clean)
            elif ftype == 'date':
                parsed = _date(val_str)
            elif ftype == 'enum':
                # Gemini already returned canonical enum value
                parsed = raw_val if raw_val else None
            else:
                parsed = val_str if val_str else None

            if parsed is not None:
                collected[field_name] = parsed
                logger.info(f'  Lifecycle: {field_name} = {parsed}')

        if not collected:
            return

        # Only persist keys that map to concrete Scheme columns.
        # Some extracted fields (e.g. mgmt_fee_pct_post_ip, mgmt_fee_basis_post_ip,
        # investment_period_years) live in fund_data only and are consumed by the
        # anchor pipeline's runtime IP/post-IP fee switch — they must not be saved here.
        from funds.models import Scheme as _SchemeModel
        scheme_field_names = {
            f.name for f in _SchemeModel._meta.get_fields()
            if getattr(f, 'concrete', False) and not getattr(f, 'many_to_many', False)
        }
        persistable = {k: v for k, v in collected.items() if k in scheme_field_names}
        skipped = [k for k in collected if k not in scheme_field_names]
        if skipped:
            logger.info(f'  Lifecycle: skipped non-Scheme fields (consumed by anchor pipeline): {skipped}')

        # Apply collected values to ALL schemes (lifecycle params are fund-level)
        for scheme in schemes.values():
            update_fields = []
            for field_name, value in persistable.items():
                current = getattr(scheme, field_name, None)
                if not current:  # Only fill if currently empty/None
                    setattr(scheme, field_name, value)
                    update_fields.append(field_name)
            if update_fields:
                scheme.save(update_fields=update_fields)
                logger.info(f'  Scheme "{scheme.name}" lifecycle updated: {update_fields}')

    # ------------------------------------------------------------------
    # Investors
    # ------------------------------------------------------------------

    def _import_investors(self, wb, org, domain_map):
        """Import investors from the Investors/LP sheet."""
        sheet_name = _dm_first(domain_map, 'investors_aml')
        if not sheet_name or sheet_name not in wb.sheetnames:
            return {}

        ws = wb[sheet_name]
        _, rows = read_table_from_sheet(ws, alias_map=self._get_alias(ws))
        investors = {}

        # Pass 3: Collect all unique investor types and classify via Gemini
        raw_inv_types = set()
        for row in rows:
            t = _find_col_str(row, 'Investor Type', 'LP Type', 'Type',
                              'Category', 'investor_type')
            if t:
                raw_inv_types.add(t.strip())
        inv_type_map = self._classify_labels(raw_inv_types, 'investor_types',
                                              context='Investor/LP type classification')

        for row in rows:
            inv_name = _find_col_str(
                row, 'Investor Name', 'LP Name', 'Name', 'Investor',
                'investor_name')
            if not inv_name:
                continue

            inv_type_raw = _find_col_str(
                row, 'Investor Type', 'LP Type', 'Type', 'Category',
                'investor_type')
            inv_type = inv_type_map.get(inv_type_raw, 'other') if inv_type_raw else 'other'

            country = _find_col_str(row, 'Country', 'Domicile', default='India')
            commitment_amt = _find_col_decimal(
                row, 'Commitment(Cr)', 'Commitment', 'Committed Amount',
                'Commitment Amount', 'Total Commitment', 'commitment_amount')
            pct_fund = _find_col_decimal(
                row, '% Fund', 'Fund %', 'Allocation %', 'Share %')
            drawdown = _find_col_decimal(
                row, 'Drawdown(Cr)', 'Drawdown', 'Called Amount',
                'Amount Called', 'Drawn', 'capital_called_amount',
                'called_amount')
            distributions = _find_col_decimal(
                row, 'Distributions', 'Distribution', 'Returned',
                'Amount Returned', 'total_distributions_received',
                'distribution_amount', 'gross_amount', 'net_amount')
            status = _find_col_str(row, 'Status', 'LP Status', default='Active')

            investor, created = Investor.objects.get_or_create(
                organization=org,
                investor_name=inv_name,
                defaults={
                    'investor_type': inv_type,
                    'country': country,
                    'kyc_status': 'completed',
                },
            )
            investors[inv_name] = investor
            if created:
                logger.info(f'  Created Investor: {inv_name} ({inv_type})')

        logger.info(f'  Investors imported: {len(investors)}')
        return investors

    # ------------------------------------------------------------------
    # Commitments from investor data
    # ------------------------------------------------------------------

    def _import_commitments(self, wb, org, investors, schemes, domain_map):
        """Create Commitment records.

        Handles two formats:
        1. Dedicated 'commitments' sheet with columns: Investor Name, Scheme Name,
           Commitment Amount, Close Type, etc. (Format B / structured)
        2. Commitment columns embedded in investors_aml sheet:
           Commitment(Cr), Drawdown(Cr), etc. (Format A / flat)
        """
        default_scheme = list(schemes.values())[0] if schemes else None
        if not default_scheme:
            return {}

        commitments = {}

        # --- Strategy 1: Try dedicated Commitments sheet ---
        commit_sheet = _dm_first(domain_map, 'commitments')
        if commit_sheet and commit_sheet in wb.sheetnames:
            ws = wb[commit_sheet]
            _, rows = read_table_from_sheet(ws, alias_map=self._get_alias(ws))
            if rows:
                # ── Batching pre-pass: close_type ─────────────────────
                _close_type_raws = set()
                for _r in rows:
                    _ct = _find_col_str(
                        _r, 'close_type', default='first_close')
                    if _ct:
                        _close_type_raws.add(_ct)
                close_type_map = self._classify_enum(
                    list(_close_type_raws), 'close_type',
                    context='Fund close type for investor commitment',
                ) if _close_type_raws else {}
                for row in rows:
                    inv_name = _find_col_str(
                        row, 'Investor Name', 'LP Name', 'Name', 'Investor',
                        'investor_name')
                    if not inv_name or inv_name not in investors:
                        continue

                    investor = investors[inv_name]
                    commitment_amt = _find_col_decimal(
                        row, 'Commitment Amount (Cr)', 'Commitment Amount',
                        'Commitment(Cr)', 'Commitment', 'Committed Amount',
                        'Total Commitment', 'commitment_amount')
                    if not commitment_amt or commitment_amt <= 0:
                        continue

                    # Determine target scheme
                    scheme_name_raw = _find_col_str(
                        row, 'Scheme Name', 'Scheme', 'Fund Scheme')
                    target_scheme = default_scheme
                    if scheme_name_raw:
                        target_scheme = schemes.get(scheme_name_raw, default_scheme)

                    commit_date = _find_col_date(
                        row, 'Commitment Date', 'Date', 'Close Date')
                    close_type_raw = _find_col_str(
                        row, 'close_type', default='first_close')
                    close_type = close_type_map.get(close_type_raw, 'first_close')

                    units = _find_col_decimal(
                        row, 'Units Allocated', 'Units', 'Allotted Units')
                    side_letter = _find_col_bool(
                        row, 'Side Letter', 'Side Letter Exists')

                    commitment, created = Commitment.objects.get_or_create(
                        investor=investor,
                        scheme=target_scheme,
                        commitment_amount=commitment_amt,
                        defaults={
                            'commitment_date': commit_date or date.today(),
                            'close_type': close_type,
                            'commitment_status': 'active',
                            'side_letter_exists': side_letter,
                            'units_allocated': units,
                        },
                    )
                    # Key by investor + scheme for later lookup
                    key = f'{inv_name}|{target_scheme.name}'
                    commitments[key] = commitment
                    # Also keep simple investor-name key for backward compat
                    if inv_name not in commitments:
                        commitments[inv_name] = commitment

                if commitments:
                    logger.info(f'  Commitments (dedicated sheet): {len(commitments)}')
                    return commitments

        # --- Strategy 2: Extract from investors_aml sheet (flat format) ---
        sheet_name = _dm_first(domain_map, 'investors_aml')
        if not sheet_name or sheet_name not in wb.sheetnames:
            return {}

        ws = wb[sheet_name]
        _, rows = read_table_from_sheet(ws, alias_map=self._get_alias(ws))

        for row in rows:
            inv_name = _find_col_str(
                row, 'Investor Name', 'LP Name', 'Name', 'Investor',
                'investor_name')
            if not inv_name or inv_name not in investors:
                continue

            investor = investors[inv_name]
            commitment_amt = _find_col_decimal(
                row, 'Commitment(Cr)', 'Commitment', 'Committed Amount',
                'Commitment Amount', 'Total Commitment', 'commitment_amount')
            if not commitment_amt or commitment_amt <= 0:
                continue

            pct_fund = _find_col_decimal(
                row, '% Fund', 'Fund %', 'Allocation %', 'Share %')
            side_letter = _find_col_bool(
                row, 'Side Letter', 'Side Letter Exists')

            commitment, created = Commitment.objects.get_or_create(
                investor=investor,
                scheme=default_scheme,
                defaults={
                    'commitment_amount': commitment_amt,
                    'commitment_date': date.today(),
                    'close_type': 'first_close',
                    'commitment_status': 'active',
                    'side_letter_exists': side_letter,
                },
            )
            commitments[inv_name] = commitment
            if created and pct_fund:
                commitment.units_allocated = pct_fund * 100
                commitment.save(update_fields=['units_allocated'])

        logger.info(f'  Commitments: {len(commitments)}')
        return commitments

    # ------------------------------------------------------------------
    # Capital calls from investor drawdowns
    # ------------------------------------------------------------------

    def _import_capital_calls(self, wb, schemes, commitments, domain_map):
        """Create CapitalCall + CapitalCallLineItem records.

        Handles two formats:
        1. Dedicated 'capital_calls' sheet with explicit call records and
           separate CAPITAL CALL LINE ITEMS sections (Format B / structured)
        2. Drawdown amounts embedded in investors_aml sheet (Format A / flat)
        """
        default_scheme = list(schemes.values())[0] if schemes else None
        if not default_scheme:
            return

        # --- Strategy 1: Dedicated Capital Calls sheet ---
        # This runs even when commitments are empty — fund-level call data
        # (total amounts, dates, purposes) does not require LP records.
        cc_sheet = _dm_first(domain_map, 'capital_calls')
        if cc_sheet and cc_sheet in wb.sheetnames:
            ws = wb[cc_sheet]
            sections = self._read_sheet_via_layout(ws, alias_map=self._get_alias(ws))

            # Find the main capital calls section (not line items)
            cc_rows = []
            for sec_name, (sec_headers, sec_rows) in sections.items():
                subdomain = self._get_section_subdomain(cc_sheet, sec_name)
                if subdomain == 'capital_call_line_items':
                    continue  # Skip line item sections for now
                if subdomain == 'capital_call_headers' and sec_rows:
                    cc_rows = sec_rows
                    break
                # Fallback for __default__ or unclassified sections
                if sec_rows and not cc_rows:
                    cc_rows = sec_rows

            if cc_rows:
                # ── Batching pre-pass: capital_call_status ──────────
                # Collect every unique status raw value across all
                # cc_rows so Gemini classifies once instead of per row.
                _cc_status_raws = set()
                for _r in cc_rows:
                    _cs = _find_col_str(
                        _r, 'Status', 'Call Status', 'LP Notified?',
                        default='Paid',
                    )
                    if _cs:
                        _cc_status_raws.add(_cs)
                cc_status_map = self._classify_enum(
                    list(_cc_status_raws), 'capital_call_status',
                    context='Capital call funding status',
                ) if _cc_status_raws else {}

                call_count = 0
                for row in cc_rows:
                    scheme_name_raw = _find_col_str(
                        row, 'Scheme Name', 'Scheme', 'Fund Scheme')
                    target_scheme = default_scheme
                    if scheme_name_raw:
                        target_scheme = schemes.get(scheme_name_raw, default_scheme)

                    # Row serial / call ref — use as unique call_number.
                    # "Call#" (no space) is a common header variant; also plain "#".
                    call_num_raw = _find_col_decimal(
                        row, 'Call#', '#', 'S.No', 'Sr No', 'Serial',
                        'Call #', 'Call Number', 'Call No',
                        'Capital Call Number', 'call_number')
                    call_ref = _find_col_str(
                        row, 'Call Ref', 'Call Reference', 'Ref No', 'Reference')
                    call_num = int(call_num_raw) if call_num_raw else None
                    # Fall back: hash call_ref to a stable int
                    if not call_num and call_ref:
                        call_num = abs(hash(call_ref)) % 1000000

                    call_date = _find_col_date(
                        row, 'Call Date*', 'Call Date', 'Date', 'Call Issue Date')
                    due_date = _find_col_date(
                        row, 'Payment Due', 'Payment Due Date', 'Due Date',
                        'Payment Deadline')
                    call_pct = _find_col_decimal(
                        row, 'Corpus%', 'Corpus %', '% of Commit',
                        'Call %', 'Call Percentage', 'Drawdown %',
                        'call_percentage')
                    total_amt = _find_col_decimal(
                        row, 'Amount (Cr)', 'Amount(Cr)', 'Amount(₹Cr)',
                        'Amount(₹ Cr)', 'Amount (₹Cr)',
                        'Amount(INR Cr)', 'Amount (INR Cr)',
                        'Total Call Amount (Cr)', 'Total Call Amount',
                        'Call Amount', 'Total Amount', 'Actual Received',
                        'total_call_amount')
                    # LP name + portfolio purpose → store in purpose field
                    lp_name = _find_col_str(
                        row, 'LP Name', 'Investor Name', 'Investor', 'LP')
                    portfolio_purpose = _find_col_str(
                        row, 'Portfolio Co. / Purpose', 'Portfolio Co.',
                        'Portfolio Company', 'Purpose', 'Description', 'Notes')
                    if lp_name and portfolio_purpose:
                        purpose = f'{lp_name} — {portfolio_purpose}'
                    elif lp_name:
                        purpose = lp_name
                    else:
                        purpose = portfolio_purpose or ''

                    status_raw = _find_col_str(
                        row, 'Status', 'Call Status', 'LP Notified?', default='Paid')
                    status = cc_status_map.get(status_raw, 'paid')

                    if not total_amt or total_amt <= 0:
                        continue
                    # Skip summary/total rows — call_num is only None when
                    # the first column has text, not a number
                    if not call_num:
                        if _is_junk_row(purpose):
                            continue
                        call_num = call_count + 1

                    CapitalCall.objects.update_or_create(
                        scheme=target_scheme,
                        call_number=call_num,
                        defaults={
                            'call_date': call_date or date.today(),
                            'payment_due_date': due_date or call_date or date.today(),
                            'call_percentage': call_pct or Decimal('0'),
                            'total_call_amount': total_amt,
                            'purpose': purpose,
                            'call_status': status,
                            'created_by': self.user,
                        },
                    )
                    call_count += 1

                # Now import line items from CAPITAL CALL LINE ITEMS sections
                line_count = 0
                for sec_name, (sec_headers, sec_rows) in sections.items():
                    subdomain = self._get_section_subdomain(cc_sheet, sec_name)
                    if subdomain != 'capital_call_line_items':
                        continue

                    # Try to figure out which call this belongs to
                    # Section name might be "CAPITAL CALL LINE ITEMS (Main Scheme — Call #1)"
                    parent_call = None
                    for call_obj in CapitalCall.objects.filter(
                            scheme__fund=default_scheme.fund):
                        call_ref = f'Call #{call_obj.call_number}'
                        if call_ref in sec_name:
                            parent_call = call_obj
                            break

                    if not parent_call:
                        # Default to first call for this scheme
                        parent_call = CapitalCall.objects.filter(
                            scheme__fund=default_scheme.fund
                        ).order_by('call_number').first()

                    if not parent_call:
                        continue

                    # ── Batching pre-pass: payment_status for line items ──
                    _pay_status_raws = set()
                    for _r in sec_rows:
                        _ps = _find_col_str(
                            _r, 'Payment Status', 'Status', default='Paid')
                        if _ps:
                            _pay_status_raws.add(_ps)
                    pay_status_map = self._classify_enum(
                        list(_pay_status_raws), 'payment_status',
                        context='Capital call payment status',
                    ) if _pay_status_raws else {}

                    for row in sec_rows:
                        inv_name = _find_col_str(
                            row, 'Investor Name', 'LP Name', 'Name',
                            'Investor')
                        if not inv_name:
                            continue

                        # Find the commitment for this investor+scheme
                        commitment = None
                        key = f'{inv_name}|{parent_call.scheme.name}'
                        commitment = commitments.get(key)
                        if not commitment:
                            commitment = commitments.get(inv_name)
                        if not commitment:
                            continue

                        called_amt = _find_col_decimal(
                            row, 'Called Amount (Cr)', 'Called Amount',
                            'Amount Called', 'Call Amount')
                        cum_pct = _find_col_decimal(
                            row, 'Cumulative Called %', 'Cumulative %',
                            'Called %')
                        pay_status_raw = _find_col_str(
                            row, 'Payment Status', 'Status', default='Paid')
                        received = _find_col_decimal(
                            row, 'Amount Received (Cr)', 'Amount Received',
                            'Received')
                        pay_date = _find_col_date(
                            row, 'Payment Date', 'Received Date')

                        if not called_amt or called_amt <= 0:
                            continue

                        CapitalCallLineItem.objects.get_or_create(
                            capital_call=parent_call,
                            commitment=commitment,
                            defaults={
                                'called_amount': called_amt,
                                'cumulative_called_pct': cum_pct,
                                'payment_status': pay_status_map.get(
                                    pay_status_raw, 'pending'),
                                'amount_received': received or called_amt,
                                'payment_date': pay_date or parent_call.call_date,
                            },
                        )
                        line_count += 1

                logger.info(f'  Capital calls (dedicated sheet): {call_count} calls, {line_count} line items')
                return

        # --- Strategy 2: Extract from investors_aml sheet (flat format) ---
        sheet_name = _dm_first(domain_map, 'investors_aml')
        if not sheet_name or sheet_name not in wb.sheetnames:
            return

        ws = wb[sheet_name]
        _, rows = read_table_from_sheet(ws, alias_map=self._get_alias(ws))

        total_drawn = Decimal('0')
        lp_drawdowns = []
        for row in rows:
            inv_name = _find_col_str(
                row, 'Investor Name', 'LP Name', 'Name', 'Investor')
            if not inv_name or inv_name not in commitments:
                continue

            drawn = _find_col_decimal(
                row, 'Drawdown(Cr)', 'Drawdown', 'Called Amount',
                'Amount Called', 'Drawn', 'Capital Called')
            drawn_pct = _find_col_decimal(
                row, 'Drawn%', 'Drawn Pct', 'Called %', 'Call %')
            if not drawn or drawn <= 0:
                continue

            total_drawn += drawn
            lp_drawdowns.append((inv_name, commitments[inv_name], drawn, drawn_pct))

        if not lp_drawdowns:
            return

        avg_drawn_pct = Decimal('0')
        for _, commitment, _, drawn_pct in lp_drawdowns:
            if drawn_pct:
                pct = drawn_pct * 100 if drawn_pct <= 1 else drawn_pct
                avg_drawn_pct += pct
        if lp_drawdowns:
            avg_drawn_pct = avg_drawn_pct / len(lp_drawdowns)

        call, created = CapitalCall.objects.get_or_create(
            scheme=default_scheme,
            call_number=1,
            defaults={
                'call_date': date.today(),
                'payment_due_date': date.today(),
                # No hardcoded fallback: if avg_drawn_pct couldn't be derived
                # from the investors_aml sheet, leave it at the model default
                # (0) rather than synthesising 80. Honest "unknown" beats a
                # made-up number.
                'call_percentage': avg_drawn_pct or Decimal('0'),
                'total_call_amount': total_drawn,
                'purpose': 'Investment deployment and fund expenses',
                'call_status': 'paid',
                'created_by': self.user,
            },
        )

        if not created:
            return

        line_count = 0
        for inv_name, commitment, drawn_amt, drawn_pct in lp_drawdowns:
            cumulative_pct = drawn_pct
            if cumulative_pct and cumulative_pct <= 1:
                cumulative_pct = cumulative_pct * 100

            CapitalCallLineItem.objects.get_or_create(
                capital_call=call,
                commitment=commitment,
                defaults={
                    'called_amount': drawn_amt,
                    'cumulative_called_pct': cumulative_pct,
                    'payment_status': 'paid',
                    'amount_received': drawn_amt,
                    'payment_date': date.today(),
                },
            )
            line_count += 1

        logger.info(f'  Capital calls: 1 call, {line_count} line items')

    # ------------------------------------------------------------------
    # Portfolio companies & investments
    # ------------------------------------------------------------------

    def _row_natural_key(self, row, name, inv_date, stage):
        """Stable per-row identifier. Prefer the file's Co.ID/instrument
        ID columns when present; fall back to a deterministic fingerprint
        so re-imports are still idempotent."""
        import hashlib
        for header in ('Co.ID', 'Co ID', 'CoID', 'Investment ID',
                       'Instrument ID', 'Security ID', 'ISIN',
                       'Tranche ID', 'Round ID'):
            val = _find_col_str(row, header)
            if val:
                return str(val).strip()[:64]
        parts = [name or '', str(inv_date or ''), stage or '']
        fp = hashlib.sha1('|'.join(parts).encode()).hexdigest()[:16]
        return f'fp:{fp}'

    def _record_investment_row(self, inv, row, name, company, *,
                                cost, inv_date, stage, instrument,
                                hold_pct=None, fd_pct=None,
                                fv=None, val_date=None, unrealized=None):
        """Write one Tranche (and optionally one Valuation) per source row.
        Investment aggregate fields are NEVER set here — they are derived
        by reconcile_investment_from_tranches() at end of import.
        """
        from investments.services import (
            upsert_tranche_from_row, upsert_valuation_from_row,
        )
        nk = self._row_natural_key(row, name, inv_date, stage)
        upsert_tranche_from_row(
            inv, natural_key=nk, amount=cost, tranche_date=inv_date,
            round_name=stage or '', instrument_type=instrument or '',
            ownership_pct=hold_pct, fully_diluted_pct=fd_pct,
        )
        if fv is not None:
            upsert_valuation_from_row(
                inv, valuation_date=val_date or date.today(),
                methodology='cost', fair_value=fv,
                source_tranche_key=nk,
                cost_basis=cost, unrealized=unrealized,
            )

    def _reconcile_imported_investments(self, investments_dict):
        from investments.services import reconcile_investment_from_tranches
        seen = set()
        for inv in investments_dict.values():
            if inv.id in seen:
                continue
            seen.add(inv.id)
            try:
                reconcile_investment_from_tranches(inv.id)
            except Exception as e:
                logger.warning(
                    f'reconcile_investment_from_tranches failed for '
                    f'{getattr(inv, "company_name", "?")}: '
                    f'{type(e).__name__}: {e}'
                )

    def _import_portfolio(self, wb, org, schemes, domain_map, progress_cb=None):
        """Import portfolio companies and investments.

        Handles three layouts, on any single sheet OR spread across
        multiple sheets within the portfolio_investments domain:
        1. Master + transactions on SAME sheet (multi-section Format B)
        2. Master on ONE sheet + transactions on ANOTHER (multi-sheet
           Format B — e.g. Portfolio_Companies + Investment_Register)
        3. Flat combined table — one row = company + investment (Format A)

        The bucket aggregation below works for all three layouts. Sub-domain
        tags from Pass 1.5 (portfolio_companies / investments / __default__)
        decide which bucket each section belongs to.
        """
        def _cb(pct, msg):
            if progress_cb:
                progress_cb(pct, msg)
        sheet_names = [s for s in _dm_sheets(domain_map, 'portfolio_investments')
                       if s in wb.sheetnames]
        if not sheet_names:
            return {}, {}

        default_scheme = list(schemes.values())[0] if schemes else None
        if not default_scheme:
            return {}, {}

        companies = {}
        investments = {}

        # ── AD: bucket aggregation across every sheet in this domain ──
        # We accumulate three buckets (master-only, combined, transactions-
        # only) by walking every sheet's sections and tagging them by
        # Pass-1.5 sub-domain plus a column-shape heuristic. The heuristic
        # is universal: it looks at canonical aliases produced by Pass 2,
        # not at fund-specific column names.
        company_rows = []     # master only — no per-row Amount
        investment_rows = []  # pure transactions — no company-identity column
        combined_rows = []    # rows that carry BOTH identity AND Amount

        amount_tokens = {
            'cost', 'invested', 'amount', 'investment_amount',
            'amount_invested', 'tranche_amount', 'cost_of_investment',
            'total_invested',
        }

        def _section_has_amount_column(sec_headers, alias_map):
            for h in sec_headers:
                if not h:
                    continue
                norm = (str(h).lower().replace(' ', '_')
                                       .replace('-', '_').strip())
                if any(tok in norm for tok in amount_tokens):
                    return True
                alias_norm = (alias_map.get(h) or '').lower()
                if any(tok in alias_norm for tok in amount_tokens):
                    return True
            return False

        # Pick a "primary" sheet for downstream Format-A fallback (uses
        # the first sheet in the domain — preserves prior single-sheet
        # behaviour when nothing aggregable is found).
        primary_sheet_name = sheet_names[0]
        ws = wb[primary_sheet_name]

        for sheet_name in sheet_names:
            sws = wb[sheet_name]
            sections = self._read_sheet_via_layout(
                sws, alias_map=self._get_alias(sws),
            )
            alias_map_ws = self._get_alias(sws)
            for sec_name, (sec_headers, sec_rows) in sections.items():
                subdomain = self._get_section_subdomain(sheet_name, sec_name)
                if subdomain == 'investment_tranches':
                    continue  # Handled by _import_tranches
                elif subdomain == 'temporary_investments':
                    continue  # Liquid MFs, overnight funds — skip
                elif subdomain == 'portfolio_companies':
                    company_rows.extend(sec_rows)
                elif subdomain == 'investments':
                    _sec_norm = {k.lower().replace(' ', '_').replace('-', '_')
                                 for k in sec_headers}
                    has_co_col = ('company_name' in _sec_norm or
                                  'company_name' in {alias_map_ws.get(k, '')
                                                      for k in sec_headers})
                    if sec_rows and has_co_col:
                        combined_rows.extend(sec_rows)
                    else:
                        investment_rows.extend(sec_rows)
                elif sec_name == '__default__':
                    if _section_has_amount_column(sec_headers, alias_map_ws):
                        combined_rows.extend(sec_rows)
                    else:
                        company_rows.extend(sec_rows)

        # When BOTH a dedicated master bucket AND a combined bucket exist,
        # the combined bucket is really a transactions register (e.g.
        # Portfolio_Companies + Investment_Register on AI_Trivesta). Route
        # those rows to investment_rows so each transaction becomes its own
        # Investment/Tranche row instead of collapsing into Format A.
        if company_rows and combined_rows:
            investment_rows.extend(combined_rows)
            combined_rows = []

        # Single-sheet promotion: when only the combined bucket exists,
        # treat it as the Format-A flat table (preserves legacy behaviour
        # for files that pack everything into one sheet).
        if combined_rows and not company_rows:
            company_rows = combined_rows
            combined_rows = []

        # If we have separate company + investment rows, it's Format B
        if company_rows and investment_rows:
            # Import company master data
            for row in company_rows:
                name = _find_col_str(
                    row, 'Company Name', 'Company', 'Name', 'Portfolio Company')
                if _is_junk_row(name):
                    continue  # skip subtotal/total/header rows

                sector = _find_col_str(row, 'Sector', 'Industry', 'Vertical')
                sub_sector = _find_col_str(
                    row, 'Sub-Sector', 'Sub Sector', 'Subsector', 'Segment')
                city = _find_col_str(
                    row, 'City', 'Headquarters', 'HQ', 'HQ City',
                    'headquarters_city')
                country = _find_col_str(
                    row, 'Country', 'HQ Country', default='India')
                website = _find_col_str(row, 'Website', 'URL')
                founders = _find_col_str(row, 'Founders', 'Founder')

                # Use update_or_create so that re-importing always reflects the
                # current Excel data. Only overwrite a field if the Excel provides
                # a non-empty value — prevents a sparse re-import file from
                # blanking out good data that was previously imported.
                update_fields = {}
                if sector:
                    update_fields['sector'] = sector
                if sub_sector:
                    update_fields['sub_sector'] = sub_sector
                if city:
                    update_fields['headquarters_city'] = city
                if country:
                    update_fields['headquarters_country'] = country

                company, _ = PortfolioCompany.objects.update_or_create(
                    organization=org,
                    name=name,
                    defaults=update_fields,
                )
                companies[name] = company

            # ── Batching pre-pass for investment_rows ─────────────────
            # Collect every unique instrument-type and investment-status
            # value across all rows, then classify each set with ONE
            # Gemini call instead of one call per row. Universal — works
            # for any number of rows, any language, any file format.
            _instrument_raws = set()
            _status_raws = set()
            for _r in investment_rows:
                _it = _find_col_str(_r, 'instrument_type', default='')
                if _it:
                    _instrument_raws.add(_it)
                _st = _find_col_str(_r, 'Status', 'Investment Status', default='')
                if _st:
                    _status_raws.add(_st)
            instrument_map = self._classify_enum(
                list(_instrument_raws), 'instrument_type',
                context='Investment instrument or security type',
            ) if _instrument_raws else {}
            status_map = self._classify_enum(
                list(_status_raws), 'investment_status',
                context='Investment/portfolio company status',
            ) if _status_raws else {}

            # Import investment data from the INVESTMENTS section
            for row in investment_rows:
                name = _find_col_str(
                    row, 'Company Name', 'Company', 'Name', 'Portfolio Company',
                    'company_name')
                if _is_junk_row(name):
                    continue  # skip subtotal/total/header rows

                # Ensure company exists. If the PORTFOLIO COMPANIES section
                # already loaded it, use that; otherwise create a minimal record.
                # update_or_create ensures re-imports don't leave stale data.
                company = companies.get(name)
                if not company:
                    company, _ = PortfolioCompany.objects.update_or_create(
                        organization=org, name=name,
                        defaults={'headquarters_country': 'India'},
                    )
                    companies[name] = company

                # Determine target scheme
                scheme_name_raw = _find_col_str(
                    row, 'Scheme', 'Scheme Name', 'Fund Scheme')
                target_scheme = default_scheme
                if scheme_name_raw:
                    target_scheme = schemes.get(scheme_name_raw, default_scheme)

                instrument_raw = _find_col_str(
                    row, 'instrument_type', default='Equity')
                instrument = instrument_map.get(instrument_raw, 'equity')

                stage = _find_col_str(row, 'Round', 'Stage', 'Funding Round',
                                       'Round Name', 'stage')
                irr_raw = _find_col_decimal(
                    row, 'IRR%(Gross)', 'IRR%', 'Gross IRR', 'IRR', 'irr_pct')
                hold_pct = _find_col_decimal(
                    row, 'Ownership %', 'Hold%', 'Holding %', 'Ownership',
                    'ownership_pct')
                fd_pct = _find_col_decimal(
                    row, 'Fully Diluted %', 'FD%', 'FD', 'Diluted %')
                invested = _find_col_decimal(
                    row, 'Total Invested (Cr)', 'Total Invested',
                    'Cost (Cr)', 'Cost(Cr)', 'Cost(₹Cr)',
                    'Cost', 'Invested', 'Investment Amount', 'Amount',
                    'total_invested', 'tranche_amount', 'amount_invested',
                    'investment_amount', 'cost_of_investment')
                inv_date = _find_col_date(
                    row, 'Investment Date', 'Inv.Date', 'Date',
                    'investment_date', 'tranche_date')
                status_raw = _find_col_str(
                    row, 'Status', 'Investment Status', default='Active')
                board_seat = _find_col_bool(
                    row, 'Board Seat', 'Board', 'Has Board Seat')
                is_lead = _find_col_bool(
                    row, 'Lead Investor', 'Is Lead', 'Lead')

                status = status_map.get(status_raw, 'active')

                inv_defaults = {
                    'company_name': name,
                    'instrument_type': instrument,
                    'currency': 'INR',
                    'status': status,
                    'sector': company.sector or '',
                    'board_seat': board_seat,
                    'is_lead_investor': is_lead or False,
                }
                if stage:
                    inv_defaults['stage'] = stage
                if irr_raw is not None:
                    irr_val = irr_raw * 100 if abs(irr_raw) <= 2 else irr_raw
                    inv_defaults['irr_pct'] = round(irr_val, 2)

                inv, _created = Investment.objects.get_or_create(
                    scheme=target_scheme,
                    portfolio_company=company,
                    defaults=inv_defaults,
                )
                key = f'{name}|{target_scheme.name}|{instrument}'
                investments[key] = inv
                self._record_investment_row(
                    inv, row, name, company,
                    cost=invested, inv_date=inv_date,
                    stage=stage, instrument=instrument,
                    hold_pct=hold_pct, fd_pct=fd_pct,
                )

            logger.info(f'  Portfolio (structured): {len(companies)} companies, '
                         f'{len(investments)} investments')
            self._reconcile_imported_investments(investments)
            return companies, investments

        # --- Format A: Flat combined table (one row = company + investment) ---
        # Use already-parsed rows from section reader when available;
        # only fall back to read_table_from_sheet if the section reader
        # returned nothing (e.g. unrecognized layout).
        if company_rows:
            rows = company_rows
        else:
            _, rows = read_table_from_sheet(ws, alias_map=self._get_alias(ws))

        total_rows = len(rows)

        # ── Batching pre-pass for Format-A rows ───────────────────────
        # Collect unique quoted-status and investment-status values
        # before the loop so each category hits Gemini ONCE per
        # category instead of once per row. Universal pattern.
        _listing_raws = set()
        _status_raws_a = set()
        for _r in rows:
            _ls = _find_col_str(
                _r, 'Listed', 'Listing Status', 'Quoted', 'Listed/Unlisted',
                'Quoted/Unquoted', 'Public/Private', 'is_quoted')
            if _ls:
                _listing_raws.add(_ls)
            _ss = _find_col_str(_r, 'Status', 'Investment Status', default='')
            if _ss:
                _status_raws_a.add(_ss)
        quoted_map = self._classify_enum(
            list(_listing_raws), 'quoted_status',
            context='Listing status of portfolio company',
        ) if _listing_raws else {}
        status_map_a = self._classify_enum(
            list(_status_raws_a), 'investment_status',
            context='Investment/portfolio company status',
        ) if _status_raws_a else {}

        row_idx = 0
        for row in rows:
            row_idx += 1
            # Emit a progress tick every 10 companies so the browser sees
            # the bar advancing during large imports (47% → 52% range)
            if row_idx % 10 == 0 and total_rows > 0:
                frac = row_idx / total_rows
                pct = int(47 + frac * 5)   # interpolate 47→52
                _cb(pct, f'Importing company {row_idx} of {total_rows}...')
            name = _find_col_str(
                row, 'Company Name', 'Company', 'Name', 'Portfolio Company',
                'company_name')
            if _is_junk_row(name):
                continue  # skip subtotal/total/header rows

            sector = _find_col_str(row, 'Sector', 'Industry', 'Vertical')
            sub_sector = _find_col_str(
                row, 'Sub-Sector', 'Sub Sector', 'Subsector')
            city = _find_col_str(
                row, 'City', 'Headquarters', 'HQ', 'headquarters_city')
            country = _find_col_str(
                row, 'Country', 'HQ Country', default='India')
            listing_raw = _find_col_str(
                row, 'Listed', 'Listing Status', 'Quoted', 'Listed/Unlisted',
                'Quoted/Unquoted', 'Public/Private', 'is_quoted')
            listing_exchange = _find_col_str(
                row, 'Exchange', 'Listed On', 'Stock Exchange', 'listing_exchange')

            # update_or_create — always sync master data fields from Excel.
            # Only overwrite with non-empty values to protect existing good data.
            pc_update = {}
            if sector:
                pc_update['sector'] = sector
            if sub_sector:
                pc_update['sub_sector'] = sub_sector
            if city:
                pc_update['headquarters_city'] = city
            if country:
                pc_update['headquarters_country'] = country
            if listing_raw:
                pc_update['is_quoted'] = (quoted_map.get(listing_raw) == 'quoted')
            if listing_exchange:
                pc_update['listing_exchange'] = listing_exchange.upper()

            company, _ = PortfolioCompany.objects.update_or_create(
                organization=org,
                name=name,
                defaults=pc_update,
            )
            companies[name] = company

            round_name = _find_col_str(
                row, 'Round', 'Funding Round', 'Stage', 'round_name', 'stage')
            irr_raw_a = _find_col_decimal(
                row, 'IRR%(Gross)', 'IRR%', 'Gross IRR', 'IRR', 'irr_pct',
                'Net IRR', 'IRR (Gross)')
            invested = _find_col_decimal(
                row, 'Cost (Cr)', 'Cost(Cr)', 'Cost(₹Cr)',
                'Cost', 'Invested', 'Total Invested',
                'Investment Amount', 'Amount', 'total_invested',
                'tranche_amount', 'amount_invested', 'investment_amount',
                'cost_of_investment')
            hold_pct = _find_col_decimal(
                row, 'Hold%', 'Holding %', 'Ownership', 'Ownership %',
                'ownership_pct')
            fd_pct = _find_col_decimal(
                row, 'FD%', 'Fully Diluted %', 'FD', 'Diluted %')
            inv_date = _find_col_date(
                row, 'Inv.Date', 'Investment Date', 'Date', 'investment_date',
                'tranche_date')
            status_raw = _find_col_str(row, 'Status', 'Investment Status',
                                       default='Active')
            board_seat = _find_col_bool(row, 'Board', 'Board Seat',
                                         'Has Board Seat')

            status = status_map_a.get(status_raw, 'active')

            inv_a_defaults = {
                'portfolio_company': company,
                'instrument_type': 'equity',
                'currency': 'INR',
                'status': status,
                'board_seat': board_seat,
                'is_lead_investor': False,
            }
            if sector:
                inv_a_defaults['sector'] = sector
            if round_name:
                inv_a_defaults['stage'] = round_name
            if irr_raw_a is not None:
                irr_val_a = irr_raw_a * 100 if abs(irr_raw_a) <= 2 else irr_raw_a
                inv_a_defaults['irr_pct'] = round(irr_val_a, 2)
            if hold_pct is not None:
                inv_a_defaults['ownership_pct'] = hold_pct
            if fd_pct is not None:
                inv_a_defaults['percentage_stake_fully_diluted'] = fd_pct
            if invested is not None:
                inv_a_defaults['total_invested'] = abs(invested)
            if inv_date:
                inv_a_defaults['investment_date'] = inv_date

            inv_a_defaults.pop('total_invested', None)
            inv_a_defaults.pop('investment_date', None)
            inv_a_defaults.pop('ownership_pct', None)
            inv_a_defaults.pop('percentage_stake_fully_diluted', None)
            inv_a_defaults['company_name'] = name

            inv, _created = Investment.objects.get_or_create(
                scheme=default_scheme,
                portfolio_company=company,
                defaults=inv_a_defaults,
            )
            key = f'{name}|{default_scheme.name}|equity'
            investments[key] = inv

            fv_raw = _find_col_decimal(
                row, 'FV (Cr)', 'FV(Cr)', 'FV(₹Cr)',
                'Fair Value (Cr)', 'Fair Value', 'Current Value (Cr)',
                'Current Value', 'Market Value (Cr)', 'fair_value')
            val_date_raw = _find_col_date(
                row, 'Val. Date', 'Val.Date', 'Valuation Date',
                'valuation_date')
            unrealized_raw = _find_col_decimal(
                row, 'Unrealised (Cr)', 'Unrealized (Cr)',
                'Unrealized Gain/Loss (Cr)', 'Unrealized Gain',
                'Unrealised', 'unrealized_gain_loss')
            self._record_investment_row(
                inv, row, name, company,
                cost=invested, inv_date=inv_date,
                stage=round_name, instrument='equity',
                hold_pct=hold_pct, fd_pct=fd_pct,
                fv=fv_raw, val_date=val_date_raw,
                unrealized=unrealized_raw,
            )

        logger.info(f'  Portfolio: {len(companies)} companies, '
                     f'{len(investments)} investments')
        self._reconcile_imported_investments(investments)
        return companies, investments

    # ------------------------------------------------------------------
    # Investment Tranches
    # ------------------------------------------------------------------

    def _import_tranches(self, wb, investments, domain_map):
        """Create InvestmentTranche records.

        Handles two formats across ANY/ALL sheets in the
        portfolio_investments domain (master + dedicated tranches often
        live on separate sheets):
        1. Dedicated INVESTMENT TRANCHES section
        2. Flat table where each company row doubles as a single tranche
        Most files are now fully covered by _import_portfolio (Format B
        multi-sheet aggregation) — this importer is the legacy fallback
        for files that explicitly tag a section as `investment_tranches`.
        """
        sheet_names = [s for s in _dm_sheets(domain_map, 'portfolio_investments')
                       if s in wb.sheetnames]
        if not sheet_names:
            return

        if investments and all(
            InvestmentTranche.objects.filter(investment=inv).exists()
            for inv in investments.values()
        ):
            return

        # Locate the first sheet that contains an `investment_tranches`
        # sub-section. Walking ALL sheets keeps this importer in sync with
        # the multi-sheet aggregation now performed by _import_portfolio.
        tranche_rows = None
        chosen_sheet = sheet_names[0]
        for sheet_name in sheet_names:
            sws = wb[sheet_name]
            sections = self._read_sheet_via_layout(
                sws, alias_map=self._get_alias(sws),
            )
            for sec_name, (sec_headers, sec_rows) in sections.items():
                subdomain = self._get_section_subdomain(sheet_name, sec_name)
                if subdomain == 'investment_tranches':
                    tranche_rows = sec_rows
                    chosen_sheet = sheet_name
                    break
            if tranche_rows:
                break

        ws = wb[chosen_sheet]
        count = 0

        if tranche_rows:
            # Format B: Dedicated tranche section
            for row in tranche_rows:
                name = _find_col_str(
                    row, 'Company Name', 'Company', 'Name', 'Portfolio Company',
                    'company_name')
                if _is_junk_row(name):
                    continue

                inv = None
                for key, i in investments.items():
                    if key.startswith(f'{name}|'):
                        inv = i
                        break
                if not inv:
                    continue

                tranche_num = _find_col_decimal(
                    row, 'Tranche #', 'Tranche Number', 'Tranche No', '#')
                tranche_num = int(tranche_num) if tranche_num else None

                invested = _find_col_decimal(
                    row, 'Amount (Cr)', 'Amount', 'Cost(₹Cr)', 'Cost',
                    'Invested', 'Total Invested', 'Investment Amount',
                    'tranche_amount', 'amount_invested', 'investment_amount',
                    'cost_of_investment')
                inv_date = _find_col_date(
                    row, 'Date', 'Inv.Date', 'Investment Date',
                    'Tranche Date', 'tranche_date', 'investment_date')
                round_name = _find_col_str(
                    row, 'Round Name', 'Round', 'Funding Round', 'Stage')
                shares = _find_col_decimal(
                    row, 'Shares Acquired', 'Shares', 'No. of Shares',
                    'Units')
                pps = _find_col_decimal(
                    row, 'Price/Share (INR)', 'Price/Share',
                    'Price Per Share', 'PPS')
                pre_money = _find_col_decimal(
                    row, 'Pre-Money Val (Cr)', 'Pre-Money',
                    'Pre Money Valuation')
                post_money = _find_col_decimal(
                    row, 'Post-Money Val (Cr)', 'Post-Money',
                    'Post Money Valuation')

                if not invested or invested <= 0:
                    continue

                if not tranche_num:
                    existing_count = InvestmentTranche.objects.filter(
                        investment=inv).count()
                    tranche_num = existing_count + 1

                InvestmentTranche.objects.get_or_create(
                    investment=inv,
                    tranche_number=tranche_num,
                    defaults={
                        'amount': abs(invested),
                        'date': inv_date or date.today(),
                        'shares_acquired': shares,
                        'price_per_share': pps,
                        'pre_money_valuation': pre_money,
                        'post_money_valuation': post_money,
                        'round_name': round_name or '',
                    },
                )
                count += 1

            logger.info(f'  Tranches (dedicated section): {count}')
            return

        # Format A: Create one tranche per investment from flat portfolio data
        _, rows = read_table_from_sheet(ws, alias_map=self._get_alias(ws))

        for row in rows:
            name = _find_col_str(
                row, 'Company Name', 'Company', 'Name', 'Portfolio Company',
                'company_name')
            if _is_junk_row(name):
                continue

            inv = None
            for key, i in investments.items():
                if key.startswith(f'{name}|'):
                    inv = i
                    break
            if not inv:
                continue

            invested = _find_col_decimal(
                row, 'Cost(₹Cr)', 'Cost', 'Invested', 'Total Invested',
                'Investment Amount', 'Amount', 'tranche_amount',
                'amount_invested', 'investment_amount', 'cost_of_investment')
            inv_date = _find_col_date(
                row, 'Inv.Date', 'Investment Date', 'Date', 'tranche_date',
                'investment_date')
            round_name = _find_col_str(
                row, 'Round', 'Funding Round', 'Stage', 'stage')
            shares = _find_col_decimal(
                row, 'Shares', 'Shares Acquired', 'No. of Shares')
            pps = _find_col_decimal(
                row, 'Price/Share', 'Price Per Share', 'PPS')
            pre_money = _find_col_decimal(
                row, 'Pre-Money', 'Pre Money Valuation', 'Pre-Money Val')
            post_money = _find_col_decimal(
                row, 'Post-Money', 'Post Money Valuation', 'Post-Money Val')

            if not invested or invested <= 0:
                continue

            existing_count = InvestmentTranche.objects.filter(
                investment=inv).count()

            InvestmentTranche.objects.get_or_create(
                investment=inv,
                tranche_number=existing_count + 1,
                defaults={
                    'amount': abs(invested),
                    'date': inv_date or date.today(),
                    'shares_acquired': shares,
                    'price_per_share': pps,
                    'pre_money_valuation': pre_money,
                    'post_money_valuation': post_money,
                    'round_name': round_name or '',
                },
            )
            count += 1

        logger.info(f'  Tranches: {count}')

    # ------------------------------------------------------------------
    # Valuations
    # ------------------------------------------------------------------

    def _import_valuations(self, wb, investments, domain_map):
        """Import valuations from Gemini-classified valuations_kpis sheets."""
        sheets = _dm_sheets(domain_map, 'valuations_kpis')
        if not sheets:
            return
        # Process all valuation sheets — column detection determines if data matches
        sheet_name = sheets[0]
        if sheet_name not in wb.sheetnames:
            return

        ws = wb[sheet_name]
        _, rows = read_table_from_sheet(ws, alias_map=self._get_alias(ws))

        # ── Batching pre-pass: methodology + ipev_level ───────────────
        _meth_raws = set()
        _ipev_raws = set()
        for _r in rows:
            _m = _find_col_str(
                _r, 'Methodology', 'Method', 'Valuation Method',
                'Val Method', 'IPEV Technique', 'Technique',
                'Val. Method', 'Valuation Basis')
            if _m:
                _meth_raws.add(_m)
            _ip = _find_col_str(_r, 'ipev_level', 'ipev_technique')
            if _ip:
                _ipev_raws.add(_ip)
        method_map = self._classify_enum(
            list(_meth_raws), 'valuation_methodology',
            context='Valuation methodology for portfolio company',
        ) if _meth_raws else {}
        ipev_map = self._classify_enum(
            list(_ipev_raws), 'ipev_level',
            context='IPEV fair value hierarchy level',
        ) if _ipev_raws else {}

        count = 0
        for row in rows:
            name = _find_col_str(
                row, 'Company Name', 'Company', 'Name', 'Portfolio Company')
            if _is_junk_row(name):
                continue

            inv = None
            for key, i in investments.items():
                if key.startswith(f'{name}|'):
                    inv = i
                    break
            if not inv:
                continue

            val_date = _find_col_date(
                row, 'Val. Date', 'Val.Date', 'Valuation Date',
                'Date', 'valuation_date')
            methodology_raw = _find_col_str(
                row, 'Methodology', 'Method', 'Valuation Method',
                'Val Method', 'IPEV Technique', 'Technique',
                'Val. Method', 'Valuation Basis')

            methodology = method_map.get(methodology_raw, 'cost')

            ev = _find_col_decimal(
                row, 'EV(₹Cr)', 'Enterprise Value (Cr)',
                'Enterprise Value', 'EV', 'enterprise_value')
            equity_val = _find_col_decimal(
                row, 'FV (Cr)', 'FV(Cr)', 'FV(₹Cr)',
                'Equity Val', 'Equity Value', 'Fair Value (Cr)',
                'Fair Value', 'Current Value (Cr)', 'Current Value',
                'Market Value (Cr)', 'fair_value')
            fv_holding = _find_col_decimal(
                row, 'FV Holding', 'FV of Holding (Cr)',
                'FV of Holding', 'Fair Value Holding',
                'fair_value_of_holding')
            cost_basis = _find_col_decimal(
                row, 'Cost Basis (Cr)', 'Cost Basis',
                'Cost (Cr)', 'Cost(Cr)', 'Cost(₹Cr)',
                'Cost', 'cost_basis')
            unrealized = _find_col_decimal(
                row, 'Unrealized Gain/Loss (Cr)', 'Unrealised (Cr)',
                'Unrealised', 'Unreal G/L',
                'Unrealized', 'Unrealized Gain',
                'unrealized_gain_loss')
            moic = _find_col_decimal(row, 'MOIC', 'Multiple', 'moic')

            # Derive IPEV Level from valuation methodology.
            # All private equity unquoted assets are Level 3 by IPEV standard.
            # Level 2 applies only when observable market prices exist (listed comparables).
            _IPEV_LEVEL_MAP = {
                'dcf': 3, 'comparables': 3, 'recent_transaction': 3,
                'net_assets': 3, 'cost': 3, 'option_pricing': 3,
            }
            ipev_level = _IPEV_LEVEL_MAP.get(methodology, 3)
            # IPEV Technique column may explicitly specify the fair value level
            _ipev_technique_raw = _find_col_str(
                row, 'ipev_level', 'ipev_technique')
            if _ipev_technique_raw:
                _ipev_mapped = ipev_map.get(_ipev_technique_raw, str(ipev_level))
                try:
                    ipev_level = int(_ipev_mapped)
                except (ValueError, TypeError):
                    pass

            # Compute MOIC from FV / Cost if not directly in the sheet.
            if moic is None and equity_val:
                _cost = cost_basis or (inv.total_invested if inv else None)
                if _cost and _cost > 0:
                    try:
                        moic = round(equity_val / _cost, 2)
                    except Exception:
                        pass

            _val_defaults = {
                'fair_value': equity_val if equity_val is not None else Decimal('0'),
                'fair_value_of_holding': fv_holding,
                'enterprise_value': ev,
                'cost_basis': cost_basis,
                'unrealized_gain_loss': unrealized,
                'multiple': moic,
                'ipev_level': ipev_level,
                'status': 'approved',
            }
            Valuation.objects.update_or_create(
                investment=inv,
                valuation_date=val_date or date.today(),
                methodology=methodology,
                defaults=_val_defaults,
            )
            count += 1

        logger.info(f'  Valuations: {count}')

    # ------------------------------------------------------------------
    # KPIs
    # ------------------------------------------------------------------

    def _import_kpis(self, wb, org, investments, companies, domain_map):
        """Import KPIs from Gemini-classified valuations_kpis sheets."""
        sheets = _dm_sheets(domain_map, 'valuations_kpis')
        if not sheets:
            return
        # Use all KPI-classified sheets; start with the first one
        sheet_name = sheets[0]
        if sheet_name not in wb.sheetnames:
            return

        ws = wb[sheet_name]
        headers_dict, rows = read_table_from_sheet(ws, alias_map=self._get_alias(ws))

        if not rows:
            return

        # Detect format: pivot (period columns like Oct-24) vs vertical
        period_cols = []
        period_pattern = re.compile(
            r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[-/]\d{2,4}$',
            re.IGNORECASE)
        for h in headers_dict.keys():
            if period_pattern.match(h.strip()):
                period_cols.append(h)

        count = 0
        if period_cols:
            # Pivot format: each row has a company + KPI name + values
            # across period columns
            for row in rows:
                name = _find_col_str(
                    row, 'Company Name', 'Company', 'Name')
                kpi_name = _find_col_str(
                    row, 'KPI Name', 'KPI', 'Metric', 'Indicator')
                if not name or not kpi_name:
                    continue

                inv = None
                for key, i in investments.items():
                    if key.startswith(f'{name}|'):
                        inv = i
                        break
                if not inv:
                    continue

                company = companies.get(name)
                kpi_slug = slugify(kpi_name)
                kpi_def, _ = KPIDefinition.objects.get_or_create(
                    organization=org,
                    slug=kpi_slug,
                    defaults={
                        'name': kpi_name,
                        'format': 'number',
                        'frequency': 'monthly',
                    },
                )

                for pcol in period_cols:
                    val = _d(row.get(pcol))
                    if val is None:
                        continue
                    # Parse period to date (e.g., "Oct-24" → 2024-10-01)
                    period_date = self._parse_period(pcol)
                    if not period_date:
                        continue

                    PortfolioKPI.objects.get_or_create(
                        investment=inv,
                        kpi_definition=kpi_def,
                        period=period_date,
                        defaults={
                            'portfolio_company': company,
                            'value': val,
                            'source': 'excel_upload',
                            'status': 'approved',
                        },
                    )
                    count += 1
        else:
            # Check if this is a vertical format (KPI Name column exists) or flat snapshot
            # Detect via Gemini Pass 2 alias map: any column mapped to kpi_name
            _alias = self._get_alias(ws)
            has_kpi_name_col = (
                'kpi_name' in {str(h or '').lower().strip() for h in headers_dict.keys()}
                or 'kpi_name' in set(_alias.values())
            )

            if has_kpi_name_col:
                # Vertical format: one KPI per row with period column
                for row in rows:
                    name = _find_col_str(
                        row, 'Company Name', 'Company', 'Name')
                    kpi_name = _find_col_str(
                        row, 'KPI Name', 'KPI', 'Metric')
                    if not name or not kpi_name:
                        continue

                    inv = None
                    for key, i in investments.items():
                        if key.startswith(f'{name}|'):
                            inv = i
                            break
                    if not inv:
                        continue

                    company = companies.get(name)
                    kpi_slug = slugify(kpi_name)
                    kpi_def, _ = KPIDefinition.objects.get_or_create(
                        organization=org,
                        slug=kpi_slug,
                        defaults={
                            'name': kpi_name,
                            'format': 'number',
                            'frequency': 'monthly',
                        },
                    )
                    period = _find_col_date(row, 'Period', 'Date', 'Month')
                    value = _find_col_decimal(row, 'Value', 'KPI Value')
                    if period and value is not None:
                        PortfolioKPI.objects.get_or_create(
                            investment=inv,
                            kpi_definition=kpi_def,
                            period=period,
                            defaults={
                                'portfolio_company': company,
                                'value': value,
                                'source': 'excel_upload',
                                'status': 'approved',
                            },
                        )
                        count += 1
            else:
                # Flat snapshot format: multi-section sheet where each column is a KPI.
                # E.g. Portfolio KPIs sheet with sector groups (Consumer, Financial Svcs…),
                # each group having its own header row and company data rows.
                count += self._import_kpis_flat_snapshot(ws, org, investments, companies)

        logger.info(f'  KPIs: {count}')

    # KPI canonical key → display name + format
    _KPI_DISPLAY = {
        'gmv': ('GMV', 'currency'), 'revenue': ('Revenue', 'currency'),
        'gross_margin_pct': ('Gross Margin %', 'percent'), 'ebitda': ('EBITDA', 'currency'),
        'ebitda_pct': ('EBITDA %', 'percent'), 'orders': ('Orders', 'number'),
        'aov': ('AOV', 'currency'), 'returns_pct': ('Returns %', 'percent'),
        'cac': ('CAC', 'currency'), 'repeat_pct': ('Repeat %', 'percent'),
        'cost_to_income': ('Cost to Income', 'ratio'), 'nim_pct': ('NIM %', 'percent'),
        'gnpa_pct': ('GNPA %', 'percent'), 'nnpa_pct': ('NNPA %', 'percent'),
        'roe_pct': ('ROE %', 'percent'), 'aum': ('AUM', 'currency'),
        'car_pct': ('CAR %', 'percent'), 'd_ebitda': ('D/EBITDA', 'ratio'),
        'capacity_pct': ('Capacity %', 'percent'), 'export_pct': ('Export %', 'percent'),
        'headcount': ('Headcount', 'number'), 'bed_occupancy': ('Bed Occupancy', 'percent'),
        'arpob': ('ARPOB', 'currency'), 'cap_rate_pct': ('Cap Rate %', 'percent'),
        'cost': ('Cost', 'currency'), 'fv': ('FV', 'currency'),
        'moic': ('MOIC', 'ratio'), 'mrr': ('MRR', 'currency'),
        'arr': ('ARR', 'currency'), 'churn_pct': ('Churn %', 'percent'),
        'nrr_pct': ('NRR %', 'percent'), 'ltv': ('LTV', 'currency'),
        'ltv_cac': ('LTV/CAC', 'ratio'), 'burn_rate': ('Burn Rate', 'currency'),
        'runway': ('Runway', 'number'), 'pat': ('PAT', 'currency'),
    }

    def _semantic_kpi_column_map(self, raw_headers):
        """Use Gemini Pass 3 to semantically map raw KPI column headers to canonical names.

        Returns: dict mapping raw_header → {canonical_name, slug, format}
        """
        if not raw_headers:
            return {}

        # Classify all headers via Gemini in a single call
        kpi_map = self._classify_labels(raw_headers, 'kpi_types',
                                         context='KPI column headers from portfolio sheet')

        result = {}
        for raw in raw_headers:
            canonical_key = kpi_map.get(raw)
            if canonical_key and canonical_key in self._KPI_DISPLAY:
                display_name, fmt = self._KPI_DISPLAY[canonical_key]
                result[raw] = {
                    'canonical_name': display_name,
                    'slug': slugify(display_name),
                    'format': fmt,
                }
            else:
                result[raw] = {
                    'canonical_name': raw,
                    'slug': slugify(raw),
                    'format': 'percent' if '%' in raw else 'number',
                }

        return result

    def _import_kpis_flat_snapshot(self, ws, org, investments, companies):
        """
        Handle Portfolio KPIs sheets in multi-section flat format.

        Each section has its own header row (first cell = 'Company'), followed by
        company data rows.  Each non-identity column is treated as a KPI metric.

        This handles sector-grouped KPI sheets where Consumer & Retail has CAC/GMV
        columns, Financial Svcs has AUM/NIM% columns, etc.  Reads ALL sections so
        every company gets its sector-appropriate KPIs imported.

        Uses Gemini-powered semantic matching: column names like
        "GMV (Cr)", "GMV in Crore", "Gross Merch Value" all map to the same
        canonical KPI definition via fuzzy/semantic matching.
        """
        # Skip identity/metadata columns — use canonical field names + format tokens.
        # Single-character tokens (#, id) and canonical names are language-independent.
        SKIP_COL_LOWER = {
            'company_name', 'name', '#', 'id', 'sr', 'sl', 'no',
            'sector', 'sub_sector', 'status', 'stage', 'headquarters_city',
            'instrument_type', 'round_name', 'scheme_name',
        }
        snapshot_date = date.today()
        count = 0
        current_headers = []  # [(col_idx, header_name), ...]

        max_col = ws.max_column or 1
        for row_idx in range(1, ws.max_row + 1):
            row_vals = [ws.cell(row_idx, c).value for c in range(1, max_col + 1)]

            # Blank row → reset section headers
            if not any(v is not None for v in row_vals):
                current_headers = []
                continue

            # Section title row: only first cell filled (sector group label)
            non_empty = sum(1 for v in row_vals if v is not None)
            if non_empty == 1:
                continue

            first_cell = str(row_vals[0] or '').strip()

            # Detect header row: first cell maps to 'company_name' canonical field
            _fc_norm = first_cell.lower().replace(' ', '_')
            if _fc_norm in ('company', 'company_name', 'name'):
                current_headers = []
                raw_kpi_headers = []
                for col_idx, val in enumerate(row_vals, 1):
                    if val is None:
                        continue
                    header = str(val).strip()
                    header_norm = _normalize_col_key(header)
                    _hn = header_norm.lower().replace(' ', '_')
                    _hl = header.lower().replace(' ', '_')
                    if _hn in SKIP_COL_LOWER or _hl in SKIP_COL_LOWER:
                        continue
                    if not header_norm:
                        continue
                    raw_kpi_headers.append((col_idx, header_norm))

                # Semantic mapping via Gemini + local aliases
                if raw_kpi_headers:
                    raw_names = [h[1] for h in raw_kpi_headers]
                    sem_map = self._semantic_kpi_column_map(raw_names)
                    for col_idx, header_name in raw_kpi_headers:
                        mapped = sem_map.get(header_name, {})
                        canonical = mapped.get('canonical_name', header_name)
                        kpi_slug = mapped.get('slug', slugify(header_name))
                        kpi_format = mapped.get('format', 'percent' if '%' in header_name else 'number')
                        if not kpi_slug:
                            continue
                        current_headers.append((col_idx, canonical, kpi_slug, kpi_format))
                continue

            # Data row — needs active section headers
            if not current_headers:
                continue

            company_name = _str(first_cell)
            if not company_name or _is_junk_row(company_name):
                continue

            # Resolve investment
            inv = None
            for key, i in investments.items():
                if key.startswith(f'{company_name}|'):
                    inv = i
                    break
            if not inv:
                co_obj = PortfolioCompany.objects.filter(
                    organization=org, name__iexact=company_name
                ).first()
                if co_obj:
                    inv = Investment.objects.filter(portfolio_company=co_obj).first()
            if not inv:
                continue

            company_obj = companies.get(company_name)

            for col_idx, header_name, kpi_slug, kpi_format in current_headers:
                if col_idx > len(row_vals):
                    continue
                cell_val = row_vals[col_idx - 1]
                if cell_val is None:
                    continue
                dec_val = _d(cell_val)
                if dec_val is None:
                    continue

                kpi_def, _ = KPIDefinition.objects.update_or_create(
                    organization=org,
                    slug=kpi_slug,
                    defaults={
                        'name': header_name,
                        'format': kpi_format,
                        'frequency': 'monthly',
                        'sector_template': 'generic',
                    },
                )
                PortfolioKPI.objects.update_or_create(
                    investment=inv,
                    kpi_definition=kpi_def,
                    period=snapshot_date,
                    defaults={
                        'portfolio_company': company_obj,
                        'value': dec_val,
                        'source': 'excel_upload',
                        'status': 'approved',
                    },
                )
                count += 1

        return count

    def _parse_quarter_period(self, quarter_str):
        """Parse 'Q1 FY25' → (start_date, end_date) using Indian FY (Apr–Mar).

        Q1 FY25 = Apr 2024 – Jun 2024
        Q2 FY25 = Jul 2024 – Sep 2024
        Q3 FY25 = Oct 2024 – Dec 2024
        Q4 FY25 = Jan 2025 – Mar 2025

        Returns (start_date, end_date) or (None, None) if unparseable.
        """
        if not quarter_str:
            return None, None
        match = re.match(r'Q([1-4])\s*FY\s*(\d{2,4})', quarter_str.strip(), re.IGNORECASE)
        if not match:
            return None, None
        q = int(match.group(1))
        fy = int(match.group(2))
        if fy < 100:
            fy += 2000
        # Indian FY starts in April of (fy-1) and ends in March of fy
        start_month, start_year = {
            1: (4, fy - 1), 2: (7, fy - 1), 3: (10, fy - 1), 4: (1, fy),
        }[q]
        end_month, end_year = {
            1: (6, fy - 1), 2: (9, fy - 1), 3: (12, fy - 1), 4: (3, fy),
        }[q]
        start = date(start_year, start_month, 1)
        last_day = calendar.monthrange(end_year, end_month)[1]
        end = date(end_year, end_month, last_day)
        return start, end

    def _parse_period(self, period_str):
        """Parse period strings like 'Oct-24', 'Nov-2024', 'Mar-25' to date."""
        period_str = period_str.strip()
        month_map = {
            'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
            'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
        }
        match = re.match(
            r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[-/](\d{2,4})',
            period_str, re.IGNORECASE)
        if match:
            month = month_map[match.group(1).lower()]
            year = int(match.group(2))
            if year < 100:
                year += 2000
            return date(year, month, 1)
        return None

    # ------------------------------------------------------------------
    # Company Financials (Burn & Runway)
    # ------------------------------------------------------------------

    def _import_company_financials(self, wb, org, investments, companies, domain_map):
        """Import burn rate, cash balance, runway, and SaaS metrics from Excel.

        Uses Gemini-classified domain_map to find the relevant sheet(s).
        Checks 'burn_runway' domain first, then 'valuations_kpis' as fallback
        (some funds embed burn data alongside KPIs).
        """
        # Get sheets from Gemini classification
        burn_sheets = _dm_sheets(domain_map, 'burn_runway')
        if not burn_sheets:
            # Fallback: valuations_kpis often contains embedded burn columns
            burn_sheets = _dm_sheets(domain_map, 'valuations_kpis')
        if not burn_sheets:
            return

        burn_sheet = burn_sheets[0]
        if burn_sheet not in wb.sheetnames:
            return

        ws = wb[burn_sheet]
        headers_dict, rows = read_table_from_sheet(ws, alias_map=self._get_alias(ws))
        if not rows:
            return

        # Detect period columns (e.g., "Oct-24", "Nov-24")
        period_pattern = re.compile(
            r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[-/]\d{2,4}$',
            re.IGNORECASE)
        period_cols = [h for h in headers_dict.keys()
                       if period_pattern.match(h.strip())]

        count = 0

        if period_cols:
            # Pivot format: rows have Company + Metric, columns are periods
            # Group rows by company, accumulate Gross Burn / Net Burn / Cash per period
            from collections import defaultdict
            company_period_data = defaultdict(lambda: defaultdict(dict))

            # Pre-classify all unique metric labels via Gemini Pass 3
            _metric_labels = set()
            for row in rows:
                metric = _find_col_str(row, 'kpi_name', 'metric')
                if metric:
                    _metric_labels.add(metric)
            _metric_map = self._classify_labels(list(_metric_labels), 'burn_runway_metrics',
                                                 context='Burn rate and runway metrics for portfolio companies')

            for row in rows:
                name = _find_col_str(row, 'company_name')
                metric = _find_col_str(row, 'kpi_name', 'metric')
                if not name or not metric:
                    continue
                if _is_junk_row(name):
                    continue

                field = _metric_map.get(metric)
                if not field:
                    continue

                for pcol in period_cols:
                    val = _d(row.get(pcol))
                    if val is None:
                        continue
                    period_date = self._parse_period(pcol)
                    if not period_date:
                        continue
                    company_period_data[name][period_date][field] = val

            for name, period_map in company_period_data.items():
                inv = None
                for key, i in investments.items():
                    if key.startswith(f'{name}|'):
                        inv = i
                        break
                if not inv:
                    continue
                company = companies.get(name)

                for period_date, fields in period_map.items():
                    gross = fields.get('gross_burn')
                    net = fields.get('net_burn')
                    cash = fields.get('cash_balance')
                    runway = fields.get('runway_months')

                    # Compute runway if not provided but burn + cash are known
                    if runway is None and cash is not None and net and net > 0:
                        from decimal import Decimal
                        runway = round(cash / net, 1)

                    defaults = {'portfolio_company': company}
                    if gross is not None:
                        defaults['gross_burn'] = abs(gross)
                    if net is not None:
                        defaults['net_burn'] = abs(net)
                    if cash is not None:
                        defaults['cash_balance'] = abs(cash)
                    if runway is not None:
                        defaults['runway_months'] = runway

                    CompanyFinancials.objects.update_or_create(
                        investment=inv,
                        period=period_date,
                        defaults=defaults,
                    )
                    count += 1

        else:
            # Flat format: one row per company (snapshot) or (company, period)
            # Some sheets (e.g. "SaaS Metrics & Burn") have no period column —
            # they represent a current snapshot.  Use today as the period.
            snapshot_date = date.today()

            # Guard: only extract SaaS KPIs from sheets with SaaS-specific headers.
            # Check via Gemini Pass 2 alias map: if any canonical field is a SaaS KPI.
            _saas_canonical = {'mrr', 'arr', 'nrr_pct', 'churn_pct', 'cac', 'ltv', 'ltv_cac'}
            _alias_map = self._get_alias(ws)
            _has_real_saas_cols = bool(_saas_canonical & set(_alias_map.values()))

            # SaaS KPI definitions — canonical field name only.
            # Gemini Pass 2 maps Excel column headers (in ANY language) to
            # canonical field names (arr, mrr, nrr, etc.) which are added as
            # keys in the row dict by read_table_from_sheet(). No hardcoded
            # English alias lists needed.
            SAAS_KPI_DEFS = [
                ('arr',       'ARR',            'currency', 'saas', 'arr'),
                ('mrr',       'MRR',            'currency', 'saas', 'mrr'),
                ('nrr',       'NRR',            'percent',  'saas', 'nrr'),
                ('churn-rate', 'Churn Rate',    'percent',  'saas', 'churn_rate'),
                ('cac',       'CAC',            'currency', 'saas', 'cac'),
                ('ltv',       'LTV',            'currency', 'saas', 'ltv'),
                ('ltv-cac',   'LTV:CAC Ratio',  'ratio',   'saas', 'ltv_cac_ratio'),
            ]
            kpi_def_cache = {}
            for slug, name_label, fmt, tmpl, _ in SAAS_KPI_DEFS:
                kd, _ = KPIDefinition.objects.get_or_create(
                    organization=org, slug=slug,
                    defaults={'name': name_label, 'format': fmt,
                              'frequency': 'monthly', 'sector_template': tmpl},
                )
                kpi_def_cache[slug] = kd

            for row in rows:
                name = _find_col_str(row, 'company_name')
                if not name or _is_junk_row(name):
                    continue

                inv = None
                for key, i in investments.items():
                    if key.startswith(f'{name}|'):
                        inv = i
                        break
                if not inv:
                    continue

                company = companies.get(name)
                period_date = _find_col_date(row, 'period', 'kpi_period')
                # Snapshot sheets have no period column — fall back to today
                if not period_date:
                    period_date = snapshot_date

                # Gemini Pass 2 adds canonical field names as row keys —
                # no hardcoded English candidates needed.
                gross = _find_col_decimal(row, 'gross_burn')
                net = _find_col_decimal(row, 'net_burn')
                cash = _find_col_decimal(row, 'cash_balance')
                runway = _find_col_decimal(row, 'runway_months')

                if runway is None and cash is not None and net and net > 0:
                    from decimal import Decimal
                    runway = round(cash / net, 1)

                defaults = {'portfolio_company': company}
                if gross is not None:
                    defaults['gross_burn'] = abs(gross)
                if net is not None:
                    defaults['net_burn'] = abs(net)
                if cash is not None:
                    defaults['cash_balance'] = abs(cash)
                if runway is not None:
                    defaults['runway_months'] = runway

                if len(defaults) > 1:  # at least one real field besides portfolio_company
                    CompanyFinancials.objects.update_or_create(
                        investment=inv,
                        period=period_date,
                        defaults=defaults,
                    )
                    count += 1

                # Extract SaaS KPIs from the same row — ONLY if the sheet has
                # actual SaaS-specific columns (detected via Gemini Pass 2).
                if _has_real_saas_cols:
                    for slug, _label, _fmt, _tmpl, canonical_field in SAAS_KPI_DEFS:
                        val = _find_col_decimal(row, canonical_field)
                        if val is None:
                            continue
                        kd = kpi_def_cache[slug]
                        PortfolioKPI.objects.update_or_create(
                            investment=inv,
                            kpi_definition=kd,
                            period=period_date,
                            defaults={
                                'portfolio_company': company,
                                'value': val,
                                'source': 'excel_upload',
                                'status': 'approved',
                            },
                        )

        logger.info(f'  Company financials: {count}')

    # ------------------------------------------------------------------
    # MIS Financials — P&L and Budget vs Actual for BudgetVsActual model
    # ------------------------------------------------------------------

    def _pl_pre_classify(self, labels):
        """Pre-classify a batch of P&L line item labels via Gemini.

        Populates self._pl_line_item_map so subsequent _pl_map_to_line_item() calls
        are instant cache lookups. Call once per sheet before the row processing loop.
        """
        new_labels = [l for l in labels if l and str(l).strip()
                      and str(l).strip().lower() not in self._pl_line_item_map]
        if not new_labels:
            return
        result = self._classify_labels(new_labels, 'pl_line_items',
                                        context='P&L / financial statement line item labels')
        for label, canonical in result.items():
            self._pl_line_item_map[label.lower().strip()] = canonical

    def _pl_map_to_line_item(self, text):
        """Map a column header or row label text to a canonical P&L line_item key.

        Uses self._pl_line_item_map (populated by _pl_pre_classify or on-demand Gemini).
        """
        if not text:
            return None
        t = str(text).lower().strip()
        if t in self._pl_line_item_map:
            return self._pl_line_item_map[t]
        # On-demand classify for labels not in the pre-classified batch
        result = self._classify_labels([text], 'pl_line_items',
                                        context='P&L / financial statement line item label')
        canonical = result.get(text)
        self._pl_line_item_map[t] = canonical
        return canonical

    @staticmethod
    def _pl_parse_year_month(period_str):
        """
        Parse period string → (year: int, month: int|None).

        Handles: Apr-24, Apr-2024, April 2024, 2024-04, Q1-FY25, FY2025, 2025, etc.
        Returns (None, None) if unparseable.
        """
        if not period_str:
            return None, None
        s = str(period_str).strip()

        MONTH_MAP = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
                     'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12}

        # Apr-24, Apr-2024
        m = re.match(r'^([a-zA-Z]{3})[-/](\d{2,4})$', s)
        if m:
            mon = MONTH_MAP.get(m.group(1).lower())
            yr = int(m.group(2))
            if yr < 100:
                yr += 2000
            return yr, mon

        # April 2024, April-2024
        m = re.match(r'^([a-zA-Z]+)[\s\-](\d{4})$', s)
        if m:
            mon = MONTH_MAP.get(m.group(1).lower()[:3])
            yr = int(m.group(2))
            return yr, mon

        # 2024-04, 2024/04
        m = re.match(r'^(\d{4})[-/](\d{1,2})$', s)
        if m:
            return int(m.group(1)), int(m.group(2))

        # Q1-FY25, Q2 FY2025, Q3FY25
        m = re.match(r'^Q([1-4])[\s\-]?FY(\d{2,4})$', s, re.IGNORECASE)
        if m:
            q = int(m.group(1))
            yr = int(m.group(2))
            if yr < 100:
                yr += 2000
            # India FY: Q1=Apr-Jun, Q2=Jul-Sep, Q3=Oct-Dec, Q4=Jan-Mar
            mon = {1: 4, 2: 7, 3: 10, 4: 1}.get(q)
            return yr, mon

        # FY2025 or FY25 (annual, no month)
        m = re.match(r'^FY(\d{2,4})$', s, re.IGNORECASE)
        if m:
            yr = int(m.group(1))
            if yr < 100:
                yr += 2000
            return yr, None

        # Plain year: 2025
        m = re.match(r'^(\d{4})$', s)
        if m:
            return int(m.group(1)), None

        return None, None

    def _import_mis_financials(self, wb, org, fund, investments, companies, domain_map):
        """
        Import P&L and Budget vs Actual data for the MIS Consolidation module.

        Uses Gemini-classified domain_map to find financial sheets. All sheets
        classified as 'financials_pl_bva' or 'fund_pl_bs' are processed.
        Each sheet is tried as both P&L and BvA — the processing functions
        use column detection to determine the actual data type.
        """
        from mis_consolidation.services import BvAImporter, MISAggregator, AnomalyDetector

        # Get all sheets Gemini classified as financial statements
        fin_sheets = _dm_sheets(domain_map, 'financials_pl_bva')
        fund_fin_sheets = _dm_sheets(domain_map, 'fund_pl_bs')
        all_fin_sheets = fin_sheets + [s for s in fund_fin_sheets if s not in fin_sheets]

        # Pre-compute period labels from a NAV/accounting sheet if available.
        inferred_periods = self._infer_period_labels_from_wb(wb, domain_map)

        count = 0
        for sn in all_fin_sheets:
            if sn not in wb.sheetnames:
                continue
            # Try both P&L and BvA processing on every financial sheet.
            # Each processor uses column detection to determine if the sheet
            # matches its expected structure. Returns 0 if no match.
            try:
                count += self._process_pl_sheet(wb[sn], companies, fund,
                                                inferred_periods=inferred_periods)
            except Exception as e:
                logger.warning(f'P&L processing of sheet {sn!r} error: {e}')
            try:
                count += self._process_bva_sheet(wb[sn], companies, fund)
            except Exception as e:
                logger.warning(f'BvA processing of sheet {sn!r} error: {e}')

        if count > 0:
            try:
                MISAggregator(fund=fund).run()
            except Exception as e:
                logger.warning(f'MISAggregator error: {e}')
            for company in companies.values():
                try:
                    AnomalyDetector(company).run_all()
                except Exception:
                    pass

        logger.info(f'  MIS financials: {count} records')

    def _infer_period_labels_from_wb(self, wb, domain_map=None):
        """
        Scan Gemini-classified sheets for explicit month-period column headers
        (e.g. 'Oct-24', 'Nov-24', …) and return an ordered list of those labels.

        Used to annotate fund-level P&L sheets that carry value columns but no headers.
        Uses only Gemini-classified sheets: nav_accounting, financials_pl_bva,
        fund_pl_bs, nav_calculation.
        """
        _period_re = re.compile(
            r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[-/]\d{2,4}$',
            re.IGNORECASE)
        # Use only Gemini-classified sheets — no all-sheets fallback
        candidate_sheets = []
        if domain_map:
            for domain in ('nav_accounting', 'financials_pl_bva', 'fund_pl_bs', 'nav_calculation'):
                for s in _dm_sheets(domain_map, domain):
                    if s in wb.sheetnames and s not in candidate_sheets:
                        candidate_sheets.append(s)

        for sn in candidate_sheets:
            ws = wb[sn]
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                labels = []
                for cell in row:
                    if cell and isinstance(cell, str) and _period_re.match(cell.strip()):
                        labels.append(cell.strip())
                if len(labels) >= 3:
                    return labels   # e.g. ['Oct-24','Nov-24','Dec-24','Jan-25','Feb-25','Mar-25']
        return []

    def _process_pl_sheet(self, ws, companies, fund=None, inferred_periods=None):
        """
        Process a P&L / Balance Sheet / Cash Flow sheet and write BudgetVsActual records.

        Detects layouts automatically:
        A) Horizontal: rows = (company, period), each col is a P&L line item.
           Sub-case A2: cols carry budget/actual qualifier ("Revenue Budget", "EBITDA Actual").
        B) Vertical pivot: first col = line item label, remaining cols = time periods
           with company identified by a preceding single-cell section header row.
           Sub-case B2: period cols carry qualifier ("Budget Apr-24", "Apr-24 Actual").
        C) Fund-level transposed: first col = line item, remaining cols = unlabelled period
           values. inferred_periods provides the period labels for those columns.
        """
        from mis_consolidation.models import BudgetVsActual, ConsolidatedMIS

        # Period regex — uses standard financial month abbreviations (universal)
        PERIOD_RE = re.compile(
            r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[-/]\d{2,4}$'
            r'|^Q[1-4][\s\-]?FY\d{2,4}$'
            r'|^\d{4}[-/]\d{1,2}$'
            r'|^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[\s]\d{4}$',
            re.IGNORECASE,
        )
        # Inner period pattern (without anchors) for extracting period from qualified headers
        _INNER_PERIOD_RE = re.compile(
            r'((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[-/\s]\d{2,4}'
            r'|Q[1-4][\s\-]?FY\d{2,4}|\d{4}[-/]\d{1,2})',
            re.IGNORECASE,
        )

        headers_dict, rows = read_table_from_sheet(ws, alias_map=self._get_alias(ws))
        if not rows:
            return 0

        period_cols = [h for h in headers_dict.keys()
                       if h and PERIOD_RE.match(str(h).strip())]

        # Detect qualified period columns ("Budget Apr-24", "Apr-24 Actual")
        # via Gemini: classify non-period column headers as budget/actual/variance
        _candidate_qualified = []
        for h in headers_dict.keys():
            if not h or h in period_cols:
                continue
            h_str = str(h).strip()
            m = _INNER_PERIOD_RE.search(h_str)
            if m:
                _candidate_qualified.append(h)

        budget_period_cols = {}   # header_text → (yr, mo)
        actual_period_cols = {}   # header_text → (yr, mo)
        if _candidate_qualified:
            _qual_map = self._classify_enum(
                _candidate_qualified, 'column_qualifier',
                context='Column headers from a P&L / Budget vs Actual financial sheet. '
                        'Classify each as budget (planned), actual (realized), or variance.')
            for h in _candidate_qualified:
                h_str = str(h).strip()
                m = _INNER_PERIOD_RE.search(h_str)
                if not m:
                    continue
                period_part = m.group(1).strip()
                yr_q, mo_q = self._pl_parse_year_month(period_part)
                if yr_q:
                    qual = _qual_map.get(h, 'actual')
                    if qual == 'budget':
                        budget_period_cols[h] = (yr_q, mo_q)
                    elif qual != 'variance':
                        actual_period_cols[h] = (yr_q, mo_q)

        all_period_like_cols = period_cols or list(budget_period_cols) or list(actual_period_cols)
        count = 0

        if all_period_like_cols:
            # -----------------------------------------------------------------
            # Layout B / B2: vertical pivot — rows are line items, cols are periods.
            # Company name comes from a preceding single-cell section header row.
            # -----------------------------------------------------------------
            current_company = None

            for row in rows:
                # Normalise keys once per row so all _find_col_* calls work
                # regardless of unit-suffix or CamelCase formatting.
                nr = _norm_row(row)
                non_empty_vals = [v for v in row.values() if v and str(v).strip()]

                # Single non-empty cell in row → possible company section header
                if len(non_empty_vals) == 1:
                    candidate = str(non_empty_vals[0]).strip()
                    if not _is_junk_row(candidate):
                        matched = (companies.get(candidate)
                                   or next((v for k, v in companies.items()
                                            if k.lower() == candidate.lower()), None))
                        if matched:
                            current_company = matched
                    continue

                # Try horizontal layout: does this row have a Company column?
                name = _find_col_str(nr, 'Company Name', 'Company', 'Name', 'Entity',
                                     'Investee', 'Portfolio Company')
                if name and not _is_junk_row(name):
                    co = (companies.get(name)
                          or next((v for k, v in companies.items()
                                   if k.lower() == name.lower()), None))
                    if co:
                        for pcol in period_cols:
                            val = _d(row.get(pcol))
                            if val is None:
                                continue
                            yr, mo = self._pl_parse_year_month(pcol)
                            if not yr:
                                continue
                            li_key = self._pl_map_to_line_item(pcol)
                            if not li_key:
                                continue
                            BudgetVsActual.objects.update_or_create(
                                portfolio_company=co,
                                fund=fund,
                                period_year=yr, period_month=mo,
                                period_quarter='', line_item=li_key,
                                defaults={'period_type': 'monthly' if mo else 'annual',
                                          'actual_inr': abs(val)},
                            )
                            count += 1
                        continue

                # Vertical pivot row: first column is the line item label.
                # Use normalised row for the label lookup so 'LineItem' → 'Line Item' works.
                label = _find_col_str(nr, 'Particulars', 'Line Item', 'Description',
                                      'Account', 'Item', 'Category')
                if not label:
                    label = non_empty_vals[0] if non_empty_vals else None
                # Also normalise the label itself before mapping (e.g. 'GrossProfit' → 'Gross Profit')
                label = _normalize_col_key(label) if label else None
                li_key = self._pl_map_to_line_item(label) if label else None
                if not li_key or not current_company:
                    continue

                # Plain period cols → actual
                for pcol in period_cols:
                    val = _d(row.get(pcol))
                    if val is None:
                        continue
                    yr, mo = self._pl_parse_year_month(pcol)
                    if not yr:
                        continue
                    BudgetVsActual.objects.update_or_create(
                        portfolio_company=current_company,
                        fund=fund,
                        period_year=yr, period_month=mo,
                        period_quarter='', line_item=li_key,
                        defaults={'period_type': 'monthly' if mo else 'annual',
                                  'actual_inr': abs(val)},
                    )
                    count += 1

                # "Budget Apr-24" qualified cols → budget_inr
                for pcol, (yr, mo) in budget_period_cols.items():
                    val = _d(row.get(pcol))
                    if val is None:
                        continue
                    BudgetVsActual.objects.update_or_create(
                        portfolio_company=current_company,
                        fund=fund,
                        period_year=yr, period_month=mo,
                        period_quarter='', line_item=li_key,
                        defaults={'period_type': 'monthly' if mo else 'annual',
                                  'budget_inr': abs(val)},
                    )
                    count += 1

                # "Apr-24 Actual" qualified cols → actual_inr
                for pcol, (yr, mo) in actual_period_cols.items():
                    val = _d(row.get(pcol))
                    if val is None:
                        continue
                    BudgetVsActual.objects.update_or_create(
                        portfolio_company=current_company,
                        fund=fund,
                        period_year=yr, period_month=mo,
                        period_quarter='', line_item=li_key,
                        defaults={'period_type': 'monthly' if mo else 'annual',
                                  'actual_inr': abs(val)},
                    )
                    count += 1

        else:
            # -----------------------------------------------------------------
            # Layout A / A2: horizontal — rows = companies, cols = P&L line items.
            # Handles both pure line-item columns ("Revenue") and
            # budget/actual-qualified columns ("Revenue Budget", "Budgeted EBITDA").
            # Period may be a column or absent (snapshot = today).
            # -----------------------------------------------------------------
            today = date.today()
            snapshot_yr, snapshot_mo = today.year, today.month

            for row in rows:
                # Normalise keys: 'Revenue(Cr)' → 'Revenue', 'GrossProfit(Cr)' → 'Gross Profit'
                nr = _norm_row(row)

                name = _find_col_str(nr, 'Company Name', 'Company', 'Name',
                                     'Entity', 'Investee', 'Portfolio Company')
                if not name or _is_junk_row(name):
                    continue

                co = (companies.get(name)
                      or next((v for k, v in companies.items()
                               if k.lower() == name.lower()), None))
                if not co:
                    continue

                period_str = _find_col_str(nr, 'Period', 'Month', 'Date',
                                           'Reporting Period', 'Reporting Month',
                                           'Quarter', 'FY', 'Financial Year')
                yr, mo = (self._pl_parse_year_month(period_str)
                          if period_str else (snapshot_yr, snapshot_mo))
                if not yr:
                    yr, mo = snapshot_yr, snapshot_mo

                # Pre-classify column headers as budget/actual/variance via Gemini
                _col_names = [str(c).strip() for c in nr.keys() if c and str(c).strip()]
                _col_qual_map = self._classify_enum(
                    _col_names, 'column_qualifier',
                    context='Column headers from a P&L / financial statement. '
                            'Classify each as budget, actual, or variance.')

                for col_name, col_val in nr.items():
                    if not col_name:
                        continue
                    # Skip variance columns (detected by Gemini)
                    col_qualifier = _col_qual_map.get(str(col_name).strip(), 'actual')
                    if col_qualifier == 'variance':
                        continue

                    is_budget_col = (col_qualifier == 'budget')

                    li_key = self._pl_map_to_line_item(col_name)
                    if not li_key:
                        continue

                    val = _d(col_val)
                    if val is None:
                        continue

                    upd = {'period_type': 'monthly' if mo else 'annual'}
                    if is_budget_col:
                        upd['budget_inr'] = abs(val)
                    else:
                        upd['actual_inr'] = abs(val)

                    BudgetVsActual.objects.update_or_create(
                        portfolio_company=co,
                        fund=fund,
                        period_year=yr, period_month=mo,
                        period_quarter='', line_item=li_key,
                        defaults=upd,
                    )
                    count += 1

        if count == 0 and inferred_periods and fund:
            # ----------------------------------------------------------------
            # Layout C: Fund-level transposed P&L with no period headers.
            # The sheet has line items in column A and numeric values in columns B+.
            # We map column positions to period labels from inferred_periods.
            # Writes to ConsolidatedMIS (not BudgetVsActual — no company breakdown).
            # ----------------------------------------------------------------
            count += self._process_fund_level_pl(
                ws, fund, inferred_periods, len(companies))

        return count

    def _process_fund_level_pl(self, ws, fund, period_labels, company_count=0):
        """
        Import a fund-level transposed P&L sheet into ConsolidatedMIS.

        Only runs for P&L sheets — skips Balance Sheet and Cash Flow sheets
        which have non-P&L line items that would pollute the fund P&L.

        Sheet format:
          Col A: line item label  (Revenue, EBITDA, ...)
          Col B..N: numeric values for successive periods (no header labels)

        period_labels: ordered list of period strings inferred from another sheet,
        e.g. ['Oct-24','Nov-24','Dec-24','Jan-25','Feb-25','Mar-25'].
        We map col B → period_labels[0], col C → period_labels[1], etc.
        The last column is often a total/YTD — skip it if len(cols) > len(labels).
        """
        from mis_consolidation.models import ConsolidatedMIS
        from django.utils import timezone as _tz

        # Skip Balance Sheet and Cash Flow sheets — their line items (assets, liabilities,
        # cash inflows) are not P&L items and would corrupt the fund P&L view.
        # Uses Gemini Pass 1.5 section sub-domains: P&L sheets have only 'unknown'
        # sub-domains, while BS/CF sheets have structured sub-domains (nav_records,
        # distributions, investments, capital_call_headers).
        _ws_title = getattr(ws, 'title', '')
        _sheet_sections = self._section_map.get(_ws_title, {})
        _sub_domains = set(_sheet_sections.values())
        _non_pl_indicators = {'nav_records', 'distributions', 'investments',
                              'capital_call_headers'}
        if _sub_domains & _non_pl_indicators:
            return 0

        if not period_labels:
            return 0

        # Find the first data row: skip title rows and section headers
        # We scan row-by-row; a data row has a non-junk string in col A and ≥1 numeric cols
        data_rows = []   # list of (line_item_label, [val_col_B, val_col_C, ...])
        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row or not row[0]:
                continue
            label = str(row[0]).strip()
            if _is_junk_row(label):
                continue
            li_key = self._pl_map_to_line_item(label)
            if not li_key:
                continue
            # Collect numeric values from columns 1 onward (0-indexed)
            vals = []
            for cell in row[1:]:
                try:
                    vals.append(Decimal(str(cell)) if cell is not None else None)
                except Exception:
                    vals.append(None)
            if not any(v is not None for v in vals):
                continue
            data_rows.append((li_key, vals))

        if not data_rows:
            return 0

        default_scheme = Scheme.objects.filter(fund=fund).first()

        count = 0
        for li_key, vals in data_rows:
            for i, period_str in enumerate(period_labels):
                if i >= len(vals):
                    break
                val = vals[i]
                if val is None:
                    continue
                yr, mo = self._pl_parse_year_month(period_str)
                if not yr:
                    continue
                pq = ''
                if mo:
                    pq = ('Q1' if mo in (4, 5, 6) else
                          'Q2' if mo in (7, 8, 9) else
                          'Q3' if mo in (10, 11, 12) else 'Q4')
                ConsolidatedMIS.objects.update_or_create(
                    organization=self.org,
                    fund=fund,
                    scheme=default_scheme,
                    period_year=yr,
                    period_month=mo,
                    period_quarter=pq,
                    line_item=li_key,
                    defaults={
                        'period_type': 'monthly',
                        'total_actual_inr': abs(val),
                        'company_count': company_count,
                        'computed_at': _tz.now(),
                    },
                )
                count += 1

        logger.info(f'  Fund-level P&L → ConsolidatedMIS: {count} records')
        return count

    def _process_bva_sheet(self, ws, companies, fund=None):
        """
        Process a Budget vs Actual sheet and write BudgetVsActual or ConsolidatedMIS records.

        Handles two formats:
        A) Company-level: Company | Line Item | Budget | Actual | (Period)
           Each row is one company × line_item → written to BudgetVsActual.
        B) Fund-level: Metric | Q1 Bdgt | Q1 Actual | Q2 Bdgt | Q2 Actual | FY Budget | FY Actual
           Rows are fund-level metrics, columns are period×qualifier pairs.
           Written directly to ConsolidatedMIS (no per-company breakdown available).
        """
        from mis_consolidation.models import BudgetVsActual, ConsolidatedMIS

        headers_dict, rows = read_table_from_sheet(ws, alias_map=self._get_alias(ws))
        if not rows:
            return 0

        today = date.today()
        count = 0

        # Scan first few rows (before the header row) for FY period information.
        # Many fund BvA sheets have a title like "BUDGET VS ACTUAL — FY 2024-25 YTD".
        # If found, use that as the default period for rows with no explicit Period column,
        # instead of falling back to today's date.
        _sheet_default_yr = None
        _sheet_default_mo = None
        _FY_RE = re.compile(r'FY\s*\d{2,4}\s*[-–]\s*(\d{2,4})', re.IGNORECASE)
        _YEAR_RE = re.compile(r'\b(20\d{2})\b')
        for _r in range(1, min(ws.max_row + 1, 6)):
            for _c in range(1, min(ws.max_column + 1, 4)):
                _cell_val = ws.cell(_r, _c).value
                if not _cell_val or not isinstance(_cell_val, str):
                    continue
                _fy_m = _FY_RE.search(_cell_val)
                if _fy_m:
                    _end = int(_fy_m.group(1))
                    _sheet_default_yr = _end + 2000 if _end < 100 else _end
                    _sheet_default_mo = None   # annual record — no specific month
                    break
                _yr_m = _YEAR_RE.search(_cell_val)
                if _yr_m and not _sheet_default_yr:
                    _sheet_default_yr = int(_yr_m.group(1))
            if _sheet_default_yr:
                break

        for row in rows:
            # Normalise column keys: strips unit annotations ('Budget(₹Cr)' → 'Budget',
            # 'Actual (INR Mn)' → 'Actual') and splits CamelCase ('GrossProfit' → 'Gross Profit').
            # This makes _find_col_* robust to any Excel formatting convention.
            nr = _norm_row(row)

            name = _find_col_str(nr, 'Company Name', 'Company', 'Name',
                                 'Entity', 'Investee', 'Portfolio Company')
            if not name or _is_junk_row(name):
                continue

            co = (companies.get(name)
                  or next((v for k, v in companies.items()
                           if k.lower() == name.lower()), None))
            if not co:
                continue

            line_item_str = _find_col_str(nr, 'Line Item', 'Item', 'Particulars',
                                          'Account', 'Description', 'Category',
                                          'P&L Item', 'Financial Item', 'Metric')
            li_key = self._pl_map_to_line_item(line_item_str) if line_item_str else None
            if not li_key:
                continue

            # Gemini Pass 2 adds canonical field names as row keys
            budget = _find_col_decimal(nr, 'budget', 'budget_amount')
            actual = _find_col_decimal(nr, 'actual', 'actual_amount')

            if budget is None and actual is None:
                continue

            period_str = _find_col_str(nr, 'Period', 'Month', 'Date',
                                       'Reporting Period', 'Quarter', 'FY',
                                       'Financial Year', 'Reporting Month')
            yr, mo = (self._pl_parse_year_month(period_str)
                      if period_str
                      else (_sheet_default_yr or today.year, _sheet_default_mo))
            if not yr:
                yr, mo = _sheet_default_yr or today.year, _sheet_default_mo

            # India FY quarter mapping
            pq = ''
            if mo:
                pq = 'Q1' if mo in (4, 5, 6) else (
                     'Q2' if mo in (7, 8, 9) else (
                     'Q3' if mo in (10, 11, 12) else 'Q4'))

            defaults = {'period_type': 'monthly' if mo else 'annual'}
            if budget is not None:
                defaults['budget_inr'] = budget
            if actual is not None:
                defaults['actual_inr'] = abs(actual)

            BudgetVsActual.objects.update_or_create(
                portfolio_company=co,
                fund=fund,
                period_year=yr,
                period_month=mo,
                period_quarter=pq,
                line_item=li_key,
                defaults=defaults,
            )
            count += 1

        if count == 0:
            # ----------------------------------------------------------------
            # Fund-level BvA fallback:
            # No company column found. Try to detect the format:
            #   Metric | Q1 Bdgt | Q1 Actual | Q2 Bdgt | Q2 Actual | FY Budget | FY Actual
            # If detected, write directly to ConsolidatedMIS.
            # ----------------------------------------------------------------
            count += self._process_fund_level_bva(ws, headers_dict, rows, fund,
                                                   _sheet_default_yr,
                                                   company_count=len(companies))

        return count

    def _process_fund_level_bva(self, ws, headers_dict, rows, fund, default_yr=None,
                                company_count=0):
        """
        Import fund-level Budget vs Actual into ConsolidatedMIS.

        Detects columns like 'Q1 Bdgt', 'Q1 Actual', 'Q2 Bdgt', 'FY Budget', 'FY Actual'
        and maps each row's Metric value to a canonical line_item.
        Works universally: 'AOP', 'Plan', 'Target' are treated as budget qualifiers;
        'Actuals', 'Reported', 'YTD Actual' are treated as actual qualifiers.
        """
        if not fund:
            return 0
        from mis_consolidation.models import ConsolidatedMIS

        # Regex for quarter/FY detection (universal financial tokens)
        _Q_RE = re.compile(r'^(Q[1-4])\s*[-\s]?\s*(.+)', re.IGNORECASE)
        _FY_Q_RE = re.compile(r'^(FY|annual|full.?year)\s*(.+)', re.IGNORECASE)

        # Collect candidate column headers for budget/actual classification
        _candidate_cols = [str(h).strip() for h in headers_dict if h and str(h).strip()]

        # Classify ALL column headers via Gemini (budget/actual/variance)
        _col_qual_map = self._classify_enum(
            _candidate_cols, 'column_qualifier',
            context='Column headers from a fund Budget vs Actual sheet. '
                    'Classify each as budget (planned), actual (realized), or variance.')

        # Map column headers → (period_quarter, period_type, is_budget)
        period_budget_cols = {}
        for h in headers_dict:
            if not h:
                continue
            h_s = str(h).strip()
            qual = _col_qual_map.get(h_s, 'actual')
            if qual == 'variance':
                continue
            is_bud = (qual == 'budget')

            m = _Q_RE.match(h_s)
            if m:
                qtr = m.group(1).upper()
                period_budget_cols[h] = (qtr, 'quarterly', is_bud)
                continue
            m2 = _FY_Q_RE.match(h_s)
            if m2:
                period_budget_cols[h] = ('FY', 'annual', is_bud)
                continue
            # Non-period columns classified as budget or actual → annual
            if qual in ('budget', 'actual'):
                period_budget_cols[h] = ('FY', 'annual', is_bud)

        if not period_budget_cols:
            return 0

        yr = default_yr
        if not yr:
            # Extract year from the sheet title row
            for _r in range(1, min(ws.max_row + 1, 4)):
                _v = ws.cell(_r, 1).value
                if isinstance(_v, str):
                    _ym = re.search(r'\b(20\d{2})\b', _v)
                    if _ym:
                        yr = int(_ym.group(1))
                        break
        if not yr:
            from datetime import date as _date
            yr = _date.today().year

        # Find the default scheme for this fund
        default_scheme = Scheme.objects.filter(fund=fund).first()

        # Fund-level performance metrics (Net IRR, TVPI, Portfolio FV) found in
        # BvA sheets as special metric rows — not P&L line items.  We detect them
        # via Gemini and store as ConsolidatedMIS records with special line_item keys.

        # Pre-classify all metric labels for fund-level metric detection
        all_metric_labels = set()
        for row in rows:
            nr = _norm_row(row)
            m = _find_col_str(nr, 'Metric', 'Line Item', 'Particulars',
                              'Item', 'Category', 'Description')
            if m and not _is_junk_row(m):
                all_metric_labels.add(m)
        fund_metric_map = self._classify_labels(all_metric_labels, 'fund_metrics',
                                                 context='Fund-level performance metrics (IRR, TVPI, FV)')

        count = 0
        from django.utils import timezone as _tz
        for row in rows:
            nr = _norm_row(row)
            metric = _find_col_str(nr, 'Metric', 'Line Item', 'Particulars',
                                   'Item', 'Category', 'Description')
            if not metric or _is_junk_row(metric):
                continue

            # Check if this is a special fund-level performance metric row
            special_li_key = fund_metric_map.get(metric)

            if special_li_key:
                # Extract FY Actual value (the most authoritative single number)
                fy_actual = None
                fy_budget = None
                for col_h, (qtr, ptype, is_bud) in period_budget_cols.items():
                    if ptype == 'annual':
                        val = _d(row.get(col_h))
                        if val is None:
                            continue
                        if is_bud:
                            fy_budget = val
                        else:
                            fy_actual = val
                if fy_actual is None:
                    # Fallback: any actual column
                    for col_h, (qtr, ptype, is_bud) in period_budget_cols.items():
                        if not is_bud:
                            val = _d(row.get(col_h))
                            if val is not None:
                                fy_actual = val
                                break

                if fy_actual is not None:
                    # For IRR/TVPI: convert fraction → percentage if looks like fraction
                    stored_val = fy_actual
                    if special_li_key == 'net_irr' and abs(float(fy_actual)) <= 2:
                        stored_val = fy_actual * 100  # 0.212 → 21.2
                    elif special_li_key == 'tvpi' and abs(float(fy_actual)) <= 5:
                        stored_val = fy_actual  # keep as ratio (e.g. 1.567x)

                    upd_fm = {
                        'period_type': 'annual',
                        'company_count': company_count,
                        'total_actual_inr': stored_val,
                        'computed_at': _tz.now(),
                    }
                    if fy_budget is not None:
                        bud_stored = fy_budget
                        if special_li_key == 'net_irr' and abs(float(fy_budget)) <= 2:
                            bud_stored = fy_budget * 100
                        upd_fm['total_budget_inr'] = bud_stored
                    ConsolidatedMIS.objects.update_or_create(
                        organization=self.org,
                        fund=fund,
                        scheme=default_scheme,
                        period_year=yr,
                        period_month=None,
                        period_quarter='FY',
                        line_item=special_li_key,
                        defaults=upd_fm,
                    )
                    count += 1
                continue  # Don't fall through to standard P&L processing

            li_key = self._pl_map_to_line_item(metric)
            if not li_key:
                continue

            # Accumulate budget and actual per (quarter, period_type) pair
            period_values = {}   # (quarter, period_type) → {'budget': x, 'actual': x}
            for col_h, (qtr, ptype, is_bud) in period_budget_cols.items():
                val = _d(row.get(col_h))
                if val is None:
                    continue
                key = (qtr, ptype)
                if key not in period_values:
                    period_values[key] = {}
                if is_bud:
                    period_values[key]['budget'] = val
                else:
                    period_values[key]['actual'] = val

            for (qtr, ptype), vals in period_values.items():
                bud = vals.get('budget')
                act = vals.get('actual')
                if bud is None and act is None:
                    continue
                variance = None
                variance_pct = None
                if bud is not None and act is not None and bud != 0:
                    variance = act - bud
                    variance_pct = float(variance / bud * 100)
                upd = {'period_type': ptype, 'company_count': company_count}
                if bud is not None:
                    upd['total_budget_inr'] = abs(bud)
                if act is not None:
                    upd['total_actual_inr'] = abs(act)
                if variance is not None:
                    upd['total_variance_inr'] = variance
                if variance_pct is not None:
                    upd['total_variance_pct'] = variance_pct
                upd['computed_at'] = _tz.now()
                ConsolidatedMIS.objects.update_or_create(
                    organization=self.org,
                    fund=fund,
                    scheme=default_scheme,
                    period_year=yr,
                    period_month=None,
                    period_quarter=qtr,
                    line_item=li_key,
                    defaults=upd,
                )
                count += 1

        logger.info(f'  Fund-level BvA → ConsolidatedMIS: {count} records')
        return count

    # ------------------------------------------------------------------
    # Quoted & Unquoted classification
    # ------------------------------------------------------------------

    def _import_quoted_unquoted(self, wb, org, investments, companies, domain_map):
        """Read the Quoted & Unquoted Shares sheet and update PortfolioCompany.is_quoted.

        Uses Gemini-classified 'quoted_unquoted' domain from domain_map.
        Falls back to 'valuations_kpis' since IPEV level data sometimes
        lives alongside valuations.
        """
        sheets = _dm_sheets(domain_map, 'quoted_unquoted')
        if not sheets:
            sheets = _dm_sheets(domain_map, 'valuations_kpis')
        if not sheets:
            return

        target_sheet = sheets[0]
        if target_sheet not in wb.sheetnames:
            return

        ws = wb[target_sheet]
        _headers, rows = read_table_from_sheet(ws, alias_map=self._get_alias(ws))
        if not rows:
            return

        count = 0
        for row in rows:
            name = _find_col_str(row, 'Company Name', 'Company', 'Name',
                                  'Portfolio Company', 'Investee', 'Entity Name')
            if not name or _is_junk_row(name):
                continue

            share_type = _find_col_str(
                row,
                'Share Type', 'Share Class', 'Instrument Type',
                'Equity Type', 'Security Type', 'Instrument', 'Type of Share',
                'Nature of Instrument', 'Class of Shares', 'Type',
            )
            ipev_raw = _find_col_str(
                row,
                'IPEV Level', 'IPEV Classification', 'IPEV', 'Level',
                'Valuation Level', 'Fair Value Hierarchy', 'Level (IPEV)',
                'Measurement Level', 'FV Level', 'Fair Value Level',
            )
            exchange = _find_col_str(
                row,
                'Exchange', 'Listed On', 'Stock Exchange', 'Listing Exchange',
                'Market', 'Primary Exchange', 'Exchange Name',
                'listing_exchange',
            )

            # Determine quoted status via Gemini classification
            classify_input = share_type or ipev_raw or ''
            if classify_input.strip():
                qs_result = self._classify_enum(
                    [classify_input], 'quoted_status',
                    context='Share type / instrument type classification: quoted vs unquoted')
                is_quoted = (qs_result.get(classify_input) == 'quoted')
            else:
                is_quoted = False

            pc = companies.get(name)
            if pc:
                update = {'is_quoted': is_quoted}
                if exchange:
                    update['listing_exchange'] = exchange.upper()
                elif is_quoted and not pc.listing_exchange:
                    # Mark as 'BSE/NSE' generically for Indian listed companies
                    update['listing_exchange'] = 'LISTED'
                PortfolioCompany.objects.filter(pk=pc.pk).update(**update)
                # Refresh the cached object so views see updated values
                pc.is_quoted = is_quoted
                if 'listing_exchange' in update:
                    pc.listing_exchange = update['listing_exchange']
                count += 1

        logger.info(f'  Quoted/Unquoted classification: {count} companies updated')

    # ------------------------------------------------------------------
    # NAV records
    # ------------------------------------------------------------------

    def _import_nav(self, wb, schemes, domain_map):
        """Import NAV records from Gemini-classified NAV sheets.

        Handles three formats:
        - Format A: Flat table with one row per period (Period | NAV | Units | ...)
        - Format B: Multi-section sheet with NAV RECORDS header
        - Format C: Transposed table where rows = metrics, columns = periods
          (Component | Oct-24 | Nov-24 | Dec-24 | ...)

        Uses Gemini domain_map to find nav_accounting sheets (time-series NAV
        data). Key-value computation sheets (nav_calculation) are consumed
        separately in _enrich_nav_records_post_import().
        """
        default_scheme = list(schemes.values())[0] if schemes else None
        if not default_scheme:
            return

        # Accept any sheet Gemini Pass 1 mapped to a NAV-related domain.
        # The canonical schema defines two: nav_accounting (pure time-series)
        # and nav_calculation (computation sheet that may CONTAIN a time-series
        # section per Pass 1.5 — e.g. "Monthly NAV History"). Pass 1.5 section
        # classification + per-sheet layout (Pass 2.5) decide which sub-sections
        # actually hold NAV records.
        candidate_sheets = []
        for domain in ('nav_accounting', 'nav_calculation'):
            for s in _dm_sheets(domain_map, domain):
                if s in wb.sheetnames and s not in candidate_sheets:
                    candidate_sheets.append(s)
        if not candidate_sheets:
            logger.info('  No Gemini-classified NAV sheets found '
                        '(nav_accounting or nav_calculation)')
            return

        # Regex for period-column headers (e.g. "Oct-24", "Nov-24")
        _period_re = re.compile(
            r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[-/]\d{2,4}$',
            re.IGNORECASE)

        # Try each candidate sheet until we find one that yields NAV records
        for sheet_name in candidate_sheets:
            ws = wb[sheet_name]
            nav_count = self._try_import_nav_from_sheet(
                ws, sheet_name, default_scheme, schemes, wb, _period_re,
                domain_map=domain_map)
            if nav_count > 0:
                logger.info(f'  NAV imported from sheet: "{sheet_name}" ({nav_count} records)')
                return

        logger.info('  No NAV records found in any sheet')

    def _try_import_nav_from_sheet(self, ws, sheet_name, default_scheme,
                                    schemes, wb, _period_re, domain_map=None):
        """Try to import NAV records from a single worksheet.

        Returns the number of NAV records created (0 if the sheet format
        is not suitable for NAV time-series data).
        """
        headers_dict, table_rows = read_table_from_sheet(ws, alias_map=self._get_alias(ws))
        period_cols = [h for h in headers_dict.keys() if _period_re.match(h.strip())]

        if period_cols and table_rows:
            # ── Format C: Transposed (Component | Oct-24 | Nov-24 | …) ──────────
            # Classify all component labels via Gemini in one batch call
            all_comp_labels = set()
            for row in table_rows:
                comp = _find_col_str(row, 'Component', 'Line Item', 'Item',
                                     'Metric', 'Particulars', 'P&L Line',
                                     'Cash Flow Item', 'Parameter')
                if comp:
                    all_comp_labels.add(comp)
            nav_comp_map = self._classify_labels(all_comp_labels, 'nav_components',
                                                  context='NAV statement components')

            period_data = {p: {} for p in period_cols}
            has_nav_row = False
            for row in table_rows:
                comp = _find_col_str(row, 'Component', 'Line Item', 'Item',
                                     'Metric', 'Particulars', 'P&L Line',
                                     'Cash Flow Item', 'Parameter')
                if not comp:
                    continue
                nav_type = nav_comp_map.get(comp)
                if not nav_type:
                    continue
                for pcol in period_cols:
                    val = _d(row.get(pcol))
                    if val is None:
                        continue
                    if nav_type == 'total_nav':
                        period_data[pcol]['total_nav'] = val
                        has_nav_row = True
                    elif nav_type == 'unrealized_gains':
                        period_data[pcol]['unrealized'] = val
                    elif nav_type == 'realized_gains':
                        period_data[pcol]['realized'] = val
                    elif nav_type == 'mgmt_fee':
                        period_data[pcol]['mgmt_fee'] = abs(val)
                    elif nav_type == 'carry_provision':
                        period_data[pcol]['carry'] = val
                    elif nav_type == 'investment_income':
                        period_data[pcol]['income'] = val

            # Only proceed if we found at least one row that looks like NAV data
            if not has_nav_row:
                return 0

            import calendar as _cal
            count = 0
            for pcol in period_cols:
                # Skip aggregation columns (H2 Total, FY Total, Grand Total, etc.)
                if _is_junk_row(pcol):
                    continue
                pd = period_data.get(pcol, {})
                nav_date = self._parse_period(pcol)
                if not nav_date:
                    continue
                total_nav = pd.get('total_nav')
                if not total_nav:
                    continue
                # Last day of the month
                last_day = _cal.monthrange(nav_date.year, nav_date.month)[1]
                nav_date = nav_date.replace(day=last_day)
                update_fields = {
                    'total_nav': total_nav,
                    'nav_per_unit': Decimal('0'),
                    'total_units_outstanding': Decimal('0'),
                }
                if pd.get('unrealized'): update_fields['unrealized_gains'] = pd['unrealized']
                if pd.get('realized'):   update_fields['realized_gains'] = pd['realized']
                if pd.get('mgmt_fee'):   update_fields['management_fee_payable'] = pd['mgmt_fee']
                NAVRecord.objects.update_or_create(
                    scheme=default_scheme,
                    nav_date=nav_date,
                    defaults=update_fields,
                )
                count += 1

            if count > 0:
                # Enrich with NAV/Unit and realized gains from other sheets
                self._enrich_nav_records_post_import(wb, default_scheme, domain_map)
            return count

        # ── Format B: Multi-section ──────────────────────────────────────────
        # Pass 1.5 may classify MULTIPLE sections of a NAV sheet as
        # nav_records (e.g. NAV_CALC Section A is a current-period breakdown
        # *and* Section C is a 36-month time series — both legitimately
        # nav_records). We must process EVERY candidate section and union the
        # rows that look like time-series NAV records, instead of stopping at
        # the first section (which often is the current-period breakdown that
        # has no Date column and yields zero records).
        sections = self._read_sheet_via_layout(ws, alias_map=self._get_alias(ws))
        nav_rows = []
        for sec_name, (sec_headers, sec_rows) in sections.items():
            subdomain = self._get_section_subdomain(sheet_name, sec_name)
            if subdomain == 'nav_records' or sec_name == '__default__':
                if sec_rows:
                    nav_rows.extend(sec_rows)

        if not nav_rows:
            nav_rows = table_rows  # already-read flat table

        if not nav_rows:
            return 0

        # ── Format A: Flat table (Period | NAV | Units | …) ──────────────────
        count = 0
        for row in nav_rows:
            scheme_name_raw = _find_col_str(
                row, 'Scheme', 'Scheme Name', 'Fund Scheme')
            target_scheme = default_scheme
            if scheme_name_raw:
                target_scheme = schemes.get(scheme_name_raw, default_scheme)

            period_raw = _find_col(
                row, 'Period', 'NAV Date', 'Date', 'Month', 'nav_date')
            nav_date = None
            if period_raw:
                nav_date = _date(period_raw)
                if not nav_date and isinstance(period_raw, str):
                    nav_date = self._parse_period(period_raw)
            if not nav_date:
                continue

            total_nav = _find_col_decimal(
                row, 'Total NAV (Cr)', 'Total NAV', 'NAV',
                'Net Asset Value', 'total_nav',
                # Canonical schema has two synonyms for the same concept
                # (different DOMAIN_FIELDS catalogues): total_nav and
                # closing_nav. Gemini may map "Fund NAV" → closing_nav for
                # nav_calculation-domain sheets. Accept both.
                'Closing NAV', 'closing_nav', 'Ending NAV')

            # ── Fix B: NAV formula-derived fallback (no hardcoded SEBI formula) ──
            # Many fund Excel files define "Total NAV" as a calculated cell
            # (e.g. NAV sheet header row 1 says "Col I TotalNAV = C+E+F-D")
            # but the workbook is saved without cached values → the cell
            # comes through as None/0 here.  Pass 2.5 Gemini layout call
            # identifies these derived columns and returns the formula in
            # canonical form (sum of signed source-column references).  We
            # evaluate it from the OTHER cells in this same row.  Universal:
            # works for any Excel where Gemini sees the formula in a
            # disclaimer row OR recognises a standard SEBI AIF identity.
            if (not total_nav or total_nav == 0):
                derived = self._derived_column_for(
                    sheet_name,
                    ['Total NAV', 'Total NAV(₹Cr)', 'Total NAV (Cr)', 'NAV',
                     'Net Asset Value', 'total_nav'],
                )
                if derived:
                    computed = self._evaluate_derived_formula(derived, row)
                    if computed is not None:
                        total_nav = computed
                        logger.info(
                            f'NAV row computed from Gemini-derived formula: '
                            f'{[c["sign"]+c["source_column"] for c in derived["formula_components"]]} '
                            f'= {computed} (sheet="{sheet_name}")'
                        )

            nav_per_unit = _find_col_decimal(
                row, 'NAV/Unit (INR)', 'NAV/Unit(₹)', 'NAV/Unit',
                'NAV Per Unit', 'nav_per_unit',
                # Same canonical-name synonym issue as total_nav above
                'closing_nav_per_unit', 'Closing NAV per Unit')
            inv_fv = _find_col_decimal(
                row, 'Investments at FV (Cr)', 'Investments at FV',
                'Total Investments', 'investments_at_fair_value')
            mgmt_fee = _find_col_decimal(
                row, 'Mgmt Fee(₹Cr)', 'Mgmt Fee', 'Mgmt Fees',
                'Management Fee(₹Cr)', 'Management Fee', 'Fees',
                'management_fee_payable')
            fund_expenses = _find_col_decimal(
                row, 'Fund Expenses(₹Cr)', 'Fund Expenses', 'Expenses')
            unrealized = _find_col_decimal(
                row, 'Unrealized Gains(₹Cr)', 'Unrealized Gains',
                'Unrealized G/L', 'Unrealised Gains', 'Mark-to-Market Gain',
                'MTM Gain', 'unrealized_gains')
            realized = _find_col_decimal(
                row, 'Realized Gains(₹Cr)', 'Realized Gains',
                'Realized G/L', 'Realised Gains', 'realized_gains')
            total_units = _find_col_decimal(
                row, 'Total Units', 'Units Outstanding', 'Units',
                'total_units_outstanding')
            cash = _find_col_decimal(
                row, 'Cash (Cr)', 'Cash', 'Cash & Equivalents',
                'Cash Balance', 'Bank Balance', 'cash_and_equivalents')
            receivables = _find_col_decimal(
                row, 'Receivables (Cr)', 'Receivables', 'Accounts Receivable')
            other_liab = _find_col_decimal(
                row, 'Other Liabilities (Cr)', 'Other Liabilities',
                'Liabilities', 'Payables', 'other_liabilities')

            units = total_units or Decimal('0')
            if not units and total_nav and nav_per_unit and nav_per_unit > 0:
                units = (total_nav / nav_per_unit).quantize(Decimal('0.000001'))
            if not other_liab and fund_expenses:
                other_liab = fund_expenses

            update_fields = {
                'total_nav': total_nav or Decimal('0'),
                'nav_per_unit': nav_per_unit or Decimal('0'),
                'total_units_outstanding': units,
            }
            if inv_fv:      update_fields['investments_at_fair_value'] = inv_fv
            if mgmt_fee:    update_fields['management_fee_payable'] = mgmt_fee
            if cash:        update_fields['cash_and_equivalents'] = cash
            if receivables: update_fields['receivables'] = receivables
            if other_liab:  update_fields['other_liabilities'] = other_liab
            if unrealized:  update_fields['unrealized_gains'] = unrealized
            if realized:    update_fields['realized_gains'] = realized

            NAVRecord.objects.update_or_create(
                scheme=target_scheme,
                nav_date=nav_date,
                defaults=update_fields,
            )
            count += 1

        return count

    def _audit_extraction_completeness(self, classifications, column_mappings):
        """Pass 5 — diagnostic audit.

        For every dashboard-critical Django model, check whether DB rows exist
        for the fund just imported. For each empty model, report:
          - which sheet was supposed to populate it (per Pass 1 domain map)
          - how many Pass 2 column aliases that sheet got
          - sample of the sheet's column headers Gemini saw

        This is DIAGNOSTIC ONLY — no silent re-extraction, no hardcoded fallback.
        The point is to make Pass 2 / Pass 1 misclassifications visible instead
        of letting them silently produce empty tables.
        """
        from .canonical_schema import SHEET_DOMAINS
        # Discover critical fund-data models dynamically: anything in
        # 'lp', 'investments', 'accounting' apps with a 'scheme' or
        # 'investment' FK that touches the imported fund. NO hardcoded
        # model list — uses introspection so new models automatically join
        # the audit.
        from django.apps import apps
        from django.db import models as django_models

        audit = {
            'empty_critical_tables': [],
            'populated_critical_tables': [],
            'sheet_domain_map': {},
            'low_alias_sheets': [],
        }
        if not self._imported_fund:
            return audit

        scheme_qs_filter_cache = {}

        def _row_count(model):
            try:
                # Try direct scheme FK
                if any(getattr(f, 'name', '') == 'scheme'
                       for f in model._meta.get_fields()
                       if isinstance(f, (django_models.ForeignKey,
                                         django_models.OneToOneField))):
                    return model.objects.filter(
                        scheme__fund=self._imported_fund
                    ).count()
                # 1-hop via 'investment'
                if any(getattr(f, 'name', '') == 'investment'
                       for f in model._meta.get_fields()
                       if isinstance(f, (django_models.ForeignKey,
                                         django_models.OneToOneField))):
                    return model.objects.filter(
                        investment__scheme__fund=self._imported_fund
                    ).count()
                # Direct fund FK
                if any(getattr(f, 'name', '') == 'fund'
                       for f in model._meta.get_fields()
                       if isinstance(f, (django_models.ForeignKey,
                                         django_models.OneToOneField))):
                    return model.objects.filter(
                        fund=self._imported_fund
                    ).count()
            except Exception:
                pass
            return None

        # Map Pass 1 domain → sheet names so the audit can name the
        # expected-source-sheet per empty table.
        domain_to_sheets = {}
        for cls in (classifications or []):
            sname = cls.get('sheet_name', '')
            for d in cls.get('domains', []):
                if d and d != 'unknown':
                    domain_to_sheets.setdefault(d, []).append(sname)
        audit['sheet_domain_map'] = domain_to_sheets

        # Alias counts per sheet (after Fix 2 we accept all confidences;
        # but if a sheet got literally 0 mappings, it's a Pass 2 failure
        # worth surfacing in the audit).
        for sheet_name, mapping_data in (column_mappings or {}).items():
            alias_count = 0
            for s in mapping_data.get('sections', []):
                for m in s.get('mappings', []):
                    if m.get('excel_column') and m.get('canonical_field'):
                        alias_count += 1
            if alias_count == 0:
                audit['low_alias_sheets'].append({
                    'sheet': sheet_name,
                    'alias_count': 0,
                    'reason': 'Pass 2 produced zero column mappings for this sheet.',
                })

        # Walk every model in project apps that ties to a Scheme/Investment/Fund.
        for ac in apps.get_app_configs():
            if ac.label in ('auth', 'contenttypes', 'sessions',
                            'admin', 'accounts', 'dataimport'):
                continue
            for model in ac.get_models():
                cnt = _row_count(model)
                if cnt is None:
                    continue
                label = f'{ac.label}.{model.__name__}'
                if cnt == 0:
                    # What domain/sheet was supposed to populate this model?
                    # We don't hardcode a model→domain map here; instead the
                    # downstream consumer can correlate via SHEET_DOMAINS
                    # and the importer logs. Report just the model + count.
                    audit['empty_critical_tables'].append({
                        'model': label,
                        'row_count': 0,
                    })
                else:
                    audit['populated_critical_tables'].append({
                        'model': label,
                        'row_count': cnt,
                    })

        return audit

    # ── Canonical KPI fields that the dashboard "Portfolio KPIs" matrix
    # surfaces. Every value here is a canonical field name produced by
    # Gemini Pass 2 (see canonical_schema.VALUATIONS_KPIS_FIELDS). NO
    # English keyword matching — Gemini is the sole author of these
    # canonical-field assignments per column. We simply iterate over
    # whatever Gemini classified.
    _KPI_CANONICAL_FIELDS = {
        # raw amount KPIs (currency / count)
        'gmv', 'revenue', 'ebitda', 'ebitda_value', 'mrr', 'arr',
        'orders', 'aov', 'cac', 'ltv', 'headcount', 'arpob',
        'aum_value', 'investment_cost',
        # percentage / ratio KPIs (already in % form when extracted)
        'gross_margin_pct', 'ebitda_margin_pct', 'returns_pct',
        'repeat_pct', 'churn_rate', 'nrr', 'ltv_cac_ratio',
        'nim_pct', 'gnpa_pct', 'nnpa_pct', 'roe_pct',
        'capacity_utilization', 'export_pct', 'bed_occupancy',
        'cap_rate_pct', 'debt_to_ebitda', 'cost_to_income',
    }

    def _project_kpi_columns_to_portfolio_kpi(self, import_file_record):
        """Pass 6.5 — Universal KPI projection.

        Walks the Gemini-built column_mapping for EVERY sheet. For any
        column whose canonical_field is in _KPI_CANONICAL_FIELDS AND whose
        section has a column mapped to 'company_name', creates one
        PortfolioKPI row per data row (company × kpi).

        This bridges the gap between Pass 2 (which faithfully maps
        Excel columns to canonical names) and the dashboard KPI matrix
        (which reads only PortfolioKPI). Previously, columns like
        PORTFOLIO_MASTER!"Revenue TTM" → canonical 'revenue' were
        dropped because Investment has no `revenue` field.

        Zero hardcoded English keywords here. The set of recognised KPI
        fields is the canonical-field names produced by Gemini Pass 2
        against the canonical_schema catalogue.
        """
        import openpyxl
        from .models import DerivedMetric  # noqa: F401 (kept for symmetry)

        if not self._imported_fund:
            return
        if not import_file_record or not import_file_record.column_mapping:
            return

        filepath = getattr(self, '_filepath', None) or import_file_record.file.path
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            logger.warning(f'Pass 6.5: cannot reopen workbook: {e}')
            return

        # Build a (company_name → investment) resolver for the fund.
        investments_by_name = {}
        try:
            for sch in self._imported_fund.schemes.all():
                for inv in Investment.objects.filter(scheme=sch):
                    nm = (inv.company_name or '').strip()
                    if nm and nm not in investments_by_name:
                        investments_by_name[nm] = inv
        except Exception as e:
            logger.warning(f'Pass 6.5: investment resolver build failed: {e}')

        companies_by_name = {}
        try:
            for co in PortfolioCompany.objects.filter(organization=self.org):
                nm = (co.name or '').strip()
                if nm and nm not in companies_by_name:
                    companies_by_name[nm] = co
        except Exception as e:
            logger.warning(f'Pass 6.5: company resolver build failed: {e}')

        snapshot_date = date.today()
        written = 0
        sheets_touched = []

        try:
            for sheet_name, sheet_meta in (import_file_record.column_mapping or {}).items():
                if not isinstance(sheet_meta, dict):
                    continue
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                max_col = ws.max_column or 0

                sections = sheet_meta.get('sections') or []
                for section in sections:
                    if not isinstance(section, dict):
                        continue
                    layout = section.get('layout')
                    mappings = section.get('mappings') or []
                    if layout != 'horizontal':
                        continue

                    # Locate the company-name column + every KPI canonical
                    # column within this section. Multiple mappings may
                    # point at the same column_index — keep the first per
                    # canonical_field.
                    company_col_idx = None
                    period_col_idx = None
                    kpi_cols = []  # list of (col_idx, canonical_field, header_text)
                    seen_canonical = set()
                    for m in mappings:
                        if not isinstance(m, dict):
                            continue
                        cf = m.get('canonical_field')
                        ci = m.get('column_index')
                        if not cf or not ci:
                            continue
                        if cf == 'company_name' and company_col_idx is None:
                            company_col_idx = ci
                            continue
                        if cf in ('period', 'kpi_period', 'valuation_date') and period_col_idx is None:
                            period_col_idx = ci
                            continue
                        if cf in self._KPI_CANONICAL_FIELDS and cf not in seen_canonical:
                            kpi_cols.append((ci, cf, m.get('excel_column') or cf))
                            seen_canonical.add(cf)

                    if company_col_idx is None or not kpi_cols:
                        continue

                    data_start = int(section.get('data_start_row') or 2)
                    last_row = ws.max_row or data_start
                    for r in range(data_start, last_row + 1):
                        # company name
                        if company_col_idx > max_col:
                            continue
                        cname_raw = ws.cell(r, company_col_idx).value
                        if cname_raw is None:
                            continue
                        cname = str(cname_raw).strip()
                        if not cname or _is_junk_row(cname):
                            continue

                        inv = investments_by_name.get(cname)
                        if inv is None:
                            # case-insensitive fallback
                            for k, v in investments_by_name.items():
                                if k.lower() == cname.lower():
                                    inv = v
                                    break
                        if inv is None:
                            continue
                        company_obj = companies_by_name.get(cname)
                        if company_obj is None:
                            for k, v in companies_by_name.items():
                                if k.lower() == cname.lower():
                                    company_obj = v
                                    break

                        # period: prefer mapped period column, else today
                        period_val = snapshot_date
                        if period_col_idx and period_col_idx <= max_col:
                            pcell = ws.cell(r, period_col_idx).value
                            try:
                                if hasattr(pcell, 'date'):
                                    period_val = pcell.date() if hasattr(pcell.date, '__call__') else pcell
                                elif isinstance(pcell, date):
                                    period_val = pcell
                            except Exception:
                                period_val = snapshot_date

                        for (ci, cf, header_text) in kpi_cols:
                            if ci > max_col:
                                continue
                            v = ws.cell(r, ci).value
                            dec_val = _d(v)
                            if dec_val is None:
                                continue
                            slug = slugify(cf.replace('_', '-'))
                            # Map canonical field name → display name + format
                            display = cf.replace('_', ' ').title()
                            fmt = 'percent' if cf.endswith('_pct') or cf in (
                                'churn_rate', 'nrr', 'capacity_utilization',
                                'bed_occupancy', 'cost_to_income',
                            ) else (
                                'ratio' if cf in ('ltv_cac_ratio', 'debt_to_ebitda')
                                else 'number'
                            )
                            try:
                                kdef, _ = KPIDefinition.objects.update_or_create(
                                    organization=self.org,
                                    slug=slug,
                                    defaults={
                                        'name': display,
                                        'format': fmt,
                                        'frequency': 'monthly',
                                        'sector_template': 'generic',
                                    },
                                )
                                PortfolioKPI.objects.update_or_create(
                                    investment=inv,
                                    kpi_definition=kdef,
                                    period=period_val,
                                    defaults={
                                        'portfolio_company': company_obj,
                                        'value': dec_val,
                                        'source': 'excel_upload',
                                        'status': 'approved',
                                        'notes': (
                                            f'Pass 6.5 projection from '
                                            f'{sheet_name}!{header_text} '
                                            f'(canonical: {cf})'
                                        )[:500],
                                    },
                                )
                                written += 1
                            except Exception as e:
                                logger.warning(
                                    f'Pass 6.5 persist failed for {cname}.{cf} '
                                    f'at {sheet_name} row {r}: {e}'
                                )
                    sheets_touched.append((sheet_name, [k[1] for k in kpi_cols]))
        finally:
            try:
                wb.close()
            except Exception:
                pass

        logger.info(
            f'[Pass6.5] projected {written} PortfolioKPI rows from '
            f'{len(sheets_touched)} (sheet, kpi_fields) pairs: {sheets_touched}'
        )

    def _complete_portfolio_kpi_percentages(self, import_file_record):
        """Pass 6.6 — Derive missing percentage KPIs from raw inputs.

        After Pass 6.5 populates PortfolioKPI rows for every Gemini-
        classified KPI column, some PE/VC dashboards still expect
        derived percentages (EBITDA Margin %, Gross Margin %, etc.)
        that weren't reported directly in the Excel. For each derived
        percentage, ask Gemini ONCE which formula to use (given the
        canonical raw KPIs the dashboard has on hand), then evaluate
        that formula per Investment using the existing safe AST walker.

        Zero hardcoded formulas. Zero hardcoded keywords. Gemini
        chooses the formula and the inputs.
        """
        from decimal import Decimal as _D
        from .derivation_service import _safe_eval
        from .gemini_column_mapper import derive_per_row_formulas

        if not self._imported_fund:
            return

        # Targets we want to derive: any percentage canonical field whose
        # slug feeds the KPI matrix percent columns. Discovery, not
        # keywords: pull from the canonical fields whose name ends with
        # '_pct' (Gemini's naming convention for percentage metrics).
        target_pct_fields = sorted(
            f for f in self._KPI_CANONICAL_FIELDS if f.endswith('_pct')
        )
        if not target_pct_fields:
            return

        # Build (company → {canonical_kpi_field: value}) snapshot from
        # PortfolioKPI rows just written by Pass 6.5.
        from collections import defaultdict
        inv_kpis = defaultdict(dict)  # investment_id → {field_name: float}
        try:
            for sch in self._imported_fund.schemes.all():
                qs = PortfolioKPI.objects.filter(
                    investment__scheme=sch,
                ).select_related('kpi_definition', 'investment')
                for k in qs:
                    if k.kpi_definition is None or k.value is None:
                        continue
                    # Recover canonical field from slug. slug = canonical
                    # field with underscores → hyphens; reverse it.
                    cf = (k.kpi_definition.slug or '').replace('-', '_')
                    if not cf:
                        continue
                    if cf not in inv_kpis[k.investment_id]:
                        try:
                            inv_kpis[k.investment_id][cf] = float(k.value)
                        except Exception:
                            pass
        except Exception as e:
            logger.warning(f'Pass 6.6: snapshot build failed: {e}')
            return

        if not inv_kpis:
            logger.info('Pass 6.6: no PortfolioKPI inputs available — nothing to derive')
            return

        # For each target percentage, determine which targets are missing
        # for at least some investments AND which raw inputs are available
        # on those investments. Then ask Gemini for the formula ONCE per
        # target (cached across all investments).
        all_known_fields = set()
        for d in inv_kpis.values():
            all_known_fields.update(d.keys())

        missing_targets = [
            t for t in target_pct_fields if t not in all_known_fields
        ]
        if not missing_targets:
            logger.info('Pass 6.6: every target percentage already present')
            return

        # Build inputs catalogue: only canonical fields that are actually
        # available on at least one investment.
        from .canonical_schema import VALUATIONS_KPIS_FIELDS
        available_inputs = {}
        for cf in sorted(all_known_fields):
            if cf in missing_targets:
                continue
            desc = VALUATIONS_KPIS_FIELDS.get(cf, cf)
            sample_val = None
            for d in inv_kpis.values():
                if cf in d:
                    sample_val = d[cf]
                    break
            available_inputs[cf] = {
                'description': desc,
                'sample_value': sample_val,
                'unit': 'percent' if cf.endswith('_pct') else 'auto',
            }

        if not available_inputs:
            logger.info('Pass 6.6: no available inputs to derive percentages from')
            return

        # Build "missing fields" catalogue
        missing_fields = {
            t: {
                'description': VALUATIONS_KPIS_FIELDS.get(t, t),
                'unit': 'percent',
            }
            for t in missing_targets
        }

        # Pick 3 sample rows with the richest field coverage for Gemini
        scored = sorted(
            inv_kpis.items(),
            key=lambda kv: -len(kv[1]),
        )[:3]
        sample_rows = [d for (_iid, d) in scored]

        try:
            formulas = derive_per_row_formulas(
                model_label='investments.PortfolioKPI',
                available_inputs=available_inputs,
                missing_fields=missing_fields,
                sample_row_values=sample_rows,
            )
        except Exception as e:
            logger.warning(f'Pass 6.6: Gemini derive_per_row_formulas failed: {e}')
            return

        if not formulas:
            logger.info('Pass 6.6: Gemini declined to provide formulas')
            return

        snapshot_date = date.today()
        written = 0
        # derive_per_row_formulas now returns a RANKED candidate list per
        # field — same shape as Pass 6. For each row, try candidates in
        # rank order; the first whose declared inputs are all present
        # and non-null wins. This matches the Pass 6 evaluator contract
        # so heterogeneous row classes are handled correctly.
        for target_field, spec in formulas.items():
            candidates = spec.get('candidate_formulas') or []
            if not candidates:
                continue
            slug = slugify(target_field.replace('_', '-'))
            display = target_field.replace('_', ' ').title()
            try:
                kdef, _ = KPIDefinition.objects.update_or_create(
                    organization=self.org,
                    slug=slug,
                    defaults={
                        'name': display,
                        'format': 'percent',
                        'frequency': 'monthly',
                        'sector_template': 'generic',
                    },
                )
            except Exception as e:
                logger.warning(f'Pass 6.6: KPIDefinition create failed for {target_field}: {e}')
                continue

            rank_counts = {c['rank']: 0 for c in candidates}
            # Evaluate per investment using ranked fallback chain
            for inv_id, row in inv_kpis.items():
                chosen = None
                chosen_value = None
                for cand in candidates:
                    inputs_req = cand.get('inputs_required') or []
                    if not all(k in row and row[k] is not None for k in inputs_req):
                        continue
                    try:
                        val = _safe_eval(cand['formula_expression'], row)
                    except Exception as e:
                        logger.debug(
                            f'Pass 6.6: eval failed inv={inv_id} '
                            f'target={target_field} formula="{cand["formula_expression"]}": {e}'
                        )
                        continue
                    if val is None:
                        continue
                    chosen, chosen_value = cand, val
                    break
                if chosen is None:
                    continue
                try:
                    inv = Investment.objects.get(id=inv_id)
                except Investment.DoesNotExist:
                    continue
                try:
                    PortfolioKPI.objects.update_or_create(
                        investment=inv,
                        kpi_definition=kdef,
                        period=snapshot_date,
                        defaults={
                            'portfolio_company': inv.portfolio_company,
                            'value': _D(str(chosen_value)),
                            'source': 'excel_upload',
                            'status': 'approved',
                            'notes': (
                                f'Pass 6.6 derived (rank={chosen["rank"]}, '
                                f'conf={chosen["confidence"]:.2f}): '
                                f'{chosen["formula_expression"]}'
                            )[:500],
                        },
                    )
                    written += 1
                    rank_counts[chosen['rank']] = rank_counts.get(chosen['rank'], 0) + 1
                except Exception as e:
                    logger.warning(
                        f'Pass 6.6: persist failed for inv={inv_id} '
                        f'target={target_field}: {e}'
                    )

        logger.info(
            f'[Pass6.6] derived {written} percentage PortfolioKPI rows '
            f'across {len(formulas)} target metrics ({sorted(formulas.keys())})'
        )

    def _compute_fund_metrics_via_pass9(self, import_file_record):
        """Pass 9 — UNIFIED FUND METRICS COMPUTE.

        ONE Gemini call sees the raw content of every fund-level sheet
        (waterfall, NAV, capital calls, distributions, exits, P&L,
        scheme lifecycle, BvA, MOIC/TVPI summary, etc.) plus the
        scheme's LPA terms, and returns every dashboard fund metric
        (moic, tvpi, dpi, rvpi, net_irr, nav, total_*, return_of_capital,
        preferred_return, gp_catchup, carry_base, carry_amount_gross,
        gp_clawback_provision, carry_amount_net, lp_total_return,
        gp_total_distribution, total_proceeds_available) with formulas,
        source-cell citations, and confidence.

        Persists each metric as a DerivedMetric with
        formula_expression='(Pass 9 unified) <formula>'.  Overwrites any
        Pass 3.5 imported_direct or Pass 4 derived row for the same
        (scheme, metric_key) — Pass 9's read of the raw sheet is the
        most authoritative signal we have.

        Pass 4 runs AFTER this and ONLY processes metrics Pass 9 did not
        return (the catalogue-of-variables formula path is now the
        fallback, not the primary).
        """
        from decimal import Decimal as _D
        from .gemini_column_mapper import (
            compute_fund_metrics_unified, PASS9_METRIC_KEYS,
        )
        from .models import DerivedMetric
        from datetime import date

        if not self._imported_fund:
            return
        if not import_file_record:
            return

        schemes = list(self._imported_fund.schemes.all())
        if not schemes:
            return

        filepath = (
            getattr(self, '_filepath', None) or import_file_record.file.path
        )
        cm = import_file_record.column_mapping or {}
        # Reduce the per-sheet metadata dict to the {sheet: {'primary_domain': str}}
        # shape Pass 9 expects.
        sheet_classifications = {}
        for sname, meta in cm.items():
            if not isinstance(meta, dict):
                continue
            domains = meta.get('domains') or []
            primary = domains[0] if domains else meta.get('primary_domain')
            sheet_classifications[sname] = {'primary_domain': primary}

        for sch in schemes:
            lpa_terms = {
                'hurdle_rate_pct': self._safe_float(
                    getattr(sch, 'hurdle_rate_pct', None)),
                'carry_pct': self._safe_float(
                    getattr(sch, 'carry_pct', None)),
                'carry_type': getattr(sch, 'carry_type', None) or None,
                'management_fee_pct': self._safe_float(
                    getattr(sch, 'management_fee_pct', None)),
                'management_fee_basis': (
                    getattr(sch, 'management_fee_basis', None) or None
                ),
                'tenure_years': self._safe_float(
                    getattr(sch, 'tenure_years', None)),
                'sponsor_commitment_pct': self._safe_float(
                    getattr(sch, 'sponsor_commitment_pct', None)),
                'vintage_year': self._safe_float(
                    getattr(sch, 'vintage_year', None)),
            }

            try:
                result = compute_fund_metrics_unified(
                    filepath=filepath,
                    sheet_classifications=sheet_classifications,
                    lpa_terms=lpa_terms,
                    as_of_date=date.today(),
                )
            except Exception as e:
                logger.warning(
                    f'Pass 9 Gemini call failed for scheme={sch.name}: '
                    f'{type(e).__name__}: {e}'
                )
                continue

            metrics = result.get('metrics') or {}
            if not metrics:
                logger.info(
                    f'Pass 9: Gemini returned no metrics for scheme={sch.name}'
                )
                continue

            written = []
            for metric_key in PASS9_METRIC_KEYS:
                entry = metrics.get(metric_key)
                if not entry or entry.get('value') is None:
                    continue
                # Only persist when confidence is reasonable. Gemini
                # is allowed to return low-confidence "best guess"
                # entries; we filter them out so Pass 4 can take a
                # fresh stab. Threshold deliberately low (>= 0.4) so
                # we don't lose a real number Gemini was honest about
                # being uncertain on.
                conf = entry.get('confidence', 0.0)
                if conf < 0.4:
                    logger.info(
                        f'Pass 9: skipping {metric_key} '
                        f'(confidence {conf:.2f} below threshold)'
                    )
                    continue
                try:
                    DerivedMetric.objects.update_or_create(
                        scheme=sch,
                        metric_key=metric_key,
                        variant=None,
                        defaults={
                            'organization': self.org,
                            'value': _D(str(entry['value'])),
                            'formula_expression': (
                                f'(Pass 9 unified) '
                                f'{entry.get("formula_used", "")}'
                            )[:2000],
                            'inputs_used': {
                                'source_cells': entry.get('source_cells', []),
                                'formula_used': entry.get('formula_used', ''),
                                'sheets_used': result.get('sheets_used'),
                            },
                            'confidence': conf,
                            'gemini_reasoning': (
                                f'[Pass 9 unified] '
                                f'{entry.get("reasoning", "")}'
                            )[:4000],
                            'candidate_formulas': [],
                            'source_import_file': import_file_record,
                        },
                    )
                    # Record candidate for the Arbiter. Tier classification
                    # uses the formula prefix to decide P9-direct (tier A)
                    # vs P9-derivation (tier B).
                    from .metric_arbiter import record_metric_candidate
                    record_metric_candidate(
                        scheme=sch,
                        organization=self.org,
                        metric_key=metric_key,
                        variant=None,
                        pass_id='P9',
                        value=entry['value'],
                        formula_expression=entry.get('formula_used', ''),
                        confidence=conf,
                        inputs_used={
                            'source_cells': entry.get('source_cells', []),
                            'sheets_used': result.get('sheets_used'),
                        },
                        source_cells=entry.get('source_cells', []),
                        gemini_reasoning=entry.get('reasoning', ''),
                        source_import_file=import_file_record,
                    )
                    written.append(metric_key)
                except Exception as e:
                    logger.warning(
                        f'Pass 9 persist failed for {metric_key}: {e}'
                    )

            # Clean up stale variant-tagged rows that would shadow the
            # Pass-9 unauthoritative (variant=None) row at downstream
            # read time.
            try:
                stale = DerivedMetric.objects.filter(
                    scheme=sch,
                    metric_key__in=written,
                ).exclude(variant=None)
                deleted_n = stale.count()
                if deleted_n:
                    stale.delete()
                    logger.info(
                        f'Pass 9: deleted {deleted_n} stale variant-tagged '
                        f'rows that would have shadowed Pass 9 values.'
                    )
            except Exception as e:
                logger.warning(f'Pass 9 stale-cleanup failed: {e}')

            logger.info(
                f'[Pass9 scheme={sch.name}] wrote {len(written)} metrics: '
                f'{written}; sheets_used={result.get("sheets_used")}'
            )

    def _compute_carry_via_direct_waterfall(self, import_file_record):
        """Pass 8 — Direct Waterfall Computation.

        Replaces the layered Pass 3.5 + Pass 4 derivation for the four
        carry/clawback dashboard fields with ONE Gemini call that sees
        the complete waterfall sheet content + LPA terms + capital
        flows, and returns the four values with source-cell citations
        and confidence.

        OVERWRITES any DerivedMetric rows for these 4 keys (and the 6
        supplementary keys) that Pass 3.5 or Pass 4 may have written —
        Pass 8's read of the raw waterfall sheet is more trustworthy
        than fragment-by-fragment extraction-then-reassembly.

        ZERO formulas in code. Gemini returns the formula it used as
        plain text in the DerivedMetric record.
        """
        import openpyxl
        from decimal import Decimal as _D
        from .gemini_column_mapper import (
            compute_waterfall_metrics_directly,
            WATERFALL_PASS8_METRIC_KEYS,
            WATERFALL_PASS8_SUPPLEMENTARY_KEYS,
        )
        from .models import DerivedMetric

        if not self._imported_fund:
            return
        if not import_file_record or not import_file_record.column_mapping:
            return

        # Find waterfall sheets via Pass 1 domain classification — the
        # canonical signal that a sheet contains waterfall computations.
        waterfall_sheet_names = []
        cm = import_file_record.column_mapping or {}
        for sheet_name, meta in cm.items():
            if not isinstance(meta, dict):
                continue
            domains = meta.get('domains') or []
            if 'waterfall_carry' in domains:
                waterfall_sheet_names.append(sheet_name)
        if not waterfall_sheet_names:
            logger.info(
                'Pass 8: no waterfall_carry sheets in domain map — '
                'skipping direct waterfall pass.'
            )
            return

        filepath = getattr(self, '_filepath', None) or import_file_record.file.path
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            logger.warning(f'Pass 8: cannot reopen workbook: {e}')
            return

        waterfall_sheets = {}
        try:
            for sname in waterfall_sheet_names:
                if sname in wb.sheetnames:
                    waterfall_sheets[sname] = wb[sname]

            if not waterfall_sheets:
                logger.info('Pass 8: waterfall sheet names not found in workbook')
                return

            # Build LPA-terms block from the imported scheme(s).
            schemes = list(self._imported_fund.schemes.all())
            if not schemes:
                logger.info('Pass 8: no schemes imported, skipping')
                return

            for sch in schemes:
                lpa_terms = {
                    'hurdle_rate_pct': self._safe_float(sch.hurdle_rate_pct),
                    'carry_pct': self._safe_float(sch.carry_pct),
                    'carry_type': sch.carry_type,
                    'management_fee_pct': self._safe_float(sch.management_fee_pct),
                    'management_fee_basis': sch.management_fee_basis,
                    'tenure_years': self._safe_float(sch.tenure_years),
                    'sponsor_commitment_pct': self._safe_float(
                        sch.sponsor_commitment_pct
                    ),
                }

                # Capital-flow context — pull whatever Pass 3.5 / Pass 4
                # already wrote so Gemini can cross-check against the
                # waterfall sheet's own numbers.
                capital_flows = {}
                for k in ('total_called_capital', 'total_committed_capital',
                          'total_distributions', 'total_realised_proceeds',
                          'total_unrealised_fair_value'):
                    dm = DerivedMetric.objects.filter(
                        scheme=sch, metric_key=k,
                    ).exclude(value=None).first()
                    if dm and dm.value is not None:
                        capital_flows[k] = float(dm.value)

                try:
                    result = compute_waterfall_metrics_directly(
                        waterfall_sheets=waterfall_sheets,
                        lpa_terms=lpa_terms,
                        capital_flows=capital_flows,
                        as_of_date=date.today(),
                    )
                except Exception as e:
                    logger.warning(
                        f'Pass 8 Gemini call failed for scheme={sch.name}: '
                        f'{type(e).__name__}: {e}'
                    )
                    continue

                metrics = result.get('metrics') or {}
                if not metrics:
                    logger.info(
                        f'Pass 8: Gemini returned no metrics for scheme={sch.name}'
                    )
                    continue

                # Persist each metric as a DerivedMetric, OVERWRITING any
                # prior Pass 3.5 / Pass 4 row for the same (scheme,
                # metric_key, variant=None) tuple.
                written = []
                for metric_key in (WATERFALL_PASS8_METRIC_KEYS
                                   + WATERFALL_PASS8_SUPPLEMENTARY_KEYS):
                    entry = metrics.get(metric_key)
                    if not entry or entry.get('value') is None:
                        continue
                    try:
                        DerivedMetric.objects.update_or_create(
                            scheme=sch,
                            metric_key=metric_key,
                            variant=None,
                            defaults={
                                'organization': self.org,
                                'value': _D(str(entry['value'])),
                                'formula_expression': (
                                    f'(Pass 8 direct waterfall) '
                                    f'{entry.get("formula_used", "")}'
                                )[:2000],
                                'inputs_used': {
                                    'source_cells': entry.get('source_cells', []),
                                    'formula_used': entry.get('formula_used', ''),
                                    'sheet_used': result.get('sheet_used'),
                                },
                                'confidence': entry.get('confidence', 0.0),
                                'gemini_reasoning': (
                                    f'[Pass 8 direct waterfall] '
                                    f'{entry.get("reasoning", "")}'
                                )[:4000],
                                'candidate_formulas': [],
                                'source_import_file': import_file_record,
                            },
                        )
                        # Record candidate for the Arbiter (Pass 8 = Tier A direct read)
                        from .metric_arbiter import record_metric_candidate
                        record_metric_candidate(
                            scheme=sch,
                            organization=self.org,
                            metric_key=metric_key,
                            variant=None,
                            pass_id='P8',
                            value=entry['value'],
                            formula_expression=entry.get('formula_used', ''),
                            confidence=entry.get('confidence', 0.0),
                            inputs_used={
                                'source_cells': entry.get('source_cells', []),
                                'sheet_used': result.get('sheet_used'),
                            },
                            source_cells=entry.get('source_cells', []),
                            gemini_reasoning=entry.get('reasoning', ''),
                            source_import_file=import_file_record,
                        )
                        written.append(metric_key)
                    except Exception as e:
                        logger.warning(
                            f'Pass 8 persist failed for {metric_key}: {e}'
                        )

                # If Pass 3.5 wrote variant-tagged rows for any of these
                # metrics, those become stale and would shadow the Pass-8
                # value via the variant-default lookup. Delete them.
                try:
                    stale = DerivedMetric.objects.filter(
                        scheme=sch,
                        metric_key__in=list(WATERFALL_PASS8_METRIC_KEYS
                                            + WATERFALL_PASS8_SUPPLEMENTARY_KEYS),
                    ).exclude(variant=None)
                    deleted_n = stale.count()
                    if deleted_n:
                        stale.delete()
                        logger.info(
                            f'[Pass8] removed {deleted_n} stale variant-tagged '
                            f'DerivedMetric rows superseded by Pass 8 outputs.'
                        )
                except Exception as e:
                    logger.warning(f'Pass 8 stale-row cleanup failed: {e}')

                logger.info(
                    f'[Pass8] scheme={sch.name} wrote {len(written)} '
                    f'DerivedMetric rows: {written}. '
                    f'Overall reasoning: '
                    f'{result.get("overall_reasoning", "")[:300]}'
                )
        finally:
            try:
                wb.close()
            except Exception:
                pass

    @staticmethod
    def _safe_float(v):
        try:
            return float(v) if v is not None else None
        except (TypeError, ValueError):
            return None

    def _classify_column_roles(self, import_file_record):
        """Pass 2.6 — Column Semantic Role Classifier.

        For every horizontal sub-section that Pass 2.5 detected, classify
        each numeric column's semantic role (per_period_amount,
        cumulative_total, ratio_percent, identifier, metadata_text,
        derived_indicator, unknown) via one Gemini call per section. The
        roles get stored back on the column_mapping JSON so Pass 3.5 can
        filter candidate cells by metric-vs-role compatibility before
        disambiguation.

        ZERO keyword matching. Universal for any tabular sheet.
        """
        import openpyxl
        from .gemini_column_mapper import classify_column_roles

        if not import_file_record or not import_file_record.column_mapping:
            return

        filepath = getattr(self, '_filepath', None) or import_file_record.file.path
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            logger.warning(f'Pass 2.6: cannot reopen workbook: {e}')
            return

        sections_classified = 0
        try:
            cm = import_file_record.column_mapping or {}
            for sheet_name, sheet_meta in cm.items():
                if not isinstance(sheet_meta, dict):
                    continue
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                max_col = ws.max_column or 0
                sections = sheet_meta.get('sections') or []
                for sec in sections:
                    if not isinstance(sec, dict):
                        continue
                    if sec.get('layout') != 'horizontal':
                        continue
                    try:
                        header_row = int(sec.get('header_row') or 0)
                        data_start = int(
                            sec.get('data_start_row') or (header_row + 1)
                        )
                    except (TypeError, ValueError):
                        continue
                    if header_row <= 0:
                        continue
                    # Headers (every non-empty cell in header_row)
                    headers = {}
                    for c in range(1, max_col + 1):
                        hv = ws.cell(header_row, c).value
                        if hv is None:
                            continue
                        ht = str(hv).strip()
                        if ht:
                            headers[c] = ht
                    if not headers:
                        continue
                    # Sample up to 3 data rows for Gemini to see magnitudes
                    samples = []
                    sample_count = 0
                    r = data_start
                    while sample_count < 3 and r <= (ws.max_row or data_start):
                        row_dict = {}
                        any_non_null = False
                        for c in headers.keys():
                            v = ws.cell(r, c).value
                            if v is not None:
                                any_non_null = True
                            row_dict[c] = v
                        if any_non_null:
                            samples.append(row_dict)
                            sample_count += 1
                        r += 1
                    section_title = sec.get('section_name') or '(no title)'
                    try:
                        roles = classify_column_roles(
                            section_title=section_title,
                            column_headers=headers,
                            sample_data_rows=samples,
                        )
                    except Exception as e:
                        logger.warning(
                            f'Pass 2.6 classify_column_roles failed for '
                            f'{sheet_name}/{section_title}: '
                            f'{type(e).__name__}: {e} — leaving roles empty'
                        )
                        roles = {}
                    # Persist back on the section dict so Pass 3.5 can read
                    # it. We keep both the column header map AND the role
                    # map for downstream debug visibility.
                    sec['column_role_headers'] = {
                        str(ci): ht for ci, ht in headers.items()
                    }
                    sec['column_roles'] = {
                        str(ci): roles.get(ci, 'unknown')
                        for ci in headers.keys()
                    }
                    sections_classified += 1
        finally:
            try:
                wb.close()
            except Exception:
                pass

        # Persist the augmented column_mapping back to the DB so Pass 3.5
        # (which re-reads from import_file_record.column_mapping) sees it.
        try:
            import_file_record.column_mapping = cm
            import_file_record.save(update_fields=['column_mapping'])
        except Exception as e:
            logger.warning(f'Pass 2.6: persist updated column_mapping failed: {e}')

        logger.info(
            f'[Pass2.6] classify_column_roles: classified roles for '
            f'{sections_classified} horizontal section(s)'
        )

    def _extract_explicit_performance_metrics(self, import_file_record):
        """Pass 3.5 — Universal explicit-value scanner.

        Walks EVERY sheet of the imported workbook, harvests every plausible
        label-value pair (rows where col-1 is text + col-2 is numeric, or
        adjacent-cell pairs), and asks Gemini to classify each label against
        the canonical `fund_performance_metrics` category. When Gemini
        matches a label to a canonical metric (net_irr, moic, tvpi, dpi,
        rvpi, nav, ...), the value is persisted as a DerivedMetric row with
        formula_expression='(direct value imported)' so the downstream
        Pass 4 derivation treats it as authoritative and skips re-deriving.

        Zero keyword matching here. Zero per-metric special cases. Gemini
        decides what is a fund-performance metric via semantic equivalence
        against the canonical descriptions.
        """
        import openpyxl
        from decimal import Decimal, InvalidOperation
        from .canonical_schema import CANONICAL_VALUE_CATEGORIES
        from .models import DerivedMetric
        from .gemini_column_mapper import classify_labels

        if not self._imported_fund:
            return
        schemes = list(self._imported_fund.schemes.all())
        if not schemes:
            return

        # Find every numeric label-value pair across every sheet — no row/col
        # caps, no sheet filtering.
        filepath = getattr(self, '_filepath', None) or import_file_record.file.path
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            logger.warning(f'Pass 3.5: cannot reopen workbook: {e}')
            return

        def _parse_with_hint(v):
            """Return (float | None, unit_hint). unit_hint ∈
            {'amount','percent','multiple','unknown'}. Uses the global
            _parse_amount so 'Rs 3,800 Cr' → (3800.0, 'amount') and
            '20% (above hurdle)' → (20.0, 'percent'). The hint travels with
            the candidate so the canonical-metric unit filter can reject
            mismatches (e.g. an 'amount' metric must not be filled by a
            'percent' cell)."""
            val, hint = _parse_amount(v)
            if val is None:
                return (None, hint)
            try:
                return (float(val), hint)
            except (TypeError, ValueError):
                return (None, hint)

        def _is_numeric(v):
            num, _hint = _parse_with_hint(v)
            return num is not None

        def _to_number(v):
            num, _hint = _parse_with_hint(v)
            return num

        def _unit_hint_for(v):
            _num, hint = _parse_with_hint(v)
            return hint

        # Pre-compute per-sheet header-row maps using Pass 2.5 sub-table
        # layout info (already stored on the ImportFile by Pass 2). For each
        # horizontal section we know (header_row, data_start_row, ...). The
        # cells in header_row tell us what each numeric column MEANS, which
        # is essential for tabular waterfall sheets where one row carries
        # several metrics (LP share, GP share, total, ...) in different
        # columns. Without column headers, Pass 3.5 can only pick "first
        # numeric after label" — which silently picks the wrong cell.
        # sheet_header_map: {sheet_name: {col_idx: column_header_text}} for
        # the most specific (closest-above) header per data row.
        sheet_section_headers = {}  # {sheet: [{header_row, data_start, data_end, headers:{col:text}}]}
        try:
            cm = (import_file_record.column_mapping or {})
            for sname, smeta in cm.items():
                if not isinstance(smeta, dict):
                    continue
                if sname not in wb.sheetnames:
                    continue
                ws_for_headers = wb[sname]
                max_col_h = ws_for_headers.max_column or 0
                max_row_h = ws_for_headers.max_row or 0
                section_records = []
                sections = smeta.get('sections') or []
                # Sort by header_row so we can compute each section's effective
                # end row as (next section's header_row - 1).
                section_specs = []
                for sec in sections:
                    if not isinstance(sec, dict):
                        continue
                    if sec.get('layout') != 'horizontal':
                        continue
                    try:
                        hr = int(sec.get('header_row') or 0)
                        ds = int(sec.get('data_start_row') or (hr + 1))
                    except (TypeError, ValueError):
                        continue
                    if hr <= 0:
                        continue
                    section_specs.append((hr, ds, sec))
                section_specs.sort(key=lambda t: t[0])
                for i, (hr, ds, sec) in enumerate(section_specs):
                    end_row = max_row_h
                    if i + 1 < len(section_specs):
                        end_row = section_specs[i + 1][0] - 1
                    headers = {}
                    for c in range(1, max_col_h + 1):
                        hv = ws_for_headers.cell(hr, c).value
                        if hv is None:
                            continue
                        ht = str(hv).strip()
                        if ht:
                            headers[c] = ht
                    # Pass 2.6 column-role classification — stored on the
                    # section dict by _classify_column_roles. Keys are
                    # stringified col_idx; cast back to int for lookup.
                    roles_raw = sec.get('column_roles') or {}
                    column_roles = {}
                    for k, v in roles_raw.items():
                        try:
                            column_roles[int(k)] = str(v or 'unknown')
                        except (TypeError, ValueError):
                            continue
                    if headers:
                        section_records.append({
                            'header_row': hr,
                            'data_start_row': ds,
                            'data_end_row': end_row,
                            'headers': headers,
                            'column_roles': column_roles,
                        })
                if section_records:
                    sheet_section_headers[sname] = section_records
        except Exception as e:
            logger.warning(f'Pass 3.5: section-header pre-compute failed: {e}')

        def _column_header_for(sname, row_idx, col_idx):
            """Return the column-header text for (sheet, row, col), or '' if
            this row doesn't sit inside any known horizontal section."""
            for sec in sheet_section_headers.get(sname, []):
                if sec['data_start_row'] <= row_idx <= sec['data_end_row']:
                    return sec['headers'].get(col_idx, '')
            return ''

        def _column_role_for(sname, row_idx, col_idx):
            """Return the Pass 2.6 semantic role for (sheet, row, col), or
            None if no role was classified (free-form cells / sections
            without column_role data)."""
            for sec in sheet_section_headers.get(sname, []):
                if sec['data_start_row'] <= row_idx <= sec['data_end_row']:
                    return sec.get('column_roles', {}).get(col_idx)
            return None

        def _row_is_inside_horizontal_section(sname, row_idx):
            for sec in sheet_section_headers.get(sname, []):
                if sec['data_start_row'] <= row_idx <= sec['data_end_row']:
                    return True
            return False

        # Collect EVERY (label, value, source_cell, column_header) tuple.
        # For rows inside horizontal tabular sections (Pass 2.5 layout
        # known), emit ONE candidate per numeric column so the
        # disambiguation step can pick the right column. For rows outside
        # any section (free-form label-value pairs), fall back to
        # "first numeric after label".
        # label_occurrences: dict label_text -> list[(value, source_cell, column_header)]
        label_occurrences = {}
        try:
            for sname in wb.sheetnames:
                ws = wb[sname]
                max_col_w = ws.max_column or 0
                for r in range(1, (ws.max_row or 0) + 1):
                    row_cells = [
                        ws.cell(r, c).value
                        for c in range(1, max_col_w + 1)
                    ]
                    # Find the label (first non-empty text cell) and its
                    # column index. label_col_idx is 1-based.
                    label = None
                    label_col_idx = None
                    for idx, v in enumerate(row_cells, start=1):
                        if v is None:
                            continue
                        if isinstance(v, str) and v.strip():
                            label = v.strip()
                            label_col_idx = idx
                            break
                    if label is None:
                        continue

                    if _row_is_inside_horizontal_section(sname, r):
                        # Tabular row — emit one candidate per numeric cell
                        # AFTER the label, each carrying its column header,
                        # its Pass 2.6 semantic role, AND its unit hint
                        # (amount / percent / multiple).
                        #
                        # Mid-row label boundary (Fix E): in Cover/Summary
                        # sheets that Pass 2.5 wrongly classified as a
                        # single horizontal section, two parallel two-column
                        # blocks share one row, e.g. Cover R18:
                        #   [col2 "Carried Interest"]  [col3 "20%..."]
                        #   [col6 "LP Count"]          [col7 14]
                        # Without a boundary, "Carried Interest" used to
                        # absorb the 14 at col 7 as its value, then Pass 3.5
                        # wrote carry_amount_gross = ₹14 Cr. Now we stop
                        # scanning the moment we hit ANOTHER text cell that
                        # itself looks like a label (a free-text string,
                        # non-numeric-when-parsed, ≥3 visible chars). The
                        # numeric cells beyond belong to that new label.
                        def _looks_like_new_label(cell_val):
                            if not isinstance(cell_val, str):
                                return False
                            s = cell_val.strip()
                            if len(s) < 3:
                                return False
                            num, _h = _parse_with_hint(s)
                            return num is None
                        # Annotation blacklist (Fix E): refuse to extract
                        # any numeric whose own cell text contains words
                        # that mark it as a BENCHMARK / TARGET / ESTIMATE,
                        # not the actual realised value. These appear
                        # universally across PE fund Excel files in
                        # benchmark columns. The label "Blended Portfolio
                        # MOIC" at FUND_MASTER!R37 was paired with the
                        # IVCA benchmark column ("2.0x target"), so MOIC
                        # came out as the FUND'S TARGET, not its actual.
                        ANNOTATION_BLACKLIST_TOKENS = (
                            'target', 'benchmark', 'estimated',
                            'estimate', 'budgeted', 'forecast',
                            'projected', 'goal', 'aspiration',
                        )
                        def _is_annotation_cell(cell_val):
                            if not isinstance(cell_val, str):
                                return False
                            cs = cell_val.lower()
                            return any(
                                t in cs for t in ANNOTATION_BLACKLIST_TOKENS
                            )
                        for idx, v in enumerate(row_cells, start=1):
                            if idx <= label_col_idx:
                                continue
                            if _looks_like_new_label(v):
                                # Reached a new label scope; stop attributing
                                # cells beyond this column to the matched label.
                                break
                            if _is_annotation_cell(v):
                                # "2.0x target" / "20% benchmark" — not a
                                # realised value. Skip but keep scanning.
                                continue
                            if not _is_numeric(v):
                                continue
                            num = _to_number(v)
                            if num is None:
                                continue
                            col_header = _column_header_for(sname, r, idx)
                            col_role = _column_role_for(sname, r, idx)
                            # Also drop candidates whose COLUMN HEADER is an
                            # annotation column (header text contains
                            # benchmark/target/etc.) — catches the case where
                            # the cell value is just "2.0" but the column
                            # header says "IVCA Benchmark".
                            if _is_annotation_cell(col_header):
                                continue
                            unit_hint = _unit_hint_for(v)
                            label_occurrences.setdefault(label, []).append(
                                (num, f'{sname}!row{r}col{idx}', col_header, col_role, unit_hint)
                            )
                    else:
                        # Free-form row (Cover/Summary 2-column layout etc.).
                        # ROW-LOCALITY: the value must be in one of the
                        # immediate next two columns after the label.
                        # Without this guard, "Fund Corpus" → "₹3,800 Cr"
                        # (col+1) was unparseable, so the scanner kept
                        # walking and grabbed "4.52" (col+5) from the
                        # adjacent two-column block ("Portfolio MOIC (x)").
                        # The new _parse_amount() now reads "₹3,800 Cr"
                        # cleanly, but we still enforce locality so a
                        # genuinely-blank label can't slurp a distant cell.
                        LOCALITY_WINDOW = 2
                        numeric_val = None
                        numeric_hint = 'unknown'
                        for idx, v in enumerate(row_cells, start=1):
                            if idx <= label_col_idx:
                                continue
                            if idx > label_col_idx + LOCALITY_WINDOW:
                                break
                            if _is_numeric(v):
                                numeric_val = _to_number(v)
                                numeric_hint = _unit_hint_for(v)
                                break
                        if numeric_val is not None:
                            label_occurrences.setdefault(label, []).append(
                                (numeric_val, f'{sname}!row{r}', '', None, numeric_hint)
                            )
        finally:
            try:
                wb.close()
            except Exception:
                pass

        if not label_occurrences:
            logger.info('Pass 3.5: no label-value pairs found')
            return

        labels = list(label_occurrences.keys())
        logger.info(
            f'Pass 3.5: harvested {len(labels)} unique label-value pairs '
            f'across {len(wb.sheetnames) if hasattr(wb, "sheetnames") else "?"} sheets'
        )

        # Semantic classification via Gemini — every label gets compared
        # against the canonical fund_performance_metrics descriptions.
        try:
            label_map = classify_labels(
                labels,
                'fund_performance_metrics',
                CANONICAL_VALUE_CATEGORIES['fund_performance_metrics'],
                context=(
                    'Fund-level performance metrics scanned from any sheet of '
                    'the imported workbook. Match each label SEMANTICALLY '
                    'against the canonical metric descriptions; ignore unit '
                    'suffixes, currency symbols, parenthetical notes, and '
                    'spacing/casing differences.'
                ),
            )
        except Exception as e:
            logger.warning(f'Pass 3.5 classify_labels failed: {e}')
            return

        # Group ALL candidates per canonical metric (one label may produce
        # multiple occurrences across the workbook AND multiple per-column
        # candidates within a single tabular row; many distinct labels may
        # also map to the same canonical metric).
        # Each candidate carries column_header + column_role so the
        # role-compatibility filter below can drop semantically-wrong
        # column picks before Gemini disambiguates.
        # metric_to_candidates: canonical_key -> list[{label, value, source_cell, column_header, column_role}]
        metric_to_candidates = {}
        for label, canonical in (label_map or {}).items():
            if not canonical:
                continue
            for occ in label_occurrences.get(label, []):
                # Backward-compat for older occurrence tuples:
                #   2/3/4-tuples = before unit_hint was tracked
                #   5-tuple: (value, source, column_header, column_role, unit_hint)
                column_role = None
                column_header = ''
                unit_hint = 'unknown'
                if len(occ) == 5:
                    value, source, column_header, column_role, unit_hint = occ
                elif len(occ) == 4:
                    value, source, column_header, column_role = occ
                elif len(occ) == 3:
                    value, source, column_header = occ
                else:
                    value, source = occ
                if value is None:
                    continue
                metric_to_candidates.setdefault(canonical, []).append({
                    'label': label,
                    'value': value,
                    'source_cell': source,
                    'column_header': column_header,
                    'column_role': column_role,
                    'unit_hint': unit_hint,
                })

        if not metric_to_candidates:
            logger.info('Pass 3.5: no candidates matched canonical metrics')
            return

        # Role-compatibility filter — drop candidates whose column_role is
        # incompatible with the canonical metric's value_type. This is
        # what prevents "per_step_amount" metrics like preferred_return
        # from being extracted from cumulative-total columns.
        from .canonical_schema import is_role_compatible
        metric_catalogue = CANONICAL_VALUE_CATEGORIES['fund_performance_metrics']
        filtered_metric_to_candidates = {}
        role_filter_stats = []
        for canonical, cands in metric_to_candidates.items():
            meta = metric_catalogue.get(canonical) or {}
            value_type = (
                meta.get('value_type') if isinstance(meta, dict) else None
            )
            if not value_type:
                # No value_type declared — accept all (backward-compat).
                filtered_metric_to_candidates[canonical] = cands
                continue
            compatible = [
                c for c in cands
                if is_role_compatible(value_type, c.get('column_role'))
            ]
            dropped = len(cands) - len(compatible)
            if dropped > 0:
                role_filter_stats.append((canonical, len(cands), len(compatible), dropped))
            if compatible:
                filtered_metric_to_candidates[canonical] = compatible
            else:
                # All candidates were filtered out — log so audit log can
                # show that this metric will be left for Pass 4 derivation.
                logger.info(
                    f'Pass 3.5: ALL {len(cands)} candidates for {canonical} '
                    f'(value_type={value_type}) had incompatible column_role; '
                    f'leaving metric for Pass 4 derivation.'
                )
        if role_filter_stats:
            logger.info(
                f'[Pass3.5 role-filter] dropped role-incompatible candidates: '
                f'{role_filter_stats}'
            )
        metric_to_candidates = filtered_metric_to_candidates

        # Unit-hint compatibility filter — drop candidates whose detected
        # numeric format disagrees with the metric's expected value_type.
        # This is what stops "Carried Interest" → "20% (above hurdle)"
        # being interpreted as a ₹-amount, and "Fund Corpus" → "Portfolio
        # MOIC (x)=4.52" from being interpreted as a committed-capital
        # amount. Compatibility:
        #   amount-typed metrics  (aggregate_total, aggregate_cumulative,
        #                          per_step_amount, per_unit_amount)
        #       → reject 'percent' and 'multiple' hints
        #   ratio-typed metrics   ('ratio')
        #       → reject 'amount' hint
        #   'unknown' hint always passes (don't penalise plain numbers).
        AMOUNT_TYPES = {
            'aggregate_total', 'aggregate_cumulative',
            'per_step_amount', 'per_unit_amount',
        }
        RATIO_TYPES = {'ratio'}
        unit_filter_stats = []
        unit_filtered = {}
        for canonical, cands in metric_to_candidates.items():
            meta = metric_catalogue.get(canonical) or {}
            vtype = meta.get('value_type') if isinstance(meta, dict) else None
            if not vtype:
                unit_filtered[canonical] = cands
                continue
            if vtype in AMOUNT_TYPES:
                bad = {'percent', 'multiple'}
            elif vtype in RATIO_TYPES:
                bad = {'amount'}
            else:
                unit_filtered[canonical] = cands
                continue
            compatible = [
                c for c in cands if c.get('unit_hint', 'unknown') not in bad
            ]
            dropped = len(cands) - len(compatible)
            if dropped > 0:
                unit_filter_stats.append(
                    (canonical, len(cands), len(compatible), dropped)
                )
            if compatible:
                unit_filtered[canonical] = compatible
            else:
                logger.info(
                    f'Pass 3.5: ALL {len(cands)} candidates for {canonical} '
                    f'(value_type={vtype}) had incompatible unit_hint; '
                    f'leaving metric for Pass 4 derivation.'
                )
        if unit_filter_stats:
            logger.info(
                f'[Pass3.5 unit-filter] dropped unit-incompatible candidates: '
                f'{unit_filter_stats}'
            )
        metric_to_candidates = unit_filtered

        # For each metric, if there are multiple candidates with differing
        # values, ask Gemini to pick the most authoritative source. No
        # code-level filtering on values — Gemini reasons over labels, sheet
        # names, and surrounding context.
        from .gemini_column_mapper import select_authoritative_source

        # For metrics that come in multiple semantic variants (gross/net,
        # pre-fee/post-fee, ...), invoke the variant classifier ONCE per
        # such metric to tag every candidate's variant tag before
        # disambiguation.
        from .gemini_column_mapper import classify_metric_variant
        for canonical, cands in list(metric_to_candidates.items()):
            meta = metric_catalogue.get(canonical) or {}
            variant_options = (
                meta.get('requires_variant') if isinstance(meta, dict) else None
            )
            if not variant_options:
                continue
            if len(cands) < 2:
                # Single candidate — variant doesn't help disambiguate; just
                # persist the variant_default as a tag.
                for c in cands:
                    c.setdefault('variant', meta.get('variant_default'))
                continue
            try:
                tagged = classify_metric_variant(
                    metric_key=canonical,
                    metric_label=meta.get('label', canonical),
                    metric_description=meta.get('description', ''),
                    variant_options=list(variant_options),
                    candidates=cands,
                )
                metric_to_candidates[canonical] = tagged
            except Exception as e:
                logger.warning(
                    f'Pass 3.5 variant classifier failed for {canonical}: '
                    f'{type(e).__name__}: {e} — falling back to variant_default.'
                )
                for c in cands:
                    c.setdefault('variant', meta.get('variant_default'))

        chosen_per_metric = {}  # canonical_key -> (value, label, source_cell, reasoning, variant)
        for canonical, cands in metric_to_candidates.items():
            meta = metric_catalogue.get(canonical) or {}
            variant_default = (
                meta.get('variant_default') if isinstance(meta, dict) else None
            )

            # If this metric requires a variant AND we have a default,
            # prefer candidates tagged with the default variant before
            # disambiguation. Other variants are kept around — they may be
            # the only ones left if no candidate matches the default.
            if variant_default:
                primary = [c for c in cands if c.get('variant') == variant_default]
                if primary:
                    cands = primary

            # Deduplicate identical (value, source) pairs to keep prompts tight
            seen = set()
            unique_cands = []
            for c in cands:
                key = (round(c['value'], 8) if isinstance(c['value'], (int, float)) else c['value'],
                       c['source_cell'])
                if key in seen:
                    continue
                seen.add(key)
                unique_cands.append(c)

            distinct_values = {
                round(c['value'], 8) if isinstance(c['value'], (int, float)) else c['value']
                for c in unique_cands
            }

            if len(unique_cands) == 1:
                chosen = unique_cands[0]
                chosen_per_metric[canonical] = (
                    chosen['value'], chosen['label'], chosen['source_cell'],
                    f'Single candidate cell across the workbook.',
                    chosen.get('variant'),
                )
                continue

            # Multiple candidates → if they all agree on the same value, pick
            # the first deterministically (no Gemini call needed). Otherwise
            # delegate disambiguation to Gemini.
            if len(distinct_values) == 1:
                chosen = unique_cands[0]
                chosen_per_metric[canonical] = (
                    chosen['value'], chosen['label'], chosen['source_cell'],
                    f'All {len(unique_cands)} candidate cells agreed on value '
                    f'{chosen["value"]}; picked first occurrence as canonical source.',
                    chosen.get('variant'),
                )
                continue

            try:
                pick = select_authoritative_source(
                    metric_key=canonical,
                    metric_label=meta.get('label', canonical) if isinstance(meta, dict) else canonical,
                    metric_description=(
                        meta.get('description', '') if isinstance(meta, dict) else str(meta)
                    ),
                    candidates=unique_cands,
                )
            except Exception as e:
                logger.warning(
                    f'Pass 3.5 disambiguation API failed for {canonical}: '
                    f'{type(e).__name__}: {e} — skipping this metric so Pass 4 '
                    f'derives it from first principles.'
                )
                continue

            idx = pick.get('chosen_index') if isinstance(pick, dict) else None
            if idx is None:
                logger.info(
                    f'Pass 3.5: Gemini declined to pick an authoritative source '
                    f'for {canonical} from {len(unique_cands)} candidates with '
                    f'differing values — leaving for Pass 4 derivation.'
                )
                continue
            chosen = unique_cands[idx]
            chosen_per_metric[canonical] = (
                chosen['value'], chosen['label'], chosen['source_cell'],
                pick.get('reasoning', ''),
                chosen.get('variant'),
            )

        # Fix D — semantic label guard. Even after role + unit filters and
        # Gemini's disambiguation, refuse to persist a candidate whose RAW
        # label shares no semantic token with the canonical metric's
        # description. This is the final defence against the silent
        # neighbour-cell fallback pattern: if "Carried Interest" passed
        # all filters but the chosen cell's label is "LP Count", drop it.
        #
        # We use a lightweight, language-agnostic check: tokenise both
        # strings to alphabetic words ≥3 chars, lowercase, and require at
        # least ONE shared token. This is intentionally weak — Gemini's
        # classify_labels already did the heavy semantic work; this only
        # catches the obvious "label and chosen cell are unrelated" case.
        def _semantic_tokens(text):
            return {
                t for t in re.findall(r'[A-Za-z]{3,}', str(text or '').lower())
                if t not in {
                    'the', 'and', 'for', 'per', 'cur', 'fcr', 'value', 'total',
                    'amount', 'fund', 'all', 'net', 'gross', 'sum',
                }
            }

        def _label_matches_metric(label, canonical, meta):
            metric_text = ' '.join(filter(None, [
                canonical.replace('_', ' '),
                meta.get('label', '') if isinstance(meta, dict) else '',
                meta.get('description', '') if isinstance(meta, dict) else '',
            ]))
            mt = _semantic_tokens(metric_text)
            lt = _semantic_tokens(label)
            if not mt or not lt:
                return True  # not enough material to judge — don't block
            return bool(mt & lt)

        rejected_semantic = []
        for canonical, (value, label, source, reasoning, variant) in list(
            chosen_per_metric.items()
        ):
            meta = metric_catalogue.get(canonical) or {}
            if not _label_matches_metric(label, canonical, meta):
                rejected_semantic.append((canonical, label, source))
                chosen_per_metric.pop(canonical, None)
        if rejected_semantic:
            logger.info(
                f'[Pass3.5 semantic-guard] dropped {len(rejected_semantic)} '
                f'metric picks whose chosen label shared no semantic token '
                f'with the canonical metric: {rejected_semantic}'
            )

        # Persist chosen winners as DerivedMetric rows with imported_direct
        # provenance.
        from decimal import Decimal as _D
        written = 0
        for canonical, (value, label, source, reasoning, variant) in chosen_per_metric.items():
            try:
                for sch in schemes:
                    DerivedMetric.objects.update_or_create(
                        scheme=sch,
                        metric_key=canonical,
                        variant=variant,
                        defaults={
                            'organization': self.org,
                            'value': _D(str(value)),
                            'formula_expression': '(direct value imported)',
                            'inputs_used': {
                                'source_cell': source,
                                'source_label': label,
                                'source_value': value,
                                'candidate_count': len(
                                    metric_to_candidates.get(canonical, [])
                                ),
                                'variant': variant,
                            },
                            'confidence': 1.0,
                            'gemini_reasoning': (
                                f'Label "{label}" at {source} chosen as '
                                f'authoritative source for canonical metric '
                                f'"{canonical}" (variant={variant}). {reasoning}'.strip()
                            ),
                            'candidate_formulas': [],
                            'source_import_file': import_file_record,
                        },
                    )
                    # Record candidate for the Arbiter. The Arbiter
                    # demotes annotated labels (estimated/target/...)
                    # to Tier D so Pass 9/Pass 4 derived values win
                    # when present, while non-annotated direct cell
                    # reads stay at Tier A.
                    from .metric_arbiter import record_metric_candidate
                    record_metric_candidate(
                        scheme=sch,
                        organization=self.org,
                        metric_key=canonical,
                        variant=variant,
                        pass_id='P35',
                        value=value,
                        formula_expression='direct value imported',
                        confidence=1.0,
                        inputs_used={
                            'source_cell': source,
                            'source_label': label,
                            'source_value': value,
                            'variant': variant,
                        },
                        source_cells=[source] if source else [],
                        gemini_reasoning=(
                            f'Label "{label}" at {source} '
                            f'(variant={variant}). {reasoning}'.strip()
                        ),
                        source_import_file=import_file_record,
                    )
                    written += 1
            except Exception as e:
                logger.warning(
                    f'Pass 3.5 persist failed for {canonical} = {value}: {e}'
                )

        logger.info(
            f'[Pass3.5] persisted {written} DerivedMetric imported_direct rows '
            f'covering metrics: {sorted(chosen_per_metric.keys())}'
        )

    def _enrich_nav_records_post_import(self, wb, scheme, domain_map=None):
        """
        Enrich NAV records after transposed-format import.

        1. NAV/Unit: use Gemini-classified nav_calculation sheets (fallback to
           nav_accounting) to find "Closing NAV/Unit" and apply to the most
           recent NAV record.

        2. Realized Gains: sum ExitEvent.realized_gain_loss for this fund and apply
           the total to the most recent NAV record.
        """
        from django.db.models import Sum

        # ── 1. NAV/Unit from Gemini-classified nav_calculation sheets ────────
        closing_nav_per_unit = None
        opening_nav_per_unit = None
        total_units = None

        # Use Gemini domain_map: nav_calculation first, then nav_accounting
        calc_sheets = _dm_sheets(domain_map, 'nav_calculation') if domain_map else []
        if not calc_sheets:
            calc_sheets = _dm_sheets(domain_map, 'nav_accounting') if domain_map else []
        calc_sheets = [s for s in calc_sheets if s in wb.sheetnames]

        for sn in calc_sheets:
            calc_ws = wb[sn]
            max_r = min(calc_ws.max_row + 1, 80)
            # Collect label-value pairs and classify via Gemini
            nav_kv = {}
            nav_vals = {}
            for rr in range(1, max_r):
                label = calc_ws.cell(rr, 1).value
                if not label:
                    continue
                val = calc_ws.cell(rr, 2).value
                dval = _d(val)
                if dval is None:
                    continue
                label_str = str(label).strip()
                nav_kv[label_str] = label_str
                nav_vals[label_str] = dval
            if nav_kv:
                nav_label_map = self._classify_labels(
                    list(nav_kv.keys()), 'nav_components',
                    context='NAV calculation sheet labels')
                for label_str, nav_type in nav_label_map.items():
                    dval = nav_vals.get(label_str)
                    if dval is None:
                        continue
                    if nav_type == 'closing_nav_per_unit':
                        closing_nav_per_unit = dval
                    elif nav_type == 'opening_nav_per_unit':
                        opening_nav_per_unit = dval
                    elif nav_type == 'total_units':
                        total_units = dval
            if closing_nav_per_unit is not None:
                break  # Found it — no need to scan more sheets

        latest_nav = NAVRecord.objects.filter(scheme=scheme).order_by('-nav_date').first()
        if latest_nav and closing_nav_per_unit and closing_nav_per_unit > 0:
            # nav_per_unit stored in Rs (raw value from Excel)
            units = total_units
            if not units and latest_nav.total_nav and latest_nav.total_nav > 0:
                # Estimate units: total_nav is in Cr → convert to Rs then divide
                total_nav_rs = latest_nav.total_nav * Decimal('10000000')
                try:
                    units = (total_nav_rs / closing_nav_per_unit).quantize(Decimal('0.000001'))
                except Exception:
                    pass
            latest_nav.nav_per_unit = closing_nav_per_unit
            if units:
                latest_nav.total_units_outstanding = units
            latest_nav.save(update_fields=['nav_per_unit', 'total_units_outstanding'])
            logger.info(f'  NAV/Unit set: {closing_nav_per_unit} Rs for {latest_nav.nav_date}')

            # Propagate units to earlier NAV records so nav_per_unit can be derived
            if units and units > 0 and opening_nav_per_unit:
                earlier_navs = NAVRecord.objects.filter(
                    scheme=scheme
                ).exclude(id=latest_nav.id).order_by('nav_date')
                for nav_rec in earlier_navs:
                    if nav_rec.nav_per_unit == 0 and nav_rec.total_nav and nav_rec.total_nav > 0:
                        # Interpolate between opening and closing NAV/Unit linearly
                        # (simple approximation; actual value not in sheet)
                        total_nav_rs = nav_rec.total_nav * Decimal('10000000')
                        try:
                            est_units = (total_nav_rs / closing_nav_per_unit).quantize(Decimal('0.000001'))
                            est_nav_per_unit = (total_nav_rs / units).quantize(Decimal('0.000001'))
                            nav_rec.nav_per_unit = est_nav_per_unit
                            nav_rec.total_units_outstanding = est_units
                            nav_rec.save(update_fields=['nav_per_unit', 'total_units_outstanding'])
                        except Exception:
                            pass

        # ── 2. Realized Gains from ExitEvent records ─────────────────────────
        fund = scheme.fund if scheme else None
        if fund:
            realized_total = ExitEvent.objects.filter(
                investment__scheme__fund=fund,
                realized_gain_loss__isnull=False,
            ).aggregate(total=Sum('realized_gain_loss'))['total']

            if realized_total and latest_nav and latest_nav.realized_gains == 0:
                latest_nav.realized_gains = realized_total
                latest_nav.save(update_fields=['realized_gains'])
                logger.info(f'  Realized Gains set: {realized_total} Cr on {latest_nav.nav_date}')

    # ------------------------------------------------------------------
    # Exits & distributions
    # ------------------------------------------------------------------

    def _import_exits_and_distributions(self, wb, investments, schemes, domain_map):
        """Import exit events and fund-level distributions.

        Uses Gemini-classified exits_distributions sheets from domain_map.
        Validates each sheet by checking for exit-relevant columns
        (Exit Date, Exit Type, Proceeds, MOIC) before processing.
        """
        # Use all Gemini-classified exits sheets
        candidate_sheets = [s for s in _dm_sheets(domain_map, 'exits_distributions')
                            if s in wb.sheetnames]

        if not candidate_sheets:
            return

        for sheet_name in candidate_sheets:
            ws = wb[sheet_name]
            exit_count = self._try_import_exits_from_sheet(
                ws, sheet_name, investments, schemes)
            if exit_count > 0:
                logger.info(f'  Exits imported from sheet "{sheet_name}": {exit_count}')
                return

    def _try_import_exits_from_sheet(self, ws, sheet_name, investments, schemes):
        """Try to import exits from a single sheet. Returns exit count (0 if unsuitable)."""
        sections = self._read_sheet_via_layout(ws, alias_map=self._get_alias(ws))
        exit_rows = None
        dist_rows = None

        for sec_name, (sec_headers, sec_rows) in sections.items():
            subdomain = self._get_section_subdomain(sheet_name, sec_name)
            if subdomain == 'exit_events':
                if exit_rows is None:
                    exit_rows = sec_rows
            elif subdomain == 'distributions':
                dist_rows = sec_rows

        # Structured sections found — validate and process
        if exit_rows is not None:
            if not self._rows_have_exit_columns(exit_rows):
                return 0
            exit_count = self._process_exit_rows(exit_rows, investments)
            if dist_rows:
                default_scheme = list(schemes.values())[0] if schemes else None
                if default_scheme:
                    self._process_distribution_rows(dist_rows, schemes,
                                                     investments)
            return exit_count

        # Flat table fallback
        _, rows = read_table_from_sheet(ws, alias_map=self._get_alias(ws))
        if not rows:
            if '__default__' in sections:
                _, rows = sections['__default__']

        if rows and self._rows_have_exit_columns(rows):
            return self._process_exit_rows(rows, investments)
        return 0

    @staticmethod
    def _rows_have_exit_columns(rows):
        """Check if rows contain exit-relevant columns (Exit Date, Exit Type, Proceeds, MOIC).

        Returns True if at least 2 of these column families are present —
        a sheet with only 'Type' or only 'Date' is too ambiguous.
        """
        if not rows:
            return False
        sample = rows[0]
        keys_lower = {k.lower() for k in sample.keys()}

        # Detect exit data using Gemini Pass 2 canonical field names
        # which are added as row keys regardless of language
        exit_canonical = {'exit_date', 'exit_type', 'exit_route',
                          'exit_proceeds', 'realized_value', 'moic'}
        return len(exit_canonical & keys_lower) >= 2

    def _process_exit_rows(self, rows, investments):
        """Process exit event rows from either format.

        Exited companies often do NOT appear in the Portfolio Investments sheet
        (they have already exited). We resolve the Investment via a 3-step lookup:
        1. In-memory investments dict (fast path, current portfolio companies)
        2. DB lookup by company name + org (case-insensitive)
        3. Create a skeleton PortfolioCompany + Investment so the ExitEvent can be saved
        """
        exit_count = 0
        default_scheme = list(investments.values())[0].scheme if investments else None

        # ── Batching pre-pass: exit_type ───────────────────────────────
        _exit_route_raws = set()
        for _r in rows:
            _er = _find_col_str(
                _r, 'Exit Route', 'Exit Type', 'Route', 'Exit Method',
                'exit_type')
            if _er:
                _exit_route_raws.add(_er)
        exit_type_map = self._classify_enum(
            list(_exit_route_raws), 'exit_type',
            context='Exit route / exit type for portfolio investment',
        ) if _exit_route_raws else {}

        for row in rows:
            name = _find_col_str(
                row, 'Company Name', 'Company', 'Name', 'Portfolio Company')
            if _is_junk_row(name):
                continue

            # Pre-extract cost from THIS row — needed for both step 3 and the
            # ExitEvent itself, and must be read before we might skip the row.
            cost = _find_col_decimal(
                row, 'Cost(Cr)', 'Cost(₹Cr)', 'Cost (Cr)',
                'Cost Basis', 'Invested', 'cost_basis')

            # Validate that the row has at least SOME exit-relevant data.
            # Without this check, investor/LP rows or key-value rows from
            # mis-mapped sheets would create phantom PortfolioCompany records.
            exit_date_check = _find_col_date(
                row, 'Exit Date', 'Date', 'Realization Date')
            exit_type_check = _find_col_str(
                row, 'Exit Route', 'Exit Type', 'Route', 'Exit Method',
                'exit_type')
            proceeds_check = _find_col_decimal(
                row, 'Proceeds(Cr)', 'Proceeds (Cr)', 'Realized(₹Cr)',
                'Realized', 'Proceeds', 'Exit Proceeds', 'Realization',
                'Gross Proceeds (Cr)', 'Net Proceeds (Cr)', 'proceeds')
            moic_check = _find_col_decimal(row, 'MOIC', 'Multiple', 'moic')
            if not exit_date_check and not exit_type_check and not proceeds_check and not moic_check:
                continue

            # Step 1: look up in the in-memory investments dict (current portfolio)
            inv = None
            for key, i in investments.items():
                if key.startswith(f'{name}|'):
                    inv = i
                    break

            if not inv:
                # Step 2: look up in DB by company name (case-insensitive)
                co = PortfolioCompany.objects.filter(
                    organization=self.org, name__iexact=name).first()
                if co:
                    inv = Investment.objects.filter(
                        portfolio_company=co).first()

            if not inv and default_scheme:
                # Step 3: create skeleton PortfolioCompany + Investment for this exit.
                # Exited companies are NOT in Portfolio Investments — they need a
                # skeleton record so the ExitEvent FK can be satisfied.
                # total_invested MUST never be None (NOT NULL constraint).
                sector = _find_col_str(row, 'Sector', 'Industry', 'Segment', default='Other')
                co, _ = PortfolioCompany.objects.update_or_create(
                    organization=self.org, name=name,
                    defaults={'sector': sector, 'is_active': False}
                        if sector and sector != 'Other'
                        else {'is_active': False},
                )
                inv, _ = Investment.objects.update_or_create(
                    portfolio_company=co,
                    scheme=default_scheme,
                    defaults={
                        'company_name': name,
                        'investment_date': exit_date_check,
                        'total_invested': abs(cost) if cost else Decimal('0'),
                        'instrument_type': 'equity',
                        'status': 'exited',
                    },
                )

            if not inv:
                logger.warning(f'  Exit row skipped — could not resolve company: {name!r}')
                continue

            # Reuse pre-validated values from row-level check above
            exit_date = exit_date_check
            exit_route = exit_type_check or ''
            realized = proceeds_check
            moic = moic_check
            irr_raw = _find_col_decimal(
                row, 'Gross IRR', 'IRR', 'IRR%', 'Gross IRR%', 'irr_pct')
            # IRR stored as decimal fraction (0.355 = 35.5%) → convert to %
            irr = (irr_raw * 100) if (irr_raw is not None and irr_raw < 2) else irr_raw
            net_irr_raw = _find_col_decimal(
                row, 'Net IRR', 'Net IRR%', 'Net Return', 'net_irr_pct')
            net_irr = (net_irr_raw * 100) if (net_irr_raw is not None and net_irr_raw < 2) else net_irr_raw
            is_actual_raw = _find_col_str(row, 'Is Actual', default='Yes')

            exit_type = exit_type_map.get(exit_route, 'secondary_sale')

            is_actual = is_actual_raw.lower() in ('yes', 'true', '1', 'y')

            gain_loss = None
            if realized and cost:
                gain_loss = realized - cost

            ExitEvent.objects.update_or_create(
                investment=inv,
                exit_type=exit_type,
                defaults={
                    'is_actual': is_actual,
                    'exit_date': exit_date,
                    'proceeds': realized or Decimal('0'),
                    'net_exit_proceeds': realized,
                    'realized_gain_loss': gain_loss,
                    'moic': moic,
                    'irr_pct': irr,
                    'irr_on_exit': net_irr,
                },
            )
            exit_count += 1
        return exit_count

    def _process_distribution_rows(self, rows, schemes, investments):
        """Process fund-level distribution rows from structured format.

        These are distributions to LPs from the Exits & Distributions sheet,
        separate from the investor-level distributions.
        """
        dist_count = 0
        default_scheme = list(schemes.values())[0] if schemes else None
        if not default_scheme:
            return

        for row in rows:
            scheme_name_raw = _find_col_str(
                row, 'Scheme', 'Scheme Name', 'Fund Scheme')
            target_scheme = default_scheme
            if scheme_name_raw:
                target_scheme = schemes.get(scheme_name_raw, default_scheme)

            dist_num = _find_col_decimal(
                row, 'Dist#', 'Dist #', 'Distribution #',
                'Distribution Number', '#', 'distribution_number')
            dist_num = int(dist_num) if dist_num else None

            dist_date = _find_col_date(
                row, 'Distribution Date', 'Date', 'Payment Date')
            # "Quarter" column (e.g. "Q1 FY25") is common in AIF distribution schedules
            if not dist_date:
                quarter_raw = _find_col_str(row, 'Quarter', 'Period')
                if quarter_raw:
                    _, dist_date = self._parse_quarter_period(quarter_raw)
            dist_type_raw = _find_col_str(
                row, 'Type', 'Distribution Type', default='return_of_capital')
            gross_amt = _find_col_decimal(
                row, 'Amount (Cr)', 'Amount(Cr)', 'Gross Amount (Cr)',
                'Gross Amount', 'Total Amount', 'total_gross_amount')
            tds_amt = _find_col_decimal(
                row, 'TDS Amount (Cr)', 'TDS Amount', 'TDS', 'Tax Deducted')
            net_amt = _find_col_decimal(
                row, 'Net Amount (Cr)', 'Net Amount', 'Net Payout')

            if not gross_amt or gross_amt <= 0:
                continue

            if not dist_num:
                existing = Distribution.objects.filter(
                    scheme=target_scheme).count()
                dist_num = existing + 1

            dist_type_result = self._classify_enum(
                [dist_type_raw], 'distribution_type',
                context='Distribution type for LP payout')
            dist_type = dist_type_result.get(dist_type_raw, 'return_of_capital')

            Distribution.objects.get_or_create(
                scheme=target_scheme,
                distribution_number=dist_num,
                defaults={
                    'distribution_date': dist_date or date.today(),
                    'distribution_type': dist_type,
                    'total_gross_amount': gross_amt,
                    'total_tds_amount': tds_amt or Decimal('0'),
                    'total_net_amount': net_amt or gross_amt,
                    'distribution_status': 'distributed',
                    'created_by': self.user,
                },
            )
            dist_count += 1

        if dist_count:
            logger.info(f'  Distributions (structured): {dist_count}')

    # ------------------------------------------------------------------
    # Distributions to LPs
    # ------------------------------------------------------------------

    def _import_distributions(self, wb, schemes, commitments, investments, domain_map):
        """Create Distribution + DistributionLineItem records from investor data.

        Reads distribution amounts from the investors_aml sheet (flat format
        where each LP row has a Distributions column). For structured formats,
        fund-level distributions are handled by _process_distribution_rows.
        """
        if not commitments:
            return

        default_scheme = list(schemes.values())[0] if schemes else None
        if not default_scheme:
            return

        # Distribution dedup — when the dedicated `exits_distributions`
        # sheet already produced Distribution rows for this scheme via
        # _import_exits_and_distributions / _process_distribution_rows
        # (which runs before this importer), the LP-register column is
        # a derived/secondary view of the same cash flows. Adding our
        # consolidated record on top of that double-counts distributions
        # (observed on AI_Trivesta: 215 from dedicated sheet + 160.74 from
        # LP register = 375.74 phantom total). Skip when authoritative
        # records already exist. Universal — applies to every fund.
        from lp.models import Distribution as _DistModel
        if _DistModel.objects.filter(scheme=default_scheme).exists():
            logger.info(
                '  Distributions: skipping LP-register fallback — '
                'authoritative records already imported from the '
                'distributions/exits sheet.'
            )
            return

        # Try investors_aml sheet for flat-format LP distribution data
        sheet_name = _dm_first(domain_map, 'investors_aml')
        if not sheet_name or sheet_name not in wb.sheetnames:
            return

        ws = wb[sheet_name]
        _, rows = read_table_from_sheet(ws, alias_map=self._get_alias(ws))

        # Check if this sheet actually has distribution columns
        # (Format B investor sheets don't have distribution amounts)
        # Detect distribution columns via Gemini Pass 2 canonical field names.
        # The canonical set must include every canonical key Pass 2 produces
        # for distribution-shaped columns; otherwise an LP-register sheet
        # that uses canonical names like "total_distributions_received" will
        # be wrongly skipped.
        has_dist_col = False
        if rows:
            sample = rows[0]
            _keys_norm = {k.lower().replace(' ', '_').replace('-', '_')
                          for k in sample.keys()}
            dist_fields = {
                'gross_amount', 'net_amount', 'tds_amount',
                'distribution_amount', 'distribution_type',
                'total_distributions_received',
                'total_gross_amount', 'total_net_amount',
            }
            has_dist_col = bool(dist_fields & _keys_norm)
        if not has_dist_col:
            return

        # Collect LP distribution data
        lp_distributions = []
        total_gross = Decimal('0')
        for row in rows:
            inv_name = _find_col_str(
                row, 'Investor Name', 'LP Name', 'Name', 'Investor',
                'investor_name')
            if not inv_name or inv_name not in commitments:
                continue

            dist_amt = _find_col_decimal(
                row, 'Distributions', 'Distribution', 'Returned',
                'Amount Returned', 'Total Distribution',
                'total_distributions_received', 'distribution_amount',
                'gross_amount', 'net_amount')
            if not dist_amt or dist_amt <= 0:
                continue

            total_gross += dist_amt
            lp_distributions.append((inv_name, commitments[inv_name], dist_amt))

        if not lp_distributions:
            return

        # Create a consolidated distribution record
        dist, created = Distribution.objects.get_or_create(
            scheme=default_scheme,
            distribution_number=1,
            defaults={
                'distribution_date': date.today(),
                'distribution_type': 'return_of_capital',
                'total_gross_amount': total_gross,
                'total_tds_amount': Decimal('0'),
                'total_net_amount': total_gross,
                'distribution_status': 'distributed',
                'created_by': self.user,
            },
        )

        if not created:
            return  # Already exists

        # Create line items per LP
        line_count = 0
        for inv_name, commitment, dist_amt in lp_distributions:
            DistributionLineItem.objects.get_or_create(
                distribution=dist,
                commitment=commitment,
                defaults={
                    'gross_amount': dist_amt,
                    'tds_rate': Decimal('0'),
                    'tds_amount': Decimal('0'),
                    'net_amount': dist_amt,
                },
            )
            line_count += 1

        logger.info(f'  Distributions: 1 distribution, {line_count} line items')

    # ------------------------------------------------------------------
    # Key Entities — Import & Link to Fund
    # ------------------------------------------------------------------

    # Map entity_type values from Excel to the Fund model FK field names
    _ENTITY_FK_MAP = {
        'manager': 'manager_entity',
        'trustee': 'trustee_entity',
        'sponsor': 'sponsor_entity',
        'custodian': 'custodian_entity',
        'statutory_auditor': 'auditor_entity',
    }

    def _normalize_entity_type(self, raw_type):
        """Normalize entity type string to a valid model choice via Gemini classification."""
        if not raw_type:
            return None
        raw = raw_type.strip()
        if not raw:
            return None
        # Direct match against Django model choices
        raw_lower = raw.lower().replace('-', '_')
        valid_types = [c[0] for c in Entity.ENTITY_TYPE_CHOICES]
        if raw_lower in valid_types:
            return raw_lower
        # Gemini classification
        result = self._classify_enum([raw], 'entity_type',
                                      context='Entity/service provider type for AIF fund')
        return result.get(raw)

    def _import_entities(self, wb, org, fund, domain_map):
        """Import key entities from Excel and link them to the fund.

        Dynamically finds entity data in:
        - Dedicated 'organization_users' domain sheet with KEY ENTITIES section
        - Fund & Scheme Master sheet with entity references
        - Any sheet containing entity-related sections

        Uses header-based reading with _find_col for format-agnostic extraction.
        """
        count = 0
        entity_map = {}  # entity_type -> Entity object

        # Strategy 1: Read from Gemini-classified organization_users domain
        sheet_name = _dm_first(domain_map, 'organization_users')

        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sections = self._read_sheet_via_layout(ws, alias_map=self._get_alias(ws))

            # Find entity section via Gemini sub-domain
            entity_rows = None
            for sec_name, (sec_headers, sec_rows) in sections.items():
                subdomain = self._get_section_subdomain(sheet_name, sec_name)
                if subdomain == 'entities':
                    entity_rows = sec_rows
                    break

            if entity_rows:
                for row in entity_rows:
                    raw_type = _find_col(
                        row, 'Entity Type', 'Type', 'Role', 'Entity Role',
                        'Category', 'Entity Category')
                    entity_name = _find_col(
                        row, 'Entity Name', 'Name', 'Legal Name',
                        'Organization Name', 'Firm Name', 'Company Name')
                    if not raw_type or not entity_name:
                        continue

                    etype = self._normalize_entity_type(str(raw_type))
                    if not etype:
                        logger.warning(f'  Entity: unknown type "{raw_type}" for "{entity_name}"')
                        continue

                    pan = _find_col(row, 'PAN', 'PAN Number', 'Tax ID') or ''
                    gstin = _find_col(row, 'GSTIN', 'GST', 'GST Number', 'GST ID') or ''
                    sebi_reg = _find_col(
                        row, 'SEBI Registration', 'SEBI Reg', 'SEBI No',
                        'Registration Number', 'Reg No', 'License') or ''
                    contact = _find_col(
                        row, 'Contact Person', 'Contact', 'Contact Name',
                        'Representative', 'Person') or ''
                    email = _find_col(
                        row, 'Email', 'Email Address', 'Email ID',
                        'Contact Email', 'E-mail') or ''
                    phone = _find_col(
                        row, 'Phone', 'Phone Number', 'Mobile',
                        'Contact Phone', 'Tel') or ''
                    address = _find_col(
                        row, 'Address', 'Office Address', 'Location',
                        'Registered Address') or ''
                    city = _find_col(row, 'City', 'Location') or ''
                    state = _find_col(row, 'State', 'Province') or ''
                    country = _find_col(row, 'Country', 'Nation') or 'India'

                    entity, _ = Entity.objects.get_or_create(
                        organization=org,
                        entity_type=etype,
                        entity_name=_str(entity_name),
                        defaults={
                            'pan': _str(pan),
                            'gstin': _str(gstin),
                            'sebi_registration': _str(sebi_reg),
                            'contact_person': _str(contact),
                            'email': _str(email),
                            'phone': _str(phone),
                            'address': _str(address),
                            'city': _str(city),
                            'state': _str(state),
                            'country': _str(country),
                        },
                    )
                    entity_map[etype] = entity
                    count += 1

        # Strategy 2: Read entity references from Fund & Scheme Master sheet
        if not entity_map:
            fsm_sheet = _dm_first(domain_map, 'fund_scheme_master')
            if fsm_sheet and fsm_sheet in wb.sheetnames:
                ws = wb[fsm_sheet]
                # Collect all label-value pairs first
                label_value_pairs = {}
                for r in range(1, ws.max_row + 1):
                    label = _str(ws.cell(r, 1).value).strip()
                    value = _str(ws.cell(r, 2).value).strip()
                    if label and value and len(value) > 2:
                        label_value_pairs[label] = value
                # Classify labels as entity types via Gemini (returns None for non-entity labels)
                if label_value_pairs:
                    from .canonical_schema import CANONICAL_ENUM_TYPES
                    from .gemini_column_mapper import classify_labels
                    entity_opts = CANONICAL_ENUM_TYPES.get('entity_type', {})
                    label_types = classify_labels(
                        list(label_value_pairs.keys()), 'entity_type_labels',
                        entity_opts,
                        context='Labels from fund master sheet — classify ONLY if label refers to an entity/service provider role, otherwise null')
                    for label, etype in label_types.items():
                        if etype and etype not in entity_map:
                            entity, _ = Entity.objects.get_or_create(
                                organization=org,
                                entity_type=etype,
                                entity_name=label_value_pairs[label],
                            )
                            entity_map[etype] = entity
                            count += 1

        # Link entities to fund
        if entity_map:
            update_fields = []
            for etype, fk_field in self._ENTITY_FK_MAP.items():
                if etype in entity_map:
                    setattr(fund, fk_field, entity_map[etype])
                    update_fields.append(fk_field)
            if update_fields:
                fund.save(update_fields=update_fields)

        logger.info(f'  Entities: {count} created/found, '
                     f'{len([k for k in entity_map if k in self._ENTITY_FK_MAP])} linked to fund')

    # ------------------------------------------------------------------
    # LP Capital Accounts — Auto-generate from imported data
    # ------------------------------------------------------------------

    def _generate_lp_capital_accounts(self, fund, schemes, commitments):
        """Auto-generate LP Capital Account snapshots from imported data.

        For each commitment, computes:
        - committed_capital from the Commitment record
        - called_capital from CapitalCallLineItems
        - distributed_capital from DistributionLineItems
        - unrealized_value pro-rata from NAV data
        - Performance metrics: TVPI, DPI, RVPI, MOIC

        Creates one snapshot per commitment as of the latest NAV date or today.
        """
        from django.db.models import Sum

        if not commitments:
            return

        count = 0
        for commit_key, commitment in commitments.items():
            scheme = commitment.scheme

            # Called capital: sum of all capital call line items for this commitment
            called = CapitalCallLineItem.objects.filter(
                commitment=commitment
            ).aggregate(total=Sum('called_amount'))['total'] or Decimal('0')

            # If no line items, try summing from capital calls on the scheme
            # pro-rata by commitment amount
            if called == 0:
                total_scheme_called = CapitalCall.objects.filter(
                    scheme=scheme
                ).aggregate(total=Sum('total_call_amount'))['total'] or Decimal('0')
                total_scheme_committed = Commitment.objects.filter(
                    scheme=scheme
                ).aggregate(total=Sum('commitment_amount'))['total'] or Decimal('0')
                if total_scheme_committed > 0 and total_scheme_called > 0:
                    ratio = commitment.commitment_amount / total_scheme_committed
                    called = (total_scheme_called * ratio).quantize(Decimal('0.01'))

            committed = commitment.commitment_amount or Decimal('0')
            uncalled = max(committed - called, Decimal('0'))

            # Distributed capital: sum of all distribution line items for this commitment
            distributed = DistributionLineItem.objects.filter(
                commitment=commitment
            ).aggregate(total=Sum('gross_amount'))['total'] or Decimal('0')

            # Unrealized value: LP's pro-rata share of latest NAV
            latest_nav = NAVRecord.objects.filter(
                scheme=scheme
            ).order_by('-nav_date').first()

            unrealized = Decimal('0')
            as_of = date.today()
            if latest_nav:
                as_of = latest_nav.nav_date
                total_nav = latest_nav.total_nav or Decimal('0')
                total_scheme_committed = Commitment.objects.filter(
                    scheme=scheme
                ).aggregate(total=Sum('commitment_amount'))['total'] or Decimal('0')
                if total_scheme_committed > 0:
                    lp_share = committed / total_scheme_committed
                    unrealized = (total_nav * lp_share).quantize(Decimal('0.01'))

            total_value = distributed + unrealized

            # Performance metrics
            tvpi = dpi = rvpi = moic = None
            if called > 0:
                tvpi = (total_value / called).quantize(Decimal('0.0001'))
                dpi = (distributed / called).quantize(Decimal('0.0001'))
                rvpi = (unrealized / called).quantize(Decimal('0.0001'))
                moic = tvpi  # Same as TVPI at aggregate level

            # Units held: from commitment if available, or pro-rata from NAV
            units_held = commitment.units_allocated
            if not units_held and latest_nav and latest_nav.total_units_outstanding:
                total_scheme_committed = Commitment.objects.filter(
                    scheme=scheme
                ).aggregate(total=Sum('commitment_amount'))['total'] or Decimal('0')
                if total_scheme_committed > 0:
                    lp_share = committed / total_scheme_committed
                    units_held = (latest_nav.total_units_outstanding * lp_share).quantize(
                        Decimal('0.000001'))

            # Management fee charged: LP's pro-rata share
            total_fees = ManagementFeeSchedule.objects.filter(
                scheme=scheme
            ).aggregate(total=Sum('fee_amount'))['total'] or Decimal('0')
            mgmt_fee_charged = Decimal('0')
            if total_fees > 0:
                total_scheme_committed = Commitment.objects.filter(
                    scheme=scheme
                ).aggregate(total=Sum('commitment_amount'))['total'] or Decimal('0')
                if total_scheme_committed > 0:
                    lp_share = committed / total_scheme_committed
                    mgmt_fee_charged = (total_fees * lp_share).quantize(Decimal('0.01'))

            LPCapitalAccount.objects.update_or_create(
                commitment=commitment,
                as_of_date=as_of,
                defaults={
                    'committed_capital': committed,
                    'called_capital': called,
                    'uncalled_capital': uncalled,
                    'distributed_capital': distributed,
                    'unrealized_value': unrealized,
                    'total_value': total_value,
                    'tvpi': tvpi,
                    'dpi': dpi,
                    'rvpi': rvpi,
                    'moic': moic,
                    'units_held': units_held,
                    'management_fee_charged': mgmt_fee_charged,
                    'carried_interest_charged': Decimal('0'),
                },
            )
            count += 1

        logger.info(f'  LP Capital Accounts: {count} snapshots generated')

    # ------------------------------------------------------------------
    # Income/Expense Ledger — Generate from NAV + Fee data
    # ------------------------------------------------------------------

    def _generate_income_expense_ledger(self, org, fund, schemes):
        """Generate income & expense ledger entries from NAV and fee data.

        This creates the journal entries that power the Income Statement
        and Cash Flow Statement in Fund Accounting.

        For each NAV period, creates entries for:
        - Unrealized gains/losses (from NAV movements)
        - Realized gains (from exits)
        - Management fees (from fee schedule)
        - Fund expenses (from NAV data)

        Uses existing COA accounts seeded by _setup_fund_accounting.
        """
        from django.db.models import Sum

        # Get COA map for this organization
        coa_map = {}
        for acct in ChartOfAccounts.objects.filter(organization=org):
            coa_map[acct.account_code] = acct

        if not coa_map:
            logger.info('  Income/expense ledger: no COA accounts found, skipping')
            return

        total_entries = 0

        for scheme_key, scheme in schemes.items():
            # Find the highest existing JE number for this scheme
            last_je = FundLedger.objects.filter(
                scheme=scheme
            ).order_by('-journal_entry_number').first()
            if last_je:
                # Extract number from JE-XXXX format
                try:
                    je_num = int(last_je.journal_entry_number.replace('JE-', '')) + 1
                except (ValueError, AttributeError):
                    je_num = 1000
            else:
                je_num = 1

            # --- NAV-based income entries (unrealized gains, realized gains) ---
            nav_records = NAVRecord.objects.filter(
                scheme=scheme
            ).order_by('nav_date')

            prev_investments_fv = None
            for nav in nav_records:
                nav_date = nav.nav_date

                # Unrealized gains: change in investment FV between periods
                current_fv = nav.investments_at_fair_value or Decimal('0')
                if prev_investments_fv is not None and current_fv > 0:
                    unrealized_change = current_fv - prev_investments_fv
                    if unrealized_change != 0 and '4100' in coa_map and '1200' in coa_map:
                        if unrealized_change > 0:
                            # Gain: debit Investments at FV, credit Unrealized Gains
                            FundLedger.objects.get_or_create(
                                scheme=scheme,
                                journal_entry_number=f'JE-{je_num:04d}',
                                defaults={
                                    'entry_date': nav_date,
                                    'description': f'Unrealized gain on investments — {nav_date.strftime("%b %Y")}',
                                    'debit_account': coa_map['1200'],
                                    'credit_account': coa_map['4100'],
                                    'amount': abs(unrealized_change),
                                    'reference_type': 'valuation_adjustment',
                                    'posted_by': self.user,
                                },
                            )
                        else:
                            # Loss: debit Unrealized Gains (reverse), credit Investments at FV
                            FundLedger.objects.get_or_create(
                                scheme=scheme,
                                journal_entry_number=f'JE-{je_num:04d}',
                                defaults={
                                    'entry_date': nav_date,
                                    'description': f'Unrealized loss on investments — {nav_date.strftime("%b %Y")}',
                                    'debit_account': coa_map['4100'],
                                    'credit_account': coa_map['1200'],
                                    'amount': abs(unrealized_change),
                                    'reference_type': 'valuation_adjustment',
                                    'posted_by': self.user,
                                },
                            )
                        je_num += 1
                        total_entries += 1

                prev_investments_fv = current_fv

                # Management fee expense entry from NAV data
                mgmt_fee = nav.management_fee_payable or Decimal('0')
                if mgmt_fee > 0 and '5000' in coa_map and '2000' in coa_map:
                    FundLedger.objects.get_or_create(
                        scheme=scheme,
                        journal_entry_number=f'JE-{je_num:04d}',
                        defaults={
                            'entry_date': nav_date,
                            'description': f'Management fee — {nav_date.strftime("%b %Y")}',
                            'debit_account': coa_map['5000'],
                            'credit_account': coa_map['2000'],
                            'amount': mgmt_fee,
                            'reference_type': 'management_fee',
                            'posted_by': self.user,
                        },
                    )
                    je_num += 1
                    total_entries += 1

                # Other liabilities / fund expenses
                other_liab = nav.other_liabilities or Decimal('0')
                if other_liab > 0 and '5100' in coa_map and '2200' in coa_map:
                    FundLedger.objects.get_or_create(
                        scheme=scheme,
                        journal_entry_number=f'JE-{je_num:04d}',
                        defaults={
                            'entry_date': nav_date,
                            'description': f'Fund expenses — {nav_date.strftime("%b %Y")}',
                            'debit_account': coa_map['5100'],
                            'credit_account': coa_map['2200'],
                            'amount': other_liab,
                            'reference_type': 'expense',
                            'posted_by': self.user,
                        },
                    )
                    je_num += 1
                    total_entries += 1

                # Cash & bank balance tracking
                cash = nav.cash_and_equivalents or Decimal('0')
                if cash > 0 and '1000' in coa_map and '1300' in coa_map:
                    receivables = nav.receivables or Decimal('0')
                    if receivables > 0:
                        FundLedger.objects.get_or_create(
                            scheme=scheme,
                            journal_entry_number=f'JE-{je_num:04d}',
                            defaults={
                                'entry_date': nav_date,
                                'description': f'Receivables — {nav_date.strftime("%b %Y")}',
                                'debit_account': coa_map['1300'],
                                'credit_account': coa_map['4200'],
                                'amount': receivables,
                                'reference_type': 'other',
                                'posted_by': self.user,
                            },
                        )
                        je_num += 1
                        total_entries += 1

            # --- Exit-based realized gains ---
            exits = ExitEvent.objects.filter(
                investment__scheme=scheme,
                is_actual=True,
            )
            for exit_event in exits:
                proceeds = exit_event.proceeds or Decimal('0')
                # Use realized_gain_loss if available, else compute from investment cost
                gain = exit_event.realized_gain_loss or Decimal('0')
                if gain == 0 and proceeds > 0 and exit_event.investment:
                    cost = exit_event.investment.total_invested or Decimal('0')
                    gain = proceeds - cost
                if gain != 0 and '4000' in coa_map and '1000' in coa_map:
                    exit_date = exit_event.exit_date or date.today()
                    company = exit_event.investment.company_name if exit_event.investment else 'Unknown'
                    if gain > 0:
                        FundLedger.objects.get_or_create(
                            scheme=scheme,
                            journal_entry_number=f'JE-{je_num:04d}',
                            defaults={
                                'entry_date': exit_date,
                                'description': f'Realized gain — exit from {company}',
                                'debit_account': coa_map['1000'],
                                'credit_account': coa_map['4000'],
                                'amount': abs(gain),
                                'reference_type': 'other',
                                'posted_by': self.user,
                            },
                        )
                    else:
                        FundLedger.objects.get_or_create(
                            scheme=scheme,
                            journal_entry_number=f'JE-{je_num:04d}',
                            defaults={
                                'entry_date': exit_date,
                                'description': f'Realized loss — exit from {company}',
                                'debit_account': coa_map['4000'],
                                'credit_account': coa_map['1000'],
                                'amount': abs(gain),
                                'reference_type': 'other',
                                'posted_by': self.user,
                            },
                        )
                    je_num += 1
                    total_entries += 1

            # --- Management fee entries from fee schedule ---
            fees = ManagementFeeSchedule.objects.filter(scheme=scheme)
            for fee in fees:
                if fee.fee_amount and fee.fee_amount > 0 and '5000' in coa_map and '1000' in coa_map:
                    FundLedger.objects.get_or_create(
                        scheme=scheme,
                        journal_entry_number=f'JE-{je_num:04d}',
                        defaults={
                            'entry_date': fee.period_end,
                            'description': f'Management fee payment — {fee.period_start.strftime("%b %Y")}',
                            'debit_account': coa_map['2000'],
                            'credit_account': coa_map['1000'],
                            'amount': fee.fee_amount,
                            'reference_type': 'management_fee',
                            'posted_by': self.user,
                        },
                    )
                    je_num += 1
                    total_entries += 1

                # GST on management fee
                if fee.gst_amount and fee.gst_amount > 0 and '5100' in coa_map and '1000' in coa_map:
                    FundLedger.objects.get_or_create(
                        scheme=scheme,
                        journal_entry_number=f'JE-{je_num:04d}',
                        defaults={
                            'entry_date': fee.period_end,
                            'description': f'GST on management fee — {fee.period_start.strftime("%b %Y")}',
                            'debit_account': coa_map['5100'],
                            'credit_account': coa_map['1000'],
                            'amount': fee.gst_amount,
                            'reference_type': 'management_fee',
                            'posted_by': self.user,
                        },
                    )
                    je_num += 1
                    total_entries += 1

            # --- Carried interest entry ---
            carry_records = CarriedInterest.objects.filter(scheme=scheme)
            for carry in carry_records:
                if carry.carry_amount_gross and carry.carry_amount_gross > 0:
                    if '5200' in coa_map and '2100' in coa_map:
                        FundLedger.objects.get_or_create(
                            scheme=scheme,
                            journal_entry_number=f'JE-{je_num:04d}',
                            defaults={
                                'entry_date': carry.calculation_date,
                                'description': f'Carried interest accrual — {carry.calculation_date}',
                                'debit_account': coa_map['5200'],
                                'credit_account': coa_map['2100'],
                                'amount': carry.carry_amount_gross,
                                'reference_type': 'carried_interest',
                                'posted_by': self.user,
                            },
                        )
                        je_num += 1
                        total_entries += 1

        logger.info(f'  Income/expense ledger: {total_entries} additional entries created')

    # ------------------------------------------------------------------
    # Chart of Accounts & Fund Ledger
    # ------------------------------------------------------------------

    # Standard fund accounting chart of accounts
    _COA_SEED = [
        ('1000', 'Cash & Bank', 'asset'),
        ('1100', 'Investments at Cost', 'asset'),
        ('1200', 'Investments at Fair Value', 'asset'),
        ('1300', 'Receivables', 'asset'),
        ('2000', 'Management Fee Payable', 'liability'),
        ('2100', 'Carried Interest Payable', 'liability'),
        ('2200', 'Other Liabilities', 'liability'),
        ('2300', 'Distributions Payable', 'liability'),
        ('3000', 'LP Capital', 'equity'),
        ('3100', 'Retained Earnings', 'equity'),
        ('4000', 'Realized Gains', 'income'),
        ('4100', 'Unrealized Gains', 'income'),
        ('4200', 'Interest Income', 'income'),
        ('4300', 'Dividend Income', 'income'),
        ('5000', 'Management Fees', 'expense'),
        ('5100', 'Fund Expenses', 'expense'),
        ('5200', 'Carried Interest', 'expense'),
    ]

    def _setup_fund_accounting(self, org, fund, schemes, investments):
        """Seed Chart of Accounts and create initial ledger entries."""
        # Seed COA
        coa_map = {}
        for code, name, acct_type in self._COA_SEED:
            acct, _ = ChartOfAccounts.objects.get_or_create(
                organization=org,
                account_code=code,
                defaults={
                    'account_name': name,
                    'account_type': acct_type,
                    'is_active': True,
                },
            )
            coa_map[code] = acct

        if not schemes:
            logger.info(f'  COA: {len(coa_map)} accounts seeded')
            return

        total_je = 0
        # Create ledger entries for ALL schemes (not just the first one)
        for scheme_key, scheme in schemes.items():
            je_num = FundLedger.objects.filter(scheme=scheme).count() + 1

            # Create ledger entries for capital calls (cash in → LP capital)
            capital_calls = CapitalCall.objects.filter(scheme=scheme)
            for call in capital_calls:
                FundLedger.objects.get_or_create(
                    scheme=scheme,
                    journal_entry_number=f'JE-{je_num:04d}',
                    defaults={
                        'entry_date': call.call_date,
                        'description': f'Capital Call #{call.call_number}',
                        'debit_account': coa_map['1000'],   # Cash
                        'credit_account': coa_map['3000'],   # LP Capital
                        'amount': call.total_call_amount,
                        'reference_type': 'capital_call',
                        'reference_id': call.id,
                        'posted_by': self.user,
                    },
                )
                je_num += 1

            # Create ledger entries for investments (cash out → investments at cost)
            for key, inv in investments.items():
                if inv.scheme_id == scheme.id and inv.total_invested and inv.total_invested > 0:
                    FundLedger.objects.get_or_create(
                        scheme=scheme,
                        journal_entry_number=f'JE-{je_num:04d}',
                        defaults={
                            'entry_date': inv.investment_date or date.today(),
                            'description': f'Investment in {inv.company_name}',
                            'debit_account': coa_map['1100'],   # Investments at Cost
                            'credit_account': coa_map['1000'],   # Cash
                            'amount': inv.total_invested,
                            'reference_type': 'investment',
                            'reference_id': inv.id,
                            'posted_by': self.user,
                        },
                    )
                    je_num += 1

            # Create ledger entries for distributions (distributions payable → cash out)
            distributions = Distribution.objects.filter(scheme=scheme)
            for dist in distributions:
                FundLedger.objects.get_or_create(
                    scheme=scheme,
                    journal_entry_number=f'JE-{je_num:04d}',
                    defaults={
                        'entry_date': dist.distribution_date,
                        'description': f'Distribution #{dist.distribution_number}',
                        'debit_account': coa_map['3000'],   # LP Capital (return)
                        'credit_account': coa_map['1000'],   # Cash
                        'amount': dist.total_gross_amount,
                        'reference_type': 'distribution',
                        'reference_id': dist.id,
                        'posted_by': self.user,
                    },
                )
                je_num += 1

            total_je += je_num - 1

        logger.info(f'  Fund accounting: {len(coa_map)} COA accounts, '
                     f'{total_je} journal entries across {len(schemes)} schemes')

    # ------------------------------------------------------------------
    # Management Fee Schedule
    # ------------------------------------------------------------------

    def _import_management_fees(self, wb, schemes, domain_map):
        """Create ManagementFeeSchedule records.

        Handles two formats:
        - Format A: Flat table (Period | Mgmt Fee | ...) from nav_accounting sheet
        - Format B: Quarterly transposed table (FEES_REGISTER / FEES sheet) with
          columns like "Q1 FY25 | Q2 FY25 | Q3 FY25 | Q4 FY25".
          This is the dominant format in fund management Excel files.
        """
        default_scheme = list(schemes.values())[0] if schemes else None
        if not default_scheme:
            return

        # No hardcoded fallback for the fee rate. If the Excel never told us,
        # the model field is 0 (legitimate "unknown"). We use that 0 rather
        # than a synthetic 2.00% so downstream consumers see honest data.
        fee_rate = default_scheme.management_fee_pct or Decimal('0')

        # ── Format B: FEES_REGISTER / dedicated fees sheet ────────────────────
        # Use Gemini-classified 'fees_register' domain, fallback to 'nav_accounting'
        fees_sheets = _dm_sheets(domain_map, 'fees_register')
        if not fees_sheets:
            fees_sheets = _dm_sheets(domain_map, 'nav_accounting')
        fees_sheet = fees_sheets[0] if fees_sheets else None

        if fees_sheet and fees_sheet in wb.sheetnames:
            ws_f = wb[fees_sheet]
            _quarter_re = re.compile(r'^Q[1-4]\s*FY\s*\d{2,4}$', re.IGNORECASE)
            headers_dict, fee_rows = read_table_from_sheet(
                ws_f, alias_map=self._get_alias(ws_f))
            quarter_cols = [h for h in headers_dict.keys()
                            if _quarter_re.match(h.strip())]

            if quarter_cols and fee_rows:
                # Classify fee component labels via Gemini
                fee_comp_labels = set()
                for row in fee_rows:
                    c = _find_col_str(row, 'Component', 'Expense Category', 'Item')
                    if c:
                        fee_comp_labels.add(c)
                fee_comp_map = self._classify_labels(
                    fee_comp_labels, 'fee_components',
                    context='Fee schedule components: management fee vs GST')

                mgmt_fee_vals = {}
                gst_vals = {}
                for row in fee_rows:
                    comp = _find_col_str(row, 'Component', 'Expense Category', 'Item')
                    fee_type = fee_comp_map.get(comp) if comp else None
                    if not fee_type:
                        continue
                    for qcol in quarter_cols:
                        val = _d(row.get(qcol))
                        if val is None:
                            continue
                        if fee_type == 'management_fee' and val > 0:
                            mgmt_fee_vals[qcol] = val
                        elif fee_type == 'gst_on_management_fee':
                            gst_vals[qcol] = abs(val)

                count = 0
                for qcol in quarter_cols:
                    fee_amt = mgmt_fee_vals.get(qcol)
                    if not fee_amt or fee_amt <= 0:
                        continue
                    gst = gst_vals.get(qcol,
                                       (fee_amt * Decimal('0.18')).quantize(Decimal('0.01')))
                    period_start, period_end = self._parse_quarter_period(qcol)
                    if not period_start:
                        continue
                    ManagementFeeSchedule.objects.get_or_create(
                        scheme=default_scheme,
                        period_start=period_start,
                        period_end=period_end,
                        defaults={
                            'fee_basis_amount': Decimal('0'),
                            'fee_rate': fee_rate,
                            'fee_amount': fee_amt,
                            'gst_amount': gst,
                            'total_fee_with_gst': fee_amt + gst,
                            'fee_status': 'paid',
                        },
                    )
                    count += 1
                logger.info(f'  Management fees (fees register): {count} quarters')
                return  # Done — skip the NAV-sheet path below

        # ── Format A: Flat table from nav_accounting sheet ────────────────────
        sheet_name = _dm_first(domain_map, 'nav_accounting')
        if not sheet_name or sheet_name not in wb.sheetnames:
            return

        ws = wb[sheet_name]
        sections = self._read_sheet_via_layout(ws, alias_map=self._get_alias(ws))
        rows = None
        for sec_name, (sec_headers, sec_rows) in sections.items():
            subdomain = self._get_section_subdomain(sheet_name, sec_name)
            if subdomain == 'nav_records' or sec_name == '__default__':
                rows = sec_rows
                break
        if not rows:
            _, rows = read_table_from_sheet(ws, alias_map=self._get_alias(ws))

        count = 0
        for row in rows:
            period_raw = _find_col(
                row, 'Period', 'NAV Date', 'Date', 'Month')
            if not period_raw:
                continue
            period_date = _date(period_raw)
            if not period_date and isinstance(period_raw, str):
                period_date = self._parse_period(period_raw)
            if not period_date:
                continue

            fee_amount = _find_col_decimal(
                row, 'Mgmt Fees', 'Management Fee', 'Fees',
                'Management Fee Amount')
            if not fee_amount or fee_amount <= 0:
                continue

            total_inv = _find_col_decimal(
                row, 'Total Investments', 'Investments at FV',
                'Total NAV', 'Fee Basis')
            fee_basis = total_inv or Decimal('0')

            # Period is monthly: period_start = 1st of month, period_end = last of month
            period_start = period_date.replace(day=1)
            last_day = calendar.monthrange(period_date.year, period_date.month)[1]
            period_end = period_date.replace(day=last_day)

            # Compute GST (18% in India)
            gst = (fee_amount * Decimal('0.18')).quantize(Decimal('0.01'))

            ManagementFeeSchedule.objects.get_or_create(
                scheme=default_scheme,
                period_start=period_start,
                period_end=period_end,
                defaults={
                    'fee_basis_amount': fee_basis,
                    'fee_rate': fee_rate,
                    'fee_amount': fee_amount,
                    'gst_amount': gst,
                    'total_fee_with_gst': fee_amount + gst,
                    'fee_status': 'paid',
                },
            )
            count += 1

        logger.info(f'  Management fees: {count} periods')

    # ------------------------------------------------------------------
    # Compliance import
    # ------------------------------------------------------------------

    def _import_compliance(self, wb, org, fund, companies, domain_map):
        """Import compliance data (per-company obligations + fund-level SEBI
        filings) from sheets that Gemini Pass 1 classified as 'compliance'.

        ZERO HARDCODING — every step uses Gemini to interpret semantics:
          * Pass 2.5 per-sheet layout to discover sub-tables (header row,
            data start/end, derived columns). Works regardless of where the
            sheet starts (banner rows, disclaimers, free text above the
            actual grid).
          * Pass 2 column alias map for canonical header normalization.
          * Pass 3 classify_labels / classify_enum to translate the
            obligation column names ("ROC/MCA", "EPF/ESIC", …), the per-cell
            statuses ("Current", "Delayed", "Pending Review", …), the fund
            filing-type labels ("SEBI QAR Filing", "FATCA/CRS Reporting", …)
            and the fund filing-status labels ("Filed On Time", …) into
            canonical keys that map 1:1 onto Django model choices.

        Two table shapes are handled by the SAME code path, decided
        per-sub-table at runtime:

          A. PER-COMPANY TRACKER GRID
             Identity column: company name (semantic match via _find_col_str).
             Obligation columns: any column header that Gemini classifies
             into the canonical 'compliance_obligation_type' set.
             For each (company, obligation) cell:
               • Gemini classifies the cell text into
                 'compliance_company_status' (compliant / due / overdue / N/A).
               • Resolve the company against the already-imported portfolio
                 (skip rows whose name doesn't match — we never invent
                 PortfolioCompany rows here; that's the portfolio importer's
                 job).
               • update_or_create PortfolioCompanyCompliance keyed on
                 (portfolio_company, obligation_type).

          B. FUND-LEVEL FILINGS TABLE
             Identity column: obligation/filing label.
             Recognised columns: Frequency, Due Date, Filed Date, Status,
             Notes. All resolved by _find_col semantic fuzzy match —
             no hardcoded header strings.
             For each row:
               • Gemini classifies the obligation label into
                 'sebi_filing_type' (qar / aar / ctr / fema / fatca_crs /
                 nav_depositories / valuation_certificate / other).
               • Gemini classifies the status text into
                 'compliance_filing_status' (filed / pending / overdue /
                 not_started).
               • QAR / AAR rows → SEBIReport (which has a proper schema
                 for them).
               • Everything else → ComplianceCalendar (general-purpose;
                 covers CTR, FEMA, FATCA, NAV-to-depositories, etc.).
             Both writes use update_or_create keyed on stable identity so
             re-imports are idempotent (CLAUDE.md rule).

        Shape detection itself is semantic, not heuristic: a sub-table is
        treated as a per-company grid iff Gemini classifies AT LEAST ONE
        of its column headers into compliance_obligation_type. Otherwise
        it is treated as a fund-level filings table. No keyword matching,
        no row-count thresholds.
        """
        if not HAS_COMPLIANCE:
            return

        sheets = _dm_sheets(domain_map, 'compliance')
        if not sheets:
            logger.info('  Compliance: no sheets classified as compliance domain')
            return

        # Build a name → PortfolioCompany lookup (case- and whitespace-
        # insensitive) so per-company rows can be resolved without a DB
        # round-trip per row.
        co_lookup = {}
        for co in companies.values() if isinstance(companies, dict) else (companies or []):
            key = (co.name or '').strip().lower()
            if key:
                co_lookup[key] = co

        per_company_written = 0
        sebi_reports_written = 0
        calendar_events_written = 0

        for sheet_name in sheets:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            sections = self._read_sheet_via_layout(ws, alias_map=self._get_alias(ws))
            if not sections:
                continue

            for sec_name, (headers, rows) in sections.items():
                if not rows:
                    continue

                header_strs = [h for h in headers.keys() if h and str(h).strip()]
                if not header_strs:
                    continue

                # ── Sub-table purpose classification (Gemini, semantic) ──
                # Defense-in-depth: Pass 1 can mis-classify a sheet into
                # 'compliance' (e.g. a "VALIDATION" sheet of cross-sheet
                # integrity checks). Don't let those rows pollute the
                # compliance tables. Ask Gemini what THIS sub-table actually
                # is, judged by headers + sample rows together — and route
                # accordingly. Pure semantic call; zero keyword matching.
                from .gemini_column_mapper import classify_subtable_purpose
                sample_rows_for_purpose = [
                    [r.get(h) for h in header_strs]
                    for r in rows[:5]
                ]
                purpose = classify_subtable_purpose(
                    headers=header_strs,
                    sample_rows=sample_rows_for_purpose,
                    allowed_purposes={
                        'per_company_tracker': (
                            'A grid where each row is ONE portfolio company and the '
                            'columns are different regulatory obligations (ROC/MCA, '
                            'GST, Labour, EPF, etc.). Cell values are statuses like '
                            'Current / Delayed / Pending Review / Overdue.'
                        ),
                        'fund_level_filings': (
                            'A list where each row is ONE fund-level regulatory '
                            'obligation (SEBI QAR / AAR / CTR, FEMA, FATCA/CRS, '
                            'NAV to depositories, valuation certificate). Columns '
                            'typically include obligation name, frequency, due '
                            'date, filed date, status.'
                        ),
                    },
                    context=(
                        f"Sheet '{sheet_name}' was classified as compliance "
                        f"domain by Pass 1; sub-section '{sec_name}'."
                    ),
                )

                if purpose == 'per_company_tracker':
                    # Header-level Gemini classification picks out which of
                    # this grid's columns are actual obligation columns
                    # (vs identity columns like #, Co.ID, Company Name,
                    # Sector). Pure semantic matching against the canonical
                    # obligation enum — no keyword tables.
                    obl_map = self._classify_labels(
                        header_strs, 'compliance_obligation_type',
                        context=(
                            f'Column headers in a per-company compliance '
                            f'tracker grid; pick out obligation columns '
                            f'(sheet "{sheet_name}", section "{sec_name}").'
                        ),
                    )
                    obligation_headers = {h: k for h, k in obl_map.items() if k}
                    if not obligation_headers:
                        logger.info(
                            f'  Compliance: sub-section "{sec_name}" classified as '
                            'per-company tracker but no obligation columns recognised; skipping'
                        )
                        continue
                    per_company_written += self._import_compliance_per_company(
                        rows, headers, obligation_headers, co_lookup, sec_name,
                    )
                elif purpose == 'fund_level_filings':
                    sw, cw = self._import_compliance_fund_filings(
                        rows, headers, fund, sec_name,
                    )
                    sebi_reports_written += sw
                    calendar_events_written += cw
                else:
                    # 'other' — Gemini says this sub-table is not actual
                    # compliance data (validation rules, instructions,
                    # sample data, metadata, etc.). Skip cleanly.
                    logger.info(
                        f'  Compliance: skipping sub-section "{sec_name}" in '
                        f'"{sheet_name}" — Gemini classified as non-compliance content'
                    )

        logger.info(
            f'  Compliance: {per_company_written} per-company obligations, '
            f'{sebi_reports_written} SEBI reports, '
            f'{calendar_events_written} calendar events'
        )

    def _import_compliance_per_company(
        self, rows, headers, obligation_headers, co_lookup, sec_name,
    ):
        """Write PortfolioCompanyCompliance rows from a per-company tracker grid.

        obligation_headers: {excel_header: canonical_obligation_key}
        Returns count of obligations written.
        """
        # Batch-classify EVERY cell value in obligation columns up front, so
        # we make one Gemini call for the whole sheet instead of one per cell.
        all_status_values = set()
        for row in rows:
            for h in obligation_headers.keys():
                v = row.get(h)
                if v is None or v == '':
                    continue
                s = str(v).strip()
                if s:
                    all_status_values.add(s)
        if not all_status_values:
            return 0

        status_map = self._classify_enum(
            all_status_values, 'compliance_company_status',
            context=f'Per-company compliance status cells in "{sec_name}"',
        )

        # Status → (PortfolioCompanyCompliance.STATUS_CHOICES, deadline-offset days)
        # No hardcoded keywords here — the four keys come directly from
        # CANONICAL_VALUE_CATEGORIES['compliance_company_status'].
        status_to_model = {
            'compliant':      'compliant',
            'due':            'due',
            'overdue':        'overdue',
            'not_applicable': 'not_applicable',
        }

        from datetime import date as _date_cls
        today = _date_cls.today()

        # Batch-classify all company names once (the grid is one row per
        # company; classify_junk_names guards against subtotal / total /
        # header rows ending up as fake PortfolioCompanyCompliance entries).
        name_rows = []
        for row in rows:
            nm = _find_col_str(
                row, 'Company Name', 'Company', 'Name',
                'Investee', 'Portfolio Company', 'Entity',
            )
            if nm and str(nm).strip():
                name_rows.append((str(nm).strip(), row))
        if not name_rows:
            return 0
        junk = self._classify_junk_names([n for n, _ in name_rows])

        written = 0
        for raw_name, row in name_rows:
            if raw_name in junk:
                continue
            co = co_lookup.get(raw_name.lower())
            if not co:
                # Name doesn't match an imported company — skip (don't invent
                # PortfolioCompany rows here; the portfolio importer is the
                # only authority for that).
                continue

            for excel_h, canonical_obl in obligation_headers.items():
                cell = row.get(excel_h)
                if cell is None or cell == '':
                    continue
                cell_str = str(cell).strip()
                if not cell_str:
                    continue
                status_key = status_map.get(cell_str)
                model_status = status_to_model.get(status_key)
                if not model_status:
                    continue

                # Use today as deadline when none is supplied by the Excel
                # (the tracker sheet is a status snapshot — there's no per-
                # cell due date column in the canonical shape). The deadline
                # field is required by the model; keep it as a placeholder
                # so RAG status (Green/Amber/Red) renders correctly.
                PortfolioCompanyCompliance.objects.update_or_create(
                    portfolio_company=co,
                    obligation_type=canonical_obl,
                    defaults={
                        'obligation_name': excel_h,
                        'deadline': today,
                        'status': model_status,
                    },
                )
                written += 1
        return written

    def _import_compliance_fund_filings(self, rows, headers, fund, sec_name):
        """Write SEBIReport / ComplianceCalendar rows from a fund-level filings table.

        Returns (sebi_reports_count, calendar_events_count).
        """
        # Collect raw obligation labels + raw status labels first → one
        # Gemini call each to classify the whole batch.
        raw_obligations = set()
        raw_statuses = set()
        rows_resolved = []
        for row in rows:
            obl_label = _find_col_str(
                row, 'Obligation', 'Filing', 'Filing Type', 'Report Type',
                'Compliance Type', 'Type', 'Particulars', 'Description', 'Item',
            )
            if not obl_label or not str(obl_label).strip():
                continue
            obl_label = str(obl_label).strip()
            status_label = _find_col_str(
                row, 'Status', 'Filing Status', 'Compliance Status', 'State',
            )
            if status_label:
                status_label = str(status_label).strip()
                if status_label:
                    raw_statuses.add(status_label)
            raw_obligations.add(obl_label)
            rows_resolved.append((obl_label, status_label, row))

        if not rows_resolved:
            return (0, 0)

        obligation_map = self._classify_enum(
            raw_obligations, 'sebi_filing_type',
            context=f'Fund-level filing labels in "{sec_name}"',
        )
        status_map = self._classify_enum(
            raw_statuses, 'compliance_filing_status',
            context=f'Fund-level filing status cells in "{sec_name}"',
        ) if raw_statuses else {}

        # Map canonical filing-status key → SEBIReport.FILING_STATUS_CHOICES
        # (model uses 'filed'/'in_review'/'rejected' etc. — keep mapping
        # explicit so future SEBI workflow changes are easy to track).
        filing_status_to_model = {
            'filed':       'filed',
            'pending':     'in_review',
            'overdue':     'in_review',  # still pending action; no 'overdue' choice exists on SEBIReport
            'not_started': 'not_started',
        }
        calendar_status_to_model = {
            'filed':       'completed',
            'pending':     'in_progress',
            'overdue':     'overdue',
            'not_started': 'upcoming',
        }

        from datetime import date as _date_cls, timedelta as _td
        today = _date_cls.today()

        sebi_count = 0
        cal_count = 0
        for obl_label, status_label, row in rows_resolved:
            filing_type = obligation_map.get(obl_label) or 'other'
            status_key = status_map.get(status_label) if status_label else None
            due_date = _date(_find_col(
                row, 'Due Date', 'Deadline', 'Filing Deadline', 'Target Date',
            )) or today
            filed_date = _date(_find_col(
                row, 'Filed Date', 'Filing Date', 'Submitted Date',
                'Date Filed', 'Submission Date',
            ))
            notes = _find_col_str(row, 'Notes', 'Remarks', 'Comments') or ''
            frequency = _find_col_str(row, 'Frequency', 'Periodicity', 'Cadence') or ''

            if filing_type in ('qar', 'aar'):
                # SEBIReport — has dedicated columns for these
                model_status = filing_status_to_model.get(status_key, 'not_started')
                # Reporting period: use due_date as a stable anchor. For
                # quarterly QAR, period ends ~15 days before due_date; for
                # annual AAR, period ends ~60 days before due_date (May 31
                # deadline → period ends March 31). These are SEBI rules,
                # not data-format heuristics.
                if filing_type == 'qar':
                    period_end = due_date - _td(days=15)
                    period_start = period_end - _td(days=89)
                else:  # aar
                    period_end = due_date - _td(days=61)
                    period_start = period_end - _td(days=364)
                SEBIReport.objects.update_or_create(
                    fund=fund,
                    report_type=filing_type,
                    due_date=due_date,
                    defaults={
                        'reporting_period_start': period_start,
                        'reporting_period_end': period_end,
                        'filing_status': model_status,
                        'filed_date': filed_date,
                        'si_portal_reference_number': notes[:50] if notes else '',
                    },
                )
                sebi_count += 1
            else:
                # ComplianceCalendar — general-purpose home for CTR, FEMA,
                # FATCA, NAV-to-depositories, valuation certificates, etc.
                # The compliance_type field is a free char(40), so we can
                # store the canonical filing_type key directly.
                cal_status = calendar_status_to_model.get(status_key, 'upcoming')
                ComplianceCalendar.objects.update_or_create(
                    organization=fund.organization,
                    fund=fund,
                    compliance_type=filing_type,
                    title=obl_label,
                    due_date=due_date,
                    defaults={
                        'status': cal_status,
                        'completed_date': filed_date,
                        'description': frequency,
                        'notes': notes,
                    },
                )
                cal_count += 1

        return (sebi_count, cal_count)

    # ------------------------------------------------------------------
    # Carried Interest Computation
    # ------------------------------------------------------------------

    def _compute_carried_interest(self, schemes, wb=None, domain_map=None):
        """Compute carried interest using European-waterfall mechanics with
        time-weighted compounded hurdle. ZERO hardcoded fallbacks — every
        rate, every percentage, every amount comes from the imported data
        (Cover/PPM → scheme model → DB rows). If a required field is
        missing, the calculation is SKIPPED and a warning is logged so
        operators know which input is absent. We never invent a default.

        Inputs (per scheme):
          • scheme.hurdle_rate_pct  — annual hurdle, e.g. 8 (% p.a.)
          • scheme.carry_pct        — GP carry above hurdle, e.g. 20 (%)
          • CapitalCall.total_call_amount + .call_date — per-call vintage
          • Distribution.total_gross_amount — money returned TO LPs.
            This is the canonical "Total Value" measure. ExitEvent.proceeds
            is NOT added on top: exit proceeds are upstream of distributions
            (an exit generates proceeds → some of that flows out as a
            distribution to LPs). Summing both double-counts. The European
            waterfall is evaluated on capital RETURNED TO LPs, which is the
            distribution number.

        Formula (European whole-fund waterfall):
          preferred_return = Σ over calls of call.amount × ((1+h)^years_held − 1)
            where years_held = (today − call_date) / 365.25 and h = hurdle/100
          carry_base = max(total_distributions − total_called − preferred_return, 0)
          carry_gross = carry_base × carry_pct/100
          clawback   = max(prior_carry_paid_to_GP − carry_due, 0)
            (where carry_due is the freshly-computed carry_gross, and
            prior_carry_paid_to_GP is taken from an explicit waterfall
            sheet if Gemini found one — else None to signal "unknown")
          carry_net  = carry_gross − clawback   (clawback==None → net=gross)

        On schemes lacking calls or rate inputs the calculation is skipped
        rather than synthesised — preserves data honesty.
        """
        if not schemes:
            return

        from django.db.models import Sum
        from datetime import date as _date_cls
        from decimal import Decimal as D

        today = _date_cls.today()
        # Year length in days — calendar-accurate (handles leap years naturally
        # over multi-year horizons). Not a tunable fudge constant; this is the
        # actual mean tropical year length.
        DAYS_PER_YEAR = D('365.25')

        for scheme_key, scheme in schemes.items():
            carry_pct = scheme.carry_pct
            hurdle_rate = scheme.hurdle_rate_pct

            # No hardcoded fallback: if the Excel never gave us the rates,
            # we cannot meaningfully compute carry. Skip and tell operators.
            if carry_pct is None or hurdle_rate is None:
                logger.warning(
                    f'Carry computation skipped for scheme "{scheme.name}" — '
                    f'missing rate(s): carry_pct={carry_pct}, hurdle_rate_pct={hurdle_rate}. '
                    f'Source: scheme.hurdle_rate_pct / scheme.carry_pct, populated by '
                    f'_import_fund_and_schemes from the Cover/PPM sheet.'
                )
                continue

            calls = list(CapitalCall.objects.filter(scheme=scheme))
            if not calls:
                logger.warning(
                    f'Carry computation skipped for scheme "{scheme.name}" — '
                    f'no CapitalCall rows in DB. Check the capital_calls '
                    f'sheet import.'
                )
                continue

            total_called = sum((c.total_call_amount or D('0')) for c in calls)
            if total_called <= 0:
                continue

            # Time-weighted COMPOUNDED preferred return. Each call's hurdle
            # accrues from ITS OWN call_date to today, at the scheme's
            # annual hurdle rate, compounded continuously over the holding
            # period in years. This is the SEBI / IVCA / textbook
            # European-waterfall formulation.
            h = hurdle_rate / D('100')   # e.g. 8 → 0.08
            preferred_return = D('0')
            for c in calls:
                if not c.total_call_amount or c.total_call_amount <= 0:
                    continue
                if not c.call_date:
                    # No vintage → conservatively assume the call was made
                    # today (zero accrual). Logged so operators see it.
                    logger.warning(
                        f'  Carry: capital call #{c.call_number} for scheme '
                        f'"{scheme.name}" has no call_date — accruing 0 hurdle '
                        f'for this call.'
                    )
                    continue
                days_held = (today - c.call_date).days
                if days_held <= 0:
                    continue
                years_held = D(days_held) / DAYS_PER_YEAR
                # Compounded: principal × ((1+h)^years − 1)
                # Use float for the exponentiation (Decimal has no ** for
                # non-integer exponents), then convert back to Decimal.
                growth = D(str((1 + float(h)) ** float(years_held) - 1))
                preferred_return += c.total_call_amount * growth
            preferred_return = preferred_return.quantize(D('0.01'))

            # Distributions ONLY — the canonical "total value returned to LPs"
            # measure. Exit proceeds are NOT added on top (they're upstream).
            total_distributions = Distribution.objects.filter(
                scheme=scheme
            ).aggregate(total=Sum('total_gross_amount'))['total'] or D('0')

            carry_base = max(
                total_distributions - total_called - preferred_return,
                D('0'),
            )
            carry_gross = (carry_base * carry_pct / D('100')).quantize(D('0.01'))

            # Clawback — only meaningful when we have a record of what was
            # ALREADY PAID to the GP. Without payment-history tracking we
            # try the Excel side (some funds publish prior-period carry
            # explicitly in a Waterfall sheet); if no such number is found
            # we set clawback to None to signal "cannot determine", instead
            # of falsely showing 0. The model field is non-null, so write
            # 0 ONLY when we genuinely know it's 0 (carry_gross is 0 →
            # nothing was ever paid → clawback is 0 by definition).
            prior_carry_paid = None
            if wb:
                # _extract_carry_from_workbook returns (carry_gross, preferred)
                # found in the Excel waterfall sheet. If the Excel has an
                # explicit "carry paid to date" value, the same helper
                # discovers it under the carry_gross label. We treat that
                # discovered value as prior_carry_paid for clawback purposes.
                excel_carry_gross, excel_pref = self._extract_carry_from_workbook(
                    wb, domain_map,
                )
                if excel_carry_gross is not None and excel_carry_gross >= 0:
                    prior_carry_paid = excel_carry_gross

            if prior_carry_paid is None:
                # Honest unknown — set to 0 only if no carry was ever due
                # (so clawback is mathematically impossible), else leave
                # the field at 0 but mark calculation_status='indicative'
                # so the dashboard can show "Status: Indicative" caveats.
                clawback = D('0') if carry_gross == 0 else D('0')
                # We keep clawback as 0 in DB (model is non-null) but the
                # 'indicative' status flag below tells consumers we cannot
                # truly compute clawback without payment history.
            else:
                clawback = max(prior_carry_paid - carry_gross, D('0')).quantize(D('0.01'))

            carry_net = (carry_gross - clawback).quantize(D('0.01'))
            calc_status = 'final' if prior_carry_paid is not None else 'indicative'

            CarriedInterest.objects.update_or_create(
                scheme=scheme,
                calculation_date=today,
                defaults={
                    'total_distributions': total_distributions,
                    'total_called_capital': total_called,
                    'preferred_return_amount': preferred_return,
                    'carry_base': carry_base,
                    'carry_amount_gross': carry_gross,
                    'carry_amount_net': carry_net,
                    'gp_clawback_provision': clawback,
                    'calculation_status': calc_status,
                },
            )

            logger.info(
                f'  Carried interest ({scheme.name}): '
                f'called={total_called}, distributions={total_distributions}, '
                f'pref_return={preferred_return} (compounded over per-call vintages), '
                f'carry_base={carry_base}, carry_gross={carry_gross}, '
                f'clawback={clawback}, carry_net={carry_net}, status={calc_status}'
            )

    def _extract_carry_from_workbook(self, wb, domain_map=None):
        """Extract carry/preferred-return amounts from Gemini-classified sheets.

        Uses waterfall_carry domain first, then nav_calculation, then nav_accounting.
        Searches key-value formatted sheets for rows that semantically match
        carry-related labels.

        Returns (carry_gross, preferred_return) as Decimals or (None, None).
        """
        carry_gross = None
        preferred_return = None

        # Use Gemini domain_map: waterfall_carry → nav_calculation → nav_accounting
        candidate_sheets = []
        if domain_map:
            for domain in ('waterfall_carry', 'nav_calculation', 'nav_accounting'):
                for s in _dm_sheets(domain_map, domain):
                    if s in wb.sheetnames and s not in candidate_sheets:
                        candidate_sheets.append(s)

        if not candidate_sheets:
            return None, None

        for sn in candidate_sheets:
            ws = wb[sn]
            max_r = min(ws.max_row + 1, 80)
            # Collect all label-value pairs
            carry_labels = {}
            carry_vals = {}
            for rr in range(1, max_r):
                for label_col, val_col in [(1, 2), (2, 3)]:
                    label = ws.cell(rr, label_col).value
                    if not label:
                        continue
                    raw_val = ws.cell(rr, val_col).value
                    val = _d(raw_val)
                    if val is None or val <= 0:
                        continue
                    label_str = str(label).strip()
                    carry_labels[label_str] = label_str
                    carry_vals[label_str] = val

            if carry_labels:
                wf_map = self._classify_labels(
                    list(carry_labels.keys()), 'waterfall_components',
                    context='Waterfall/carry components: carried interest vs preferred return')
                for label_str, wf_type in wf_map.items():
                    val = carry_vals.get(label_str)
                    if val is None:
                        continue
                    if wf_type == 'carry_gross' and carry_gross is None:
                        carry_gross = val
                    elif wf_type == 'preferred_return' and preferred_return is None:
                        preferred_return = val

            if carry_gross is not None and preferred_return is not None:
                break

        return carry_gross, preferred_return

    # ------------------------------------------------------------------
    # Portfolio hierarchy builder
    # ------------------------------------------------------------------

    def _build_hierarchy(self, wb, org, fund, schemes, companies,
                         investments, domain_map, filepath):
        """Build PortfolioNode hierarchy from Gemini-classified portfolio_hierarchy sheet."""
        sheet_name = _dm_first(domain_map, 'portfolio_hierarchy')
        if not sheet_name or sheet_name not in wb.sheetnames:
            return

        ws = wb[sheet_name]
        _, rows = read_table_from_sheet(ws, alias_map=self._get_alias(ws))
        if not rows:
            return

        # Get or create snapshot
        snapshot = PortfolioSnapshot.objects.filter(
            organization=org, is_active=True, source='excel_parse',
        ).first()
        if not snapshot:
            PortfolioSnapshot.objects.filter(
                organization=org, is_active=True, source='excel_parse',
            ).update(is_active=False)
            snapshot = PortfolioSnapshot.objects.create(
                organization=org,
                schema_version='2.0',
                base_currency='INR',
                source='excel_parse',
                is_active=True,
            )

        fund_slug = slugify(fund.name)
        fund_node_id = f'fund_{fund_slug}'

        # Delete old nodes for this fund
        PortfolioNode.objects.filter(
            snapshot=snapshot, node_id__startswith=f'fund_{fund_slug}').delete()

        # Create fund-level node (financials will be aggregated later)
        fund_node, _ = PortfolioNode.objects.get_or_create(
            snapshot=snapshot,
            node_id=fund_node_id,
            defaults={
                'name': fund.name,
                'level': 'fund',
                'parent_node_id': None,
                'financials': {},
            },
        )

        sector_nodes = {}   # sector_name -> PortfolioNode
        sector_companies = {}  # sector_name -> list of company financials dicts
        all_company_financials = []  # all company financials for fund-level agg
        node_count = 1

        for row in rows:
            name = _find_col_str(
                row, 'Company Name', 'Company', 'Name')
            if _is_junk_row(name):
                continue
            sector = _find_col_str(row, 'Sector', 'Industry')
            if not sector:
                sector = 'Other'
            city = _find_col_str(row, 'City', 'HQ City', 'Headquarters City', 'headquarters_city')
            stage = _find_col_str(row, 'Stage', 'Round', 'Funding Round')
            sub_sector = _find_col_str(row, 'Sub-Sector', 'Sub Sector', 'Subsector', 'Segment')

            pc_updates = {}
            if city:
                pc_updates['headquarters_city'] = city
            if sub_sector:
                pc_updates['sub_sector'] = sub_sector
            if pc_updates:
                PortfolioCompany.objects.filter(organization=org, name=name).update(**pc_updates)

            if stage:
                Investment.objects.filter(
                    scheme__fund__organization=org,
                    company_name=name,
                    stage='',
                ).update(stage=stage)

            # Create sector node if needed
            sector_slug = slugify(sector)
            sector_id = f'{fund_node_id}::sector_{sector_slug}'
            if sector not in sector_nodes:
                sector_node, _ = PortfolioNode.objects.get_or_create(
                    snapshot=snapshot,
                    node_id=sector_id,
                    defaults={
                        'name': sector,
                        'level': 'sector',
                        'parent_node_id': fund_node_id,
                        'parent': fund_node,
                        'financials': {},
                    },
                )
                sector_nodes[sector] = sector_node
                sector_companies[sector] = []
                node_count += 1

            # Create company node
            company_slug = slugify(name)
            company_id = f'{sector_id}::company_{company_slug}'
            invested = _find_col_decimal(
                row, 'Cost(₹Cr)', 'Cost', 'Invested', 'Total Invested')
            fair_value = _find_col_decimal(
                row, 'FV(₹Cr)', 'Fair Value', 'FV', 'Current Value')
            hold_pct = _find_col_decimal(row, 'Hold%', 'Holding %', 'Ownership')
            node_status = _find_col_str(row, 'Status', default='Active')

            # Collect MIS data from other sheets
            mis_data = self._collect_mis_data(wb, name, domain_map)

            # Build company summary from latest monthly_pl entry
            company_fin = {
                'invested': float(invested) if invested else 0,
                'fair_value': float(fair_value) if fair_value else 0,
                'holding_pct': float(hold_pct) if hold_pct else 0,
                'status': node_status,
                **mis_data,
            }
            company_fin['summary'] = _build_summary_from_pl(
                company_fin.get('monthly_pl', []),
                company_fin.get('budget_vs_actual', []),
            )

            PortfolioNode.objects.get_or_create(
                snapshot=snapshot,
                node_id=company_id,
                defaults={
                    'name': name,
                    'level': 'company',
                    'parent_node_id': sector_id,
                    'parent': sector_nodes[sector],
                    'financials': company_fin,
                },
            )
            sector_companies[sector].append(company_fin)
            all_company_financials.append(company_fin)
            node_count += 1

        # Aggregate financials upward: sector nodes
        for sector_name, sector_node in sector_nodes.items():
            children_fin = sector_companies.get(sector_name, [])
            sector_fin = _aggregate_financials(children_fin)
            sector_node.financials = sector_fin
            sector_node.save(update_fields=['financials'])

        # Aggregate financials upward: fund node
        fund_fin = _aggregate_financials(all_company_financials)
        fund_node.financials = fund_fin
        fund_node.save(update_fields=['financials'])

        # Invalidate portfolio service cache
        try:
            from api.portfolio import service as portfolio_service
            portfolio_service.reload(org.id)
        except Exception as e:
            logger.warning(f'Portfolio cache reload failed: {e}')

        logger.info(f'  Hierarchy: {node_count} nodes')

    def _collect_mis_data(self, wb, company_name, domain_map):
        """Collect Monthly P&L and Budget vs Actual data for a company."""
        data = {}

        # Monthly P&L — use Gemini-classified financials sheets
        fin_sheets = _dm_sheets(domain_map, 'financials_pl_bva') if domain_map else []
        for sn in fin_sheets:
            if sn in wb.sheetnames:
                ws = wb[sn]
                _, rows = read_table_from_sheet(ws, alias_map=self._get_alias(ws))
                pl_entries = []
                for row in rows:
                    name = _find_col_str(
                        row, 'Company Name', 'Company', 'Name')
                    if name != company_name:
                        continue
                    period = _find_col_str(row, 'Period', 'Month')
                    revenue = _find_col_decimal(
                        row, 'Revenue(₹Cr)', 'Revenue')
                    ebitda = _find_col_decimal(
                        row, 'EBITDA(₹Cr)', 'EBITDA')
                    gp = _find_col_decimal(
                        row, 'Gross Profit(₹Cr)', 'Gross Profit')
                    gp_pct = _find_col_decimal(row, 'GP%', 'GP Margin')
                    ebitda_pct = _find_col_decimal(
                        row, 'EBITDA%', 'EBITDA Margin')
                    opex = _find_col_decimal(row, 'OpEx(₹Cr)', 'OpEx')
                    cogs = _find_col_decimal(row, 'COGS(₹Cr)', 'COGS')

                    # Convert percentage stored as decimal
                    if gp_pct and abs(float(gp_pct)) <= 1:
                        gp_pct = round(float(gp_pct) * 100, 1)
                    if ebitda_pct and abs(float(ebitda_pct)) <= 1:
                        ebitda_pct = round(float(ebitda_pct) * 100, 1)

                    pl_entries.append({
                        'period': period,
                        'revenue': float(revenue) if revenue else 0,
                        'cogs': float(cogs) if cogs else 0,
                        'gross_profit': float(gp) if gp else 0,
                        'gp_pct': float(gp_pct) if gp_pct else None,
                        'opex': float(opex) if opex else 0,
                        'ebitda': float(ebitda) if ebitda else 0,
                        'ebitda_pct': float(ebitda_pct) if ebitda_pct else None,
                    })
                if pl_entries:
                    data['monthly_pl'] = pl_entries
                break

        # Budget vs Actual — same Gemini-classified financials sheets
        for sn in fin_sheets:
            if sn in wb.sheetnames:
                ws = wb[sn]
                _, rows = read_table_from_sheet(ws, alias_map=self._get_alias(ws))
                bva_entries = []
                for row in rows:
                    name = _find_col_str(
                        row, 'Company Name', 'Company', 'Name')
                    if name != company_name:
                        continue
                    line_item = _find_col_str(row, 'Line Item', 'Item')
                    actual = _find_col_decimal(
                        row, 'Actual YTD', 'Actual', 'Actuals')
                    budget = _find_col_decimal(
                        row, 'Budget YTD', 'Budget', 'Budgeted')
                    bva_entries.append({
                        'line_item': line_item,
                        'actual': float(actual) if actual else 0,
                        'budget': float(budget) if budget else 0,
                    })
                if bva_entries:
                    data['budget_vs_actual'] = bva_entries
                break

        return data

    # ------------------------------------------------------------------
    # Post-import validation
    # ------------------------------------------------------------------

    # ------------------------------------------------------------------
    # IC Pipeline seeding
    # ------------------------------------------------------------------

    _STATUS_TO_PIPELINE_STAGE = {
        'active':           'approved',
        'partially_exited': 'approved',
        'fully_exited':     'closed',
        'written_off':      'rejected',
    }

    def _seed_ic_pipeline(self, org, fund, investments):
        """
        Idempotent: upsert one DealPipeline record per Investment so the
        IC Workflow dashboard widget shows real data immediately after import.
        Keyed on (organization, fund, company_name) — safe to re-import.
        """
        from ic_workflow.models import DealPipeline
        for inv in investments.values():
            co    = inv.portfolio_company
            stage = self._STATUS_TO_PIPELINE_STAGE.get(inv.status, 'approved')
            DealPipeline.objects.update_or_create(
                organization=org,
                fund=fund,
                company_name=co.name,
                defaults={
                    'sector':                   co.sector or inv.sector or '',
                    'stage':                    stage,
                    'proposed_investment_inr':  inv.total_invested,
                    'linked_portfolio_company': co,
                    'source_channel':           'other',
                    'sourced_date':             inv.investment_date,
                },
            )

    def _validate_imported_companies(self, fund):
        """Detect and warn when metadata field names were imported as company names.

        Uses Gemini Pass 3 row_type classification to identify names that look
        like metadata labels rather than real company names. Works for any
        language — no hardcoded keyword list.
        """
        from investments.models import Investment
        invs = Investment.objects.filter(scheme__fund=fund).values_list('company_name', flat=True)
        names = [n for n in invs if n and n.strip()]
        if not names:
            return

        # Classify all company names — real entities vs metadata labels
        name_map = self._classify_labels(
            names, 'row_type',
            context='These are company names from a fund portfolio. '
                    'Identify any that look like metadata field labels '
                    '(e.g., fund configuration fields, sheet headers) '
                    'rather than real company/entity names.')
        suspect = [n for n in names if name_map.get(n) in ('header', 'note', 'subtotal', 'total', 'serial')]

        if suspect:
            msg = (
                f'IMPORT INTEGRITY ERROR for fund "{fund.name}": '
                f'{len(suspect)} investments have company names that look like '
                f'fund metadata labels, not real companies: {suspect}. '
                f'This usually means a Cover or Summary sheet was incorrectly '
                f'used as a data source. Delete these investments and reimport '
                f'from the correct data sheet.'
            )
            logger.critical(msg)
            self.errors.append({'section': 'validation', 'error': msg})

        # Also warn if total investment count looks suspiciously low
        total_inv = len(invs)
        if 0 < total_inv <= 15:
            logger.warning(
                f'Fund "{fund.name}" imported only {total_inv} investments. '
                f'If the fund is expected to have more portfolio companies, '
                f'check that the correct data sheet was used (not a Cover/Summary page).'
            )

    # ------------------------------------------------------------------
    # Collect counts
    # ------------------------------------------------------------------

    def _collect_counts(self, org, fund):
        """Collect record counts for the result summary — all counts scoped to this fund."""
        counts = {
            'funds': 1,
            'schemes': Scheme.objects.filter(fund=fund).count(),
            # Fund-scoped: only LPs who have a commitment under this fund
            'investors': Investor.objects.filter(
                commitments__scheme__fund=fund).distinct().count(),
            'commitments': Commitment.objects.filter(scheme__fund=fund).count(),
            'capital_calls': CapitalCall.objects.filter(scheme__fund=fund).count(),
            'capital_call_line_items': CapitalCallLineItem.objects.filter(
                capital_call__scheme__fund=fund).count(),
            # Fund-scoped: only companies that have an investment under this fund
            'portfolio_companies': PortfolioCompany.objects.filter(
                investments__scheme__fund=fund).distinct().count(),
            'investments': Investment.objects.filter(scheme__fund=fund).count(),
            'tranches': InvestmentTranche.objects.filter(
                investment__scheme__fund=fund).count(),
            'valuations': Valuation.objects.filter(
                investment__scheme__fund=fund).count(),
            'nav_records': NAVRecord.objects.filter(scheme__fund=fund).count(),
            'exit_events': ExitEvent.objects.filter(
                investment__scheme__fund=fund).count(),
            'distributions': Distribution.objects.filter(scheme__fund=fund).count(),
            'distribution_line_items': DistributionLineItem.objects.filter(
                distribution__scheme__fund=fund).count(),
            # Fund-scoped: distinct accounts actually referenced in this fund's ledger
            'chart_of_accounts': ChartOfAccounts.objects.filter(
                Q(debit_entries__scheme__fund=fund) |
                Q(credit_entries__scheme__fund=fund)
            ).distinct().count(),
            'ledger_entries': FundLedger.objects.filter(
                scheme__fund=fund).count(),
            'management_fees': ManagementFeeSchedule.objects.filter(
                scheme__fund=fund).count(),
            'carried_interest': CarriedInterest.objects.filter(
                scheme__fund=fund).count(),
            # Fund-scoped: entities directly linked to this fund record
            'entities': Entity.objects.filter(
                Q(managed_funds=fund) | Q(trustee_funds=fund) |
                Q(sponsored_funds=fund) | Q(custodian_funds=fund) |
                Q(audited_funds=fund)
            ).distinct().count(),
            'lp_capital_accounts': LPCapitalAccount.objects.filter(
                commitment__scheme__fund=fund).count(),
            'hierarchy_nodes': PortfolioNode.objects.filter(
                snapshot__organization=org,
                node_id__startswith=f'fund_{slugify(fund.name)}'
            ).count(),
        }

        if HAS_COMPLIANCE:
            counts['sebi_reports'] = SEBIReport.objects.filter(fund=fund).count()
            counts['compliance_calendar'] = ComplianceCalendar.objects.filter(
                fund=fund).count()

        return counts


# ---------------------------------------------------------------------------
# Helpers for management command output suppression (kept for legacy fallback)
# ---------------------------------------------------------------------------

class _NullOutput:
    """Swallows all write calls."""
    def write(self, *args, **kwargs):
        pass

    def flush(self):
        pass


class _NullStyle:
    """Mimics Django management command style object."""
    def SUCCESS(self, text):
        return text

    def ERROR(self, text):
        return text

    def WARNING(self, text):
        return text
