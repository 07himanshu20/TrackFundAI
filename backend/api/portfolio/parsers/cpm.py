"""
cpm.py
CPM / Chemopharm Sdn Bhd Group parser (File 2: 0625 CPM Group ECPM June 25.xlsx).
Group consolidated, multi-country (MY, SG, VN, TH, ID, PHP).
Native currency: MYR.
Period: YTD 6 months ended 30 June 2025 (current "snapshot" = Jun 2025).
Segments mapped from `Segment sales & gp` sheet: Analytical, JGS, O&G, Healthcare,
Life Science, Service, Medigene, MRI (and others per-country).

Strategy: use the cleaner `Country (YTD)` sheet for the headline P&L (it has
Actual + Budget columns side by side), and `Segment sales & gp` for the segment
breakdown. PL BS CF gives us monthly trend per segment (rows 10-19) + monthly REVENUE
total (row 22).
"""

import openpyxl
from datetime import datetime

from api.portfolio.schema import empty_financials


COMPANY_NAME = "Chemopharm Group (CPM)"
COMPANY_SLUG = "cpm"
CURRENCY = "MYR"

# Top-level segments we care about (from Segment sales & gp sheet)
TOP_SEGMENTS = ["Analytical", "O&G", "Healthcare", "Life Science", "Service", "Medigene", "MRI"]


def parse(filepath: str) -> dict:
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    fin = empty_financials()
    segment_children = []

    # ------------------------------------------------------------------
    # 1) Headline P&L from Country (YTD) sheet — includes BUDGET columns
    # ------------------------------------------------------------------
    if "Country (YTD)" in wb.sheetnames:
        ws = wb["Country (YTD)"]
        rows = list(ws.iter_rows(values_only=True))

        # Row 4: header "CY25 Actual" ... "CY 25 Budget" ...
        # Row 5: country columns under each
        # Total (Excl Hausen) is in column index 7, Total (Incl Hausen) in col 9 (TOTAL)
        # Budget Total is in col 17 ("Total" under "CY 25 Budget" section)

        def _get(label):
            """Find a row by label and return list of cell values."""
            for r in rows:
                if r[0] and isinstance(r[0], str) and r[0].strip() == label:
                    return r
            return None

        # Actual columns layout (observed): 1=MY 2=SG 3=VN 4=TH 5=ID 6=PHP 7=Total Excl Hausen 8=Hausen 9=TOTAL Inc Hausen
        # Budget columns: 10=MY 11=SG 12=VN 13=TH 14=ID 15=PHP 16=Hausen? 17=Total
        ACT_TOTAL = 9    # TOTAL inc Hausen (Actual YTD)
        BUD_TOTAL = 18   # Try col 18 first (Total Budget)

        def _val(row, idx):
            if row is None or idx >= len(row):
                return None
            v = row[idx]
            return v if isinstance(v, (int, float)) else None

        rev_row = _get("Revenue")
        cogs_row = _get("(-) COGS")
        gp_row = _get("Gross Profit")
        opex_row = _get("TOTAL OPEX")
        ebitda_row = _get("EBITDA")

        revenue = _val(rev_row, ACT_TOTAL)
        cogs = _val(cogs_row, ACT_TOTAL)
        if cogs is not None:
            cogs = abs(cogs)
        gp = _val(gp_row, ACT_TOTAL)
        opex = _val(opex_row, ACT_TOTAL)
        if opex is not None:
            opex = abs(opex)
        ebitda = _val(ebitda_row, ACT_TOTAL)

        bud_revenue = _val(rev_row, BUD_TOTAL)
        bud_ebitda = _val(ebitda_row, BUD_TOTAL)

        gp_pct = round(gp / revenue * 100, 2) if (gp is not None and revenue) else None

        fin["summary"] = {
            "period": "Jun 2025 YTD",
            "revenue": revenue,
            "cogs": cogs,
            "gross_profit": gp,
            "gp_pct": gp_pct,
            "opex": opex,
            "ebitda": ebitda,
            "ytd_revenue": revenue,
            "ytd_gross_profit": gp,
            "ytd_ebitda": ebitda,
            "ytd_budget_revenue": bud_revenue,
            "ytd_budget_ebitda": bud_ebitda,
        }

        # Cost structure
        if revenue:
            fin["cost_structure"] = {
                "cogs_pct": round((cogs or 0) / revenue * 100, 2),
                "gp_pct": gp_pct,
                "opex_pct": round((opex or 0) / revenue * 100, 2),
                "ebitda_pct": round((ebitda or 0) / revenue * 100, 2),
            }

        # Budget vs Actual rows (YTD)
        for li_key, li_label, act, bud in [
            ("revenue", "Revenue", revenue, bud_revenue),
            ("ebitda", "EBITDA", ebitda, bud_ebitda),
        ]:
            if act is None and bud is None:
                continue
            var = (act - bud) if (act is not None and bud is not None) else None
            var_pct = round(var / bud * 100, 2) if (var is not None and bud) else None
            fin["budget_vs_actual"].append({
                "period": "Jun 2025 YTD",
                "line_item": li_label,
                "budget": bud,
                "actual": act,
                "variance": var,
                "variance_pct": var_pct,
            })

        # Sales by geography (countries)
        country_cols = [
            ("Malaysia", 1), ("Singapore", 2), ("Vietnam", 3),
            ("Thailand", 4), ("Indonesia", 5), ("Philippines", 6),
        ]
        for name, idx in country_cols:
            c_rev = _val(rev_row, idx) if rev_row else None
            c_gp = _val(gp_row, idx) if gp_row else None
            c_gp_pct = round(c_gp / c_rev * 100, 2) if (c_gp is not None and c_rev) else None
            if c_rev is not None:
                fin["sales_by_geo"].append({
                    "label": name,
                    "revenue": c_rev,
                    "gross_margin": c_gp,
                    "gm_pct": c_gp_pct,
                })

    # ------------------------------------------------------------------
    # 2) Monthly P&L trend from PL BS CF sheet
    #    Row 22 = total REVENUE; need to find COGS / GP / OPEX / EBITDA rows.
    # ------------------------------------------------------------------
    if "PL BS CF" in wb.sheetnames:
        ws = wb["PL BS CF"]
        rows = list(ws.iter_rows(values_only=True))

        # Row 6 has month labels: January, February, March, April, May, June
        # Each month spans 2 cols (value + blank or Inc/Excl); MYR col is at indices 3, 5, 7, 9, 11, 13
        # Verify by reading row 6:
        month_header = rows[6] if len(rows) > 6 else []
        month_indices = []  # list of (month_name, col_idx)
        for j, cell in enumerate(month_header):
            if isinstance(cell, str) and cell.strip() in ("January", "February", "March", "April", "May", "June",
                                                            "July", "August", "September", "October", "November", "December"):
                month_indices.append((cell.strip(), j))

        # Find KPI rows by label — labels can be in col 0 OR col 1 (this sheet uses col 1)
        def _find_row(*labels):
            wanted = {l.upper() for l in labels}
            for r in rows:
                if not r:
                    continue
                for col_idx in (0, 1):
                    if col_idx < len(r) and isinstance(r[col_idx], str) and r[col_idx].strip().upper() in wanted:
                        return r
            return None

        rev_r = _find_row("REVENUE")
        cogs_r = _find_row("COGS", "(-) COGS", "(-) Cogs", "Total COGS")
        gp_r = _find_row("Gross Profit", "GP")
        opex_r = _find_row("TOTAL OPEX", "OPEX", "Total Operating Expenses")
        ebitda_r = _find_row("EBITDA")

        month_num_map = {
            "January": 1, "February": 2, "March": 3, "April": 4, "May": 5, "June": 6,
            "July": 7, "August": 8, "September": 9, "October": 10, "November": 11, "December": 12,
        }

        for month_name, col in month_indices:
            mn = month_num_map[month_name]
            period = f"2025-{mn:02d}"

            def _v(r):
                if r is None or col >= len(r):
                    return None
                v = r[col]
                return v if isinstance(v, (int, float)) else None

            rev = _v(rev_r)
            cogs = _v(cogs_r)
            if cogs is not None:
                cogs = abs(cogs)
            gp = _v(gp_r)
            opex = _v(opex_r)
            if opex is not None:
                opex = abs(opex)
            ebitda = _v(ebitda_r)

            # Skip months with no actual data (e.g. Jul-Dec when reporting through Jun)
            if not rev or rev == 0:
                continue

            gp_pct = round(gp / rev * 100, 2) if (gp is not None and rev) else None
            ebitda_pct = round(ebitda / rev * 100, 2) if (ebitda is not None and rev) else None

            fin["monthly_pl"].append({
                "period": period,
                "revenue": rev,
                "cogs": cogs,
                "gross_profit": gp,
                "gp_pct": gp_pct,
                "opex": opex,
                "ebitda": ebitda,
                "ebitda_pct": ebitda_pct,
            })

    # ------------------------------------------------------------------
    # 3) Segment breakdown from `Segment sales & gp`
    #    Rows are nested: country rows (Malaysia, Singapore, Vietnam, ...) followed
    #    by segment rows (Analytical, O&G, Healthcare, Life Science, Service, Medigene, MRI).
    #    We aggregate across countries to get group-level segment totals.
    # ------------------------------------------------------------------
    if "Segment sales & gp" in wb.sheetnames:
        ws = wb["Segment sales & gp"]
        rows = list(ws.iter_rows(values_only=True))

        # Cols (observed): 2=Label, 3=CY25A Revenue, 6=CY25A GP
        REV_COL = 3
        GP_COL = 6

        seg_totals = {}  # segment_name -> {revenue, gp}

        for r in rows:
            if not r or len(r) < 7:
                continue
            label = r[2] if r[2] else None
            if not isinstance(label, str):
                continue
            label = label.strip()
            if label in TOP_SEGMENTS:
                rev = r[REV_COL] if isinstance(r[REV_COL], (int, float)) else None
                gp = r[GP_COL] if isinstance(r[GP_COL], (int, float)) else None
                if rev is None:
                    continue
                d = seg_totals.setdefault(label, {"revenue": 0, "gp": 0})
                d["revenue"] += rev
                if gp is not None:
                    d["gp"] += gp

        for seg_name, vals in seg_totals.items():
            rev = vals["revenue"]
            gp = vals["gp"]
            gp_pct = round(gp / rev * 100, 2) if rev else None

            fin["sales_by_segment"].append({
                "label": seg_name,
                "revenue": rev,
                "gross_margin": gp,
                "gm_pct": gp_pct,
            })

            seg_fin = empty_financials()
            seg_fin["summary"] = {
                "period": "Jun 2025 YTD",
                "revenue": rev,
                "gross_profit": gp,
                "gp_pct": gp_pct,
            }

            slug = (seg_name.lower()
                    .replace(" ", "_")
                    .replace("&", "and")
                    .replace("/", "_"))
            segment_children.append({
                "id_slug": slug,
                "name": seg_name,
                "financials": seg_fin,
            })

    # ------------------------------------------------------------------
    # 4) Cash flow — pull from Conso or PL BS CF if a CF block exists.
    #    For now, derive a simple closing-cash trend from monthly REVENUE
    #    minus a proxy operating cost (we don't have a clean CF sheet here).
    #    Skip for v1; the cash flow waterfall comparison will rely on Analisa
    #    + mock data. We can revisit if the user needs CPM CF.
    # ------------------------------------------------------------------

    wb.close()

    return {
        "company_meta": {
            "name": COMPANY_NAME,
            "slug": COMPANY_SLUG,
            "currency": CURRENCY,
            "report_month": "Jun 2025 YTD",
            "description": "Multi-country (MY, SG, VN, TH, ID, PH) life-science/healthcare distribution group. 6 segments + Hausen sub-entity.",
        },
        "financials": fin,
        "segments": segment_children,
    }
