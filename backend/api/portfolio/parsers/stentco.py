"""
stentco.py
Stent-Co (Translumina-style cardiology medical-device distributor) parser.
Files 6+7: Sale Report_Board_May-25.xlsx, Sale Report_Team_May-25.xlsx (identical structure).
Native currency: EUR.
3 categories (DES, DCB, PTCA) × ~15 brands.

Strategy: parse Summary_Brand-YTD for the YTD revenue+GM by brand, and
roll up by category (DES/DCB/PTCA) for the segment level. Use Summary_Brand
for current-period (July-25) actual vs AOP comparison.
"""

import openpyxl
from api.portfolio.schema import empty_financials


CURRENCY = "EUR"
SCALE = 1_000   # values in column "Revenue €'000" — scale to full EUR

# Brand → category map (observed in Summary_Brand-YTD)
BRAND_CATEGORY = {
    "Chrome": "DES", "Choice PC": "DES", "Vivo ISAR": "DES", "Flex": "DES",
    "Racer": "DES", "Ultima PC": "DES", "BMS-CC": "DES", "Isar Summit": "DES",
    "Protégé": "DCB",
    "Bentley": "PTCA", "BMD Balloon": "PTCA", "Cape Cross": "PTCA",
    "Boosting Catheter": "PTCA", "Cathy": "PTCA", "Yukon/Optima": "PTCA",
    "Others": "PTCA",
}

CATEGORY_TOTALS = {"DES", "DCB", "PTCA", "Grand Total"}


def parse(filepath: str, view: str = "board") -> dict:
    """
    Parse a Stent-Co sales workbook.
    view: "board" or "team" — used to set the company name/slug.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    fin = empty_financials()
    segment_children = []

    if view == "board":
        company_name = "Stent-Co (Board View)"
        company_slug = "stentco_board"
    else:
        company_name = "Stent-Co (Team View)"
        company_slug = "stentco_team"

    if "Summary_Brand-YTD" not in wb.sheetnames:
        wb.close()
        return {
            "company_meta": {"name": company_name, "slug": company_slug, "currency": CURRENCY},
            "financials": fin,
            "segments": [],
        }

    # ----------------------------------------------------------------
    # Summary_Brand-YTD layout (observed):
    #   col 2 = Units (Actual YTD)
    #   col 3 = ASP €  (Actual YTD)
    #   col 4 = Revenue €'000 (Actual YTD)
    #   col 5 = GM €'000 (Actual YTD)
    #   col 7 = Brand name
    #   col 9 = Units (AOP)
    #   col 11 = Revenue €'000 (AOP)
    #   col 12 = GM €'000 (AOP)
    # ----------------------------------------------------------------
    ws = wb["Summary_Brand-YTD"]
    rows = list(ws.iter_rows(values_only=True))

    BRAND_COL = 7
    REV_ACT_COL = 4
    GM_ACT_COL = 5
    REV_AOP_COL = 11
    GM_AOP_COL = 12

    by_category = {"DES": {"rev": 0, "gm": 0, "brands": []},
                   "DCB": {"rev": 0, "gm": 0, "brands": []},
                   "PTCA": {"rev": 0, "gm": 0, "brands": []}}
    grand_rev = 0
    grand_gm = 0
    grand_rev_aop = 0
    grand_gm_aop = 0

    for r in rows:
        if not r or len(r) <= GM_AOP_COL:
            continue
        brand = r[BRAND_COL]
        if not isinstance(brand, str):
            continue
        brand = brand.strip()

        if brand == "Grand Total":
            grand_rev = (r[REV_ACT_COL] or 0) * SCALE
            grand_gm = (r[GM_ACT_COL] or 0) * SCALE
            grand_rev_aop = (r[REV_AOP_COL] or 0) * SCALE
            grand_gm_aop = (r[GM_AOP_COL] or 0) * SCALE
            continue

        if brand in CATEGORY_TOTALS:
            continue

        cat = BRAND_CATEGORY.get(brand)
        if not cat:
            continue

        b_rev = (r[REV_ACT_COL] or 0) * SCALE
        b_gm = (r[GM_ACT_COL] or 0) * SCALE

        by_category[cat]["rev"] += b_rev
        by_category[cat]["gm"] += b_gm
        by_category[cat]["brands"].append({
            "label": brand,
            "revenue": b_rev,
            "gross_margin": b_gm,
            "gm_pct": round(b_gm / b_rev * 100, 2) if b_rev else None,
        })

    # Top-level summary
    fin["summary"] = {
        "period": "YTD July 2025",
        "revenue": grand_rev,
        "gross_profit": grand_gm,
        "gp_pct": round(grand_gm / grand_rev * 100, 2) if grand_rev else None,
        "ytd_revenue": grand_rev,
        "ytd_gross_profit": grand_gm,
        "ytd_budget_revenue": grand_rev_aop,
    }

    # Cost structure
    if grand_rev:
        fin["cost_structure"] = {
            "cogs_pct": round((grand_rev - grand_gm) / grand_rev * 100, 2),
            "gp_pct": round(grand_gm / grand_rev * 100, 2),
        }

    # Budget vs Actual (YTD)
    var = grand_rev - grand_rev_aop if grand_rev_aop else None
    var_pct = round(var / grand_rev_aop * 100, 2) if (var is not None and grand_rev_aop) else None
    fin["budget_vs_actual"].append({
        "period": "YTD July 2025",
        "line_item": "Revenue",
        "budget": grand_rev_aop,
        "actual": grand_rev,
        "variance": var,
        "variance_pct": var_pct,
    })

    # Sales-by-segment = sales by category (DES/DCB/PTCA)
    for cat, d in by_category.items():
        if d["rev"] == 0:
            continue
        gm_pct = round(d["gm"] / d["rev"] * 100, 2) if d["rev"] else None
        fin["sales_by_segment"].append({
            "label": cat,
            "revenue": d["rev"],
            "gross_margin": d["gm"],
            "gm_pct": gm_pct,
        })

        # Build a segment child with brands as sub-entries
        seg_fin = empty_financials()
        seg_fin["summary"] = {
            "period": "YTD July 2025",
            "revenue": d["rev"],
            "gross_profit": d["gm"],
            "gp_pct": gm_pct,
        }
        seg_fin["sales_by_segment"] = d["brands"]   # brand-level breakdown
        segment_children.append({
            "id_slug": cat.lower(),
            "name": cat,
            "financials": seg_fin,
        })

    wb.close()

    return {
        "company_meta": {
            "name": company_name,
            "slug": company_slug,
            "currency": CURRENCY,
            "report_month": "YTD July 2025",
            "description": "Cardiology medical-device distributor (DES/DCB/PTCA stents & balloons), global sales (~30 countries).",
        },
        "financials": fin,
        "segments": segment_children,
    }
