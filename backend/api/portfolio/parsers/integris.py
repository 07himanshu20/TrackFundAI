"""
integris.py
Integris EL Group parser (File 5: Integris EL MIS - EL May'25.xlsx).
Native currency: USD ($M — millions).
Multi-country (MY, SG, IN, PH, TH, ID/VN) × 5 segments
(Clinical Diagnostics, Analytical, Lifescience, Service, Covid).

The Finance sheet has Actual / AOP / Prior-year columns side by side for both
MTD (May 2025) and YTD FY26.
"""

import openpyxl
from api.portfolio.schema import empty_financials


COMPANY_NAME = "Integris EL Group"
COMPANY_SLUG = "integris"
CURRENCY = "USD"
# All numbers on the Finance sheet are in $M — multiply to get full USD
SCALE = 1_000_000


def parse(filepath: str) -> dict:
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    fin = empty_financials()
    segment_children = []

    if "Finance" not in wb.sheetnames:
        wb.close()
        return {
            "company_meta": {"name": COMPANY_NAME, "slug": COMPANY_SLUG, "currency": CURRENCY},
            "financials": fin,
            "segments": [],
        }

    ws = wb["Finance"]
    rows = list(ws.iter_rows(values_only=True))

    # Column map (observed in dump):
    # col 1 = Actual MTD May'25
    # col 2 = AOP MTD May'25
    # col 3 = Prior MTD May'24
    # col 7 = Actual YTD FY26
    # col 8 = AOP YTD FY26
    # col 9 = Prior YTD FY25
    # col 12 = FY26 AOP (full year)

    def _row_by_label(*labels):
        wanted = {l.upper().strip() for l in labels}
        for r in rows:
            if r and isinstance(r[0], str) and r[0].strip().upper() in wanted:
                return r
        return None

    def _val(r, idx):
        if r is None or idx >= len(r):
            return None
        v = r[idx]
        return v if isinstance(v, (int, float)) else None

    def _scaled(r, idx):
        v = _val(r, idx)
        return v * SCALE if v is not None else None

    rev_r = _row_by_label("Total Revenue")
    gm_r = _row_by_label("Total GM")
    gm_pct_r = _row_by_label("GM%")
    sga_r = _row_by_label("SG&A")
    ebitda_r = _row_by_label("EBITDA")
    ebitda_pct_r = _row_by_label("EBITDA %")

    revenue = _scaled(rev_r, 1)        # MTD Actual
    gm = _scaled(gm_r, 1)
    sga = _scaled(sga_r, 1)
    ebitda = _scaled(ebitda_r, 1)

    rev_ytd = _scaled(rev_r, 7)        # YTD FY26 Actual
    gm_ytd = _scaled(gm_r, 7)
    ebitda_ytd = _scaled(ebitda_r, 7)

    bud_rev_mtd = _scaled(rev_r, 2)
    bud_ebitda_mtd = _scaled(ebitda_r, 2)
    bud_rev_ytd = _scaled(rev_r, 8)
    bud_ebitda_ytd = _scaled(ebitda_r, 8)

    gm_pct = (_val(gm_pct_r, 1) or 0) * 100  # already a fraction

    fin["summary"] = {
        "period": "May 2025",
        "revenue": revenue,
        "gross_profit": gm,
        "gp_pct": round(gm_pct, 2) if gm_pct else None,
        "opex": abs(sga) if sga else None,
        "ebitda": ebitda,
        "ytd_revenue": rev_ytd,
        "ytd_gross_profit": gm_ytd,
        "ytd_ebitda": ebitda_ytd,
        "budget_revenue": bud_rev_mtd,
        "budget_ebitda": bud_ebitda_mtd,
        "ytd_budget_revenue": bud_rev_ytd,
        "ytd_budget_ebitda": bud_ebitda_ytd,
    }

    # Cost structure (using YTD)
    if rev_ytd:
        cogs_ytd = (rev_ytd - gm_ytd) if gm_ytd is not None else None
        opex_pct = abs(sga) / revenue * 100 if (sga and revenue) else None
        fin["cost_structure"] = {
            "cogs_pct": round(cogs_ytd / rev_ytd * 100, 2) if cogs_ytd is not None else None,
            "gp_pct": round(gm_ytd / rev_ytd * 100, 2) if gm_ytd is not None else None,
            "opex_pct": round(opex_pct, 2) if opex_pct else None,
            "ebitda_pct": round(ebitda_ytd / rev_ytd * 100, 2) if ebitda_ytd is not None else None,
        }

    # Budget vs Actual (YTD)
    for label, act, bud in [
        ("Revenue", rev_ytd, bud_rev_ytd),
        ("EBITDA", ebitda_ytd, bud_ebitda_ytd),
    ]:
        if act is None and bud is None:
            continue
        var = (act - bud) if (act is not None and bud is not None) else None
        var_pct = round(var / bud * 100, 2) if (var is not None and bud) else None
        fin["budget_vs_actual"].append({
            "period": "YTD FY26 (through May 2025)",
            "line_item": label,
            "budget": bud,
            "actual": act,
            "variance": var,
            "variance_pct": var_pct,
        })

    # Country breakdown (rows 16-21 are countries)
    country_labels = ["Malaysia", "Singapore", "India", "Philippines", "Thailand", "Others (ID & VN)"]
    for cname in country_labels:
        r = _row_by_label(cname)
        if r is None:
            continue
        c_rev = _scaled(r, 7)  # YTD
        if c_rev is None:
            continue
        # GM% is in the "Country Margin %" block — find it
        # Margin row label is the same name; we'll find the second occurrence
        margin_pct = None
        seen_first = False
        for rr in rows:
            if rr and isinstance(rr[0], str) and rr[0].strip() == cname:
                if seen_first:
                    margin_pct = _val(rr, 7)  # YTD margin
                    if margin_pct is not None:
                        margin_pct = round(margin_pct * 100, 2)
                    break
                else:
                    seen_first = True

        c_gm = c_rev * (margin_pct / 100) if margin_pct else None
        fin["sales_by_geo"].append({
            "label": cname,
            "revenue": c_rev,
            "gross_margin": c_gm,
            "gm_pct": margin_pct,
        })

    # Segment breakdown (rows 25-29: Clinical Diagnostics, Analytical, Lifescience, Service, Covid)
    segment_labels = ["Clinical Diagnostics", "Analytical", "Lifescience", "Service", "Covid"]
    for sname in segment_labels:
        r = _row_by_label(sname)
        if r is None:
            continue
        s_rev = _scaled(r, 7)  # YTD
        if s_rev is None:
            continue

        # Find segment margin %
        margin_pct = None
        seen_first = False
        for rr in rows:
            if rr and isinstance(rr[0], str) and rr[0].strip() == sname:
                if seen_first:
                    margin_pct = _val(rr, 7)
                    if margin_pct is not None:
                        margin_pct = round(margin_pct * 100, 2)
                    break
                else:
                    seen_first = True

        s_gm = s_rev * (margin_pct / 100) if margin_pct else None

        fin["sales_by_segment"].append({
            "label": sname,
            "revenue": s_rev,
            "gross_margin": s_gm,
            "gm_pct": margin_pct,
        })

        seg_fin = empty_financials()
        seg_fin["summary"] = {
            "period": "YTD FY26",
            "revenue": s_rev,
            "gross_profit": s_gm,
            "gp_pct": margin_pct,
        }
        slug = sname.lower().replace(" ", "_")
        segment_children.append({
            "id_slug": slug,
            "name": sname,
            "financials": seg_fin,
        })

    wb.close()

    return {
        "company_meta": {
            "name": COMPANY_NAME,
            "slug": COMPANY_SLUG,
            "currency": CURRENCY,
            "report_month": "May 2025 (FY26)",
            "description": "Multi-country (MY/SG/IN/PH/TH/ID-VN) clinical diagnostics & life-science distribution group.",
        },
        "financials": fin,
        "segments": segment_children,
    }
