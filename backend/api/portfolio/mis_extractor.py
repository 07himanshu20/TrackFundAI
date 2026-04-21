"""
mis_extractor.py
================
Converts any Excel MIS workbook into a compact, token-efficient JSON
representation that Gemini can read and semantically analyse.

Design goals:
  - Preserve ALL cell content — no hardcoded sheet/row/column assumptions
  - Keep size manageable: skip fully-empty rows, truncate sheets >500 rows
  - Represent dates as ISO strings, keep numbers as numbers
  - Flag metadata: sheet names, row count, any datetime headers found

Output shape per sheet:
  {
    "name": "05 Summary P&L (2)",
    "nrows": 42,
    "ncols": 12,
    "rows": [
      {"r": 0, "cells": [{"c": 0, "v": "In MYR '000"}, {"c": 3, "v": "Actual"}, ...]},
      ...
    ],
    "datetime_header_rows": [4],   # row indices that contain >=2 datetime values
    "has_budget_keyword": true,
    "has_revenue_keyword": true
  }
"""

from __future__ import annotations

import json
import logging
from datetime import datetime, date
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

# Max rows per sheet sent to Gemini (prevents token overflow on huge sheets)
MAX_ROWS_PER_SHEET = 500
# Max sheets to extract (skip obvious non-financial tabs)
MAX_SHEETS = 60

# Keywords that indicate financial relevance (used to rank/filter sheets)
FINANCE_KEYWORDS = {
    "revenue", "sales", "profit", "loss", "ebitda", "ebit", "cogs", "cost",
    "opex", "budget", "actual", "forecast", "p&l", "pl", "income", "expense",
    "cash", "flow", "balance", "sheet", "working", "capital", "dso", "margin",
    "gp", "gross", "net", "ytd", "mtd", "q1", "q2", "q3", "q4",
}

NON_FINANCIAL_KEYWORDS = {
    "cover", "index", "contents", "chart", "graph", "image", "photo",
    "instructions", "readme", "template", "waterfall_chart",
}


def _serialize_cell(value: Any) -> Any:
    """Convert a cell value to a JSON-serializable form."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        if value != value:  # NaN check
            return None
        return value
    if isinstance(value, datetime):
        return {"__dt": value.strftime("%Y-%m-%dT%H:%M:%S")}
    if isinstance(value, date):
        return {"__dt": value.strftime("%Y-%m-%d")}
    if isinstance(value, str):
        s = value.strip()
        # Normalize error strings
        if s in ("#DIV/0!", "#REF!", "#N/A", "#VALUE!", "#NAME?", "#NULL!", ""):
            return None
        return s
    return str(value)


def _is_financial_sheet(name: str) -> bool:
    """Heuristic: should we include this sheet?"""
    n = name.lower().replace(" ", "").replace("_", "").replace("-", "")
    if any(kw in n for kw in NON_FINANCIAL_KEYWORDS):
        return False
    # Include everything by default — Gemini will decide relevance
    return True


def extract_workbook(filepath: str) -> dict:
    """
    Open an Excel workbook and return a compact dict of all sheets.

    Returns:
    {
        "filepath": str,
        "sheet_names": [str, ...],
        "sheets": [SheetData, ...]   # ordered by relevance
    }
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        raise RuntimeError(f"Cannot open workbook '{filepath}': {e}") from e

    all_sheet_names = wb.sheetnames
    sheets = []

    for sheet_name in all_sheet_names[:MAX_SHEETS]:
        if not _is_financial_sheet(sheet_name):
            logger.debug("Skipping non-financial sheet: %s", sheet_name)
            continue
        try:
            sheet_data = _extract_sheet(wb[sheet_name], sheet_name)
            if sheet_data["nrows"] == 0:
                continue  # completely empty
            sheets.append(sheet_data)
        except Exception as e:
            logger.warning("Error extracting sheet '%s': %s", sheet_name, e)

    wb.close()

    # Sort: sheets with more finance keywords first
    sheets.sort(key=lambda s: s["finance_score"], reverse=True)

    return {
        "filepath": filepath,
        "sheet_names": all_sheet_names,
        "sheets": sheets,
    }


def _extract_sheet(ws, name: str) -> dict:
    """Extract a single worksheet into a compact row/cell structure."""
    rows_out = []
    datetime_header_rows = []
    finance_score = 0
    has_budget = False
    has_revenue = False
    has_monthly_dates = False

    row_idx = 0
    for row in ws.iter_rows(values_only=True):
        if row_idx >= MAX_ROWS_PER_SHEET:
            break

        # Skip entirely empty rows
        if not any(c is not None for c in row):
            row_idx += 1
            continue

        cells = []
        dt_count = 0
        for col_idx, raw_val in enumerate(row):
            sv = _serialize_cell(raw_val)
            if sv is None:
                continue
            cells.append({"c": col_idx, "v": sv})

            # Heuristic checks
            if isinstance(sv, dict) and "__dt" in sv:
                dt_count += 1
            if isinstance(sv, str):
                sl = sv.lower()
                for kw in FINANCE_KEYWORDS:
                    if kw in sl:
                        finance_score += 1
                if "budget" in sl or "aop" in sl or "forecast" in sl:
                    has_budget = True
                if "revenue" in sl or "sales" in sl or "turnover" in sl or "income" in sl:
                    has_revenue = True

        if cells:
            rows_out.append({"r": row_idx, "cells": cells})

        if dt_count >= 2:
            datetime_header_rows.append(row_idx)
            has_monthly_dates = True

        row_idx += 1

    # Estimate ncols from max col index seen
    max_col = 0
    for r in rows_out:
        for c in r["cells"]:
            if c["c"] > max_col:
                max_col = c["c"]

    return {
        "name": name,
        "nrows": len(rows_out),
        "ncols": max_col + 1,
        "rows": rows_out,
        "datetime_header_rows": datetime_header_rows,
        "has_budget_keyword": has_budget,
        "has_revenue_keyword": has_revenue,
        "has_monthly_dates": has_monthly_dates,
        "finance_score": finance_score,
    }


def extract_to_json(filepath: str) -> str:
    """Extract workbook and return compact JSON string."""
    data = extract_workbook(filepath)
    return json.dumps(data, ensure_ascii=False, separators=(",", ":"))


def extract_summary_for_gemini(filepath: str, max_tokens_approx: int = 180_000) -> dict:
    """
    Extract workbook and apply token-budget management.
    Gemini 2.5 Flash has a 1M token context — we use up to ~180K tokens
    (~720K chars) to stay well within limits while capturing all sheets.

    Strategy when over budget:
      1. Truncate rows in lowest-scored sheets first (keep headers + totals)
      2. Then drop lowest-scored sheets entirely
    Returns the dict (not JSON string) — caller decides serialization.
    """
    data = extract_workbook(filepath)
    sheets = data["sheets"]

    # Rough token estimate: 1 char ≈ 0.25 tokens
    def _size(obj) -> int:
        return len(json.dumps(obj, ensure_ascii=False, separators=(",", ":")))

    char_budget = max_tokens_approx * 4  # 4 chars per token approx

    if _size(data) <= char_budget:
        return data

    # Step 1: For sheets with many rows, keep first 40 rows (headers + structure)
    # and last 15 rows (totals), drop the middle
    for sheet in reversed(sheets):  # start with lowest-scored
        if _size(data) <= char_budget:
            break
        rows = sheet["rows"]
        if len(rows) > 60:
            sheet["rows"] = rows[:40] + rows[-15:]

    if _size(data) <= char_budget:
        return data

    # Step 2: Drop lowest-scored sheets entirely
    while len(sheets) > 5 and _size(data) > char_budget:
        sheets.pop()  # removes lowest-scored (already sorted desc)
        data["sheets"] = sheets

    # Step 3: Harder truncation on remaining sheets
    for sheet in reversed(sheets):
        if _size(data) <= char_budget:
            break
        rows = sheet["rows"]
        if len(rows) > 30:
            sheet["rows"] = rows[:25] + rows[-10:]

    return data
