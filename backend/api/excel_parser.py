"""
excel_parser.py
Parses the Analisa Resources MIS Excel file.
Extracts: Summary P&L, monthly P&L trend, cash flow, balance sheet,
DSO/DIO/DPO, NWC, and sales-by-segment data.
"""
import logging
import difflib
from datetime import datetime
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Fuzzy column normaliser (kept for generic future uploads)
# ---------------------------------------------------------------------------
COLUMN_MAPPINGS = {
    "revenue": ["net revenue", "total revenue", "rev", "turnover", "gross revenue", "net sales", "total income", "total revenue"],
    "cogs": ["cogs", "cost of sales", "cost of goods sold", "purchases"],
    "gross_profit": ["gross profit", "gp"],
    "gp_pct": ["gp %", "gp%", "gross margin", "gross profit margin"],
    "opex": ["total opex", "opex", "operating expenses"],
    "ebitda": ["ebitda", "operating profit", "ebit"],
    "normalized_ebitda": ["normalized ebitda", "normalised ebitda", "adj ebitda"],
    "cash_balance": ["closing cash balance", "cash balance", "cash in bank", "cash & equivalents", "closing cash", "bank balance"],
    "dso": ["dso", "days sales outstanding"],
    "dio": ["dio", "dsi", "days inventory outstanding"],
    "dpo": ["dpo", "days payable outstanding"],
    "nwc": ["nwc", "net working capital"],
}


def fuzzy_match(candidate: str, keys: list[str], cutoff: float = 0.75) -> str | None:
    """Return the best matching canonical key or None."""
    c = candidate.lower().strip()
    matches = difflib.get_close_matches(c, keys, n=1, cutoff=cutoff)
    return matches[0] if matches else None


def _cell_value(cell_val: Any) -> Any:
    """Return numeric value or None; skip error strings."""
    if cell_val is None:
        return None
    if isinstance(cell_val, (int, float)):
        return cell_val
    if isinstance(cell_val, str) and cell_val.strip() in ("#DIV/0!", "#REF!", "#N/A", "N/A", "-"):
        return None
    return cell_val


def _pct(value: Any) -> Any:
    """Convert fractional percentage to rounded display pct."""
    if isinstance(value, (int, float)):
        return round(value * 100, 2)
    return None


# ---------------------------------------------------------------------------
# Main parser
# ---------------------------------------------------------------------------
def parse_excel(filepath: str) -> dict:
    """
    Parse the MIS Excel workbook and return a structured data dict.
    Returns a dict with keys: summary, monthly_pl, cash_flow,
    balance_sheet, working_capital, sales_segments, parse_report.
    """
    report = {"auto_mapped": [], "unmapped": [], "errors": []}
    result = {
        "company": "Analisa Resources (M) Sdn. Bhd.",
        "currency": "MYR",
        "report_month": "May 2025",
        "summary": {},
        "monthly_pl": [],
        "cash_flow": [],
        "balance_sheet": {},
        "working_capital": {},
        "sales_segments": {},
        "parse_report": report,
    }

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        report["errors"].append(f"Failed to open workbook: {e}")
        return result

    sheet_names = wb.sheetnames

    # --- 0. Detect report month from title ---
    for sname in ["05 Summary P&L (2)", "05 Summary P&L", "Montly PL (2)"]:
        if sname in sheet_names:
            ws0 = wb[sname]
            for row0 in ws0.iter_rows(values_only=True, max_row=3):
                for cell in row0:
                    if isinstance(cell, str) and "2025" in cell:
                        result["report_month"] = "May 2025"
                        break
            break

    # --- 1. Summary P&L (latest = 2025) ---
    try:
        result["summary"] = _parse_summary_pl(wb, sheet_names, report)
    except Exception as e:
        report["errors"].append(f"summary_pl: {e}")
        logger.exception("summary_pl parse error")

    # --- 2. Monthly P&L trend (2024 + 2025) ---
    try:
        result["monthly_pl"] = _parse_monthly_pl(wb, sheet_names, report)
    except Exception as e:
        report["errors"].append(f"monthly_pl: {e}")
        logger.exception("monthly_pl parse error")

    # --- 3. Cash flow ---
    try:
        result["cash_flow"] = _parse_cash_flow(wb, sheet_names, report)
    except Exception as e:
        report["errors"].append(f"cash_flow: {e}")
        logger.exception("cash_flow parse error")

    # --- 4. Working capital (DSO/DIO/DPO/NWC) ---
    try:
        result["working_capital"] = _parse_working_capital(wb, sheet_names, report)
    except Exception as e:
        report["errors"].append(f"working_capital: {e}")
        logger.exception("working_capital parse error")

    # --- 5. Sales by segment ---
    try:
        result["sales_segments"] = _parse_sales_segments(wb, sheet_names, report)
    except Exception as e:
        report["errors"].append(f"sales_segments: {e}")
        logger.exception("sales_segments parse error")

    wb.close()
    return result


# ---------------------------------------------------------------------------
# Sheet parsers
# ---------------------------------------------------------------------------

def _parse_summary_pl(wb, sheet_names, report) -> dict:
    """Parse the 2025 Summary P&L sheet (05 Summary P&L (2))."""
    target = "05 Summary P&L (2)"
    fallback = "05 Summary P&L"
    sheet_name = target if target in sheet_names else (fallback if fallback in sheet_names else None)
    if not sheet_name:
        report["unmapped"].append("summary_pl sheet not found")
        return {}

    ws = wb[sheet_name]
    rows = [row for row in ws.iter_rows(values_only=True) if any(c is not None for c in row)]

    summary = {}
    label_map = {
        "Revenue": "revenue",
        "(-) COGS": "cogs",
        "Gross Profit": "gross_profit",
        "GP %": "gp_pct",
        "Total OPEX": "opex",
        "EBITDA": "ebitda",
        "One-time adjustment": "one_time_adj",
        "Normalized EBITDA": "normalized_ebitda",
    }

    # Header row tells us column positions: Actual, Budget, Prior Year Actual, YTD Actual, YTD Budget, YTD Prior
    # Row 4 (idx 3): headers like 'May 25', 'May 25', 'May 24', 'vs Budget', 'vs May 24' ...
    # Row 5 (idx 4): 'Actual', 'Budget', 'Actual', '%', '%' ...
    # Data starts row 6 (idx 5)

    # Find header rows
    col_map = {}  # canonical_name -> col_index
    for i, row in enumerate(rows):
        first = str(row[0]).strip() if row[0] else ""
        if first.startswith("In MYR") or "vs Budget" in str(row):
            # Period row — next row has Actual/Budget
            if i + 1 < len(rows):
                sub = rows[i + 1]
                for j, v in enumerate(sub):
                    if v == "Actual" and "actual_month" not in col_map:
                        col_map["actual_month"] = j
                    elif v == "Budget" and "budget_month" not in col_map:
                        col_map["budget_month"] = j
            if i + 2 < len(rows):
                sub2 = rows[i + 2]
                for j, v in enumerate(sub2):
                    if v == "Actual" and "actual_ytd" not in col_map:
                        col_map["actual_ytd"] = j
                    elif v == "Budget" and "budget_ytd" not in col_map:
                        col_map["budget_ytd"] = j
            break

    # Collect metric rows
    for row in rows:
        label = str(row[0]).strip() if row[0] else ""
        canonical = label_map.get(label)
        if not canonical:
            continue

        entry = {}
        for key, col in col_map.items():
            if col < len(row):
                val = _cell_value(row[col])
                if canonical == "gp_pct":
                    entry[key] = _pct(val)
                else:
                    entry[key] = round(val, 2) if isinstance(val, float) else val
        summary[canonical] = entry

    # Determine report month from sheet title row
    for row in rows[:5]:
        for cell in row:
            if isinstance(cell, str) and "May 2025" in cell:
                summary["report_month"] = "May 2025"
                break
            if isinstance(cell, str) and "2025" in cell:
                summary["report_year"] = "2025"

    report["auto_mapped"].append(f"summary_pl from '{sheet_name}'")
    return summary


def _parse_monthly_pl(wb, sheet_names, report) -> list:
    """
    Parse monthly P&L from Montly PL (2025) and Montly PL (2024) sheets.
    Returns list of {month, year, revenue, cogs, gross_profit, opex, ebitda, normalized_ebitda}.
    """
    monthly = []

    sheet_pairs = [
        ("Montly PL (2)", 2025),
        ("Montly PL", 2024),
        ("ProfitLoss (2)", 2025),
        ("ProfitLoss (24)", 2024),
        # Intentionally exclude 2022/2023 sheets — they contain extreme outliers
        # (e.g. Mar 2022 MYR 5.3M) that distort chart scaling for 2024-2025 view
    ]

    # Cat2/Cat1 values observed in Montly PL sheets:
    # Revenue, COGS, Employee Related, Establishment Cost, Marketing Related,
    # Other OPEX, Professional Fees, Travel & Entertainment, D & A,
    # Other Income, One-Off, Finance Costs, Income Tax/ WHT
    OPEX_CATS = {
        "Employee Related", "Establishment Cost", "Marketing Related",
        "Other OPEX", "Professional Fees", "Travel & Entertainment",
        "D & A", "TOTAL OPEX",
    }

    for sheet_name, year in sheet_pairs:
        if sheet_name not in sheet_names:
            continue

        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))

        # Find header row: contains datetime objects for months
        header_row_idx = None
        month_cols = {}  # col_index -> datetime

        for i, row in enumerate(all_rows):
            dt_count = sum(1 for c in row if isinstance(c, datetime))
            if dt_count >= 3:
                header_row_idx = i
                for j, cell in enumerate(row):
                    if isinstance(cell, datetime):
                        month_cols[j] = cell
                break

        if header_row_idx is None:
            report["unmapped"].append(f"{sheet_name}: no month header row found")
            continue

        # Cat2 column: col index 4 based on observed structure (0=GL, 1=CY2021, 2=SAP GL, 3=SAP GL Desc, 4=Cat2, 5=Cat1)
        CAT2_COL = 4
        CAT1_COL = 5

        # Aggregate by month
        month_totals = {col: {"revenue": 0.0, "cogs": 0.0, "opex": 0.0, "one_off": 0.0} for col in month_cols}

        for row in all_rows[header_row_idx + 1:]:
            if not any(c is not None for c in row):
                continue
            cat2 = str(row[CAT2_COL]).strip() if CAT2_COL < len(row) and row[CAT2_COL] else ""
            cat1 = str(row[CAT1_COL]).strip() if CAT1_COL < len(row) and row[CAT1_COL] else ""

            # Classify row into a bucket
            cat2_lower = cat2.lower()
            cat1_lower = cat1.lower()
            if cat2 == "Revenue" or cat1 == "Revenue":
                target_bucket = "revenue"
            elif cat2 == "COGS" or cat1 == "COGS":
                target_bucket = "cogs"
            elif cat2 in OPEX_CATS or cat1 in OPEX_CATS:
                target_bucket = "opex"
            elif cat2 == "One-Off" or cat1 == "One-Off":
                target_bucket = "one_off"
            else:
                continue

            for col, dt in month_cols.items():
                if col >= len(row):
                    continue
                val = _cell_value(row[col])
                if not isinstance(val, (int, float)):
                    continue

                if target_bucket == "revenue":
                    month_totals[col]["revenue"] += val
                elif target_bucket == "cogs":
                    month_totals[col]["cogs"] += val
                elif target_bucket == "opex":
                    month_totals[col]["opex"] += val
                elif target_bucket == "one_off":
                    month_totals[col]["one_off"] += val

        for col, dt in sorted(month_cols.items(), key=lambda x: x[1]):
            # Only include 2024 and 2025 data
            if dt.year < 2024:
                continue
            # Skip future months beyond May 2025
            if dt > datetime(2025, 5, 31):
                continue

            t = month_totals[col]
            rev = t["revenue"]
            cogs = t["cogs"]
            opex = t["opex"]
            gp = rev - cogs
            ebitda = gp - opex
            norm_ebitda = ebitda + t["one_off"]
            gp_pct = round(gp / rev * 100, 2) if rev else 0

            # Skip future months with no data (all zeros)
            if rev == 0 and cogs == 0 and opex == 0:
                continue

            monthly.append({
                "month": dt.strftime("%b"),
                "month_num": dt.month,
                "year": dt.year,
                "period": dt.strftime("%b %Y"),
                "revenue": round(rev, 2),
                "cogs": round(cogs, 2),
                "gross_profit": round(gp, 2),
                "gp_pct": gp_pct,
                "opex": round(opex, 2),
                "ebitda": round(ebitda, 2),
                "normalized_ebitda": round(norm_ebitda, 2),
            })

        report["auto_mapped"].append(f"monthly_pl {year} from '{sheet_name}'")

    # Sort by date
    monthly.sort(key=lambda x: (x["year"], x["month_num"]))
    # Deduplicate by period (keep latest parsed)
    seen = {}
    for m in monthly:
        seen[m["period"]] = m
    return list(seen.values())


def _parse_cash_flow(wb, sheet_names, report) -> list:
    """
    Parse cash flow from '09 Cash Flow (Jan-May)' and 'CF (Jan-May)'.
    Returns list of {period, ebitda, net_cash_ops, net_cash_inv, net_cash_fin, net_cash_flow, closing_cash}.
    """
    cash_data = []
    targets = ["09 Cash Flow (Jan-May)", "CF (Jan-May)", "CF (24)"]

    LABEL_MAP = {
        "EBITDA": "ebitda",
        "Net cash generated from operating activities": "net_cash_ops",
        "Net cash used in investing activities": "net_cash_inv",
        "Net cash used in financing activities": "net_cash_fin",
        "Net Cash Flow": "net_cash_flow",
        "Closing Cash Balance": "closing_cash",
        "Opening Cash Balance": "opening_cash",
    }

    for sheet_name in targets:
        if sheet_name not in sheet_names:
            continue

        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue

        # Row 0 has period headers (datetime objects)
        header_row = all_rows[0]
        month_cols = {}
        for j, cell in enumerate(header_row):
            if isinstance(cell, datetime):
                month_cols[j] = cell

        if not month_cols:
            continue

        period_data = {col: {"period": dt.strftime("%b %Y")} for col, dt in month_cols.items()}

        for row in all_rows[1:]:
            if not row or row[0] is None:
                continue
            label = str(row[0]).strip()
            canonical = LABEL_MAP.get(label)
            if not canonical:
                continue
            for col, dt in month_cols.items():
                if col >= len(row):
                    continue
                val = _cell_value(row[col])
                if isinstance(val, (int, float)):
                    period_data[col][canonical] = round(val, 4)

        for col, dt in sorted(month_cols.items(), key=lambda x: x[1]):
            # Skip future months beyond May 2025
            if dt > datetime(2025, 5, 31):
                continue
            entry = period_data[col]
            if len(entry) > 1:  # has at least one metric
                cash_data.append(entry)

        report["auto_mapped"].append(f"cash_flow from '{sheet_name}'")

    # Deduplicate by period — prefer entries with valid (non-zero) opening_cash
    seen = {}
    for entry in cash_data:
        p = entry["period"]
        opening = entry.get("opening_cash", 0) or 0
        closing = entry.get("closing_cash", 0) or 0
        existing = seen.get(p)
        if existing is None:
            seen[p] = entry
        else:
            # Prefer the entry with a non-zero opening_cash (more complete row)
            existing_opening = existing.get("opening_cash", 0) or 0
            if opening != 0 and existing_opening == 0:
                seen[p] = entry
            elif len(entry) > len(existing):
                seen[p] = entry

    # Remove entries where closing_cash equals net_cash_flow with opening=0
    # (these are incomplete rows where only part of the CF statement was filled)
    cleaned = []
    for entry in seen.values():
        opening = entry.get("opening_cash", 0) or 0
        closing = entry.get("closing_cash")
        net = entry.get("net_cash_flow")
        # Skip if opening is 0 and closing == net_cash_flow (means opening was missing)
        if opening == 0 and closing is not None and net is not None and abs(closing - net) < 0.01:
            continue
        cleaned.append(entry)

    result = sorted(cleaned, key=lambda x: datetime.strptime(x["period"], "%b %Y"))
    return result


def _parse_working_capital(wb, sheet_names, report) -> dict:
    """Parse DSO/DIO/DPO/NWC from '12. DSO DSI DIO' and 'NWC Report 2025'."""
    wc = {"dso_dio_dpo": [], "nwc_trend": []}

    # --- DSO/DIO/DPO sheet ---
    # Structure: Row 4 = 'CY2025', Row 5 = month headers (Apr, May, Jun…),
    # Row 6 onward = AR, DSO, Inventory, DIO, AP, DPO, NWC, CCC
    dso_sheet = "12. DSO DSI DIO"
    if dso_sheet in sheet_names:
        ws = wb[dso_sheet]
        rows = list(ws.iter_rows(values_only=True))

        LABEL_MAP = {
            "AR": "ar", "DSO": "dso",
            "Inventory": "inventory", "DIO": "dio",
            "AP": "ap", "DPO": "dpo",
            "NWC": "nwc", "CCC": "ccc",
        }

        # Row 5 (idx 4, 0-based) has month labels in cols 1+
        month_row_idx = None
        months = {}  # col_idx -> month_label
        for i, row in enumerate(rows):
            # Look for the row starting with "In MYR'000" or None with short month strings
            first = str(row[0]).strip() if row[0] else ""
            if "In MYR" in first or first == "":
                # Check if next cols are month abbreviations
                short_strs = [c for c in row[1:7] if isinstance(c, str) and 2 <= len(c.strip()) <= 5]
                if len(short_strs) >= 2:
                    month_row_idx = i
                    for j, c in enumerate(row):
                        if j > 0 and isinstance(c, str) and 2 <= len(c.strip()) <= 5:
                            months[j] = c.strip()
                    break

        if month_row_idx is not None and months:
            period_data = {j: {"month": m} for j, m in months.items()}
            for row in rows[month_row_idx + 1:]:
                if not row or row[0] is None:
                    continue
                label = str(row[0]).strip()
                canonical = LABEL_MAP.get(label)
                if not canonical:
                    continue
                for j, m in months.items():
                    if j < len(row):
                        val = _cell_value(row[j])
                        if isinstance(val, (int, float)):
                            period_data[j][canonical] = round(val, 4)
            wc["dso_dio_dpo"] = [v for v in period_data.values() if len(v) > 1]

        report["auto_mapped"].append(f"dso_dio_dpo from '{dso_sheet}'")

    # --- NWC trend from NWC Report 2025 ---
    nwc_sheet = "NWC Report 2025"
    if nwc_sheet in sheet_names:
        ws = wb[nwc_sheet]
        rows = list(ws.iter_rows(values_only=True))

        # Row 6 (idx 5) has datetime month headers
        header_idx = None
        month_cols = {}
        for i, row in enumerate(rows):
            if sum(1 for c in row if isinstance(c, datetime)) >= 6:
                header_idx = i
                for j, c in enumerate(row):
                    if isinstance(c, datetime):
                        month_cols[j] = c
                break

        if header_idx is not None:
            NWC_LABELS = {
                "Net Sales": "net_sales",
                "Cost of Goods Sold": "cogs",
                "Accounts Receivable -Trade": "ar",
                "DSO": "dso",
                "Inventories": "inventory",
                "DSI ": "dsi",
                "Accounts Payable -Trade": "ap",
                "DPO": "dpo",
                "Net Working Capital": "nwc",
            }
            period_data = {col: {"period": dt.strftime("%b %Y")} for col, dt in month_cols.items()}

            for row in rows[header_idx + 1:]:
                label = str(row[0]).strip() if row[0] else ""
                canonical = NWC_LABELS.get(label)
                if not canonical:
                    continue
                for col, dt in month_cols.items():
                    if col < len(row):
                        val = _cell_value(row[col])
                        if isinstance(val, (int, float)):
                            period_data[col][canonical] = round(val, 4)

            wc["nwc_trend"] = sorted(
                [v for v in period_data.values() if len(v) > 2],
                key=lambda x: datetime.strptime(x["period"], "%b %Y")
            )

        report["auto_mapped"].append(f"nwc_trend from '{nwc_sheet}'")

    return wc


def _parse_sales_segments(wb, sheet_names, report) -> dict:
    """Parse sales by segment from '03 Sales update' and '02 Business Segment'."""
    segments = {"ytd_comparison": {}, "monthly": []}

    sales_sheet = "03 Sales update "
    if sales_sheet not in sheet_names:
        sales_sheet = next((s for s in sheet_names if "Sales" in s and "update" in s), None)

    if sales_sheet and sales_sheet in sheet_names:
        ws = wb[sales_sheet]
        rows = list(ws.iter_rows(values_only=True))

        # YTD comparison: rows 6–13 (idx 5–12)
        # Cols: None, Segment, FYC 2024 YTD May, FYC 2025 YTD May, YoY ratio
        SEGMENT_ROWS_END = 20
        for row in rows[:SEGMENT_ROWS_END]:
            if not row or row[1] is None:
                continue
            seg = str(row[1]).strip()
            if seg in ("", "HID", "LabFriend", "Project/NGS", "Sci.Lab", "Sci.Lab-Qiagen", "Service", "Thermo", "Total", "check"):
                if len(row) >= 4 and isinstance(row[2], (int, float)) and seg != "check":
                    segments["ytd_comparison"][seg] = {
                        "ytd_2024": round(row[2], 2) if isinstance(row[2], (int, float)) else None,
                        "ytd_2025": round(row[3], 2) if isinstance(row[3], (int, float)) else None,
                        "yoy_ratio": round(row[4], 4) if isinstance(row[4], (int, float)) else None,
                    }

        # Monthly detail: find rows with April/May datetime headers
        for i, row in enumerate(rows):
            dt_count = sum(1 for c in row if isinstance(c, datetime))
            if dt_count >= 2:
                # monthly breakdown rows below
                month_cols = {}
                for j, c in enumerate(row):
                    if isinstance(c, datetime):
                        month_cols[j] = c

                SEGMENT_NAMES = ["HID", "LabFriend", "Project/NGS", "Sci.Lab (Total)", "Sci.Lab-Qiagen", "Service"]
                for row2 in rows[i + 1: i + 10]:
                    if not row2 or row2[1] is None:
                        continue
                    seg = str(row2[1]).strip()
                    if seg not in SEGMENT_NAMES:
                        continue
                    for j, dt in month_cols.items():
                        if j < len(row2) and isinstance(row2[j], (int, float)):
                            segments["monthly"].append({
                                "segment": seg,
                                "period": dt.strftime("%b %Y"),
                                "revenue": round(row2[j], 2),
                            })
                break

        report["auto_mapped"].append(f"sales_segments from '{sales_sheet}'")

    return segments
