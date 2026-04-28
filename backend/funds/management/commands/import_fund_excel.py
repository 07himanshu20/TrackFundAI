"""
Import mock fund Excel data into the TrackFundAI database.

Reads Excel files from mock_fund_data/ and creates records across all apps:
  accounts (Organization, User)
  funds (FundCategory, Entity, Fund, Scheme)
  lp (BankAccount, Investor, Commitment, CapitalCall, CapitalCallLineItem,
      Distribution, DistributionLineItem)
  investments (PortfolioCompany, Investment, InvestmentTranche, Valuation,
               KPIDefinition, PortfolioKPI, ExitEvent, BoardMeeting)
  accounting (ChartOfAccounts, NAVRecord, CarriedInterest, FundLedger,
              ManagementFeeSchedule)
  compliance (SEBIReport, ComplianceCalendar, AMLDueDiligence)
  portfolio (PortfolioSnapshot, PortfolioNode)

Usage:
  python manage.py import_fund_excel
  python manage.py import_fund_excel --file 01_Avendus_Future_Leaders_Fund_II.xlsx
  python manage.py import_fund_excel --reset  # wipe and re-import all
"""

import os
from datetime import date
from decimal import Decimal, InvalidOperation

import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils.text import slugify

from accounts.models import Organization, User, FundAccess
from funds.models import FundCategory, Entity, Fund, Scheme
from lp.models import (BankAccount, Investor, Commitment, CapitalCall,
                        CapitalCallLineItem, Distribution, DistributionLineItem)
from investments.models import (PortfolioCompany, Investment, InvestmentTranche,
                                 Valuation, KPIDefinition, PortfolioKPI,
                                 ExitEvent, BoardMeeting)
from accounting.models import (ChartOfAccounts, NAVRecord, CarriedInterest,
                                FundLedger, ManagementFeeSchedule)
from portfolio.models import PortfolioSnapshot, PortfolioNode

# Attempt compliance imports — may not all exist
try:
    from compliance.models import (SEBIReport, AMLDueDiligence,
                                    ComplianceCalendar, ComplianceTestReport,
                                    CTRChecklistItem, PPMAmendment,
                                    SEBICircular, CircularAction)
    HAS_COMPLIANCE = True
except ImportError:
    HAS_COMPLIANCE = False


MOCK_DATA_DIR = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    '..', '..', '..', '..', 'mock_fund_data',
)

# SEBI category codes
CATEGORY_MAP = {
    'CAT_I_VCF': ('Category I AIF', 'Venture Capital Fund', False),
    'CAT_II': ('Category II AIF', 'Private Equity Fund', False),
    'CAT_III_LVF': ('Category III AIF', 'Long-Short Equity Fund', True),
}

# Investor type mapping from Excel to Django choices
INVESTOR_TYPE_MAP = {
    'insurance': 'insurance',
    'pension': 'pension',
    'huf': 'huf',
    'trust': 'trust',
    'individual': 'individual',
    'fund_of_funds': 'fund_of_funds',
    'fpi': 'fpi',
    'company': 'company',
    'nri': 'nri',
    'family_office': 'family_office',
    'endowment': 'endowment',
    'llp': 'llp',
    'sovereign': 'sovereign',
    'bank': 'bank',
}


def _d(val, default=None):
    """Convert a value to Decimal, handling None/empty."""
    if val is None or val == '' or val == 'None':
        return default
    try:
        return Decimal(str(val))
    except (InvalidOperation, ValueError):
        return default


def _date(val):
    """Convert a value to date, handling datetime objects."""
    if val is None:
        return None
    if hasattr(val, 'date'):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _str(val, default=''):
    """Safe string conversion."""
    if val is None:
        return default
    return str(val).strip()


def _bool(val):
    """Convert Yes/No/True/False to bool."""
    if isinstance(val, bool):
        return val
    if val is None:
        return False
    return str(val).strip().lower() in ('yes', 'true', '1')


def read_table(ws, start_row=1, max_rows=None):
    """
    Read rows from a worksheet starting at a header row.
    Returns list of dicts keyed by header names.
    Stops at first fully empty row or section header.
    """
    # Find header row (first row with data in column 1)
    header_row = start_row
    for r in range(start_row, min(ws.max_row + 1, start_row + 10)):
        val = ws.cell(r, 1).value
        if val and not _is_section_header(val):
            header_row = r
            break
        elif val and _is_section_header(val) and r == start_row:
            continue

    headers = []
    for c in range(1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        if h:
            headers.append((c, str(h).strip()))

    if not headers:
        return []

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        if max_rows and len(rows) >= max_rows:
            break
        row_data = {}
        all_empty = True
        for col, name in headers:
            val = ws.cell(r, col).value
            if val is not None:
                all_empty = False
            row_data[name] = val
        if all_empty:
            break
        # Stop if we hit another section header
        first_val = ws.cell(r, 1).value
        if _is_section_header(first_val):
            break
        rows.append(row_data)

    return rows


def find_section_rows(ws, section_name):
    """Find the start row of a named section (e.g., 'INVESTMENTS')."""
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val and section_name in str(val):
            return r
    return None


# Known section headers that appear in Excel sheets.
# Used to detect where data sections end.
_SECTION_HEADERS = {
    'ORGANIZATION MASTER', 'KEY ENTITIES', 'GP USERS', 'FUND ACCESS MATRIX',
    'FUND MASTER DATA', 'SCHEMES', 'PORTFOLIO HIERARCHY',
    'CROSS-FUND SECTOR MAPPING', 'PORTFOLIO COMPANIES', 'INVESTMENTS',
    'INVESTMENT TRANCHES', 'CAPITAL CALLS', 'CAPITAL CALL LINE ITEMS',
    'NAV RECORDS', 'EXIT EVENTS', 'DISTRIBUTIONS', 'DISTRIBUTION LINE ITEMS',
    'CHART OF ACCOUNTS', 'DOUBLE-ENTRY', 'CARRIED INTEREST',
    'COMPLIANCE CALENDAR', 'SEBI REPORT FILINGS', 'AML DUE DILIGENCE',
    'COMPLIANCE TEST REPORT', 'SEBI CIRCULARS', 'PPM AMENDMENTS',
}


def _is_section_header(val):
    """Check if a cell value looks like a section header, not data."""
    if not val:
        return False
    s = str(val).strip()
    # Check against known headers
    for header in _SECTION_HEADERS:
        if header in s.upper():
            return True
    # Heuristic: all-caps AND long AND contains spaces → likely a section header
    # Short all-caps like "EPFO", "BFSI", "ADIA" are valid entity names
    return s.isupper() and len(s) > 15 and ' ' in s


class Command(BaseCommand):
    help = 'Import mock fund Excel data into TrackFundAI database'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file', type=str, default=None,
            help='Import a specific Excel file (filename only)',
        )
        parser.add_argument(
            '--reset', action='store_true',
            help='Delete ALL imported data before re-importing',
        )
        parser.add_argument(
            '--data-dir', type=str, default=None,
            help='Path to mock_fund_data directory',
        )

    def handle(self, *args, **options):
        data_dir = options['data_dir'] or os.path.abspath(MOCK_DATA_DIR)
        if not os.path.isdir(data_dir):
            self.stderr.write(self.style.ERROR(f'Directory not found: {data_dir}'))
            return

        if options['reset']:
            self._reset_data()

        # Get or create organization (shared across all funds)
        org = self._ensure_organization()

        # Ensure fund categories exist
        self._ensure_fund_categories()

        # Determine which files to import
        if options['file']:
            files = [options['file']]
        else:
            files = sorted([f for f in os.listdir(data_dir)
                           if f.endswith('.xlsx') and not f.startswith('~')])

        self.stdout.write(f'\nImporting {len(files)} fund file(s) from {data_dir}')
        self.stdout.write(f'Organization: {org.name}\n')

        for filename in files:
            filepath = os.path.join(data_dir, filename)
            if not os.path.exists(filepath):
                self.stderr.write(self.style.ERROR(f'File not found: {filepath}'))
                continue
            self._import_fund_file(filepath, org)

        # Build portfolio hierarchy from all funds
        self._build_portfolio_hierarchy(org, data_dir, files)

        self.stdout.write(self.style.SUCCESS('\nImport complete!'))
        self._print_summary(org)

    def _reset_data(self):
        """Delete all data that was imported."""
        self.stdout.write(self.style.WARNING('Resetting all imported data...'))
        org = Organization.objects.filter(slug='trivesta-capital').first()
        if org:
            # Delete in reverse FK order
            PortfolioNode.objects.filter(snapshot__source='excel_parse').delete()
            PortfolioSnapshot.objects.filter(source='excel_parse').delete()
            if HAS_COMPLIANCE:
                CircularAction.objects.filter(
                    circular__organization=org).delete()
                SEBICircular.objects.filter(organization=org).delete()
                PPMAmendment.objects.filter(fund__organization=org).delete()
                CTRChecklistItem.objects.filter(
                    compliance_test_report__scheme__fund__organization=org).delete()
                ComplianceTestReport.objects.filter(
                    scheme__fund__organization=org).delete()
                AMLDueDiligence.objects.filter(
                    investor__organization=org).delete()
                SEBIReport.objects.filter(fund__organization=org).delete()
                ComplianceCalendar.objects.filter(organization=org).delete()
            ManagementFeeSchedule.objects.filter(scheme__fund__organization=org).delete()
            FundLedger.objects.filter(scheme__fund__organization=org).delete()
            CarriedInterest.objects.filter(scheme__fund__organization=org).delete()
            NAVRecord.objects.filter(scheme__fund__organization=org).delete()
            ChartOfAccounts.objects.filter(organization=org).delete()
            BoardMeeting.objects.filter(
                investment__scheme__fund__organization=org).delete()
            ExitEvent.objects.filter(
                investment__scheme__fund__organization=org).delete()
            PortfolioKPI.objects.filter(
                investment__scheme__fund__organization=org).delete()
            KPIDefinition.objects.filter(organization=org).delete()
            Valuation.objects.filter(
                investment__scheme__fund__organization=org).delete()
            InvestmentTranche.objects.filter(
                investment__scheme__fund__organization=org).delete()
            Investment.objects.filter(scheme__fund__organization=org).delete()
            PortfolioCompany.objects.filter(organization=org).delete()
            DistributionLineItem.objects.filter(
                distribution__scheme__fund__organization=org).delete()
            Distribution.objects.filter(scheme__fund__organization=org).delete()
            CapitalCallLineItem.objects.filter(
                capital_call__scheme__fund__organization=org).delete()
            CapitalCall.objects.filter(scheme__fund__organization=org).delete()
            Commitment.objects.filter(scheme__fund__organization=org).delete()
            Investor.objects.filter(organization=org).delete()
            BankAccount.objects.filter(organization=org).delete()
            Scheme.objects.filter(fund__organization=org).delete()
            Fund.objects.filter(organization=org).delete()
            Entity.objects.filter(organization=org).delete()
            FundAccess.objects.filter(fund__organization=org).delete()
            User.objects.filter(organization=org).exclude(
                username__in=['admin', 'himanshu']).delete()
            org.delete()
        self.stdout.write(self.style.SUCCESS('  Reset complete'))

    def _ensure_organization(self):
        """Create or get the Trivesta Capital organization."""
        org, created = Organization.objects.get_or_create(
            slug='trivesta-capital',
            defaults={
                'name': 'Trivesta Capital Advisors LLP',
                'subscription_tier': 'enterprise',
            },
        )
        if created:
            self.stdout.write(self.style.SUCCESS(
                f'  Created Organization: {org.name}'))
        return org

    def _ensure_fund_categories(self):
        """Create SEBI AIF category records."""
        for code, (name, sub_cat, leverage) in CATEGORY_MAP.items():
            FundCategory.objects.get_or_create(
                sebi_category_code=code,
                defaults={
                    'name': name,
                    'sub_category': sub_cat,
                    'leverage_permitted': leverage,
                },
            )

    @transaction.atomic
    def _import_fund_file(self, filepath, org):
        """Import a single fund Excel file."""
        filename = os.path.basename(filepath)
        self.stdout.write(f'\n{"="*60}')
        self.stdout.write(f'Importing: {filename}')
        self.stdout.write(f'{"="*60}')

        wb = openpyxl.load_workbook(filepath, data_only=True)

        # 1. Create entities (from Organization & Users sheet)
        users = self._import_users(wb, org)

        # 2. Create Fund & Schemes
        fund, schemes = self._import_fund_and_schemes(wb, org)

        # 3. Grant fund access to users
        for user in users:
            FundAccess.objects.get_or_create(
                user=user, fund=fund,
                defaults={'access_level': 'admin'},
            )

        # 4. Import investors
        investors = self._import_investors(wb, org)

        # 5. Import commitments
        commitments = self._import_commitments(wb, org, schemes, investors)

        # 6. Import capital calls
        self._import_capital_calls(wb, schemes, commitments)

        # 7. Import portfolio companies & investments
        companies, investments = self._import_portfolio(wb, org, schemes)

        # 8. Import valuations
        self._import_valuations(wb, investments)

        # 9. Import KPIs
        self._import_kpis(wb, org, investments, companies)

        # 10. Import NAV records
        self._import_nav(wb, schemes)

        # 11. Import exits & distributions
        self._import_exits_and_distributions(wb, investments, schemes, commitments)

        # 12. Import accounting (chart of accounts, ledger, carried interest, mgmt fees)
        self._import_accounting(wb, org, schemes)

        # 13. Import compliance data
        if HAS_COMPLIANCE:
            self._import_compliance(wb, fund, schemes)

        # 14. Import board meetings
        self._import_board_meetings(wb, investments)

        self.stdout.write(self.style.SUCCESS(f'  Done: {filename}'))

    def _import_users(self, wb, org):
        """Import GP users from Organization & Users sheet."""
        if 'Organization & Users' not in wb.sheetnames:
            return []

        ws = wb['Organization & Users']
        # Find GP USERS section
        users_start = find_section_rows(ws, 'GP USERS')
        if not users_start:
            return []

        users = []
        # Read from the header row after section title
        for r in range(users_start + 1, ws.max_row + 1):
            username = ws.cell(r, 1).value
            if not username or str(username).strip() == 'Username':
                continue
            if _is_section_header(username):
                break
            full_name = _str(ws.cell(r, 2).value)
            role = _str(ws.cell(r, 3).value)
            email = _str(ws.cell(r, 4).value)

            parts = full_name.split(' ', 1)
            first_name = parts[0] if parts else username
            last_name = parts[1] if len(parts) > 1 else ''

            user, created = User.objects.get_or_create(
                username=username,
                defaults={
                    'email': email,
                    'first_name': first_name,
                    'last_name': last_name,
                    'organization': org,
                    'role': role,
                },
            )
            if created:
                user.set_password(f'{username}123')
                user.save()
                self.stdout.write(f'    Created User: {username} ({role})')
            users.append(user)

        # Also import entities
        entities_start = find_section_rows(ws, 'KEY ENTITIES')
        if entities_start:
            for r in range(entities_start + 1, ws.max_row + 1):
                entity_type = ws.cell(r, 1).value
                if not entity_type or str(entity_type).strip() == 'Entity Type':
                    continue
                if _is_section_header(entity_type):
                    break
                entity_name = _str(ws.cell(r, 2).value)
                if not entity_name:
                    continue

                Entity.objects.get_or_create(
                    organization=org,
                    entity_type=_str(entity_type),
                    entity_name=entity_name,
                    defaults={
                        'pan': _str(ws.cell(r, 3).value),
                        'gstin': _str(ws.cell(r, 4).value),
                        'sebi_registration': _str(ws.cell(r, 5).value),
                        'contact_person': _str(ws.cell(r, 6).value),
                        'email': _str(ws.cell(r, 7).value),
                    },
                )

        return users

    def _import_fund_and_schemes(self, wb, org):
        """Import fund master data and schemes."""
        ws = wb['Fund & Scheme Master']

        # Read key-value pairs from fund master section
        fund_data = {}
        for r in range(1, 30):
            label = _str(ws.cell(r, 1).value)
            value = ws.cell(r, 2).value
            if label:
                fund_data[label] = value

        fund_name = _str(fund_data.get('Fund Name', ''))
        sebi_reg = _str(fund_data.get('SEBI Registration', ''))
        cat_code = _str(fund_data.get('SEBI Category Code', 'CAT_II'))
        is_gift = _bool(fund_data.get('GIFT City', 'No'))

        fund_category = FundCategory.objects.filter(
            sebi_category_code=cat_code).first()

        # Link entities
        manager = Entity.objects.filter(
            organization=org, entity_type='manager').first()
        trustee = Entity.objects.filter(
            organization=org, entity_type='trustee').first()
        custodian = Entity.objects.filter(
            organization=org, entity_type='custodian').first()
        auditor = Entity.objects.filter(
            organization=org, entity_type='statutory_auditor').first()
        sponsor = Entity.objects.filter(
            organization=org, entity_type='sponsor').first()

        fund, created = Fund.objects.get_or_create(
            organization=org,
            name=fund_name,
            defaults={
                'sebi_registration_number': sebi_reg,
                'fund_category': fund_category,
                'structure_type': _str(fund_data.get('Structure', 'trust')).lower(),
                'pan': _str(fund_data.get('PAN', '')),
                'gstin': _str(fund_data.get('GSTIN', '')),
                'inception_date': _date(fund_data.get('Inception Date')),
                'corpus_target': _d(fund_data.get('Target Corpus (Cr)',
                                    fund_data.get('Target Corpus (USD Mn)', 0))),
                'base_currency': _str(fund_data.get('Base Currency', 'INR')),
                'is_gift_city': is_gift,
                'manager_entity': manager,
                'trustee_entity': trustee,
                'custodian_entity': custodian,
                'auditor_entity': auditor,
                'sponsor_entity': sponsor,
            },
        )
        self.stdout.write(f'  {"Created" if created else "Found"} Fund: {fund.name}')

        # Import schemes
        schemes_start = find_section_rows(ws, 'SCHEMES')
        schemes = {}
        if schemes_start:
            for r in range(schemes_start + 1, ws.max_row + 1):
                scheme_name = ws.cell(r, 1).value
                if not scheme_name or str(scheme_name).strip() == 'Scheme Name':
                    continue
                if _is_section_header(scheme_name):
                    break

                status_map = {
                    'Fundraising': 'fundraising',
                    'Investing': 'investing',
                    'Harvesting': 'harvesting',
                    'Dissolved': 'dissolved',
                }
                raw_status = _str(ws.cell(r, 13).value)
                scheme_status = status_map.get(raw_status, 'investing')

                scheme, _ = Scheme.objects.get_or_create(
                    fund=fund,
                    name=_str(scheme_name),
                    defaults={
                        'vintage_year': ws.cell(r, 2).value,
                        'first_close_date': _date(ws.cell(r, 3).value),
                        'final_close_date': _date(ws.cell(r, 4).value),
                        'scheme_size': _d(ws.cell(r, 5).value),
                        'tenure_years': ws.cell(r, 6).value,
                        'hurdle_rate_pct': _d(ws.cell(r, 7).value),
                        'carry_pct': _d(ws.cell(r, 8).value),
                        'carry_type': _str(ws.cell(r, 9).value).lower().split(' ')[0],
                        'management_fee_basis': _str(ws.cell(r, 10).value).lower(),
                        'management_fee_pct': _d(ws.cell(r, 11).value),
                        'sponsor_commitment_pct': _d(ws.cell(r, 12).value),
                        'scheme_status': scheme_status,
                        'is_active': scheme_status != 'dissolved',
                    },
                )
                schemes[_str(scheme_name)] = scheme
                self.stdout.write(f'    Scheme: {scheme.name} ({scheme_status})')

        return fund, schemes

    def _import_investors(self, wb, org):
        """Import investors from the Investors sheet."""
        sheet_name = None
        for name in ['Investors', 'Investors & AML']:
            if name in wb.sheetnames:
                sheet_name = name
                break
        if not sheet_name:
            return {}

        ws = wb[sheet_name]
        investors = {}

        for r in range(1, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if not name or str(name).strip() == 'Investor Name':
                continue
            if _is_section_header(name):
                break

            inv_type = _str(ws.cell(r, 2).value).lower()
            inv_type = INVESTOR_TYPE_MAP.get(inv_type, 'other')

            kyc_map = {
                'Completed': 'completed', 'Pending': 'pending',
                'Expired': 'expired', 'In Progress': 'in_progress',
                'Rejected': 'rejected',
            }
            fatca_map = {
                'Not Applicable': 'not_applicable', 'Compliant': 'compliant',
                'Pending': 'pending', 'Non-Compliant': 'non_compliant',
            }

            inv_name = _str(name)

            # Create bank account first
            bank_name = _str(ws.cell(r, 22).value)
            acct_no = _str(ws.cell(r, 23).value)
            bank_account = None
            if bank_name and acct_no:
                # Column 24 is IFSC or SWIFT depending on fund
                ifsc_or_swift = _str(ws.cell(r, 24).value)
                acct_type_raw = _str(ws.cell(r, 25).value).lower()
                acct_type_map = {
                    'current': 'current', 'savings': 'savings',
                    'nre': 'nre', 'nro': 'nro', 'fcnr': 'fcnr',
                }
                acct_type = acct_type_map.get(acct_type_raw, 'current')

                bank_account, _ = BankAccount.objects.get_or_create(
                    organization=org,
                    account_number=acct_no,
                    defaults={
                        'account_holder_name': inv_name,
                        'bank_name': bank_name,
                        'ifsc_code': ifsc_or_swift if len(ifsc_or_swift) == 11 else '',
                        'swift_code': ifsc_or_swift if len(ifsc_or_swift) != 11 else '',
                        'account_type': acct_type,
                        'is_primary': True,
                    },
                )

            investor, created = Investor.objects.get_or_create(
                organization=org,
                investor_name=inv_name,
                defaults={
                    'investor_type': inv_type,
                    'contact_person': _str(ws.cell(r, 3).value),
                    'email': _str(ws.cell(r, 4).value),
                    'phone': _str(ws.cell(r, 5).value),
                    'address': _str(ws.cell(r, 6).value),
                    'city': _str(ws.cell(r, 7).value),
                    'state': _str(ws.cell(r, 8).value),
                    'country': _str(ws.cell(r, 9).value, 'India'),
                    'pan': _str(ws.cell(r, 10).value),
                    'aadhaar_last_4': _str(ws.cell(r, 11).value),
                    'ckyc_number': _str(ws.cell(r, 12).value),
                    'kyc_status': kyc_map.get(_str(ws.cell(r, 13).value), 'pending'),
                    'kyc_completed_date': _date(ws.cell(r, 14).value),
                    'kyc_expiry_date': _date(ws.cell(r, 15).value),
                    'is_accredited_investor': _bool(ws.cell(r, 16).value),
                    'accreditation_date': _date(ws.cell(r, 17).value),
                    'is_land_border_country': _bool(ws.cell(r, 18).value),
                    'land_border_country_name': _str(ws.cell(r, 19).value),
                    'is_politically_exposed': _bool(ws.cell(r, 20).value),
                    'fatca_status': fatca_map.get(_str(ws.cell(r, 21).value), 'not_applicable'),
                    'primary_bank_account': bank_account,
                },
            )
            investors[inv_name] = investor
            if created:
                self.stdout.write(f'    Investor: {inv_name} ({inv_type})')

        return investors

    def _import_commitments(self, wb, org, schemes, investors):
        """Import commitments from the Commitments sheet."""
        if 'Commitments' not in wb.sheetnames:
            return {}
        ws = wb['Commitments']

        commitments = {}
        for r in range(1, ws.max_row + 1):
            inv_name = ws.cell(r, 1).value
            if not inv_name or str(inv_name).strip() == 'Investor Name':
                continue
            if _is_section_header(inv_name):
                break

            inv_name = _str(inv_name)
            scheme_name = _str(ws.cell(r, 2).value)

            investor = investors.get(inv_name)
            scheme = schemes.get(scheme_name)
            if not investor or not scheme:
                continue

            close_map = {
                'First Close': 'first_close',
                'Subsequent Close': 'subsequent_close',
                'Final Close': 'final_close',
            }
            status_map = {
                'Active': 'active', 'Defaulted': 'defaulted',
                'Transferred': 'transferred', 'Cancelled': 'cancelled',
            }

            commitment, created = Commitment.objects.get_or_create(
                investor=investor,
                scheme=scheme,
                defaults={
                    'commitment_amount': _d(ws.cell(r, 3).value, 0),
                    'commitment_date': _date(ws.cell(r, 4).value),
                    'close_type': close_map.get(_str(ws.cell(r, 5).value), 'first_close'),
                    'units_allocated': _d(ws.cell(r, 6).value),
                    'side_letter_exists': _bool(ws.cell(r, 7).value),
                    'commitment_status': status_map.get(_str(ws.cell(r, 8).value), 'active'),
                },
            )
            key = f'{inv_name}|{scheme_name}'
            commitments[key] = commitment

        self.stdout.write(f'    Commitments: {len(commitments)}')
        return commitments

    def _import_capital_calls(self, wb, schemes, commitments):
        """Import capital calls from the Capital Calls sheet."""
        if 'Capital Calls' not in wb.sheetnames:
            return
        ws = wb['Capital Calls']

        # Find CAPITAL CALLS section
        calls_start = find_section_rows(ws, 'CAPITAL CALLS')
        if not calls_start:
            calls_start = 0

        call_count = 0
        for r in range(calls_start + 1, ws.max_row + 1):
            scheme_name = ws.cell(r, 1).value
            if not scheme_name or str(scheme_name).strip() in ('Scheme Name', 'Scheme'):
                continue
            if _is_section_header(scheme_name):
                break

            scheme = schemes.get(_str(scheme_name))
            if not scheme:
                continue

            call_num = ws.cell(r, 2).value
            if not call_num:
                continue

            status_map = {
                'Draft': 'draft', 'Approved': 'approved',
                'Sent': 'sent', 'Paid': 'paid', 'Defaulted': 'defaulted',
            }

            CapitalCall.objects.get_or_create(
                scheme=scheme,
                call_number=int(call_num),
                defaults={
                    'call_date': _date(ws.cell(r, 3).value) or date.today(),
                    'payment_due_date': _date(ws.cell(r, 4).value) or date.today(),
                    'call_percentage': _d(ws.cell(r, 5).value, 0),
                    'total_call_amount': _d(ws.cell(r, 6).value, 0),
                    'purpose': _str(ws.cell(r, 7).value),
                    'call_status': status_map.get(_str(ws.cell(r, 8).value), 'paid'),
                },
            )
            call_count += 1

        self.stdout.write(f'    Capital Calls: {call_count}')

    def _import_portfolio(self, wb, org, schemes):
        """Import portfolio companies and investments."""
        if 'Portfolio & Investments' not in wb.sheetnames:
            return {}, {}
        ws = wb['Portfolio & Investments']

        companies = {}
        investments = {}

        # Import PORTFOLIO COMPANIES
        co_start = find_section_rows(ws, 'PORTFOLIO COMPANIES')
        if co_start:
            for r in range(co_start + 1, ws.max_row + 1):
                name = ws.cell(r, 1).value
                if not name or str(name).strip() == 'Company Name':
                    continue
                if _is_section_header(name):
                    break

                name = _str(name)
                company, _ = PortfolioCompany.objects.get_or_create(
                    organization=org,
                    name=name,
                    defaults={
                        'cin': _str(ws.cell(r, 2).value),
                        'pan': _str(ws.cell(r, 3).value),
                        'sector': _str(ws.cell(r, 4).value),
                        'sub_sector': _str(ws.cell(r, 5).value),
                        'incorporation_date': _date(ws.cell(r, 6).value),
                        'headquarters_city': _str(ws.cell(r, 7).value),
                        'headquarters_country': _str(ws.cell(r, 8).value, 'India'),
                        'website': _str(ws.cell(r, 9).value),
                        'founder_names': [n.strip() for n in _str(ws.cell(r, 10).value).split(',') if n.strip()],
                        'description': _str(ws.cell(r, 11).value),
                    },
                )
                companies[name] = company

        # Import INVESTMENTS
        inv_start = find_section_rows(ws, 'INVESTMENTS')
        if inv_start:
            for r in range(inv_start + 1, ws.max_row + 1):
                name = ws.cell(r, 1).value
                if not name or str(name).strip() == 'Company Name':
                    continue
                if _is_section_header(name):
                    break

                name = _str(name)
                scheme_name = _str(ws.cell(r, 2).value)
                scheme = schemes.get(scheme_name)
                company = companies.get(name)
                if not scheme:
                    continue

                instrument_raw = _str(ws.cell(r, 3).value).lower()
                instrument_map = {
                    'equity': 'equity', 'ccps': 'ccps', 'ccd': 'ccd',
                    'ncd': 'ncd', 'safe': 'safe',
                    'convertible note': 'convertible_note',
                    'term loan': 'term_loan', 'odi': 'odi',
                }
                instrument = instrument_map.get(instrument_raw, 'equity')

                status_map = {
                    'Active': 'active', 'Partially Exited': 'partially_exited',
                    'Fully Exited': 'fully_exited', 'Written Off': 'written_off',
                }

                # Column layout varies between funds — detect by headers
                # Standard: col3=instrument, col4=ownership, col5=fd%, col6=invested
                # Fund 5 (Cat III): col3=instrument, col4=position, col5=ownership, col6=fd%
                header_row = inv_start + 1
                col4_header = _str(ws.cell(header_row, 4).value)

                if col4_header == 'Position':
                    # Cat III layout
                    ownership = _d(ws.cell(r, 5).value)
                    fd_pct = _d(ws.cell(r, 6).value)
                    invested = _d(ws.cell(r, 7).value, 0)
                    inv_date = _date(ws.cell(r, 8).value)
                    raw_status = _str(ws.cell(r, 9).value)
                    sector = _str(ws.cell(r, 10).value)
                else:
                    # Standard layout
                    ownership = _d(ws.cell(r, 4).value)
                    fd_pct = _d(ws.cell(r, 5).value)
                    invested = _d(ws.cell(r, 6).value, 0)
                    inv_date = _date(ws.cell(r, 7).value)
                    raw_status = _str(ws.cell(r, 8).value)
                    sector = _str(ws.cell(r, 11).value)

                status = status_map.get(raw_status, 'active')

                inv, created = Investment.objects.get_or_create(
                    scheme=scheme,
                    company_name=name,
                    instrument_type=instrument,
                    defaults={
                        'portfolio_company': company,
                        'ownership_pct': ownership,
                        'percentage_stake_fully_diluted': fd_pct,
                        'total_invested': abs(invested),
                        'investment_date': inv_date,
                        'currency': scheme.fund.base_currency,
                        'status': status,
                        'sector': sector,
                    },
                )
                key = f'{name}|{scheme_name}|{instrument}'
                investments[key] = inv

        # Import TRANCHES
        tr_start = find_section_rows(ws, 'INVESTMENT TRANCHES')
        tr_count = 0
        if tr_start:
            for r in range(tr_start + 1, ws.max_row + 1):
                name = ws.cell(r, 1).value
                if not name or str(name).strip() == 'Company Name':
                    continue
                if _is_section_header(name):
                    break

                name = _str(name)
                tranche_num = ws.cell(r, 2).value
                if not tranche_num:
                    continue

                # Find matching investment
                inv = None
                for key, i in investments.items():
                    if key.startswith(f'{name}|'):
                        inv = i
                        break
                if not inv:
                    continue

                InvestmentTranche.objects.get_or_create(
                    investment=inv,
                    tranche_number=int(tranche_num),
                    defaults={
                        'amount': _d(ws.cell(r, 3).value, 0),
                        'date': _date(ws.cell(r, 4).value) or date.today(),
                        'shares_acquired': _d(ws.cell(r, 5).value),
                        'price_per_share': _d(ws.cell(r, 6).value),
                        'pre_money_valuation': _d(ws.cell(r, 7).value),
                        'post_money_valuation': _d(ws.cell(r, 8).value),
                        'round_name': _str(ws.cell(r, 9).value),
                    },
                )
                tr_count += 1

        self.stdout.write(
            f'    Portfolio: {len(companies)} companies, '
            f'{len(investments)} investments, {tr_count} tranches')
        return companies, investments

    def _import_valuations(self, wb, investments):
        """Import valuations."""
        if 'Valuations' not in wb.sheetnames:
            return
        ws = wb['Valuations']

        count = 0
        for r in range(1, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if not name or str(name).strip() == 'Company Name':
                continue
            if _is_section_header(name):
                break

            name = _str(name)
            inv = None
            for key, i in investments.items():
                if key.startswith(f'{name}|'):
                    inv = i
                    break
            if not inv:
                continue

            method_map = {
                'DCF': 'dcf', 'Comparables': 'comparables',
                'Market Comparables': 'comparables',
                'Recent Transaction': 'recent_transaction',
                'Net Assets': 'net_assets', 'Cost': 'cost',
                'Option Pricing Model': 'option_pricing',
            }
            status_map = {
                'Draft': 'draft', 'Submitted': 'submitted',
                'Approved': 'approved', 'Rejected': 'rejected',
            }

            val_date = _date(ws.cell(r, 2).value)
            methodology = method_map.get(_str(ws.cell(r, 3).value), 'cost')

            Valuation.objects.get_or_create(
                investment=inv,
                valuation_date=val_date or date.today(),
                methodology=methodology,
                defaults={
                    'fair_value': _d(ws.cell(r, 4).value, 0),
                    'fair_value_of_holding': _d(ws.cell(r, 5).value),
                    'enterprise_value': _d(ws.cell(r, 6).value),
                    'cost_basis': _d(ws.cell(r, 7).value),
                    'unrealized_gain_loss': _d(ws.cell(r, 8).value),
                    'multiple': _d(ws.cell(r, 9).value),
                    'fvtpl_movement': _d(ws.cell(r, 10).value),
                    'status': status_map.get(_str(ws.cell(r, -3 + ws.max_column).value), 'approved'),
                },
            )
            count += 1

        self.stdout.write(f'    Valuations: {count}')

    def _import_kpis(self, wb, org, investments, companies):
        """Import KPIs."""
        if 'Portfolio KPIs' not in wb.sheetnames:
            return
        ws = wb['Portfolio KPIs']

        count = 0
        for r in range(1, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if not name or str(name).strip() == 'Company Name':
                continue
            if _is_section_header(name):
                break

            name = _str(name)
            kpi_name = _str(ws.cell(r, 2).value)

            # Find matching investment
            inv = None
            for key, i in investments.items():
                if key.startswith(f'{name}|'):
                    inv = i
                    break
            if not inv:
                continue

            company = companies.get(name)

            # Get or create KPI definition
            format_map = {
                'Currency': 'currency', 'Number': 'number',
                'Percentage': 'percent', 'Ratio': 'ratio',
            }
            source_map = {
                'Manual Entry': 'manual', 'Excel Upload': 'excel_upload',
                'Tally Import': 'tally_import', 'API Integration': 'api_integration',
            }
            status_map = {
                'Draft': 'draft', 'Submitted': 'submitted',
                'Approved': 'approved', 'Rejected': 'rejected',
            }

            kpi_format = format_map.get(_str(ws.cell(r, 6).value), 'number')
            kpi_slug = slugify(kpi_name)

            kpi_def, _ = KPIDefinition.objects.get_or_create(
                organization=org,
                slug=kpi_slug,
                defaults={
                    'name': kpi_name,
                    'format': kpi_format,
                    'frequency': 'monthly' if 'MRR' in kpi_name else 'quarterly',
                },
            )

            period = _date(ws.cell(r, 3).value) or date.today()
            value = _d(ws.cell(r, 5).value, 0)

            PortfolioKPI.objects.get_or_create(
                investment=inv,
                kpi_definition=kpi_def,
                period=period,
                defaults={
                    'portfolio_company': company,
                    'period_end_date': _date(ws.cell(r, 4).value),
                    'value': value,
                    'source': source_map.get(_str(ws.cell(r, 7).value), 'manual'),
                    'status': status_map.get(_str(ws.cell(r, 8).value), 'draft'),
                    'notes': _str(ws.cell(r, 9).value),
                },
            )
            count += 1

        self.stdout.write(f'    KPIs: {count}')

    def _import_nav(self, wb, schemes):
        """Import NAV records."""
        nav_sheet = None
        for name in ['NAV & Accounting', 'NAV Monthly', 'NAV & Carried Interest']:
            if name in wb.sheetnames:
                nav_sheet = name
                break
        if not nav_sheet:
            return

        ws = wb[nav_sheet]
        nav_start = find_section_rows(ws, 'NAV RECORDS')
        if not nav_start:
            return

        count = 0
        for r in range(nav_start + 1, ws.max_row + 1):
            scheme_name = ws.cell(r, 1).value
            if not scheme_name or str(scheme_name).strip() in ('Scheme', 'Scheme Name'):
                continue
            if _is_section_header(scheme_name):
                break

            scheme = schemes.get(_str(scheme_name))
            if not scheme:
                continue

            nav_date = _date(ws.cell(r, 2).value)
            if not nav_date:
                continue

            # Detect column layout from number of NAV columns
            # Standard: 13 cols, Cat III: 15 cols (has Short Positions, Leverage)
            has_short = ws.cell(nav_start + 1, 9).value and 'Short' in _str(
                ws.cell(nav_start + 1, 9).value)

            if has_short:
                # Cat III layout (15 cols)
                total_nav = _d(ws.cell(r, 3).value, 0)
                total_units = _d(ws.cell(r, 4).value, 0)
                nav_per_unit = _d(ws.cell(r, 5).value, 0)
                inv_fv = _d(ws.cell(r, 6).value, 0)
                cash = _d(ws.cell(r, 7).value, 0)
                recv = _d(ws.cell(r, 8).value, 0)
                mgmt_fee = _d(ws.cell(r, 11).value, 0)
                other_liab = _d(ws.cell(r, 12).value, 0)
                dep_type = _str(ws.cell(r, 13).value).lower()
                reconciled = _bool(ws.cell(r, 14).value)
                variance = _d(ws.cell(r, 15).value, 0)
            else:
                # Standard layout (13 cols)
                total_nav = _d(ws.cell(r, 3).value, 0)
                total_units = _d(ws.cell(r, 4).value, 0)
                nav_per_unit = _d(ws.cell(r, 5).value, 0)
                inv_fv = _d(ws.cell(r, 6).value, 0)
                cash = _d(ws.cell(r, 7).value, 0)
                recv = _d(ws.cell(r, 8).value, 0)
                mgmt_fee = _d(ws.cell(r, 9).value, 0)
                other_liab = _d(ws.cell(r, 10).value, 0)
                dep_type = _str(ws.cell(r, 11).value).lower()
                reconciled = _bool(ws.cell(r, 12).value)
                variance = _d(ws.cell(r, 13).value, 0)

            NAVRecord.objects.get_or_create(
                scheme=scheme,
                nav_date=nav_date,
                defaults={
                    'total_nav': total_nav,
                    'total_units_outstanding': total_units,
                    'nav_per_unit': nav_per_unit,
                    'investments_at_fair_value': inv_fv,
                    'cash_and_equivalents': cash,
                    'receivables': recv,
                    'management_fee_payable': mgmt_fee,
                    'other_liabilities': other_liab,
                    'depository_type': dep_type if dep_type in ('cdsl', 'nsdl') else '',
                    'depository_reconciled': reconciled,
                    'depository_variance_amount': variance,
                },
            )
            count += 1

        self.stdout.write(f'    NAV Records: {count}')

    def _import_exits_and_distributions(self, wb, investments, schemes, commitments):
        """Import exit events and distributions."""
        exit_sheet = None
        for name in ['Distributions & Exits', 'Exits & Distributions']:
            if name in wb.sheetnames:
                exit_sheet = name
                break
        if not exit_sheet:
            return

        ws = wb[exit_sheet]

        # Import EXIT EVENTS
        exit_start = find_section_rows(ws, 'EXIT EVENTS')
        exit_count = 0
        if exit_start:
            for r in range(exit_start + 1, ws.max_row + 1):
                name = ws.cell(r, 1).value
                if not name or str(name).strip() == 'Company Name':
                    continue
                if _is_section_header(name):
                    break

                name = _str(name)
                inv = None
                for key, i in investments.items():
                    if key.startswith(f'{name}|'):
                        inv = i
                        break
                if not inv:
                    continue

                exit_type_map = {
                    'IPO': 'ipo', 'Merger & Acquisition': 'merger_acquisition',
                    'Secondary Sale': 'secondary_sale', 'Buyback': 'buyback',
                    'Write-Off': 'write_off',
                }
                gain_map = {
                    'LTCG': 'ltcg', 'Long Term Capital Gain': 'ltcg',
                    'STCG': 'stcg', 'Short Term Capital Gain': 'stcg',
                    'Long Term Loss': 'long_term_loss',
                    'Short Term Loss': 'short_term_loss',
                    'Not Applicable': 'na',
                }

                ExitEvent.objects.get_or_create(
                    investment=inv,
                    exit_type=exit_type_map.get(_str(ws.cell(r, 2).value), 'secondary_sale'),
                    defaults={
                        'is_actual': _bool(ws.cell(r, 3).value),
                        'exit_date': _date(ws.cell(r, 4).value),
                        'exit_valuation': _d(ws.cell(r, 5).value),
                        'proceeds': _d(ws.cell(r, 6).value),
                        'net_exit_proceeds': _d(ws.cell(r, 7).value),
                        'realized_gain_loss': _d(ws.cell(r, 8).value),
                        'gain_loss_nature': gain_map.get(_str(ws.cell(r, 9).value), 'na'),
                        'moic': _d(ws.cell(r, 10).value),
                        'irr_pct': _d(ws.cell(r, 11).value),
                        'buyer_name': _str(ws.cell(r, 12).value),
                    },
                )
                exit_count += 1

        # Import DISTRIBUTIONS
        dist_start = find_section_rows(ws, 'DISTRIBUTIONS')
        dist_count = 0
        if dist_start:
            for r in range(dist_start + 1, ws.max_row + 1):
                scheme_name = ws.cell(r, 1).value
                if not scheme_name or str(scheme_name).strip() in ('Scheme', 'Scheme Name'):
                    continue
                if _is_section_header(scheme_name):
                    break

                scheme = schemes.get(_str(scheme_name))
                if not scheme:
                    continue

                dist_num = ws.cell(r, 2).value
                if not dist_num:
                    continue

                dist_type_map = {
                    'Return of Capital': 'return_of_capital',
                    'STCG': 'stcg', 'LTCG': 'ltcg',
                    'Interest': 'interest', 'Dividend': 'dividend',
                    'Carried Interest Distribution': 'carry',
                }

                Distribution.objects.get_or_create(
                    scheme=scheme,
                    distribution_number=int(dist_num),
                    defaults={
                        'distribution_date': _date(ws.cell(r, 3).value) or date.today(),
                        'distribution_type': dist_type_map.get(
                            _str(ws.cell(r, 4).value), 'other'),
                        'total_gross_amount': _d(ws.cell(r, 5).value, 0),
                        'total_tds_amount': _d(ws.cell(r, 6).value, 0),
                        'total_net_amount': _d(ws.cell(r, 7).value),
                        'notes': _str(ws.cell(r, 8).value),
                        'distribution_status': 'distributed',
                    },
                )
                dist_count += 1

        self.stdout.write(
            f'    Exits: {exit_count}, Distributions: {dist_count}')

    def _import_accounting(self, wb, org, schemes):
        """Import chart of accounts, fund ledger, carried interest, mgmt fees."""
        # Chart of Accounts
        coa_count = 0
        for sheet_name in ['NAV & Accounting', 'Fund Ledger']:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            coa_start = find_section_rows(ws, 'CHART OF ACCOUNTS')
            if not coa_start:
                continue

            for r in range(coa_start + 1, ws.max_row + 1):
                code = ws.cell(r, 1).value
                if not code or str(code).strip() == 'Account Code':
                    continue
                if _is_section_header(code):
                    break

                acct_type_map = {
                    'Asset': 'asset', 'Liability': 'liability',
                    'Equity': 'equity', 'Income': 'income',
                    'Expense': 'expense',
                }

                ChartOfAccounts.objects.get_or_create(
                    organization=org,
                    account_code=_str(code),
                    defaults={
                        'account_name': _str(ws.cell(r, 2).value),
                        'account_type': acct_type_map.get(
                            _str(ws.cell(r, 3).value), 'asset'),
                        'description': _str(ws.cell(r, 4).value),
                    },
                )
                coa_count += 1

        if coa_count:
            self.stdout.write(f'    Chart of Accounts: {coa_count}')

        # Fund Ledger
        if 'Fund Ledger' in wb.sheetnames:
            ws = wb['Fund Ledger']
            ledger_start = find_section_rows(ws, 'DOUBLE-ENTRY')
            if ledger_start:
                ledger_count = 0
                for r in range(ledger_start + 1, ws.max_row + 1):
                    je_num = ws.cell(r, 1).value
                    if not je_num or str(je_num).strip() == 'JE Number':
                        continue
                    if _is_section_header(je_num):
                        break

                    # Find scheme for this entry (use first available scheme)
                    scheme = list(schemes.values())[0] if schemes else None
                    if not scheme:
                        continue

                    # Parse debit/credit account codes
                    debit_str = _str(ws.cell(r, 4).value)
                    credit_str = _str(ws.cell(r, 5).value)
                    debit_code = debit_str.split(' ')[0] if debit_str else ''
                    credit_code = credit_str.split(' ')[0] if credit_str else ''

                    debit_acct = ChartOfAccounts.objects.filter(
                        organization=org, account_code=debit_code).first()
                    credit_acct = ChartOfAccounts.objects.filter(
                        organization=org, account_code=credit_code).first()

                    if not debit_acct or not credit_acct:
                        continue

                    ref_type_map = {
                        'Capital Call': 'capital_call',
                        'Investment': 'investment',
                        'Distribution': 'distribution',
                        'Management Fee': 'management_fee',
                        'Carried Interest': 'carried_interest',
                        'Valuation Adjustment': 'valuation_adjustment',
                        'Expense': 'expense',
                        'Other': 'other',
                    }

                    FundLedger.objects.get_or_create(
                        scheme=scheme,
                        journal_entry_number=_str(je_num),
                        defaults={
                            'entry_date': _date(ws.cell(r, 2).value) or date.today(),
                            'description': _str(ws.cell(r, 3).value),
                            'debit_account': debit_acct,
                            'credit_account': credit_acct,
                            'amount': _d(ws.cell(r, 6).value, 0),
                            'reference_type': ref_type_map.get(
                                _str(ws.cell(r, 7).value), 'other'),
                            'is_reversed': _bool(ws.cell(r, 9).value),
                        },
                    )
                    ledger_count += 1

                self.stdout.write(f'    Fund Ledger: {ledger_count} entries')

        # Carried Interest
        for sheet_name in wb.sheetnames:
            if 'Carried Interest' not in sheet_name:
                continue
            ws = wb[sheet_name]
            ci_start = find_section_rows(ws, 'CARRIED INTEREST')
            if not ci_start:
                continue
            ci_count = 0
            for r in range(ci_start + 1, ws.max_row + 1):
                scheme_name = ws.cell(r, 1).value
                if not scheme_name or str(scheme_name).strip() in ('Scheme', 'Scheme Name'):
                    continue
                if _is_section_header(scheme_name):
                    break

                scheme = schemes.get(_str(scheme_name))
                if not scheme:
                    continue

                status_map = {
                    'Indicative': 'indicative',
                    'Crystallised': 'crystallised',
                    'Paid': 'paid',
                }

                CarriedInterest.objects.get_or_create(
                    scheme=scheme,
                    calculation_date=_date(ws.cell(r, 2).value) or date.today(),
                    defaults={
                        'total_distributions': _d(ws.cell(r, 3).value, 0),
                        'total_called_capital': _d(ws.cell(r, 4).value, 0),
                        'preferred_return_amount': _d(ws.cell(r, 5).value, 0),
                        'carry_base': _d(ws.cell(r, 6).value, 0),
                        'carry_amount_gross': _d(ws.cell(r, 7).value, 0),
                        'carry_amount_net': _d(ws.cell(r, 8).value, 0),
                        'gp_clawback_provision': _d(ws.cell(r, 9).value, 0),
                        'calculation_status': status_map.get(
                            _str(ws.cell(r, 10).value), 'indicative'),
                        'notes': _str(ws.cell(r, 11).value),
                    },
                )
                ci_count += 1
            if ci_count:
                self.stdout.write(f'    Carried Interest: {ci_count}')

        # Management Fee Schedule
        if 'Management Fee Schedule' in wb.sheetnames:
            ws = wb['Management Fee Schedule']
            fee_count = 0
            for r in range(1, ws.max_row + 1):
                scheme_name = ws.cell(r, 1).value
                if not scheme_name or str(scheme_name).strip() in ('Scheme', 'Scheme Name'):
                    continue
                if _is_section_header(scheme_name) or 'NOTE' in _str(scheme_name):
                    break

                scheme = schemes.get(_str(scheme_name))
                if not scheme:
                    continue

                period_start = _date(ws.cell(r, 2).value)
                period_end = _date(ws.cell(r, 3).value)
                if not period_start or not period_end:
                    continue

                fee_status_map = {
                    'Calculated': 'calculated',
                    'Invoiced': 'invoiced',
                    'Paid': 'paid',
                }

                ManagementFeeSchedule.objects.get_or_create(
                    scheme=scheme,
                    period_start=period_start,
                    period_end=period_end,
                    defaults={
                        'fee_basis_amount': _d(ws.cell(r, 4).value, 0),
                        'fee_rate': _d(ws.cell(r, 5).value, 0),
                        'fee_amount': _d(ws.cell(r, 6).value, 0),
                        'gst_amount': _d(ws.cell(r, 7).value, 0),
                        'total_fee_with_gst': _d(ws.cell(r, 8).value),
                        'fee_status': fee_status_map.get(
                            _str(ws.cell(r, 9).value), 'calculated'),
                        'invoice_number': _str(ws.cell(r, 10).value),
                        'invoice_date': _date(ws.cell(r, 11).value),
                    },
                )
                fee_count += 1
            if fee_count:
                self.stdout.write(f'    Management Fees: {fee_count}')

    def _import_compliance(self, wb, fund, schemes):
        """Import compliance data (SEBI reports, AML, calendar)."""
        # SEBI Reports
        for sheet_name in ['Compliance & SEBI Reports', 'Compliance']:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]

            # SEBI Report Filings
            rpt_start = find_section_rows(ws, 'SEBI REPORT FILINGS')
            if rpt_start:
                rpt_count = 0
                for r in range(rpt_start + 1, ws.max_row + 1):
                    rpt_type = ws.cell(r, 1).value
                    if not rpt_type or str(rpt_type).strip() == 'Report Type':
                        continue
                    if _is_section_header(rpt_type):
                        break

                    type_map = {'QAR': 'qar', 'AAR': 'aar'}
                    status_map = {
                        'Filed': 'filed', 'Pending': 'not_started',
                        'Rejected': 'rejected',
                    }

                    due_date = _date(ws.cell(r, 3).value)
                    if not due_date:
                        continue

                    SEBIReport.objects.get_or_create(
                        fund=fund,
                        report_type=type_map.get(_str(rpt_type), 'qar'),
                        due_date=due_date,
                        defaults={
                            'reporting_period_start': due_date.replace(day=1),
                            'reporting_period_end': due_date,
                            'filing_status': status_map.get(
                                _str(ws.cell(r, 5).value), 'not_started'),
                            'filed_date': _date(ws.cell(r, 4).value),
                            'si_portal_reference_number': _str(ws.cell(r, 6).value),
                        },
                    )
                    rpt_count += 1
                if rpt_count:
                    self.stdout.write(f'    SEBI Reports: {rpt_count}')

            # Compliance Calendar
            cal_start = find_section_rows(ws, 'COMPLIANCE CALENDAR')
            if cal_start:
                cal_count = 0
                type_map = {
                    'Quarterly AIF Report': 'sebi_qar',
                    'Annual AIF Report': 'sebi_aar',
                    'CTR Preparation': 'ctr_preparation',
                    'GST Filing': 'gst_filing',
                    'TDS Filing': 'tds_filing',
                    'Custodian Report': 'custodian_report',
                    'NAV Declaration': 'nav_declaration',
                    'IFSCA Compliance': 'other',
                }
                status_map_cal = {
                    'Filed': 'completed', 'Completed': 'completed',
                    'Overdue': 'overdue', 'Pending': 'upcoming',
                    'In Progress': 'in_progress',
                }
                for r in range(cal_start + 1, ws.max_row + 1):
                    event_name = ws.cell(r, 1).value
                    if not event_name or str(event_name).strip() == 'Event Name':
                        continue
                    if _is_section_header(event_name):
                        break

                    due_date = _date(ws.cell(r, 3).value)
                    if not due_date:
                        continue

                    raw_type = _str(ws.cell(r, 2).value)
                    ComplianceCalendar.objects.get_or_create(
                        organization=fund.organization,
                        fund=fund,
                        title=_str(event_name),
                        due_date=due_date,
                        defaults={
                            'compliance_type': type_map.get(raw_type, 'other'),
                            'status': status_map_cal.get(
                                _str(ws.cell(r, 4).value), 'upcoming'),
                            'completed_date': _date(ws.cell(r, 5).value),
                            'notes': _str(ws.cell(r, 6).value),
                        },
                    )
                    cal_count += 1
                if cal_count:
                    self.stdout.write(f'    Compliance Calendar: {cal_count}')

        # AML Due Diligence
        for sheet_name in ['Investors & AML']:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            aml_start = find_section_rows(ws, 'AML DUE DILIGENCE')
            if not aml_start:
                continue

            aml_count = 0
            for r in range(aml_start + 1, ws.max_row + 1):
                inv_name = ws.cell(r, 1).value
                if not inv_name or str(inv_name).strip() == 'Investor Name':
                    continue
                if _is_section_header(inv_name):
                    break

                investor = Investor.objects.filter(
                    organization=fund.organization,
                    investor_name=_str(inv_name)).first()
                if not investor:
                    continue

                risk_map = {
                    'High': 'high', 'Medium': 'normal', 'Low': 'low',
                    'Very High': 'very_high',
                }

                ubo_name = _str(ws.cell(r, 5).value)
                AMLDueDiligence.objects.get_or_create(
                    investor=investor,
                    defaults={
                        'risk_rating': risk_map.get(
                            _str(ws.cell(r, 2).value), 'normal'),
                        'beneficial_owner_identified': _bool(ws.cell(r, 4).value),
                        'beneficial_owner_details': (
                            {'name': ubo_name} if ubo_name else {}),
                        'is_land_border_country_investor': bool(
                            _d(ws.cell(r, 6).value, 0)),
                        'exceeds_50pct_threshold': _bool(ws.cell(r, 7).value),
                        'str_filed': _bool(ws.cell(r, 8).value),
                        'str_reference': _str(ws.cell(r, 9).value),
                        'risk_assessment_date': _date(ws.cell(r, 10).value),
                        'risk_notes': _str(ws.cell(r, 12).value),
                    },
                )
                aml_count += 1
            if aml_count:
                self.stdout.write(f'    AML Due Diligence: {aml_count}')

        # Compliance Test Reports (CTR)
        for sheet_name in ['Compliance & SEBI Reports', 'Compliance']:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            ctr_start = find_section_rows(ws, 'COMPLIANCE TEST REPORT')
            if not ctr_start:
                continue

            # Group by financial year to create CTR + checklist items
            first_scheme = list(schemes.values())[0] if schemes else None
            if not first_scheme:
                continue

            ctr_items = {}
            for r in range(ctr_start + 1, ws.max_row + 1):
                fy = ws.cell(r, 1).value
                if not fy or str(fy).strip() == 'Test Year':
                    continue
                if _is_section_header(fy):
                    break

                fy_str = _str(fy)
                if fy_str not in ctr_items:
                    ctr_items[fy_str] = []
                ctr_items[fy_str].append(r)

            ctr_count = 0
            for fy_str, rows in ctr_items.items():
                # Determine overall status
                results = [_str(ws.cell(r, 4).value) for r in rows]
                if any(s == 'Fail' for s in results):
                    overall = 'non_compliant'
                elif all(s == 'Pass' for s in results):
                    overall = 'compliant'
                else:
                    overall = 'partially_compliant'

                ctr, _ = ComplianceTestReport.objects.get_or_create(
                    scheme=first_scheme,
                    financial_year=fy_str,
                    defaults={
                        'overall_compliance_status': overall,
                        'report_status': 'finalized',
                    },
                )

                for idx, r in enumerate(rows, 1):
                    result_map = {
                        'Pass': 'compliant', 'Fail': 'non_compliant',
                        'N/A': 'not_applicable',
                    }
                    CTRChecklistItem.objects.get_or_create(
                        compliance_test_report=ctr,
                        check_number=idx,
                        defaults={
                            'regulation_reference': f'CTR-{fy_str}-{idx}',
                            'description': _str(ws.cell(r, 2).value),
                            'compliance_status': result_map.get(
                                _str(ws.cell(r, 4).value), 'pending_review'),
                            'remarks': _str(ws.cell(r, 7).value),
                        },
                    )
                ctr_count += 1
            if ctr_count:
                self.stdout.write(f'    Compliance Test Reports: {ctr_count}')

        # SEBI Circulars (Fund 4 has these)
        for sheet_name in ['Compliance & SEBI Reports', 'Compliance']:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            circ_start = find_section_rows(ws, 'SEBI CIRCULARS')
            if not circ_start:
                continue

            circ_count = 0
            for r in range(circ_start + 1, ws.max_row + 1):
                circ_num = ws.cell(r, 1).value
                if not circ_num or str(circ_num).strip() == 'Circular Number':
                    continue
                if _is_section_header(circ_num):
                    break

                priority_map = {
                    'Critical': 'critical', 'High': 'high',
                    'Medium': 'medium', 'Low': 'low',
                }
                status_map_circ = {
                    'Completed': 'completed', 'Pending': 'pending',
                    'In Progress': 'in_progress',
                }

                circular, _ = SEBICircular.objects.get_or_create(
                    organization=fund.organization,
                    circular_number=_str(circ_num),
                    defaults={
                        'title': _str(ws.cell(r, 2).value),
                        'circular_date': _date(ws.cell(r, 3).value) or date.today(),
                        'impact_level': priority_map.get(
                            _str(ws.cell(r, 5).value), 'medium'),
                        'compliance_deadline': _date(ws.cell(r, 9).value),
                    },
                )

                # Create action item
                action_desc = _str(ws.cell(r, 7).value)
                if action_desc:
                    CircularAction.objects.get_or_create(
                        circular=circular,
                        fund=fund,
                        action_title=action_desc[:255],
                        defaults={
                            'action_description': action_desc,
                            'priority': priority_map.get(
                                _str(ws.cell(r, 5).value), 'medium'),
                            'status': status_map_circ.get(
                                _str(ws.cell(r, 6).value), 'pending'),
                            'due_date': _date(ws.cell(r, 9).value),
                            'completion_date': _date(ws.cell(r, 10).value),
                            'completion_notes': _str(ws.cell(r, 11).value),
                        },
                    )
                circ_count += 1
            if circ_count:
                self.stdout.write(f'    SEBI Circulars: {circ_count}')

        # PPM Amendments (Fund 4 has these)
        for sheet_name in ['Compliance & SEBI Reports', 'Compliance']:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            ppm_start = find_section_rows(ws, 'PPM AMENDMENTS')
            if not ppm_start:
                continue

            ppm_count = 0
            for r in range(ppm_start + 1, ws.max_row + 1):
                amend_num = ws.cell(r, 1).value
                if not amend_num or str(amend_num).strip() == 'Amendment #':
                    continue
                if _is_section_header(amend_num):
                    break
                try:
                    amend_num = int(amend_num)
                except (ValueError, TypeError):
                    continue

                PPMAmendment.objects.get_or_create(
                    fund=fund,
                    amendment_number=amend_num,
                    defaults={
                        'amendment_type': 'other',
                        'title': _str(ws.cell(r, 3).value)[:255],
                        'description': _str(ws.cell(r, 3).value),
                        'board_approval_date': _date(ws.cell(r, 2).value),
                        'effective_date': _date(ws.cell(r, 5).value),
                        'approval_status': 'effective',
                        'notes': _str(ws.cell(r, 7).value),
                    },
                )
                ppm_count += 1
            if ppm_count:
                self.stdout.write(f'    PPM Amendments: {ppm_count}')

    def _import_board_meetings(self, wb, investments):
        """Import board meetings."""
        if 'Board Meetings' not in wb.sheetnames:
            return
        ws = wb['Board Meetings']

        count = 0
        for r in range(1, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if not name or str(name).strip() == 'Company Name':
                continue
            if _is_section_header(name):
                break

            name = _str(name)
            inv = None
            for key, i in investments.items():
                if key.startswith(f'{name}|'):
                    inv = i
                    break
            if not inv:
                continue

            meeting_date = _date(ws.cell(r, 2).value)
            if not meeting_date:
                continue

            BoardMeeting.objects.get_or_create(
                investment=inv,
                meeting_date=meeting_date,
                defaults={
                    'meeting_number': ws.cell(r, 3).value,
                    'agenda': _str(ws.cell(r, 4).value),
                    'attendees': [a.strip() for a in _str(ws.cell(r, 5).value).split(',') if a.strip()],
                    'resolutions': [_str(ws.cell(r, 6).value)] if ws.cell(r, 6).value else [],
                    'next_meeting_date': _date(ws.cell(r, 7).value),
                },
            )
            count += 1

        if count:
            self.stdout.write(f'    Board Meetings: {count}')

    def _build_portfolio_hierarchy(self, org, data_dir, files):
        """Build PortfolioNode hierarchy from all fund hierarchy sheets.

        Uses a single org-scoped snapshot that accumulates nodes across
        multiple file imports. When a fund's hierarchy is re-imported, the
        old nodes for that fund are replaced (not duplicated).
        """
        self.stdout.write(f'\n{"="*60}')
        self.stdout.write('Building Portfolio Hierarchy (PortfolioNode tree)')
        self.stdout.write(f'{"="*60}')

        # Get or create the org's active snapshot (one per org, not per file)
        snapshot = PortfolioSnapshot.objects.filter(
            organization=org, is_active=True, source='excel_parse',
        ).first()
        if not snapshot:
            # Deactivate any old orphan snapshots for this org
            PortfolioSnapshot.objects.filter(
                organization=org, is_active=True, source='excel_parse',
            ).update(is_active=False)
            snapshot = PortfolioSnapshot.objects.create(
                organization=org,
                schema_version='2.0',
                base_currency='INR',
                source='excel_parse',
                is_active=True,
            )

        node_count = 0
        node_map = {}  # node_id → PortfolioNode

        # Collect fund node_ids from incoming files so we can remove old nodes
        # for the same fund(s) before inserting fresh ones.
        fund_ids_in_files = set()

        # Collect MIS data (Monthly P&L + Budget vs Actual) from all files
        # keyed by company name → {monthly_pl: [...], budget_vs_actual: [...]}
        mis_data = {}

        for filename in files:
            filepath = os.path.join(data_dir, filename)
            if not os.path.exists(filepath):
                continue

            wb = openpyxl.load_workbook(filepath, data_only=True)
            if 'Portfolio Hierarchy' not in wb.sheetnames:
                continue

            # Parse Monthly P&L sheet if present
            if 'Monthly P&L' in wb.sheetnames:
                ws_pl = wb['Monthly P&L']
                for r in range(2, ws_pl.max_row + 1):
                    co_name = _str(ws_pl.cell(r, 1).value)
                    period = _str(ws_pl.cell(r, 2).value)
                    if not co_name or not period:
                        continue
                    if co_name not in mis_data:
                        mis_data[co_name] = {'monthly_pl': [], 'budget_vs_actual': []}
                    revenue = ws_pl.cell(r, 3).value
                    cogs = ws_pl.cell(r, 4).value
                    gross_profit = ws_pl.cell(r, 5).value
                    gp_pct_raw = ws_pl.cell(r, 6).value
                    opex = ws_pl.cell(r, 7).value
                    ebitda = ws_pl.cell(r, 8).value
                    ebitda_pct_raw = ws_pl.cell(r, 9).value
                    # GP% and EBITDA% may be stored as decimal (0.65) or percent (65)
                    gp_pct = (float(gp_pct_raw) * 100 if gp_pct_raw is not None
                              and abs(float(gp_pct_raw)) <= 1 else
                              float(gp_pct_raw) if gp_pct_raw is not None else None)
                    ebitda_pct = (float(ebitda_pct_raw) * 100 if ebitda_pct_raw is not None
                                  and abs(float(ebitda_pct_raw)) <= 1 else
                                  float(ebitda_pct_raw) if ebitda_pct_raw is not None else None)
                    mis_data[co_name]['monthly_pl'].append({
                        'period': period,
                        'revenue': float(revenue) if revenue else 0,
                        'cogs': float(cogs) if cogs else 0,
                        'gross_profit': float(gross_profit) if gross_profit else 0,
                        'gp_pct': round(gp_pct, 1) if gp_pct is not None else None,
                        'opex': float(opex) if opex else 0,
                        'ebitda': float(ebitda) if ebitda else 0,
                        'ebitda_pct': round(ebitda_pct, 1) if ebitda_pct is not None else None,
                    })

            # Parse Budget vs Actual sheet if present
            if 'Budget vs Actual' in wb.sheetnames:
                ws_bva = wb['Budget vs Actual']
                for r in range(2, ws_bva.max_row + 1):
                    co_name = _str(ws_bva.cell(r, 1).value)
                    line_item = _str(ws_bva.cell(r, 2).value)
                    if not co_name or not line_item:
                        continue
                    if co_name not in mis_data:
                        mis_data[co_name] = {'monthly_pl': [], 'budget_vs_actual': []}
                    actual = ws_bva.cell(r, 3).value
                    budget = ws_bva.cell(r, 4).value
                    mis_data[co_name]['budget_vs_actual'].append({
                        'line_item': line_item,
                        'actual': float(actual) if actual else 0,
                        'budget': float(budget) if budget else 0,
                    })

            if mis_data:
                self.stdout.write(f'  Parsed MIS data for {len(mis_data)} companies from {filename}')

            ws = wb['Portfolio Hierarchy']

            # First pass: collect fund-level node_ids from this file
            file_fund_ids = set()
            for r in range(1, ws.max_row + 1):
                level = ws.cell(r, 1).value
                if level == 'Fund':
                    nid = _str(ws.cell(r, 2).value)
                    if nid.startswith('fund_'):
                        file_fund_ids.add(nid)

            # Remove old nodes for these fund(s) from the snapshot
            # (so re-imports replace rather than duplicate)
            for fund_nid in file_fund_ids:
                PortfolioNode.objects.filter(
                    snapshot=snapshot, node_id__startswith=fund_nid,
                ).delete()
                # Also delete the fund node itself
                PortfolioNode.objects.filter(
                    snapshot=snapshot, node_id=fund_nid,
                ).delete()

            fund_ids_in_files.update(file_fund_ids)

            # Second pass: create nodes
            for r in range(1, ws.max_row + 1):
                level = ws.cell(r, 1).value
                if not level or level not in ('Fund', 'Sector', 'Segment', 'Company'):
                    continue

                node_id = _str(ws.cell(r, 2).value)
                # Skip cross-fund mapping rows (e.g., "Also in Fund(s)")
                if not node_id.startswith('fund_'):
                    continue
                name = _str(ws.cell(r, 3).value)
                parent_node_id = _str(ws.cell(r, 4).value)
                sort_order = ws.cell(r, 5).value or 0

                level_lower = level.lower()

                # Get financials from investment/company data + MIS sheets
                financials = self._compute_node_financials(
                    org, level_lower, name, node_id, mis_data)

                db_node = PortfolioNode.objects.create(
                    snapshot=snapshot,
                    node_id=node_id,
                    name=name,
                    level=level_lower,
                    parent_node_id=parent_node_id if parent_node_id else None,
                    sort_order=sort_order,
                    financials=financials,
                )
                node_map[node_id] = db_node
                node_count += 1

        # Set parent FK references — include existing nodes in the snapshot
        # so cross-file parent references resolve correctly
        existing_nodes = {
            n.node_id: n
            for n in PortfolioNode.objects.filter(snapshot=snapshot)
        }
        existing_nodes.update(node_map)

        for node_id, db_node in node_map.items():
            if db_node.parent_node_id:
                parent = existing_nodes.get(db_node.parent_node_id)
                if parent:
                    db_node.parent = parent
                    db_node.save(update_fields=['parent'])

        # Aggregate financials bottom-up: company → segment → sector → fund
        self._aggregate_hierarchy_financials(snapshot)

        self.stdout.write(self.style.SUCCESS(
            f'  Created {node_count} nodes in snapshot {snapshot.id}'))
        self.stdout.write(
            f'    Funds: {PortfolioNode.objects.filter(snapshot=snapshot, level="fund").count()}')
        self.stdout.write(
            f'    Sectors: {PortfolioNode.objects.filter(snapshot=snapshot, level="sector").count()}')
        self.stdout.write(
            f'    Segments: {PortfolioNode.objects.filter(snapshot=snapshot, level="segment").count()}')
        self.stdout.write(
            f'    Companies: {PortfolioNode.objects.filter(snapshot=snapshot, level="company").count()}')

    def _compute_node_financials(self, org, level, name, node_id, mis_data=None):
        """Compute financials for a node from Investment/Valuation/KPI + MIS data."""
        financials = {}
        mis_data = mis_data or {}

        if level == 'company':
            company = PortfolioCompany.objects.filter(
                organization=org, name=name).first()
            if not company:
                return financials

            financials['sector'] = company.sector
            financials['sub_sector'] = company.sub_sector

            # Gather investment data
            investments = Investment.objects.filter(portfolio_company=company)
            tranches = InvestmentTranche.objects.filter(
                investment__portfolio_company=company)
            cost_basis = float(sum(t.amount for t in tranches if t.amount))

            # Latest valuation
            latest_val = Valuation.objects.filter(
                investment__portfolio_company=company
            ).order_by('-valuation_date').first()

            fair_value = float(latest_val.fair_value) if latest_val and latest_val.fair_value else 0

            # ── MIS data (Monthly P&L + Budget vs Actual) ──
            company_mis = mis_data.get(name, {})
            monthly_pl = company_mis.get('monthly_pl', [])
            bva_rows = company_mis.get('budget_vs_actual', [])

            if monthly_pl:
                financials['monthly_pl'] = monthly_pl
            if bva_rows:
                financials['budget_vs_actual'] = bva_rows

            # ── Build summary from MIS data (preferred) or KPI fallback ──
            summary = {}

            if monthly_pl:
                # Compute YTD from CY2025 monthly rows
                ytd_rows = [r for r in monthly_pl if r['period'].startswith('2025')]
                if ytd_rows:
                    ytd_revenue = sum(r.get('revenue', 0) for r in ytd_rows)
                    ytd_cogs = sum(r.get('cogs', 0) for r in ytd_rows)
                    ytd_gp = sum(r.get('gross_profit', 0) for r in ytd_rows)
                    ytd_opex = sum(r.get('opex', 0) for r in ytd_rows)
                    ytd_ebitda = sum(r.get('ebitda', 0) for r in ytd_rows)

                    summary['ytd_revenue'] = round(ytd_revenue, 2)
                    summary['ytd_cogs'] = round(ytd_cogs, 2)
                    summary['ytd_gross_profit'] = round(ytd_gp, 2)
                    summary['ytd_opex'] = round(ytd_opex, 2)
                    summary['ytd_ebitda'] = round(ytd_ebitda, 2)
                    if ytd_revenue:
                        summary['gp_pct'] = round(ytd_gp / ytd_revenue * 100, 1)
                        summary['ebitda_pct'] = round(ytd_ebitda / ytd_revenue * 100, 1)

                # Latest month as current revenue/ebitda
                latest = monthly_pl[-1]
                summary['revenue'] = round(latest.get('revenue', 0), 2)
                summary['cogs'] = round(latest.get('cogs', 0), 2)
                summary['gross_profit'] = round(latest.get('gross_profit', 0), 2)
                summary['opex'] = round(latest.get('opex', 0), 2)
                summary['ebitda'] = round(latest.get('ebitda', 0), 2)

            else:
                # KPI fallback (original logic)
                revenue = None
                ebitda_margin = None

                rev_kpis = PortfolioKPI.objects.filter(
                    portfolio_company=company,
                    kpi_definition__name='Revenue',
                ).order_by('-period_end_date')
                if rev_kpis.exists():
                    revenue = float(rev_kpis.first().value)

                if revenue is None:
                    mrr_kpis = PortfolioKPI.objects.filter(
                        portfolio_company=company,
                        kpi_definition__name='MRR',
                    ).order_by('-period_end_date')
                    if mrr_kpis.exists():
                        revenue = float(mrr_kpis.first().value) * 12

                if revenue is None:
                    arr_kpis = PortfolioKPI.objects.filter(
                        portfolio_company=company,
                        kpi_definition__name='ARR',
                    ).order_by('-period_end_date')
                    if arr_kpis.exists():
                        revenue = float(arr_kpis.first().value)

                ebitda_kpis = PortfolioKPI.objects.filter(
                    portfolio_company=company,
                    kpi_definition__name='EBITDA Margin',
                ).order_by('-period_end_date')
                if ebitda_kpis.exists():
                    ebitda_margin = float(ebitda_kpis.first().value)

                if revenue is not None:
                    summary['revenue'] = round(revenue, 2)
                    if ebitda_margin is not None:
                        ebitda = revenue * ebitda_margin / 100
                        summary['ebitda'] = round(ebitda, 2)
                        summary['ebitda_pct'] = round(ebitda_margin, 2)

            # Budget data for summary
            if bva_rows:
                for bva in bva_rows:
                    li = bva.get('line_item', '').lower()
                    if li == 'revenue':
                        summary['ytd_budget_revenue'] = round(bva.get('budget', 0), 2)
                        summary['budget_revenue'] = round(bva.get('budget', 0), 2)
                    elif li == 'ebitda':
                        summary['ytd_budget_ebitda'] = round(bva.get('budget', 0), 2)
                        summary['budget_ebitda'] = round(bva.get('budget', 0), 2)

            # Investment metrics
            summary['cost_basis'] = round(cost_basis, 2)
            summary['fair_value'] = round(fair_value, 2)
            summary['investment_count'] = investments.count()

            if cost_basis and fair_value:
                summary['moic'] = round(fair_value / cost_basis, 2)

            if summary:
                financials['summary'] = summary

        return financials

    def _aggregate_hierarchy_financials(self, snapshot):
        """Aggregate financials bottom-up through the hierarchy.

        For each non-company node (segment, sector, fund), sum the
        summary fields, monthly_pl, and budget_vs_actual from children.
        """
        SUMMARY_FIELDS = [
            'revenue', 'cogs', 'gross_profit', 'opex', 'ebitda',
            'cost_basis', 'fair_value', 'investment_count',
        ]
        YTD_FIELDS = [
            'ytd_revenue', 'ytd_cogs', 'ytd_gross_profit',
            'ytd_opex', 'ytd_ebitda',
            'ytd_budget_revenue', 'budget_revenue',
            'ytd_budget_ebitda', 'budget_ebitda',
        ]

        # Process in bottom-up order: segment → sector → fund
        for level in ('segment', 'sector', 'fund'):
            parent_nodes = PortfolioNode.objects.filter(
                snapshot=snapshot, level=level)

            for parent in parent_nodes:
                children = PortfolioNode.objects.filter(
                    snapshot=snapshot, parent_node_id=parent.node_id)

                if not children.exists():
                    continue

                agg = {}
                agg_monthly = {}  # period → {revenue, cogs, ...}
                agg_bva = {}      # line_item → {actual, budget}

                for child in children:
                    child_fin = child.financials or {}
                    child_summary = child_fin.get('summary', {})

                    for field in SUMMARY_FIELDS + YTD_FIELDS:
                        val = child_summary.get(field)
                        if isinstance(val, (int, float)):
                            agg[field] = agg.get(field, 0) + val

                    # Aggregate monthly_pl by period
                    for row in child_fin.get('monthly_pl', []) or []:
                        period = row.get('period')
                        if not period:
                            continue
                        if period not in agg_monthly:
                            agg_monthly[period] = {
                                'period': period, 'revenue': 0, 'cogs': 0,
                                'gross_profit': 0, 'opex': 0, 'ebitda': 0,
                            }
                        for k in ('revenue', 'cogs', 'gross_profit', 'opex', 'ebitda'):
                            agg_monthly[period][k] += (row.get(k) or 0)

                    # Aggregate budget_vs_actual by line_item
                    for row in child_fin.get('budget_vs_actual', []) or []:
                        li = row.get('line_item', '')
                        if not li:
                            continue
                        if li not in agg_bva:
                            agg_bva[li] = {'line_item': li, 'actual': 0, 'budget': 0}
                        agg_bva[li]['actual'] += (row.get('actual') or 0)
                        agg_bva[li]['budget'] += (row.get('budget') or 0)

                # Compute derived percentages for summary
                rev = agg.get('revenue')
                if rev:
                    gp = agg.get('gross_profit')
                    ebitda = agg.get('ebitda')
                    if gp is not None:
                        agg['gp_pct'] = round(gp / rev * 100, 2)
                    if ebitda is not None:
                        agg['ebitda_pct'] = round(ebitda / rev * 100, 2)

                # YTD-based percentages
                ytd_rev = agg.get('ytd_revenue')
                if ytd_rev:
                    ytd_gp = agg.get('ytd_gross_profit')
                    ytd_ebitda = agg.get('ytd_ebitda')
                    if ytd_gp is not None and 'gp_pct' not in agg:
                        agg['gp_pct'] = round(ytd_gp / ytd_rev * 100, 2)
                    if ytd_ebitda is not None and 'ebitda_pct' not in agg:
                        agg['ebitda_pct'] = round(ytd_ebitda / ytd_rev * 100, 2)

                cost = agg.get('cost_basis')
                fv = agg.get('fair_value')
                if cost and fv:
                    agg['moic'] = round(fv / cost, 2)

                # Round all numeric values
                for k, v in agg.items():
                    if isinstance(v, float):
                        agg[k] = round(v, 2)

                # Build aggregated monthly_pl with derived percentages
                monthly_pl_agg = []
                for period in sorted(agg_monthly.keys()):
                    row = agg_monthly[period]
                    for k in ('revenue', 'cogs', 'gross_profit', 'opex', 'ebitda'):
                        row[k] = round(row[k], 2)
                    if row['revenue']:
                        row['gp_pct'] = round(row['gross_profit'] / row['revenue'] * 100, 1)
                        row['ebitda_pct'] = round(row['ebitda'] / row['revenue'] * 100, 1)
                    monthly_pl_agg.append(row)

                # Build aggregated budget_vs_actual
                bva_agg = []
                for li in ('Revenue', 'COGS', 'Gross Profit', 'Operating Expenses', 'EBITDA'):
                    if li in agg_bva:
                        row = agg_bva[li]
                        row['actual'] = round(row['actual'], 2)
                        row['budget'] = round(row['budget'], 2)
                        bva_agg.append(row)

                # Update the parent node's financials
                fin = parent.financials or {}
                fin['summary'] = agg
                if monthly_pl_agg:
                    fin['monthly_pl'] = monthly_pl_agg
                if bva_agg:
                    fin['budget_vs_actual'] = bva_agg
                parent.financials = fin
                parent.save(update_fields=['financials'])

    def _print_summary(self, org):
        """Print summary of imported data."""
        self.stdout.write(f'\n{"="*60}')
        self.stdout.write('IMPORT SUMMARY')
        self.stdout.write(f'{"="*60}')
        self.stdout.write(f'  Organization: {org.name}')
        self.stdout.write(f'  Users: {User.objects.filter(organization=org).count()}')
        self.stdout.write(f'  Entities: {Entity.objects.filter(organization=org).count()}')
        self.stdout.write(f'  Funds: {Fund.objects.filter(organization=org).count()}')
        self.stdout.write(f'  Schemes: {Scheme.objects.filter(fund__organization=org).count()}')
        self.stdout.write(f'  Investors: {Investor.objects.filter(organization=org).count()}')
        self.stdout.write(f'  Commitments: {Commitment.objects.filter(scheme__fund__organization=org).count()}')
        self.stdout.write(f'  Capital Calls: {CapitalCall.objects.filter(scheme__fund__organization=org).count()}')
        self.stdout.write(f'  Portfolio Companies: {PortfolioCompany.objects.filter(organization=org).count()}')
        self.stdout.write(f'  Investments: {Investment.objects.filter(scheme__fund__organization=org).count()}')
        self.stdout.write(f'  Tranches: {InvestmentTranche.objects.filter(investment__scheme__fund__organization=org).count()}')
        self.stdout.write(f'  Valuations: {Valuation.objects.filter(investment__scheme__fund__organization=org).count()}')
        self.stdout.write(f'  KPI Definitions: {KPIDefinition.objects.filter(organization=org).count()}')
        self.stdout.write(f'  KPI Values: {PortfolioKPI.objects.filter(investment__scheme__fund__organization=org).count()}')
        self.stdout.write(f'  NAV Records: {NAVRecord.objects.filter(scheme__fund__organization=org).count()}')
        self.stdout.write(f'  Exit Events: {ExitEvent.objects.filter(investment__scheme__fund__organization=org).count()}')
        self.stdout.write(f'  Distributions: {Distribution.objects.filter(scheme__fund__organization=org).count()}')
        self.stdout.write(f'  Chart of Accounts: {ChartOfAccounts.objects.filter(organization=org).count()}')
        self.stdout.write(f'  Portfolio Nodes: {PortfolioNode.objects.filter(snapshot__source="excel_parse", snapshot__is_active=True).count()}')
        self.stdout.write(f'\n  Login as: rajesh.trivedi / rajesh.trivedi123')
        self.stdout.write(f'  This user has admin access to ALL 5 funds')
