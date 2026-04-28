"""
Add Monthly P&L and Budget vs Actual sheets to all 5 mock fund Excel files.

This populates the financial data needed by the compare panel:
  - financials.monthly_pl  → Monthly P&L sheet
  - financials.budget_vs_actual → Budget vs Actual sheet
  - financials.summary (enhanced with COGS, gross_profit, opex, ytd_* fields)

Data is realistic and proportional to each company's existing KPI data
(Revenue, MRR, ARR, EBITDA Margin) where available. For companies without
KPI data, reasonable industry-typical figures are generated.
"""

import os
import random
from copy import copy

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers

MOCK_DIR = os.path.dirname(os.path.abspath(__file__))

# ─── Company financial profiles ──────────────────────────────────────────
# For each company: annual_revenue (INR), cogs_pct, opex_pct, growth_trend
# Revenue in INR (same unit as KPIs). Budget is ~5-10% above/below actual.
# monthly revenue = annual / 12 * seasonal_factor * noise

COMPANY_PROFILES = {
    # ──── Fund 1: Avendus Future Leaders Fund II ────
    "TechCorp India Pvt Ltd": {
        "annual_revenue": 62_400_000,  # MRR ~5.2M → ARR 62.4M
        "cogs_pct": 0.35,
        "opex_pct": 0.40,
        "growth": 1.04,  # 4% QoQ
        "sector": "Enterprise SaaS",
    },
    "GreenEnergy Solutions Ltd": {
        "annual_revenue": 25_000_000,  # ARR=25M
        "cogs_pct": 0.55,
        "opex_pct": 0.25,
        "growth": 1.03,
        "sector": "Clean Energy",
    },
    "HealthFirst Diagnostics Pvt Ltd": {
        "annual_revenue": 45_500_000,  # Revenue KPI
        "cogs_pct": 0.50,
        "opex_pct": 0.38,  # EBITDA Margin 12% => opex=38%
        "growth": 1.02,
        "sector": "Healthcare Diagnostics",
    },
    "FinStack Technologies Pvt Ltd": {
        "annual_revenue": 1_440_000,  # MRR 120K → ARR 1.44M (early stage)
        "cogs_pct": 0.20,
        "opex_pct": 0.65,  # pre-profit, burn
        "growth": 1.08,
        "sector": "Lending Infra",
    },
    "AgriVista Farms Pvt Ltd": {
        "annual_revenue": 8_500_000,
        "cogs_pct": 0.60,
        "opex_pct": 0.25,
        "growth": 1.02,
        "sector": "AgriTech",
    },
    # ──── Fund 2: Blume Ventures Fund IV ────
    "BatchOne AI Pvt Ltd": {
        "annual_revenue": 46_200_000,  # MRR ~3.85M → ARR 46.2M
        "cogs_pct": 0.30,
        "opex_pct": 0.45,
        "growth": 1.06,
        "sector": "AI/ML",
    },
    "CloudKitchen Co Pvt Ltd": {
        "annual_revenue": 12_000_000,  # Revenue KPI
        "cogs_pct": 0.65,
        "opex_pct": 0.30,  # pre-profit, high COGS (food)
        "growth": 1.03,
        "sector": "Cloud Kitchen",
    },
    "PayGrid Technologies Pvt Ltd": {
        "annual_revenue": 18_000_000,
        "cogs_pct": 0.25,
        "opex_pct": 0.50,
        "growth": 1.05,
        "sector": "B2B Payments",
    },
    "NanoMed Biosciences Pvt Ltd": {
        "annual_revenue": 2_500_000,  # Grant Revenue
        "cogs_pct": 0.40,
        "opex_pct": 0.50,
        "growth": 1.01,
        "sector": "Biotech",
    },
    "Lingua AI Pte Ltd": {
        "annual_revenue": 6_000_000,
        "cogs_pct": 0.25,
        "opex_pct": 0.55,
        "growth": 1.07,
        "sector": "NLP",
    },
    "FarmStack Agri Pvt Ltd": {
        "annual_revenue": 45_000_000,  # GMV=45M, take rate ~10%
        "cogs_pct": 0.70,
        "opex_pct": 0.20,
        "growth": 1.04,
        "sector": "Agri Supply Chain",
    },
    # ──── Fund 3: Motilal Oswal PE Fund IV ────
    "Metropolis Healthcare Ltd": {
        "annual_revenue": 180_000_000,
        "cogs_pct": 0.45,
        "opex_pct": 0.30,
        "growth": 1.02,
        "sector": "Healthcare Diagnostics",
    },
    "Go Fashion (India) Ltd": {
        "annual_revenue": 95_000_000,
        "cogs_pct": 0.50,
        "opex_pct": 0.28,
        "growth": 1.03,
        "sector": "Fashion Retail",
    },
    "Prataap Snacks Ltd": {
        "annual_revenue": 450_000_000,  # Revenue KPI
        "cogs_pct": 0.62,
        "opex_pct": 0.285,  # EBITDA Margin 9.5%
        "growth": 1.01,
        "sector": "FMCG",
    },
    "Suryoday Small Finance Bank": {
        "annual_revenue": 280_000_000,  # NII = Revenue proxy for banks
        "cogs_pct": 0.0,  # Banks don't have COGS; use provisions
        "opex_pct": 0.65,
        "growth": 1.02,
        "sector": "Small Finance Bank",
        "is_bank": True,
    },
    "Clean Science & Technology Ltd": {
        "annual_revenue": 120_000_000,
        "cogs_pct": 0.45,
        "opex_pct": 0.22,
        "growth": 1.02,
        "sector": "Specialty Chemicals",
    },
    "AutoParts India Pvt Ltd": {
        "annual_revenue": 112_500_000,  # Revenue KPI
        "cogs_pct": 0.52,
        "opex_pct": 0.26,  # EBITDA Margin 22%
        "growth": 1.01,
        "sector": "Auto Components",
    },
    "PharmaCo Generics Pvt Ltd": {
        "annual_revenue": 85_000_000,
        "cogs_pct": 0.55,
        "opex_pct": 0.25,
        "growth": 1.02,
        "sector": "Generic Drugs",
    },
    # ──── Fund 4: Stellaris VP Fund III (GIFT City) ────
    "Ather Energy Pvt Ltd": {
        "annual_revenue": 250_000_000,
        "cogs_pct": 0.70,
        "opex_pct": 0.25,
        "growth": 1.05,
        "sector": "Electric Vehicles",
    },
    "Perfios Software Solutions Pvt Ltd": {
        "annual_revenue": 40_000_000,
        "cogs_pct": 0.20,
        "opex_pct": 0.50,
        "growth": 1.06,
        "sector": "Credit Infrastructure",
    },
    "Licious (Delightful Gourmet Pvt Ltd)": {
        "annual_revenue": 80_000_000,
        "cogs_pct": 0.65,
        "opex_pct": 0.28,
        "growth": 1.03,
        "sector": "D2C Fresh",
    },
    "Acko Technology & Services Pvt Ltd": {
        "annual_revenue": 55_000_000,
        "cogs_pct": 0.60,
        "opex_pct": 0.30,
        "growth": 1.04,
        "sector": "Digital Insurance",
    },
    "Navi Technologies Ltd": {
        "annual_revenue": 35_000_000,
        "cogs_pct": 0.15,
        "opex_pct": 0.55,
        "growth": 1.05,
        "sector": "Digital Lending",
    },
    # ──── Fund 5: Nippon India AIF Cat III ────
    "HDFC Bank Ltd": {
        "annual_revenue": 2_500_000_000,
        "cogs_pct": 0.0,
        "opex_pct": 0.55,
        "growth": 1.01,
        "sector": "Private Banks",
        "is_bank": True,
    },
    "Reliance Industries Ltd": {
        "annual_revenue": 9_500_000_000,
        "cogs_pct": 0.65,
        "opex_pct": 0.18,
        "growth": 1.02,
        "sector": "Diversified Conglomerate",
    },
    "Infosys Ltd": {
        "annual_revenue": 1_800_000_000,
        "cogs_pct": 0.55,
        "opex_pct": 0.20,
        "growth": 1.01,
        "sector": "IT Services",
    },
    "Bharti Airtel Ltd": {
        "annual_revenue": 1_500_000_000,
        "cogs_pct": 0.35,
        "opex_pct": 0.25,
        "growth": 1.02,
        "sector": "Telecom",
    },
    "Tata Motors Ltd": {
        "annual_revenue": 4_200_000_000,
        "cogs_pct": 0.72,
        "opex_pct": 0.18,
        "growth": 1.02,
        "sector": "Automotive",
    },
    "ITC Ltd": {
        "annual_revenue": 700_000_000,
        "cogs_pct": 0.50,
        "opex_pct": 0.20,
        "growth": 1.01,
        "sector": "FMCG",
    },
    "Adani Enterprises Ltd": {
        "annual_revenue": 3_000_000_000,
        "cogs_pct": 0.70,
        "opex_pct": 0.18,
        "growth": 1.03,
        "sector": "Infrastructure",
    },
    "Deepak Nitrite Ltd": {
        "annual_revenue": 85_000_000,
        "cogs_pct": 0.55,
        "opex_pct": 0.18,
        "growth": 1.02,
        "sector": "Specialty Chemicals",
    },
    "Muthoot Finance Ltd": {
        "annual_revenue": 120_000_000,
        "cogs_pct": 0.0,
        "opex_pct": 0.45,
        "growth": 1.02,
        "sector": "NBFC",
        "is_bank": True,
    },
    "Shriram Transport Finance": {
        "annual_revenue": 95_000_000,
        "cogs_pct": 0.0,
        "opex_pct": 0.50,
        "growth": 1.01,
        "sector": "NBFC",
        "is_bank": True,
    },
}

# Seasonal factors (Jan=1 .. Dec=12) — mild seasonality
SEASONAL = [0.85, 0.88, 0.95, 1.0, 1.02, 1.05, 0.98, 1.0, 1.05, 1.08, 1.10, 1.15]

# Reporting months (CY2024 + CY2025 YTD through May)
MONTHS_2024 = [f"2024-{m:02d}" for m in range(1, 13)]
MONTHS_2025 = [f"2025-{m:02d}" for m in range(1, 6)]  # Jan-May 2025
ALL_MONTHS = MONTHS_2024 + MONTHS_2025

# Styling
HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=10)
MONEY_FMT = '#,##0'
PCT_FMT = '0.0%'
THIN_BORDER = Border(
    left=Side(style='thin', color='333333'),
    right=Side(style='thin', color='333333'),
    top=Side(style='thin', color='333333'),
    bottom=Side(style='thin', color='333333'),
)


def _noise(base, pct=0.05):
    """Add random noise ±pct to a base value."""
    return base * (1 + random.uniform(-pct, pct))


def generate_monthly_pl(company_name: str) -> list[dict]:
    """Generate 17 months of P&L data (Jan 2024 – May 2025) for a company."""
    profile = COMPANY_PROFILES.get(company_name)
    if not profile:
        return []

    annual_rev = profile["annual_revenue"]
    cogs_pct = profile["cogs_pct"]
    opex_pct = profile["opex_pct"]
    growth = profile["growth"]
    is_bank = profile.get("is_bank", False)

    monthly_base = annual_rev / 12
    rows = []

    for i, period in enumerate(ALL_MONTHS):
        month_num = int(period.split("-")[1])
        year = int(period.split("-")[0])
        seasonal = SEASONAL[month_num - 1]

        # Apply growth: compounding from Jan 2024
        months_elapsed = i
        growth_factor = growth ** (months_elapsed / 3)  # quarterly compounding

        revenue = round(_noise(monthly_base * seasonal * growth_factor, 0.04))

        if is_bank:
            # Banks: Revenue = NII, no COGS, provisions as pseudo-COGS
            cogs = round(revenue * _noise(0.15, 0.03))  # provision for bad debts
        else:
            cogs = round(revenue * _noise(cogs_pct, 0.02))

        gross_profit = revenue - cogs
        gp_pct = round(gross_profit / revenue * 100, 1) if revenue else 0

        opex = round(revenue * _noise(opex_pct, 0.03))
        ebitda = gross_profit - opex
        ebitda_pct = round(ebitda / revenue * 100, 1) if revenue else 0

        rows.append({
            "company": company_name,
            "period": period,
            "revenue": revenue,
            "cogs": cogs,
            "gross_profit": gross_profit,
            "gp_pct": gp_pct,
            "opex": opex,
            "ebitda": ebitda,
            "ebitda_pct": ebitda_pct,
        })

    return rows


def generate_budget_vs_actual(company_name: str, monthly_rows: list[dict]) -> list[dict]:
    """Generate Budget vs Actual rows from monthly P&L data.

    Budget is set at ~95-108% of actual (some over, some under).
    We produce YTD figures for CY2025 (Jan-May 2025).
    """
    if not monthly_rows:
        return []

    # Sum CY2025 YTD actuals
    ytd_rows = [r for r in monthly_rows if r["period"].startswith("2025")]
    if not ytd_rows:
        return []

    ytd_revenue = sum(r["revenue"] for r in ytd_rows)
    ytd_cogs = sum(r["cogs"] for r in ytd_rows)
    ytd_gross_profit = sum(r["gross_profit"] for r in ytd_rows)
    ytd_opex = sum(r["opex"] for r in ytd_rows)
    ytd_ebitda = sum(r["ebitda"] for r in ytd_rows)

    # Budget = actual * factor (some favorable, some adverse)
    bva = []
    for line_item, actual in [
        ("Revenue", ytd_revenue),
        ("COGS", ytd_cogs),
        ("Gross Profit", ytd_gross_profit),
        ("Operating Expenses", ytd_opex),
        ("EBITDA", ytd_ebitda),
    ]:
        # Revenue budget slightly higher (target), COGS budget slightly lower (target)
        if line_item == "Revenue":
            budget = round(actual * random.uniform(0.92, 1.05))
        elif line_item in ("COGS", "Operating Expenses"):
            budget = round(actual * random.uniform(0.95, 1.08))
        elif line_item == "Gross Profit":
            budget = round(actual * random.uniform(0.90, 1.06))
        else:  # EBITDA
            budget = round(actual * random.uniform(0.88, 1.10))

        variance = actual - budget
        variance_pct = round(variance / budget * 100, 1) if budget else 0

        bva.append({
            "company": company_name,
            "line_item": line_item,
            "actual": actual,
            "budget": budget,
            "variance": variance,
            "variance_pct": variance_pct,
        })

    return bva


def style_header_row(ws, row_num, col_count):
    """Apply dark header styling to a row."""
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def write_monthly_pl_sheet(wb, companies: list[str]):
    """Add 'Monthly P&L' sheet to the workbook."""
    ws = wb.create_sheet("Monthly P&L")

    # Headers
    headers = [
        "Company Name", "Period", "Revenue", "COGS",
        "Gross Profit", "GP %", "Operating Expenses",
        "EBITDA", "EBITDA %"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    row_num = 2
    for company in companies:
        monthly_data = generate_monthly_pl(company)
        for m in monthly_data:
            ws.cell(row=row_num, column=1, value=m["company"])
            ws.cell(row=row_num, column=2, value=m["period"])
            ws.cell(row=row_num, column=3, value=m["revenue"])
            ws.cell(row=row_num, column=4, value=m["cogs"])
            ws.cell(row=row_num, column=5, value=m["gross_profit"])
            ws.cell(row=row_num, column=6, value=m["gp_pct"] / 100)
            ws.cell(row=row_num, column=7, value=m["opex"])
            ws.cell(row=row_num, column=8, value=m["ebitda"])
            ws.cell(row=row_num, column=9, value=m["ebitda_pct"] / 100)

            # Formatting
            for c in (3, 4, 5, 7, 8):
                ws.cell(row=row_num, column=c).number_format = MONEY_FMT
            for c in (6, 9):
                ws.cell(row=row_num, column=c).number_format = PCT_FMT
            for c in range(1, 10):
                ws.cell(row=row_num, column=c).font = DATA_FONT
                ws.cell(row=row_num, column=c).border = THIN_BORDER

            row_num += 1

    # Column widths
    widths = [35, 12, 18, 18, 18, 10, 20, 18, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    return ws


def write_budget_vs_actual_sheet(wb, companies: list[str]):
    """Add 'Budget vs Actual' sheet to the workbook."""
    ws = wb.create_sheet("Budget vs Actual")

    headers = [
        "Company Name", "Line Item", "Actual (YTD)",
        "Budget (YTD)", "Variance", "Variance %"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    row_num = 2
    for company in companies:
        monthly_data = generate_monthly_pl(company)
        bva_data = generate_budget_vs_actual(company, monthly_data)
        for b in bva_data:
            ws.cell(row=row_num, column=1, value=b["company"])
            ws.cell(row=row_num, column=2, value=b["line_item"])
            ws.cell(row=row_num, column=3, value=b["actual"])
            ws.cell(row=row_num, column=4, value=b["budget"])
            ws.cell(row=row_num, column=5, value=b["variance"])
            ws.cell(row=row_num, column=6, value=b["variance_pct"] / 100)

            for c in (3, 4, 5):
                ws.cell(row=row_num, column=c).number_format = MONEY_FMT
            ws.cell(row=row_num, column=6).number_format = PCT_FMT
            for c in range(1, 7):
                ws.cell(row=row_num, column=c).font = DATA_FONT
                ws.cell(row=row_num, column=c).border = THIN_BORDER

            row_num += 1

    widths = [35, 22, 18, 18, 18, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    return ws


def get_companies_from_hierarchy(wb) -> list[str]:
    """Extract company names from the Portfolio Hierarchy sheet."""
    if "Portfolio Hierarchy" not in wb.sheetnames:
        return []

    ws = wb["Portfolio Hierarchy"]
    companies = []
    for r in range(1, ws.max_row + 1):
        level = ws.cell(r, 1).value
        if level == "Company":
            name = ws.cell(r, 3).value
            if name:
                companies.append(name)
    return companies


def process_file(filepath: str):
    """Add MIS sheets to a single Excel file."""
    fname = os.path.basename(filepath)
    print(f"\nProcessing: {fname}")

    wb = openpyxl.load_workbook(filepath)

    # Remove existing MIS sheets if re-running
    for sheet_name in ["Monthly P&L", "Budget vs Actual"]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            print(f"  Removed existing '{sheet_name}' sheet")

    companies = get_companies_from_hierarchy(wb)
    if not companies:
        print(f"  WARNING: No companies found in hierarchy, skipping.")
        wb.close()
        return

    print(f"  Companies: {companies}")

    # Use a fixed seed per file for reproducibility
    seed = hash(fname) % 2**32
    random.seed(seed)

    write_monthly_pl_sheet(wb, companies)
    print(f"  Added 'Monthly P&L' sheet ({len(companies)} companies × {len(ALL_MONTHS)} months = {len(companies) * len(ALL_MONTHS)} rows)")

    # Re-seed so BVA data matches monthly data
    random.seed(seed)
    write_budget_vs_actual_sheet(wb, companies)
    print(f"  Added 'Budget vs Actual' sheet ({len(companies)} companies × 5 line items = {len(companies) * 5} rows)")

    wb.save(filepath)
    print(f"  Saved: {filepath}")
    wb.close()


def main():
    files = sorted([
        f for f in os.listdir(MOCK_DIR)
        if f.endswith('.xlsx') and not f.startswith('~') and f != 'Archive.zip'
    ])

    print(f"Found {len(files)} Excel files in {MOCK_DIR}")
    print(f"Months covered: {ALL_MONTHS[0]} to {ALL_MONTHS[-1]} ({len(ALL_MONTHS)} months)")

    for fname in files:
        filepath = os.path.join(MOCK_DIR, fname)
        process_file(filepath)

    print("\n" + "=" * 60)
    print("DONE — All files updated with Monthly P&L and Budget vs Actual sheets.")
    print("=" * 60)


if __name__ == "__main__":
    main()
